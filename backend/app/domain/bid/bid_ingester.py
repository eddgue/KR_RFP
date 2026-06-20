"""Ingest a returned bid template back to `bid.bid_line` (+ components) rows (D20/D21 ingest-end).

This is the INGEST end of the round-trip. It reads OUR owned template (`template_schema.py`) —
NOT arbitrary legacy layouts (D20: "no universal-format guessing for the live product").

D21 — KEY-VALIDATED load, never a name resolve (for OUR template):
  Each generated `Bids` row embeds the system-owned surrogate KEY IDs (cycle/round/tf/lot/item/
  dc/supplier). `ingest_template(data, scope)` READS those embedded keys and VALIDATES each row's
  full key tuple against the cycle scope's known key set (`CycleScope.key_set()`):
    * exact match  -> accept; the `ParsedBidLine` carries those keys DIRECTLY (no resolution).
    * blank key    -> quarantine MISSING_KEY (the locked identity cell was cleared).
    * unknown/tampered key -> quarantine UNKNOWN_KEY / KEY_MISMATCH (NEVER fall back to guessing
      an identity from the display names).
  The human names are NOT the join key — at most a cross-check that emits a WARNING (the row still
  loads on its keys), never a re-resolve.

The legacy name-resolution path (`IdentityResolver`, `ingest_template_resolved`) is retained ONLY
for the migration bridge in `legacy_adapter.py`, where the inputs predate embedded keys. The live
product never name-resolves. See the LEGACY-ONLY banner below `ingest_template`.

For every accepted row, regardless of path, it then:
  * Constructs the per-line cost via the engine's §7 rule (All-In primary; fallback =
    FOB + Delivery + VegCool - Lot Discount) WITH the double-subtract guard.
  * Flags completeness: `bid` (priced), `no_bid` (all price cells blank), `incomplete` (partial).

Output is an in-memory `IngestResult` of parsed lines + quarantined rows (+ name-mismatch
warnings); persisting to the DB (`bid.bid_line` / the component columns added by migration 0007)
is the caller's unit of work. The parsed line carries exactly the columns the engine's cost
construction needs, so a generated template round-trips to the same `bid_line` grain — and now the
same KEY IDs — it was built from (the D20/D21 proof).
"""

from __future__ import annotations

from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from enum import StrEnum
from io import BytesIO
from typing import Protocol

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.domain.bid.template_schema import (
    HEADER_ROW,
    KEY_ID_COLUMNS,
    SHEET_BIDS,
    BidColumn,
    CycleScope,
    ScopeRow,
)

# The full embedded key tuple (KEY_ID_COLUMNS order): cycle, round, tf, lot, item, dc, supplier.
KeyGrain = tuple[str, str, str, str, str, str, str]


class Completeness(StrEnum):
    """Per-line completeness flag (CYCLE_FIELDTOMATO_STRUCTURE.md §4 No-Bid handling)."""

    BID = "bid"  # a usable price is present (All-In, or enough components for the fallback)
    NO_BID = "no_bid"  # every price/volume cell blank — a declined cell, NOT a zero price
    INCOMPLETE = "incomplete"  # partial: some price intent but not enough to construct a price


class QuarantineReason(StrEnum):
    """Why a row could not become a valid, costed bid line (we quarantine, never guess)."""

    # --- D21 key-validation reasons (OUR template — key-validated load). ---
    MISSING_KEY = "missing_key"  # a locked KEY-ID cell was blank/cleared
    UNKNOWN_KEY = "unknown_key"  # the embedded key tuple is not in the cycle scope's key set
    KEY_MISMATCH = "key_mismatch"  # alias of UNKNOWN_KEY — a tampered/foreign embedded key
    # --- Legacy name-resolution reasons (LEGACY ONLY — used by legacy_adapter). ---
    UNRESOLVED_SUPPLIER = "unresolved_supplier"
    UNRESOLVED_DC = "unresolved_dc"
    UNRESOLVED_LOT = "unresolved_lot"
    UNRESOLVED_ITEM = "unresolved_item"
    UNRESOLVED_TF = "unresolved_tf"
    MISSING_IDENTITY = "missing_identity"  # a required scope cell was blank
    # --- Shared (both paths). ---
    DOUBLE_SUBTRACT = "double_subtract"  # All-In AND a discount both populated (§7 footgun)
    BAD_NUMERIC = "bad_numeric"  # a price cell held a non-numeric value


@dataclass(frozen=True)
class ResolvedIdentity:
    """The system-owned key IDs a parsed row carries.

    For OUR template (D21) these are the embedded keys, VALIDATED against the cycle scope (not
    resolved). For the legacy path they are the ids the name resolver mapped the labels to. The
    `cycle_id`/`round_id`/`tf_id` fields default to "" so the legacy path (which has no embedded
    cycle/round/tf key, only resolvable codes) is unaffected; `tf_code` remains the engine's TF
    period token used everywhere downstream.
    """

    supplier_id: str
    dc_id: str
    lot_id: str
    item_id: str
    tf_code: str
    cycle_id: str = ""
    round_id: str = ""
    tf_id: str = ""


class IdentityResolver(Protocol):
    """Resolves the human labels in a template back to canonical store ids.

    The real implementation queries `ref.supplier_alias` / `ref.dc_alias` / `ref.item_alias`
    (KEEP #6, the alias layer) and the cycle's `cyc.cycle_lot` / `cyc.cycle_timeframe` scope.
    Each method returns the canonical id, or None when the label cannot be resolved (-> quarantine,
    never a guess).
    """

    def resolve_supplier(self, label: str) -> str | None: ...
    def resolve_dc(self, label: str) -> str | None: ...
    def resolve_lot(self, label: str) -> str | None: ...
    def resolve_item(self, label: str) -> str | None: ...
    def resolve_tf(self, code: str) -> str | None: ...


@dataclass
class StubIdentityResolver:
    """A simple dict-backed resolver (stub) — maps normalized labels to canonical ids.

    Stands in for the ref/alias layer for the prototype and tests. Normalization is the same
    case/space-folding the alias layer applies (`normalized_alias_text`), so an exact or
    alias-style label resolves and an unknown one returns None (-> quarantine).
    """

    suppliers: dict[str, str] = field(default_factory=dict)
    dcs: dict[str, str] = field(default_factory=dict)
    lots: dict[str, str] = field(default_factory=dict)
    items: dict[str, str] = field(default_factory=dict)
    tfs: dict[str, str] = field(default_factory=dict)

    @staticmethod
    def _norm(text: str) -> str:
        return " ".join(text.strip().lower().split())

    def resolve_supplier(self, label: str) -> str | None:
        return self.suppliers.get(self._norm(label))

    def resolve_dc(self, label: str) -> str | None:
        return self.dcs.get(self._norm(label))

    def resolve_lot(self, label: str) -> str | None:
        return self.lots.get(self._norm(label))

    def resolve_item(self, label: str) -> str | None:
        return self.items.get(self._norm(label))

    def resolve_tf(self, code: str) -> str | None:
        return self.tfs.get(self._norm(code))


@dataclass(frozen=True)
class ParsedComponents:
    """The cost components parsed from one bid row (maps 1:1 to the engine's BidComponents / §7)."""

    all_in: Decimal | None
    fob: Decimal | None
    delivery_surcharge: Decimal
    vegcool_surcharge: Decimal
    lot_discount: Decimal


@dataclass(frozen=True)
class ParsedBidLine:
    """One resolved, costed bid line — the round-trip target (maps onto bid.bid_line + components).

    `landed_cost_per_case` is the §7-constructed price (the engine's Price); `components` are the
    raw parts persisted to the bid_line component columns (migration 0007). `price_basis` records
    which §7 branch produced the price (ALL_IN vs COMPONENT_FALLBACK), so the store is honest about
    provenance and the no-double-discount CHECK has a basis to key on.
    """

    round_code: str
    bid_type: str
    identity: ResolvedIdentity
    components: ParsedComponents
    landed_cost_per_case: Decimal | None
    price_basis: str
    weekly_vol_offered: Decimal | None
    total_vol_offered: Decimal | None
    invested_r1: bool | None
    pricing_comments: str | None
    transit_days: int | None  # supplier-stated lane transit (origin→DC); None when not provided
    completeness: Completeness
    source_row_number: int


@dataclass(frozen=True)
class QuarantinedRow:
    """A row that could not be ingested — kept verbatim with its reason (we do not guess)."""

    source_row_number: int
    reason: QuarantineReason
    detail: str
    raw: dict[str, str]


@dataclass(frozen=True)
class NameMismatchWarning:
    """A display-name that disagrees with the validated keyed identity (D21).

    The row STILL loads on its embedded keys — the name is an attribute, not the join key. We only
    WARN (e.g. someone hand-edited a display name); we never re-resolve the identity from the name.
    """

    source_row_number: int
    column: str
    expected_name: str
    found_name: str


@dataclass
class IngestResult:
    """The outcome of ingesting one returned template."""

    lines: list[ParsedBidLine] = field(default_factory=list)
    quarantined: list[QuarantinedRow] = field(default_factory=list)
    name_warnings: list[NameMismatchWarning] = field(default_factory=list)

    @property
    def bid_count(self) -> int:
        return sum(1 for line in self.lines if line.completeness is Completeness.BID)

    @property
    def no_bid_count(self) -> int:
        return sum(1 for line in self.lines if line.completeness is Completeness.NO_BID)

    @property
    def incomplete_count(self) -> int:
        return sum(1 for line in self.lines if line.completeness is Completeness.INCOMPLETE)


# Price basis labels (mirror the bid_line.price_basis vocabulary).
PRICE_BASIS_ALL_IN = "ALL_IN"
PRICE_BASIS_FALLBACK = "COMPONENT_FALLBACK"


def _to_decimal(value: object) -> Decimal | None:
    """Coerce a cell to Decimal; None/blank -> None; non-numeric -> raises ValueError."""

    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None
        try:
            return Decimal(stripped)
        except InvalidOperation as exc:
            raise ValueError(f"non-numeric price cell: {value!r}") from exc
    if isinstance(value, bool):  # guard: bool is an int subclass, never a price
        raise ValueError(f"non-numeric price cell: {value!r}")
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    raise ValueError(f"non-numeric price cell: {value!r}")


def _to_bool(value: object) -> bool | None:
    if value is None:
        return None
    text = str(value).strip().lower()
    if text in {"", "-"}:
        return None
    return text in {"y", "yes", "true", "1", "invested"}


def _header_index(ws: Worksheet) -> dict[str, int]:
    """Map header text -> 1-based column index from the template's header row (by NAME, §2 rule)."""

    index: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=HEADER_ROW, column=col).value
        if cell is not None:
            index[str(cell).strip()] = col
    return index


def construct_price(components: ParsedComponents) -> tuple[Decimal | None, str | None, str | None]:
    """Apply the engine's §7 cost construction + double-subtract guard.

    Returns (price, basis, error). Price = All-In if present; else
    FOB + Delivery + VegCool - Lot Discount. The guard: if All-In is present AND a discount is
    also populated, that is the double-subtract footgun (V3 §7) — we do NOT recompute and do NOT
    silently drop the discount; we surface it for quarantine (the store's `no_double_discount`
    CHECK is the persistence-side backstop).
    """

    if components.all_in is not None:
        # All-In is taken verbatim (already net of discounts). A populated Lot Discount alongside
        # it is the ambiguous double-subtract case — block it rather than guess which is canonical.
        if components.lot_discount != Decimal("0"):
            return None, None, "All-In present together with a Lot Discount (double-subtract risk)"
        return components.all_in, PRICE_BASIS_ALL_IN, None

    if components.fob is not None:
        price = (
            components.fob
            + components.delivery_surcharge
            + components.vegcool_surcharge
            - components.lot_discount
        )
        return price, PRICE_BASIS_FALLBACK, None

    # No All-In and no FOB -> nothing to construct (a no-bid or incomplete row).
    return None, None, None


def _classify(
    components: ParsedComponents,
    price: Decimal | None,
    weekly: Decimal | None,
    total: Decimal | None,
) -> Completeness:
    """bid / no_bid / incomplete classification (§4 No-Bid handling)."""

    zero = Decimal("0")
    has_any_price_intent = (
        components.all_in is not None
        or components.fob is not None
        or components.delivery_surcharge != zero
        or components.vegcool_surcharge != zero
        or components.lot_discount != zero
    )
    has_vol = weekly is not None or total is not None
    if price is not None and price > 0:
        return Completeness.BID
    if not has_any_price_intent and not has_vol:
        return Completeness.NO_BID
    return Completeness.INCOMPLETE


def _iter_body_rows(data: bytes) -> list[tuple[int, dict[str, str]]]:
    """Load the `Bids` sheet and return [(row_number, raw-dict)] for each non-blank body row."""

    wb = load_workbook(BytesIO(data), data_only=True, read_only=True)
    if SHEET_BIDS not in wb.sheetnames:
        wb.close()
        raise ValueError(f"template missing required sheet {SHEET_BIDS!r}")
    ws = wb[SHEET_BIDS]
    headers = _header_index(ws)

    rows: list[tuple[int, dict[str, str]]] = []
    for row_number, row_cells in enumerate(
        ws.iter_rows(min_row=HEADER_ROW + 1, values_only=False), start=HEADER_ROW + 1
    ):
        raw = _row_to_dict(row_cells, headers)
        if not any(raw.values()):
            continue  # skip fully blank rows
        rows.append((row_number, raw))
    wb.close()
    return rows


def ingest_template(data: bytes, scope: CycleScope) -> IngestResult:
    """Ingest OUR returned template — KEY-VALIDATED, never name-resolved (D21).

    For each row it reads the EMBEDDED key IDs and validates the full key tuple against
    `scope.key_set()`. Exact match -> the line carries those keys directly. Missing key ->
    MISSING_KEY. Unknown/tampered key -> UNKNOWN_KEY (alias KEY_MISMATCH). We NEVER fall back to
    resolving identity from the display names; a name that disagrees with the keyed identity is a
    WARNING (`result.name_warnings`), not a re-resolve.
    """

    rows = _iter_body_rows(data)
    valid_keys = scope.key_set()
    # Map each key tuple back to its scope row so we can cross-check the display names (warn-only).
    by_key = {row.key_grain(scope.cycle_id): row for row in scope.rows}

    result = IngestResult()
    for row_number, raw in rows:
        outcome = _parse_keyed_row(raw, row_number, scope.cycle_id, valid_keys, by_key)
        if isinstance(outcome, QuarantinedRow):
            result.quarantined.append(outcome)
            continue
        line, warnings = outcome
        result.lines.append(line)
        result.name_warnings.extend(warnings)
    return result


def ingest_template_resolved(data: bytes, resolver: IdentityResolver) -> IngestResult:
    """LEGACY-ONLY: ingest a template whose identity must be NAME-RESOLVED (no embedded keys).

    Retained solely for the migration bridge (`legacy_adapter.ingest_legacy`). Legacy inputs
    predate D21's embedded keys, so the only available identity is the human labels, resolved via
    the ref/alias layer (`IdentityResolver`). The LIVE product uses `ingest_template(data, scope)`
    and never reaches this path. Do not call this for OUR generated template.
    """

    rows = _iter_body_rows(data)
    result = IngestResult()
    for row_number, raw in rows:
        parsed = _parse_resolved_row(raw, row_number, resolver)
        if isinstance(parsed, QuarantinedRow):
            result.quarantined.append(parsed)
        else:
            result.lines.append(parsed)
    return result


def _row_to_dict(row_cells: object, headers: dict[str, int]) -> dict[str, str]:
    """Build {header: stringified value} for a row, keyed by the template's header names."""

    # row_cells is a tuple of cells; index by (col - 1).
    cells = list(row_cells)  # type: ignore[call-overload]
    out: dict[str, str] = {}
    for header, col_index in headers.items():
        idx = col_index - 1
        value = cells[idx].value if 0 <= idx < len(cells) else None
        out[header] = "" if value is None else str(value).strip()
    return out


def _cell(raw: dict[str, str], column: BidColumn) -> str:
    return raw.get(column.value, "").strip()


def _parse_pricing_and_build(
    raw: dict[str, str], row_number: int, identity: ResolvedIdentity
) -> ParsedBidLine | QuarantinedRow:
    """Shared tail: construct price (§7 + guard), classify completeness, build the line.

    Identity is already established (validated keys for OUR template, or resolved labels for the
    legacy path) — this step is identical for both and never touches identity again.
    """

    try:
        components = ParsedComponents(
            all_in=_to_decimal(raw.get(BidColumn.ALL_IN.value)),
            fob=_to_decimal(raw.get(BidColumn.FOB.value)),
            delivery_surcharge=_to_decimal(raw.get(BidColumn.DELIVERY_SURCHARGE.value))
            or Decimal("0"),
            vegcool_surcharge=_to_decimal(raw.get(BidColumn.VEGCOOL_SURCHARGE.value))
            or Decimal("0"),
            lot_discount=_to_decimal(raw.get(BidColumn.LOT_DISCOUNT.value)) or Decimal("0"),
        )
        weekly = _to_decimal(raw.get(BidColumn.WEEKLY_VOL_OFFERED.value))
        total = _to_decimal(raw.get(BidColumn.TOTAL_VOL_OFFERED.value))
        transit = _to_decimal(raw.get(BidColumn.TRANSIT_DAYS.value))
    except ValueError as exc:
        return QuarantinedRow(row_number, QuarantineReason.BAD_NUMERIC, str(exc), raw)

    price, basis, price_error = construct_price(components)
    if price_error is not None:
        return QuarantinedRow(row_number, QuarantineReason.DOUBLE_SUBTRACT, price_error, raw)

    completeness = _classify(components, price, weekly, total)
    comments = _cell(raw, BidColumn.PRICING_COMMENTS) or None
    return ParsedBidLine(
        round_code=_cell(raw, BidColumn.ROUND),
        bid_type=_cell(raw, BidColumn.BID_TYPE),
        identity=identity,
        components=components,
        landed_cost_per_case=price,
        price_basis=basis or "",
        weekly_vol_offered=weekly,
        total_vol_offered=total,
        invested_r1=_to_bool(raw.get(BidColumn.INVESTED_R1.value)),
        pricing_comments=comments,
        transit_days=int(transit) if transit is not None else None,
        completeness=completeness,
        source_row_number=row_number,
    )


# The display-name columns cross-checked (warn-only) against the keyed identity (D21).
_KEY_TO_NAME_COLUMN: tuple[tuple[BidColumn, BidColumn], ...] = (
    (BidColumn.SUPPLIER_ID, BidColumn.SUPPLIER),
    (BidColumn.DC_ID, BidColumn.DC),
    (BidColumn.LOT_ID, BidColumn.LOT),
    (BidColumn.ITEM_ID, BidColumn.ITEM),
)


def _parse_keyed_row(
    raw: dict[str, str],
    row_number: int,
    cycle_id: str,
    valid_keys: frozenset[KeyGrain],
    by_key: dict[KeyGrain, ScopeRow],
) -> tuple[ParsedBidLine, list[NameMismatchWarning]] | QuarantinedRow:
    """D21 key-validated parse for OUR template: validate embedded keys, never resolve names."""

    # Read the embedded KEY-ID cells. The tuple order MUST match KEY_ID_COLUMNS / key_grain().
    keys: KeyGrain = tuple(_cell(raw, col) for col in KEY_ID_COLUMNS)  # type: ignore[assignment]
    # A blank/cleared locked key cell -> MISSING_KEY (the identity cell was tampered/emptied).
    for col, value in zip(KEY_ID_COLUMNS, keys, strict=True):
        if not value:
            return QuarantinedRow(
                row_number, QuarantineReason.MISSING_KEY, f"blank {col.value}", raw
            )

    # The embedded cycle key must be THIS cycle, and the full key tuple must be a known scope cell.
    embedded_cycle = _cell(raw, BidColumn.CYCLE_ID)
    if embedded_cycle != cycle_id or keys not in valid_keys:
        # Unknown / tampered / foreign key -> quarantine, NEVER guess from the display names.
        return QuarantinedRow(
            row_number,
            QuarantineReason.UNKNOWN_KEY,
            f"embedded keys not in cycle scope: {keys}",
            raw,
        )

    # Accepted on keys. The keys carry the identity DIRECTLY (no resolution).
    cycle_k, round_k, tf_k, lot_k, item_k, dc_k, supplier_k = keys
    identity = ResolvedIdentity(
        supplier_id=supplier_k,
        dc_id=dc_k,
        lot_id=lot_k,
        item_id=item_k,
        tf_code=_cell(raw, BidColumn.TF),  # the engine's TF period token (display TF code)
        cycle_id=cycle_k,
        round_id=round_k,
        tf_id=tf_k,
    )

    # Cross-check the display names against the scope (WARN-only — names are attributes, D21).
    warnings = _name_cross_check(raw, row_number, by_key.get(keys))

    built = _parse_pricing_and_build(raw, row_number, identity)
    if isinstance(built, QuarantinedRow):
        return built
    return built, warnings


def _name_cross_check(
    raw: dict[str, str], row_number: int, scope_row: ScopeRow | None
) -> list[NameMismatchWarning]:
    """Warn (never re-resolve) when a row's display name disagrees with the keyed scope identity."""

    if scope_row is None:
        return []
    warnings: list[NameMismatchWarning] = []
    expected = {
        BidColumn.SUPPLIER: scope_row.supplier_label,
        BidColumn.DC: scope_row.dc_label,
        BidColumn.LOT: scope_row.lot_label,
        BidColumn.ITEM: scope_row.item_label,
    }
    for _key_col, name_col in _KEY_TO_NAME_COLUMN:
        found = _cell(raw, name_col)
        want = expected[name_col]
        if found and want and found != want:
            warnings.append(
                NameMismatchWarning(
                    row_number, name_col.value, expected_name=want, found_name=found
                )
            )
    return warnings


def _parse_resolved_row(
    raw: dict[str, str], row_number: int, resolver: IdentityResolver
) -> ParsedBidLine | QuarantinedRow:
    """LEGACY-ONLY name resolution: map labels -> ids via the alias layer, or quarantine.

    Used solely by the migration bridge (legacy inputs have no embedded keys). The live product
    never reaches here — OUR template is key-validated in `_parse_keyed_row`.
    """

    supplier_label = _cell(raw, BidColumn.SUPPLIER)
    dc_label = _cell(raw, BidColumn.DC)
    lot_label = _cell(raw, BidColumn.LOT)
    item_label = _cell(raw, BidColumn.ITEM)
    tf_code = _cell(raw, BidColumn.TF)

    # A required scope cell missing -> cannot place the row in the grain.
    for label in (supplier_label, dc_label, lot_label, item_label, tf_code):
        if not label:
            return QuarantinedRow(
                row_number, QuarantineReason.MISSING_IDENTITY, "blank scope cell", raw
            )

    supplier_id = resolver.resolve_supplier(supplier_label)
    if supplier_id is None:
        return QuarantinedRow(
            row_number, QuarantineReason.UNRESOLVED_SUPPLIER, supplier_label, raw
        )
    dc_id = resolver.resolve_dc(dc_label)
    if dc_id is None:
        return QuarantinedRow(row_number, QuarantineReason.UNRESOLVED_DC, dc_label, raw)
    lot_id = resolver.resolve_lot(lot_label)
    if lot_id is None:
        return QuarantinedRow(row_number, QuarantineReason.UNRESOLVED_LOT, lot_label, raw)
    item_id = resolver.resolve_item(item_label)
    if item_id is None:
        return QuarantinedRow(row_number, QuarantineReason.UNRESOLVED_ITEM, item_label, raw)
    resolved_tf = resolver.resolve_tf(tf_code)
    if resolved_tf is None:
        return QuarantinedRow(row_number, QuarantineReason.UNRESOLVED_TF, tf_code, raw)

    identity = ResolvedIdentity(
        supplier_id=supplier_id, dc_id=dc_id, lot_id=lot_id, item_id=item_id, tf_code=resolved_tf
    )
    return _parse_pricing_and_build(raw, row_number, identity)
