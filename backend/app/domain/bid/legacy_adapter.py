"""Adapt a SYNTHETIC legacy-shaped workbook into our owned template, then ingest (resilience).

D20: our owned template is the LIVE contract; the messy reference formats (.xlsb, 14-tab legacy)
are TEST/REFERENCE inputs only — proof of migration resilience, not the live path. This adapter
is the migration bridge: it reads a *different-shaped* (synthetic, our own — NOT the quarantined
real files) workbook, maps its columns onto our owned `Bids` grain, and hands the normalized bytes
to `bid_ingester.ingest_template`. Anything it cannot map flows through to the ingester's
quarantine (we never guess).

This deliberately mirrors the §2/§4 reality of the reference corpus WITHOUT importing it: a header
row that is NOT at our row 2, supplier/DC/item down the rows, a primary block of cost sub-fields,
and `No Bid` cells. It proves the same parsed grain results from a foreign layout, or quarantines
cleanly. The ingestion logic itself is unchanged — only the *shape* is bridged.
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook, load_workbook

from app.domain.bid.bid_ingester import IdentityResolver, IngestResult, ingest_template
from app.domain.bid.template_schema import (
    BID_HEADERS,
    BODY_START_ROW,
    HEADER_ROW,
    SHEET_BIDS,
    TITLE_ROW,
    BidColumn,
)

# Mapping from a legacy header label -> our owned BidColumn. The legacy fixture uses the
# reference's wording (FOB $/Case (Corrugate), Dlvd Cost, Comments, ...); we map the subset that
# carries the IN_Bids contract and let the rest fall away (the trailing-width inflation, §2).
LEGACY_TO_OWNED: dict[str, BidColumn] = {
    "Round ID": BidColumn.ROUND,
    "Bid Type": BidColumn.BID_TYPE,
    "Supplier": BidColumn.SUPPLIER,
    "DC Name": BidColumn.DC,
    "Lot_ID": BidColumn.LOT,
    "Product Description": BidColumn.ITEM,
    "TF": BidColumn.TF,
    "All-In $/Case": BidColumn.ALL_IN,
    "FOB $/Case (Corrugate)": BidColumn.FOB,
    "Delivery Surcharge": BidColumn.DELIVERY_SURCHARGE,
    "VegCool Surcharge": BidColumn.VEGCOOL_SURCHARGE,
    "Lot Discount": BidColumn.LOT_DISCOUNT,
    "Comments": BidColumn.PRICING_COMMENTS,
    "Weekly Vol Cap": BidColumn.WEEKLY_VOL_OFFERED,
    "Total Vol Cap": BidColumn.TOTAL_VOL_OFFERED,
}

# The token a legacy sheet uses to decline a cell (§4) — normalized to blank price cells.
LEGACY_NO_BID = "No Bid"


def adapt_legacy_to_owned(
    legacy_bytes: bytes,
    *,
    sheet_name: str,
    legacy_header_row: int,
) -> bytes:
    """Rewrite a legacy-shaped sheet into our owned single-sheet template bytes.

    `legacy_header_row` is the 1-based row where the legacy headers live (NOT trusted to be our
    row 2 — the reference puts them at row 4/17). We read by header NAME there, then re-emit on
    our owned grain. `No Bid` cells become blank price cells (the ingester then flags `no_bid`).
    """

    src = load_workbook(BytesIO(legacy_bytes), data_only=True, read_only=True)
    ws = src[sheet_name]

    # Build the legacy header -> column index map at the legacy header row.
    legacy_headers: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=legacy_header_row, column=col).value
        if value is not None:
            legacy_headers[str(value).strip()] = col

    out = Workbook()
    dest = out.active
    assert dest is not None  # noqa: S101
    dest.title = SHEET_BIDS
    dest.cell(row=TITLE_ROW, column=1, value="Bids (migrated from legacy)")
    owned_col_index = {header: i for i, header in enumerate(BID_HEADERS, start=1)}
    for i, header in enumerate(BID_HEADERS, start=1):
        dest.cell(row=HEADER_ROW, column=i, value=header)

    dest_row = BODY_START_ROW
    for row in range(legacy_header_row + 1, ws.max_row + 1):
        wrote_any = False
        for legacy_header, legacy_col in legacy_headers.items():
            owned = LEGACY_TO_OWNED.get(legacy_header)
            if owned is None:
                continue  # an unmapped legacy column (trailing inflation) — dropped, not guessed
            raw_value = ws.cell(row=row, column=legacy_col).value
            value = "" if raw_value is None else str(raw_value).strip()
            if value == LEGACY_NO_BID:
                value = ""  # decline -> blank price cell (-> ingester flags no_bid)
            if value:
                dest.cell(row=dest_row, column=owned_col_index[owned.value], value=value)
                wrote_any = True
        if wrote_any:
            dest_row += 1
    src.close()

    buffer = BytesIO()
    out.save(buffer)
    return buffer.getvalue()


def ingest_legacy(
    legacy_bytes: bytes,
    resolver: IdentityResolver,
    *,
    sheet_name: str,
    legacy_header_row: int,
) -> IngestResult:
    """Adapt a synthetic legacy workbook to our owned template, then ingest it normally."""

    owned_bytes = adapt_legacy_to_owned(
        legacy_bytes, sheet_name=sheet_name, legacy_header_row=legacy_header_row
    )
    return ingest_template(owned_bytes, resolver)
