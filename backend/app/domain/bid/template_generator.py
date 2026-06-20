"""Generate the owned bid-template xlsx from a cycle scope (D20 generate-end, D17 our design).

Given a `CycleScope` (synthetic for now — in production it is read from `cyc.*`: the lots, DCs,
items, TFs, and rounds in scope), produce a multi-sheet workbook:
  * `Instructions` — cycle identity + window + the supplier rules + the template-version token.
  * `Bids`         — one pre-filled scope row per cell, with blank price/volume cells.
  * `Capacity`     — one pre-filled row per supplier x DC x item x TF, blank capacity cells.

This is the GENERATE end of the round-trip; `bid_ingester.py` is the INGEST end. Both share
`template_schema.py`, so they cannot drift (D20). File IO via openpyxl is fine here — this is a
service, not the pure engine.
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.domain.bid.template_preset import FULL_PRESET, BidTemplatePreset
from app.domain.bid.template_schema import (
    BID_STATUS_HEADER,
    BODY_START_ROW,
    CAPACITY_ENTRY_COLUMNS,
    CAPACITY_HEADERS,
    HEADER_ROW,
    KEY_ID_COLUMNS,
    SHEET_BIDS,
    SHEET_CAPACITY,
    SHEET_INSTRUCTIONS,
    TEMPLATE_PROTECT_PASSWORD,
    TITLE_ROW,
    BidColumn,
    CapacityColumn,
    CycleScope,
    ScopeRow,
)

# Form styling: entry cells (the only editable ones) get a pale "type here" fill; the readiness
# traffic light is coloured by state. Excel applies the cell-locking only once the sheet is
# protected — openpyxl writes through locks freely, so the automated fill/ingest path is unaffected.
_ENTRY_FILL = PatternFill("solid", fgColor="FFF7DC")  # pale yellow — "enter here"
_UNLOCKED = Protection(locked=False)
_STATUS_RULES = (
    ("Complete", "C6EFCE", "006100"),
    ("Incomplete", "FFEB9C", "9C6500"),
    ("Not bid", "E7E6E6", "808080"),
)


def _protect_form(ws: Worksheet) -> None:
    """Make the sheet a true form: protect it (password) so only unlocked cells stay editable."""

    ws.protection.sheet = True
    ws.protection.password = TEMPLATE_PROTECT_PASSWORD
    ws.protection.selectLockedCells = True  # readable
    ws.protection.selectUnlockedCells = True  # the entry cells are tab-to-able
    ws.protection.formatCells = False


def _unlock_entry_cells(
    ws: Worksheet, header_index: dict[str, int], entry_headers: tuple[str, ...], n_rows: int
) -> None:
    """Unlock + highlight the supplier-owned entry cells (the only editable cells on the form)."""

    cols = [header_index[h] for h in entry_headers]
    for r in range(BODY_START_ROW, BODY_START_ROW + n_rows):
        for c in cols:
            cell = ws.cell(row=r, column=c)
            cell.protection = _UNLOCKED
            cell.fill = _ENTRY_FILL


def _hide_columns(ws: Worksheet, header_index: dict[str, int], headers: tuple[str, ...]) -> None:
    """Hide the given columns (the raw key IDs) so the supplier sees names, not UUIDs (D23)."""

    for h in headers:
        ws.column_dimensions[get_column_letter(header_index[h])].hidden = True


# The supplier rules written onto the Instructions sheet (label, value) — our wording, D17.
_INSTRUCTION_RULES: tuple[tuple[str, str], ...] = (
    ("Template version", ""),  # filled from scope at build time
    ("Cycle", ""),
    ("Window", ""),
    (
        "Pricing",
        "Enter EITHER an All-In $/case (already net of all discounts) OR the components "
        "(FOB + Delivery Surcharge + VegCool Surcharge - Lot Discount). Do not enter both an "
        "All-In and a Lot Discount on the same row.",
    ),
    (
        "No Bid",
        "Leave ALL price cells blank on a row you decline. Do not enter 0 for a no-bid.",
    ),
    (
        "Volume",
        "Enter Weekly and/or Total Vol Offered in cases for rows you bid.",
    ),
    (
        "Grain",
        "One row per DC x Lot x Item x TF. Prices are per period (TF) — fill each row you serve.",
    ),
)


def _write_headers(ws: Worksheet, title: str, headers: tuple[str, ...]) -> None:
    """Write the row-1 title band and the row-2 header row."""

    ws.cell(row=TITLE_ROW, column=1, value=title)
    for col_index, header in enumerate(headers, start=1):
        ws.cell(row=HEADER_ROW, column=col_index, value=header)


def _build_instructions(ws: Worksheet, scope: CycleScope) -> None:
    ws.cell(row=1, column=1, value="Kroger Produce RFP — Bid Template")
    rows: list[tuple[str, str]] = []
    for label, value in _INSTRUCTION_RULES:
        if label == "Template version":
            value = scope.template_version
        elif label == "Cycle":
            value = f"{scope.cycle_code} — {scope.cycle_name}"
        elif label == "Window":
            value = scope.window_label
        rows.append((label, value))
    for offset, (label, value) in enumerate(rows, start=3):
        ws.cell(row=offset, column=1, value=label)
        ws.cell(row=offset, column=2, value=value)


def _scope_cell_values(row: ScopeRow, cycle_id: str) -> dict[BidColumn, str]:
    """The system-owned scope cells written into a `Bids` row.

    D21: the KEY-ID columns carry the system-owned surrogate UUIDs (the join identity); the display
    columns carry the human names (attributes only). Both are system-owned/locked — a supplier
    never edits them — but only the keys are the identity the ingester validates against the scope.
    """

    return {
        # --- Key IDs (D21) — the join identity (validated, never resolved, on ingest). ---
        BidColumn.CYCLE_ID: cycle_id,
        BidColumn.ROUND_ID: row.round_id,
        BidColumn.TF_ID: row.tf_id,
        BidColumn.LOT_ID: row.lot_id,
        BidColumn.ITEM_ID: row.item_id,
        BidColumn.DC_ID: row.dc_id,
        BidColumn.SUPPLIER_ID: row.supplier_id,
        # --- Display names (attributes only; a mismatch warns, never re-resolves). ---
        BidColumn.ROUND: row.round_code,
        BidColumn.BID_TYPE: row.bid_type,
        BidColumn.SUPPLIER: row.supplier_label,
        BidColumn.DC: row.dc_label,
        BidColumn.LOT: row.lot_label,
        BidColumn.ITEM: row.item_label,
        BidColumn.TF: row.tf_code,
    }


def _add_bid_status_column(
    ws: Worksheet, status_col: int, n_rows: int, preset: BidTemplatePreset
) -> None:
    """Append the per-row readiness traffic light (Not bid / Incomplete / Complete), locked.

    A LOCKED formula column (generator-owned, NOT ingested): Not bid when no price/volume is
    entered; Complete when there is a usable price (All-In or FOB) AND a volume (Weekly or Total);
    otherwise Incomplete. Built from whatever price/volume columns the PRESET carries (the preset
    guarantees at least one usable price and one volume), referenced by their letters in the
    preset's header order. Coloured by state via conditional formatting.
    """

    headers = preset.bid_headers()

    def col(c: BidColumn) -> str:
        return str(get_column_letter(headers.index(c.value) + 1))

    price_components = [
        c
        for c in (
            BidColumn.ALL_IN,
            BidColumn.FOB,
            BidColumn.DELIVERY_SURCHARGE,
            BidColumn.VEGCOOL_SURCHARGE,
            BidColumn.LOT_DISCOUNT,
        )
        if c in preset.entry_columns
    ]
    usable_price = [c for c in (BidColumn.ALL_IN, BidColumn.FOB) if c in preset.entry_columns]
    vols = [
        c
        for c in (BidColumn.WEEKLY_VOL_OFFERED, BidColumn.TOTAL_VOL_OFFERED)
        if c in preset.entry_columns
    ]
    letter = get_column_letter(status_col)
    ws.cell(row=HEADER_ROW, column=status_col, value=BID_STATUS_HEADER)
    for r in range(BODY_START_ROW, BODY_START_ROW + n_rows):
        any_price = "COUNTA(" + ",".join(f"{col(c)}{r}" for c in price_components) + ")>0"
        any_vol = "COUNTA(" + ",".join(f"{col(c)}{r}" for c in vols) + ")>0"
        has_price = "OR(" + ",".join(f'{col(c)}{r}<>""' for c in usable_price) + ")"
        has_vol = "OR(" + ",".join(f'{col(c)}{r}<>""' for c in vols) + ")"
        ws.cell(row=r, column=status_col).value = (
            f'=IF(AND(NOT({any_price}),NOT({any_vol})),"Not bid",'
            f'IF(AND({has_price},{has_vol}),"Complete","Incomplete"))'
        )
    rng = f"{letter}{BODY_START_ROW}:{letter}{BODY_START_ROW + n_rows - 1}"
    for text, fill_hex, font_hex in _STATUS_RULES:
        ws.conditional_formatting.add(
            rng,
            CellIsRule(
                operator="equal",
                formula=[f'"{text}"'],
                fill=PatternFill("solid", fgColor=fill_hex),
                font=Font(color=font_hex),
            ),
        )


def _build_bids(ws: Worksheet, scope: CycleScope, preset: BidTemplatePreset) -> None:
    headers = preset.bid_headers()
    _write_headers(ws, f"Bids — {scope.cycle_name}", headers)
    header_index = {header: i for i, header in enumerate(headers, start=1)}
    for offset, scope_row in enumerate(scope.rows, start=BODY_START_ROW):
        for column, value in _scope_cell_values(scope_row, scope.cycle_id).items():
            ws.cell(row=offset, column=header_index[column.value], value=value)
        # Price/volume cells are intentionally left blank for the supplier to fill.
    n_rows = len(scope.rows)
    # Form treatment: unlock + highlight only this preset's price/volume entry cells; hide the raw
    # key IDs so the supplier reads names not UUIDs; add the traffic light; protect the sheet.
    _unlock_entry_cells(ws, header_index, tuple(c.value for c in preset.entry_columns), n_rows)
    _hide_columns(ws, header_index, tuple(c.value for c in KEY_ID_COLUMNS))
    _add_bid_status_column(ws, len(headers) + 1, n_rows, preset)
    ws.freeze_panes = ws.cell(row=BODY_START_ROW, column=1)
    _protect_form(ws)


def _build_capacity(ws: Worksheet, scope: CycleScope) -> None:
    _write_headers(ws, f"Capacity — {scope.cycle_name}", CAPACITY_HEADERS)
    header_index = {header: i for i, header in enumerate(CAPACITY_HEADERS, start=1)}
    # Capacity grain is supplier x DC x item x TF (round-independent) — dedupe the bid scope.
    seen: set[tuple[str, str, str, str]] = set()
    offset = BODY_START_ROW
    for scope_row in scope.rows:
        key = (
            scope_row.supplier_label,
            scope_row.dc_label,
            scope_row.item_label,
            scope_row.tf_code,
        )
        if key in seen:
            continue
        seen.add(key)
        ws.cell(row=offset, column=header_index[CapacityColumn.SUPPLIER.value], value=key[0])
        ws.cell(row=offset, column=header_index[CapacityColumn.DC.value], value=key[1])
        ws.cell(row=offset, column=header_index[CapacityColumn.ITEM.value], value=key[2])
        ws.cell(row=offset, column=header_index[CapacityColumn.TF.value], value=key[3])
        offset += 1
    # Same form treatment: only the capacity entry cells are editable, the sheet is protected.
    _unlock_entry_cells(
        ws, header_index, tuple(c.value for c in CAPACITY_ENTRY_COLUMNS), offset - BODY_START_ROW
    )
    _protect_form(ws)


def build_template_workbook(scope: CycleScope, preset: BidTemplatePreset = FULL_PRESET) -> Workbook:
    """Build the in-memory multi-sheet template workbook for a cycle scope.

    `preset` selects which supplier-entry columns the Bids sheet carries (default = the full column
    superset, so existing behaviour is unchanged); scope columns are always included.
    """

    wb = Workbook()
    # openpyxl seeds a default sheet; repurpose it as Instructions.
    instructions = wb.active
    assert instructions is not None  # noqa: S101 — Workbook() always has an active sheet
    instructions.title = SHEET_INSTRUCTIONS
    _build_instructions(instructions, scope)
    _protect_form(instructions)  # read-only cover sheet (no entry cells)

    bids = wb.create_sheet(SHEET_BIDS)
    _build_bids(bids, scope, preset)

    capacity = wb.create_sheet(SHEET_CAPACITY)
    _build_capacity(capacity, scope)

    return wb


def generate_template_bytes(scope: CycleScope, preset: BidTemplatePreset = FULL_PRESET) -> bytes:
    """Generate the template and return it as xlsx bytes (the artifact the system releases)."""

    wb = build_template_workbook(scope, preset)
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
