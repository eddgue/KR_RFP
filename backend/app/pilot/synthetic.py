"""Synthetic TOMATO cycle builders — the canonical in-memory setup + bid fixtures.

These build a small, fully-priced synthetic cycle (2 DCs, 2 lots, 2 suppliers, 1 TF, 2 rounds) as
in-memory .xlsx bytes: a filled Setup/Kickoff workbook and a filled owned bid template. They are
PURE (no DB, no pytest, no I/O beyond openpyxl in-memory), so they can be reused anywhere a
synthetic cycle is needed:

  * the end-to-end test (`tests/pilot/test_pilot_cycle_e2e.py`) imports them as its fixtures;
  * the deploy seed (`deploy/gcp/seed.py`) imports them to seed the demo TOMATO cycle.

Keeping them here (not in the test module) means the seed never has to import pytest at runtime —
the one source of truth lives in the app package, and the test just re-exports it.
"""

from __future__ import annotations

from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.domain.bid.template_schema import BODY_START_ROW, HEADER_ROW, SHEET_BIDS, BidColumn
from app.pilot.setup_template import (
    EXAMPLE_START_ROW,
    TAB_CYCLE,
    TAB_DCS,
    TAB_INCUMBENTS,
    TAB_LOTS,
    TAB_SUPPLIERS,
    TAB_TIMEFRAMES,
    TAB_VOLUMES,
    build_setup_workbook,
)
from app.pilot.setup_template import (
    HEADER_ROW as SETUP_HEADER_ROW,
)


# ---------------------------------------------------------------------------
# build a synthetic filled SETUP workbook in-memory
# ---------------------------------------------------------------------------
def _setup_header_col(ws: Worksheet, header: str) -> int:
    for col in range(1, ws.max_column + 1):
        if str(ws.cell(row=SETUP_HEADER_ROW, column=col).value or "").strip() == header:
            return col
    raise AssertionError(f"header {header!r} not found on tab {ws.title!r}")


def _write_setup_rows(ws: Worksheet, rows: list[dict[str, object]]) -> None:
    for ri, row in enumerate(rows):
        excel_row = EXAMPLE_START_ROW + ri
        for header, value in row.items():
            ws.cell(row=excel_row, column=_setup_header_col(ws, header), value=value)
    for extra in range(len(rows), len(rows) + 3):
        excel_row = EXAMPLE_START_ROW + extra
        for col in range(1, ws.max_column + 1):
            ws.cell(row=excel_row, column=col, value=None)


def build_filled_setup(premium_ceiling: float = 0.12, weight_preset: str = "balanced") -> bytes:
    """A synthetic setup: 2 DCs, 2 lots, 2 suppliers, 1 TF, 2 rounds, volumes, incumbents.

    `premium_ceiling` and `weight_preset` are buyer-adjustable strategy knobs written to the Cycle
    tab (defaults match the preset); callers pass distinctive values to prove they flow to the
    engine.
    """

    wb = load_workbook(BytesIO(build_setup_workbook()))
    _write_setup_rows(
        wb[TAB_CYCLE],
        [
            {
                "Cycle Label": "E2E Tomatoes Cycle",
                "Commodity": "Field Tomatoes",
                "Sub-commodity": "Round/Vine",
                "Horizon (weeks)": 13,
                "Rounds": 2,
                "Target Effective Date": "2026-12-31",
                "Weight Preset": weight_preset,
                "Max Suppliers / DC": 2,
                "Premium Ceiling": premium_ceiling,
                "Concentration Threshold": 0.40,
                "Coverage Floor": 0.80,
            }
        ],
    )
    _write_setup_rows(
        wb[TAB_DCS],
        [
            {"DC Name": "Atlanta DC", "Region": "East", "State": "GA"},
            {"DC Name": "Dallas DC", "Region": "South", "State": "TX"},
        ],
    )
    _write_setup_rows(
        wb[TAB_LOTS],
        [
            {
                "Lot Name": "Lot 1 - Grape",
                "Item Description": "Premium Grape Tomato 10oz",
                "Pack Size / UOM": "10oz clamshell",
                "Product Type": "Conventional",
                "Category": "Tomatoes",
            },
            {
                "Lot Name": "Lot 2 - Roma",
                "Item Description": "Roma Tomato Bulk 25lb",
                "Pack Size / UOM": "25lb carton",
                "Product Type": "Organic",
                "Category": "Tomatoes",
            },
        ],
    )
    _write_setup_rows(
        wb[TAB_SUPPLIERS],
        [
            {"Supplier Name": "Green Valley Farms", "Region / Origin": "East", "Notes": "inc"},
            {"Supplier Name": "Sunbelt Produce", "Region / Origin": "South", "Notes": "new"},
        ],
    )
    _write_setup_rows(
        wb[TAB_TIMEFRAMES],
        [
            {
                "Timeframe Label": "Spring 2026",
                "Start Date": "2026-04-01",
                "End Date": "2026-06-30",
                "Week Count": 13,
            }
        ],
    )
    vol_rows: list[dict[str, object]] = [
        {
            "DC Name": dc,
            "Lot Name": lot,
            "Timeframe": "Spring 2026",
            "Method": "WEEKLY_X_WEEKS",
            "Weekly Cases": 400,
            "Weeks": 13,
        }
        for dc in ("Atlanta DC", "Dallas DC")
        for lot in ("Lot 1 - Grape", "Lot 2 - Roma")
    ]
    _write_setup_rows(wb[TAB_VOLUMES], vol_rows)
    inc_rows: list[dict[str, object]] = [
        {
            "DC Name": dc,
            "Lot Name": lot,
            "Incumbent Supplier": "Green Valley Farms",
            "Routing Baseline $/case": 13.50,
            "Contract Notes": "auto",
        }
        for dc in ("Atlanta DC", "Dallas DC")
        for lot in ("Lot 1 - Grape", "Lot 2 - Roma")
    ]
    _write_setup_rows(wb[TAB_INCUMBENTS], inc_rows)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# fill the generated OWNED bid template in-memory with synthetic bids
# ---------------------------------------------------------------------------
def _header_map(ws: Worksheet) -> dict[str, int]:
    out: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=HEADER_ROW, column=col).value
        if value is not None:
            out[str(value).strip()] = col
    return out


def fill_bid_template(template_bytes: bytes) -> bytes:
    """Write synthetic All-In + volume into every scope row so the engine has priced bids.

    Two suppliers, varied so Scenario B has a real split. Green Valley (incumbent) is keener on Lot
    1; Sunbelt is keener on Lot 2 — so each DC's two lots split across two suppliers.
    """

    wb = load_workbook(BytesIO(template_bytes))
    ws = wb[SHEET_BIDS]
    headers = _header_map(ws)
    all_in_col = headers[BidColumn.ALL_IN.value]
    fob_col = headers[BidColumn.FOB.value]
    weekly_col = headers[BidColumn.WEEKLY_VOL_OFFERED.value]
    total_col = headers[BidColumn.TOTAL_VOL_OFFERED.value]
    sup_col = headers[BidColumn.SUPPLIER.value]
    lot_col = headers[BidColumn.LOT.value]

    for row in range(BODY_START_ROW, ws.max_row + 1):
        sup = str(ws.cell(row=row, column=sup_col).value or "").strip()
        lot = str(ws.cell(row=row, column=lot_col).value or "").strip()
        if not sup or not lot:
            continue
        base = Decimal("12.00")
        # Each supplier is keenest on a different lot (a real 2-supplier DC split).
        if "Grape" in lot:
            price = base - (Decimal("1.50") if "Green Valley" in sup else Decimal("0.20"))
        else:
            price = base - (Decimal("1.50") if "Sunbelt" in sup else Decimal("0.20"))
        ws.cell(row=row, column=all_in_col, value=float(price))
        ws.cell(row=row, column=fob_col, value=float(price - Decimal("1.00")))
        ws.cell(row=row, column=weekly_col, value=600)
        ws.cell(row=row, column=total_col, value=600 * 13)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
