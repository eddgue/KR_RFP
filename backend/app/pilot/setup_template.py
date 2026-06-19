"""The blank Setup/Kickoff workbook the sponsor fills (PILOT_INPUT_DOCS_SPEC step 0).

`build_setup_workbook()` writes a presentation-quality (D24), VALIDATED, one-tab-per-dimension
workbook — Cycle, DCs, Lots/Items, Suppliers, Volumes, Incumbents, Timeframes (plus the optional
Premiums, Scenario rules, Safeties tabs). The sponsor types only NAMES (D23); the pilot generates
keys on ingest. Each tab carries:

  * a titled banner (the D24 presentation language, reused from `app.output.formatting`),
  * a header row of field names,
  * an italic inline NOTE row describing each field,
  * a couple of GREYED example rows the sponsor overwrites, and
  * data-validation dropdowns where the domain is closed (product type, region, volume method).

`setup_ingest.ingest_setup_workbook` reads exactly these tabs/columns back. The two modules share
the tab/column names via the constants here so the template and the ingester never drift.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from app.output.formatting import (
    NUMFMT_INT,
    NUMFMT_MONEY,
    Col,
    _title_block,
)

# ---------------------------------------------------------------------------
# Closed domains (the dropdowns) — shared with the ingester for validation.
# ---------------------------------------------------------------------------
PRODUCT_TYPES = ("Conventional", "Organic")
REGIONS = ("East", "South", "West", "Midwest", "Central")
VOLUME_METHODS = ("WEEKLY_X_WEEKS", "PERIOD_TOTAL")
WEIGHT_PRESETS = ("balanced", "price_focus", "coverage_focus", "risk_averse", "custom")

# ---------------------------------------------------------------------------
# Tab names (one per dimension) — the single source of truth for both modules.
# ---------------------------------------------------------------------------
TAB_CYCLE = "Cycle"
TAB_DCS = "DCs"
TAB_LOTS = "Lots and Items"
TAB_SUPPLIERS = "Suppliers"
TAB_VOLUMES = "Volumes"
TAB_INCUMBENTS = "Incumbents"
TAB_TIMEFRAMES = "Timeframes"
TAB_PREMIUMS = "Premiums (optional)"
TAB_SCENARIO_RULES = "Scenario rules (optional)"
TAB_SAFETIES = "Safeties (optional)"

# The banner row geometry: title block occupies rows 1-3, then the header, then the note row,
# then example rows, then where the sponsor types. Both modules agree on these.
HEADER_ROW = 5
NOTE_ROW = HEADER_ROW + 1
EXAMPLE_START_ROW = NOTE_ROW + 1

_GREY_FILL = PatternFill("solid", fgColor="F2F2F2")
_GREY_FONT = Font(italic=True, color="808080", size=10)
_NOTE_FONT = Font(italic=True, color="595959", size=9)
_NOTE_FILL = PatternFill("solid", fgColor="EDEDED")
_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)


@dataclass(frozen=True)
class FieldSpec:
    """One column on a setup tab: header text, a note, width, format, and an optional dropdown."""

    header: str
    note: str
    width: int = 22
    number_format: str | None = None
    choices: tuple[str, ...] | None = None  # closed domain -> data-validation dropdown


@dataclass(frozen=True)
class TabSpec:
    """One dimension tab: its name, its columns, and a couple of greyed example rows."""

    name: str
    title: str
    fields: tuple[FieldSpec, ...]
    examples: tuple[tuple[str | int | float, ...], ...] = field(default_factory=tuple)


# ---------------------------------------------------------------------------
# The tab specifications (PILOT_INPUT_DOCS_SPEC step 0 table).
# ---------------------------------------------------------------------------
def _tab_specs() -> tuple[TabSpec, ...]:
    return (
        TabSpec(
            name=TAB_CYCLE,
            title="CYCLE — what we are sourcing & how we score it",
            fields=(
                FieldSpec("Cycle Label", "Plain-language name for this RFP cycle.", 34),
                FieldSpec("Commodity", "The commodity (e.g. Field Tomatoes).", 24),
                FieldSpec("Sub-commodity", "The sub-commodity / category.", 24),
                FieldSpec("Horizon (weeks)", "Weeks the award covers.", 16, NUMFMT_INT),
                FieldSpec("Rounds", "How many bidding rounds (2-6).", 12, NUMFMT_INT),
                FieldSpec(
                    "Target Effective Date",
                    "When the award takes effect (YYYY-MM-DD).",
                    20,
                ),
                FieldSpec(
                    "Weight Preset",
                    "Scoring emphasis.",
                    18,
                    choices=WEIGHT_PRESETS,
                ),
                FieldSpec("Max Suppliers / DC", "Split cap per DC.", 16, NUMFMT_INT),
                FieldSpec("Premium Ceiling", "Max premium vs lowest (e.g. 0.12).", 16),
                FieldSpec("Concentration Threshold", "Category concentration flag.", 18),
                FieldSpec("Coverage Floor", "Min coverage to be eligible (e.g. 0.80).", 16),
            ),
            examples=(
                (
                    "Field Tomatoes — Spring/Summer (EXAMPLE)",
                    "Field Tomatoes",
                    "Round/Vine",
                    13,
                    3,
                    "2026-12-31",
                    "balanced",
                    2,
                    0.12,
                    0.40,
                    0.80,
                ),
            ),
        ),
        TabSpec(
            name=TAB_DCS,
            title="DCs — the distribution centers in scope",
            fields=(
                FieldSpec("DC Name", "Distribution center display name.", 28),
                FieldSpec("Region", "Shipping region.", 16, choices=REGIONS),
                FieldSpec("State", "Two-letter state code.", 12),
            ),
            examples=(
                ("Atlanta DC (EXAMPLE)", "East", "GA"),
                ("Dallas DC (EXAMPLE)", "South", "TX"),
            ),
        ),
        TabSpec(
            name=TAB_LOTS,
            title="LOTS & ITEMS — one item per lot (the engine grain)",
            fields=(
                FieldSpec("Lot Name", "The lot display name.", 28),
                FieldSpec("Item Description", "What the lot is.", 30),
                FieldSpec("Pack Size / UOM", "Pack descriptor.", 20),
                FieldSpec(
                    "Product Type",
                    "Conventional or Organic.",
                    16,
                    choices=PRODUCT_TYPES,
                ),
                FieldSpec("Category", "Free-text category.", 20),
            ),
            examples=(
                (
                    "Lot 1 — Grape Tomato (EXAMPLE)",
                    "Premium Grape Tomato 10oz",
                    "10oz clamshell",
                    "Conventional",
                    "Tomatoes",
                ),
                (
                    "Lot 2 — Roma Bulk (EXAMPLE)",
                    "Roma Tomato Bulk 25lb",
                    "25lb carton",
                    "Conventional",
                    "Tomatoes",
                ),
            ),
        ),
        TabSpec(
            name=TAB_SUPPLIERS,
            title="SUPPLIERS — the invited suppliers",
            fields=(
                FieldSpec("Supplier Name", "Supplier display name.", 30),
                FieldSpec("Region / Origin", "Where they ship from.", 20),
                FieldSpec("Notes", "Any notes.", 30),
            ),
            examples=(
                ("Green Valley Farms (EXAMPLE)", "East", "Incumbent on grape"),
                ("Sunbelt Produce (EXAMPLE)", "South", "New entrant"),
            ),
        ),
        TabSpec(
            name=TAB_VOLUMES,
            title="VOLUMES — projected cases per DC × Lot × Timeframe",
            fields=(
                FieldSpec("DC Name", "Must match a DC tab name.", 28),
                FieldSpec("Lot Name", "Must match a Lots & Items name.", 28),
                FieldSpec("Timeframe", "Must match a Timeframes name.", 24),
                FieldSpec(
                    "Method",
                    "How volume is given.",
                    18,
                    choices=VOLUME_METHODS,
                ),
                FieldSpec("Weekly Cases", "Cases per week.", 16, NUMFMT_INT),
                FieldSpec("Weeks", "Weeks in the timeframe.", 12, NUMFMT_INT),
            ),
            examples=(
                (
                    "Atlanta DC (EXAMPLE)",
                    "Lot 1 — Grape Tomato (EXAMPLE)",
                    "Spring 2026 (EXAMPLE)",
                    "WEEKLY_X_WEEKS",
                    400,
                    13,
                ),
            ),
        ),
        TabSpec(
            name=TAB_INCUMBENTS,
            title="INCUMBENTS — current supplier & routing baseline per DC × Lot",
            fields=(
                FieldSpec("DC Name", "Must match a DC tab name.", 28),
                FieldSpec("Lot Name", "Must match a Lots & Items name.", 28),
                FieldSpec("Incumbent Supplier", "Must match a Suppliers name.", 28),
                FieldSpec(
                    "Routing Baseline $/case",
                    "Prior-period actual-paid $/case.",
                    22,
                    NUMFMT_MONEY,
                ),
                FieldSpec("Contract Notes", "Any notes.", 26),
            ),
            examples=(
                (
                    "Atlanta DC (EXAMPLE)",
                    "Lot 1 — Grape Tomato (EXAMPLE)",
                    "Green Valley Farms (EXAMPLE)",
                    11.20,
                    "Auto-renew",
                ),
            ),
        ),
        TabSpec(
            name=TAB_TIMEFRAMES,
            title="TIMEFRAMES — the season windows",
            fields=(
                FieldSpec("Timeframe Label", "Season window display name.", 26),
                FieldSpec("Start Date", "YYYY-MM-DD.", 16),
                FieldSpec("End Date", "YYYY-MM-DD.", 16),
                FieldSpec("Week Count", "Weeks in the window.", 14, NUMFMT_INT),
            ),
            examples=(
                ("Spring 2026 (EXAMPLE)", "2026-04-01", "2026-06-30", 13),
            ),
        ),
        TabSpec(
            name=TAB_PREMIUMS,
            title="PREMIUMS (optional) — per-lot premium override",
            fields=(
                FieldSpec("Lot Name", "Must match a Lots & Items name.", 28),
                FieldSpec("Premium Threshold", "Override premium ceiling (e.g. 0.10).", 18),
                FieldSpec("Rationale", "Why.", 30),
            ),
        ),
        TabSpec(
            name=TAB_SCENARIO_RULES,
            title="SCENARIO RULES (optional) — exclusions / preferred / limits (lenses E/F/G)",
            fields=(
                FieldSpec(
                    "Rule Type",
                    "Exclusion / Preferred / VolumeLimit.",
                    18,
                    choices=("Exclusion", "Preferred", "VolumeLimit"),
                ),
                FieldSpec("Supplier Name", "Supplier the rule applies to.", 28),
                FieldSpec("Scope (DC/Lot)", "Where it applies (optional).", 22),
                FieldSpec("Value", "Limit / note.", 20),
            ),
        ),
        TabSpec(
            name=TAB_SAFETIES,
            title="SAFETIES (optional, ADR-0014) — collars / cadence / tolerances",
            fields=(
                FieldSpec("Scope (Cell/Lot)", "Where the safety applies.", 22),
                FieldSpec("Collar Cap", "Upper price collar.", 16, NUMFMT_MONEY),
                FieldSpec("Collar Floor", "Lower price collar.", 16, NUMFMT_MONEY),
                FieldSpec("Cadence", "Rolling-midpoint cadence.", 18),
                FieldSpec("Tolerance Band", "Tolerance band.", 16),
                FieldSpec("Disaster Triggers", "Triggers.", 24),
            ),
        ),
    )


# ---------------------------------------------------------------------------
# rendering one tab
# ---------------------------------------------------------------------------
def _render_tab(ws: Worksheet, spec: TabSpec) -> None:
    span = len(spec.fields)

    _title_block(
        ws,
        title=spec.title,
        subtitle_lines=[
            "Type NAMES only — the pilot generates the keys on ingest.",
            "Grey rows are EXAMPLES — overwrite or delete them. Dropdown cells have a fixed list.",
        ],
        span=span,
        start_row=1,
    )

    # Header row — reuse the D24 header styling via format-of-Col widths, but write headers here so
    # the note row + examples sit directly underneath (format_table assumes a clean body; this tab
    # is a fill-out form, so we lay it out directly).
    from app.output.formatting import (  # local import to keep the public surface tidy
        _BORDER,
        _HEADER_FILL,
        _HEADER_FONT,
        _WRAP_CENTER,
    )

    for ci, fld in enumerate(spec.fields, start=1):
        cell = ws.cell(row=HEADER_ROW, column=ci, value=fld.header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _WRAP_CENTER
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(ci)].width = fld.width
    ws.row_dimensions[HEADER_ROW].height = 30

    # Inline NOTE row (italic grey, wrapped) — describes each field.
    for ci, fld in enumerate(spec.fields, start=1):
        cell = ws.cell(row=NOTE_ROW, column=ci, value=fld.note)
        cell.font = _NOTE_FONT
        cell.fill = _NOTE_FILL
        cell.alignment = _WRAP
        cell.border = _BORDER
    ws.row_dimensions[NOTE_ROW].height = 28

    # Greyed example rows the sponsor overwrites.
    for ri, example in enumerate(spec.examples):
        row = EXAMPLE_START_ROW + ri
        for ci, fld in enumerate(spec.fields, start=1):
            value = example[ci - 1] if ci - 1 < len(example) else None
            cell = ws.cell(row=row, column=ci, value=value)
            cell.font = _GREY_FONT
            cell.fill = _GREY_FILL
            cell.border = _BORDER
            if fld.number_format and isinstance(value, int | float):
                cell.number_format = fld.number_format

    # Data-validation dropdowns where the domain is closed — applied to a generous body range so
    # the sponsor's typed rows below the examples still get the list.
    body_start = EXAMPLE_START_ROW
    body_end = body_start + 200
    for ci, fld in enumerate(spec.fields, start=1):
        if not fld.choices:
            continue
        letter = get_column_letter(ci)
        joined = ",".join(fld.choices)
        dv = DataValidation(
            type="list",
            formula1=f'"{joined}"',
            allow_blank=True,
            showDropDown=False,  # False => the arrow IS shown (openpyxl/Excel quirk)
        )
        dv.promptTitle = fld.header
        dv.prompt = f"Pick from: {', '.join(fld.choices)}"
        dv.errorTitle = "Not a valid choice"
        dv.error = f"{fld.header} must be one of: {', '.join(fld.choices)}"
        ws.add_data_validation(dv)
        dv.add(f"{letter}{body_start}:{letter}{body_end}")

    ws.freeze_panes = ws.cell(row=EXAMPLE_START_ROW, column=1)


# ---------------------------------------------------------------------------
# public entry point
# ---------------------------------------------------------------------------
def build_setup_workbook() -> bytes:
    """Build the blank, validated, presentation-quality Setup/Kickoff workbook -> xlsx bytes."""

    wb = Workbook()
    default = wb.active  # the auto-created first sheet; replaced by the Cycle tab
    specs = _tab_specs()
    for i, spec in enumerate(specs):
        ws = default if i == 0 and default is not None else wb.create_sheet()
        ws.title = spec.name
        _render_tab(ws, spec)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# Re-export the Col symbol so callers importing the template's presentation vocabulary find it here
# too (the setup tabs reuse the same D24 widths/number-formats vocabulary as the output surfaces).
__all__ = [
    "PRODUCT_TYPES",
    "REGIONS",
    "TAB_CYCLE",
    "TAB_DCS",
    "TAB_INCUMBENTS",
    "TAB_LOTS",
    "TAB_SUPPLIERS",
    "TAB_TIMEFRAMES",
    "TAB_VOLUMES",
    "VOLUME_METHODS",
    "WEIGHT_PRESETS",
    "Col",
    "build_setup_workbook",
]
