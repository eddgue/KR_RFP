"""Ingest a filled Setup/Kickoff workbook into the governed store (PILOT_INPUT_DOCS_SPEC step 0).

`ingest_setup_workbook(session, data, …) -> cycle_id` parses the sponsor-filled setup workbook and
writes the cycle + its full scope into the SAME governed tables the demo's `seed_cycle` writes
(ref.client/commodity/subcommodity/item/dc/supplier, cyc.cycle/cycle_lot/cycle_item_scope/
cycle_lot_item/cycle_timeframe/cycle_round/cycle_invited_supplier/cycle_projected_volume, and
perf.historical_award_assignment + perf.historical_awarded_price_basis for the routing baseline) —
but from WORKBOOK ROWS instead of synthetic constants. Keys are generated here and embedded; the
sponsor only ever typed names (D23).

VALIDATION is strict (the spec's "never silently drops"): the ingester collects every row it cannot
resolve (missing required cells, cross-references that don't match a named DC/Lot/Supplier/TF, bad
numbers) and raises `SetupIngestError` listing ALL problems at once. A clean workbook ingests
atomically; a dirty one ingests nothing.

Transaction discipline (PLAN §7): this is a service-style function — it executes + flushes inside
the caller's unit of work and never commits. `app.cycle.loader.load_cycle` reads the result back,
so a setup workbook round-trips: template -> fill -> ingest -> load_cycle.
"""

from __future__ import annotations

import uuid
from dataclasses import dataclass, field
from datetime import UTC, date, datetime
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.pilot.setup_template import (
    EXAMPLE_START_ROW,
    HEADER_ROW,
    PRODUCT_TYPES,
    TAB_CYCLE,
    TAB_DCS,
    TAB_INCUMBENTS,
    TAB_LOTS,
    TAB_SUPPLIERS,
    TAB_TIMEFRAMES,
    TAB_VOLUMES,
)

# The first data row the sponsor types into (examples start at EXAMPLE_START_ROW but are greyed and
# carry an "(EXAMPLE)" marker; we skip any row whose key cell still contains that marker).
_BODY_START = EXAMPLE_START_ROW
_EXAMPLE_MARKER = "(EXAMPLE)"

# The platform's bidding-round range (the cyc.cycle round_count check constraint is 2..6).
_MIN_ROUNDS = 2
_MAX_ROUNDS = 6


class SetupIngestError(ValueError):
    """Raised when the filled setup workbook has rows/tabs that cannot be resolved.

    Carries the full list of problems so the sponsor gets every fix at once, not one at a time.
    """

    def __init__(self, problems: list[str]) -> None:
        self.problems = problems
        joined = "\n".join(f"  - {p}" for p in problems)
        super().__init__(
            f"Setup workbook could not be ingested ({len(problems)} problem(s)):\n{joined}"
        )


@dataclass
class _Problems:
    """Accumulates resolution problems so we can report them all together (never silent-drop)."""

    items: list[str] = field(default_factory=list)

    def add(self, tab: str, row: int | None, message: str) -> None:
        loc = f"tab '{tab}'" + (f", row {row}" if row is not None else "")
        self.items.append(f"{loc}: {message}")

    def raise_if_any(self) -> None:
        if self.items:
            raise SetupIngestError(self.items)


def _id() -> str:
    return str(uuid.uuid4())


# ---------------------------------------------------------------------------
# cell readers
# ---------------------------------------------------------------------------
def _header_map(ws: Worksheet) -> dict[str, int]:
    out: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=HEADER_ROW, column=col).value
        if value is not None:
            out[str(value).strip()] = col
    return out


def _cell(ws: Worksheet, row: int, col: int) -> str:
    value = ws.cell(row=row, column=col).value
    return "" if value is None else str(value).strip()


def _is_example_or_blank(ws: Worksheet, row: int, key_col: int) -> bool:
    """A row is skipped if its key cell is blank or still holds the greyed '(EXAMPLE)' marker."""

    value = _cell(ws, row, key_col)
    return value == "" or _EXAMPLE_MARKER in value


def _data_rows(ws: Worksheet, key_col: int) -> list[int]:
    """Row indices the sponsor actually filled (skip greyed examples + trailing blanks)."""

    rows: list[int] = []
    for row in range(_BODY_START, ws.max_row + 1):
        if _is_example_or_blank(ws, row, key_col):
            continue
        rows.append(row)
    return rows


def _to_decimal(raw: str) -> Decimal | None:
    try:
        return Decimal(raw.replace(",", "").replace("$", "").strip())
    except (InvalidOperation, ValueError):
        return None


def _to_int(raw: str) -> int | None:
    dec = _to_decimal(raw)
    if dec is None:
        return None
    try:
        return int(dec)
    except (ValueError, OverflowError):
        return None


@dataclass
class _Named:
    """A parsed named entity (DC / lot / supplier / timeframe): generated id + display name."""

    id: str
    name: str


# ---------------------------------------------------------------------------
# the ingest
# ---------------------------------------------------------------------------
def ingest_setup_workbook(
    session: Session, data: bytes, *, created_by: str = "pilot"
) -> str:
    """Parse the filled setup workbook and write the cycle + scope; return the new cycle_id.

    Validates first (collecting ALL problems), then writes atomically within the caller's unit of
    work. Raises `SetupIngestError` if anything can't be resolved — nothing is written in that case.
    """

    from io import BytesIO

    wb = load_workbook(BytesIO(data), data_only=True)
    problems = _Problems()

    required_tabs = (
        TAB_CYCLE,
        TAB_DCS,
        TAB_LOTS,
        TAB_SUPPLIERS,
        TAB_VOLUMES,
        TAB_INCUMBENTS,
        TAB_TIMEFRAMES,
    )
    for tab in required_tabs:
        if tab not in wb.sheetnames:
            problems.add(tab, None, "required tab is missing")
    problems.raise_if_any()

    # --- Cycle tab (single row) ---
    cyc_ws = wb[TAB_CYCLE]
    cyc_hdr = _header_map(cyc_ws)
    cycle_rows = _data_rows(cyc_ws, cyc_hdr.get("Cycle Label", 1))
    if not cycle_rows:
        problems.add(TAB_CYCLE, None, "no cycle row filled in (need the cycle label etc.)")
        problems.raise_if_any()
    crow = cycle_rows[0]
    cycle_label = _cell(cyc_ws, crow, cyc_hdr["Cycle Label"])
    commodity_name = _cell(cyc_ws, crow, cyc_hdr.get("Commodity", 0)) or "Commodity"
    subcommodity_name = _cell(cyc_ws, crow, cyc_hdr.get("Sub-commodity", 0)) or commodity_name
    # The platform requires 2..6 bidding rounds (the cyc.cycle round_count check constraint).
    rounds = _to_int(_cell(cyc_ws, crow, cyc_hdr.get("Rounds", 0))) or _MIN_ROUNDS
    if not (_MIN_ROUNDS <= rounds <= _MAX_ROUNDS):
        problems.add(
            TAB_CYCLE,
            crow,
            f"Rounds must be between {_MIN_ROUNDS} and {_MAX_ROUNDS} (got {rounds})",
        )
    ted_raw = _cell(cyc_ws, crow, cyc_hdr.get("Target Effective Date", 0))
    target_effective = _parse_date(ted_raw) if ted_raw else None

    # --- DCs ---
    dcs = _parse_named_tab(wb[TAB_DCS], "DC Name", TAB_DCS, problems)
    # --- Suppliers ---
    suppliers = _parse_named_tab(wb[TAB_SUPPLIERS], "Supplier Name", TAB_SUPPLIERS, problems)
    # --- Timeframes ---
    tf_ws = wb[TAB_TIMEFRAMES]
    tf_hdr = _header_map(tf_ws)
    timeframes: dict[str, _Named] = {}
    tf_meta: dict[str, tuple[date | None, date | None, int]] = {}
    for row in _data_rows(tf_ws, tf_hdr.get("Timeframe Label", 1)):
        name = _cell(tf_ws, row, tf_hdr["Timeframe Label"])
        if name in timeframes:
            problems.add(TAB_TIMEFRAMES, row, f"duplicate timeframe '{name}'")
            continue
        start = _parse_date(_cell(tf_ws, row, tf_hdr.get("Start Date", 0)))
        end = _parse_date(_cell(tf_ws, row, tf_hdr.get("End Date", 0)))
        weeks = _to_int(_cell(tf_ws, row, tf_hdr.get("Week Count", 0))) or 13
        timeframes[name] = _Named(_id(), name)
        tf_meta[name] = (start, end, weeks)
    if not timeframes:
        problems.add(TAB_TIMEFRAMES, None, "no timeframes filled in")

    # --- Lots & Items ---
    lots_ws = wb[TAB_LOTS]
    lots_hdr = _header_map(lots_ws)
    lots: dict[str, _Named] = {}
    items_by_lot: dict[str, _Named] = {}
    lot_item_meta: dict[str, tuple[str, str, str, str]] = {}  # lot -> (desc, pack, ptype, category)
    for row in _data_rows(lots_ws, lots_hdr.get("Lot Name", 1)):
        lot_name = _cell(lots_ws, row, lots_hdr["Lot Name"])
        if lot_name in lots:
            problems.add(TAB_LOTS, row, f"duplicate lot '{lot_name}'")
            continue
        desc = _cell(lots_ws, row, lots_hdr.get("Item Description", 0)) or lot_name
        pack = _cell(lots_ws, row, lots_hdr.get("Pack Size / UOM", 0))
        ptype = _cell(lots_ws, row, lots_hdr.get("Product Type", 0))
        if ptype and ptype not in PRODUCT_TYPES:
            problems.add(
                TAB_LOTS, row, f"Product Type '{ptype}' not in {', '.join(PRODUCT_TYPES)}"
            )
        category = _cell(lots_ws, row, lots_hdr.get("Category", 0))
        lots[lot_name] = _Named(_id(), lot_name)
        items_by_lot[lot_name] = _Named(_id(), desc)
        lot_item_meta[lot_name] = (desc, pack, ptype, category)
    if not lots:
        problems.add(TAB_LOTS, None, "no lots filled in")

    # --- Volumes (cross-references DC / Lot / Timeframe) ---
    vol_ws = wb[TAB_VOLUMES]
    vol_hdr = _header_map(vol_ws)
    volume_rows: list[tuple[str, str, str, Decimal]] = []  # (dc, lot, tf, period_cases)
    for row in _data_rows(vol_ws, vol_hdr.get("DC Name", 1)):
        dc_name = _cell(vol_ws, row, vol_hdr["DC Name"])
        lot_name = _cell(vol_ws, row, vol_hdr.get("Lot Name", 0))
        tf_name = _cell(vol_ws, row, vol_hdr.get("Timeframe", 0))
        if dc_name not in dcs:
            problems.add(TAB_VOLUMES, row, f"DC '{dc_name}' not found on the DCs tab")
        if lot_name not in lots:
            problems.add(TAB_VOLUMES, row, f"Lot '{lot_name}' not found on the Lots tab")
        if tf_name not in timeframes:
            problems.add(TAB_VOLUMES, row, f"Timeframe '{tf_name}' not found on the Timeframes tab")
        weekly = _to_decimal(_cell(vol_ws, row, vol_hdr.get("Weekly Cases", 0)))
        vol_weeks = _to_int(_cell(vol_ws, row, vol_hdr.get("Weeks", 0)))
        if weekly is None or vol_weeks is None:
            problems.add(TAB_VOLUMES, row, "Weekly Cases and Weeks must both be numbers")
            continue
        if dc_name in dcs and lot_name in lots and tf_name in timeframes:
            volume_rows.append((dc_name, lot_name, tf_name, weekly * Decimal(vol_weeks)))

    # --- Incumbents (cross-references DC / Lot / Supplier) ---
    inc_ws = wb[TAB_INCUMBENTS]
    inc_hdr = _header_map(inc_ws)
    incumbent_rows: list[tuple[str, str, str, Decimal]] = []  # (dc, lot, supplier, routing)
    for row in _data_rows(inc_ws, inc_hdr.get("DC Name", 1)):
        dc_name = _cell(inc_ws, row, inc_hdr["DC Name"])
        lot_name = _cell(inc_ws, row, inc_hdr.get("Lot Name", 0))
        sup_name = _cell(inc_ws, row, inc_hdr.get("Incumbent Supplier", 0))
        if dc_name not in dcs:
            problems.add(TAB_INCUMBENTS, row, f"DC '{dc_name}' not found on the DCs tab")
        if lot_name not in lots:
            problems.add(TAB_INCUMBENTS, row, f"Lot '{lot_name}' not found on the Lots tab")
        if sup_name not in suppliers:
            problems.add(
                TAB_INCUMBENTS, row, f"Supplier '{sup_name}' not found on the Suppliers tab"
            )
        routing = _to_decimal(_cell(inc_ws, row, inc_hdr.get("Routing Baseline $/case", 0)))
        if routing is None:
            problems.add(TAB_INCUMBENTS, row, "Routing Baseline $/case must be a number")
            continue
        if dc_name in dcs and lot_name in lots and sup_name in suppliers:
            incumbent_rows.append((dc_name, lot_name, sup_name, routing))

    problems.raise_if_any()

    # === All resolved — WRITE the cycle + scope (mirrors seed_cycle, by key) ===
    return _write_cycle(
        session,
        created_by=created_by,
        cycle_label=cycle_label,
        commodity_name=commodity_name,
        subcommodity_name=subcommodity_name,
        rounds=rounds,
        target_effective=target_effective,
        dcs=dcs,
        suppliers=suppliers,
        timeframes=timeframes,
        tf_meta=tf_meta,
        lots=lots,
        items_by_lot=items_by_lot,
        lot_item_meta=lot_item_meta,
        volume_rows=volume_rows,
        incumbent_rows=incumbent_rows,
    )


def _parse_named_tab(
    ws: Worksheet, key_header: str, tab: str, problems: _Problems
) -> dict[str, _Named]:
    """Parse a simple name-keyed tab (DCs / Suppliers) into {name: _Named}, flagging duplicates."""

    hdr = _header_map(ws)
    key_col = hdr.get(key_header, 1)
    out: dict[str, _Named] = {}
    for row in _data_rows(ws, key_col):
        name = _cell(ws, row, key_col)
        if name in out:
            problems.add(tab, row, f"duplicate '{name}'")
            continue
        out[name] = _Named(_id(), name)
    if not out:
        problems.add(tab, None, f"no rows filled in (need at least one {key_header})")
    return out


def _parse_date(raw: str) -> date | None:
    raw = raw.strip()
    if not raw:
        return None
    # openpyxl may already hand back a datetime string; accept ISO and common forms.
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


# ---------------------------------------------------------------------------
# the writer (mirrors seed_cycle's inserts, FK-heavy governed spine)
# ---------------------------------------------------------------------------
def _write_cycle(  # noqa: PLR0913 — one cohesive writer; args are the parsed scope
    session: Session,
    *,
    created_by: str,
    cycle_label: str,
    commodity_name: str,
    subcommodity_name: str,
    rounds: int,
    target_effective: date | None,
    dcs: dict[str, _Named],
    suppliers: dict[str, _Named],
    timeframes: dict[str, _Named],
    tf_meta: dict[str, tuple[date | None, date | None, int]],
    lots: dict[str, _Named],
    items_by_lot: dict[str, _Named],
    lot_item_meta: dict[str, tuple[str, str, str, str]],
    volume_rows: list[tuple[str, str, str, Decimal]],
    incumbent_rows: list[tuple[str, str, str, Decimal]],
) -> str:
    now = datetime.now(UTC).replace(tzinfo=None)

    commodity_id = _id()
    subcommodity_id = _id()
    cycle_id = _id()
    cycle_code = f"CYC-{now:%Y%m%d}-{cycle_id[:4].upper()}"
    effective = target_effective or date(now.year, 12, 31)

    # ref.client / commodity / subcommodity — the FK chain the cyc spine hangs off.
    session.execute(
        text(
            "INSERT INTO ref.client (id, client_code, client_name, is_active) "
            "VALUES (gen_random_uuid(), :code, :name, true)"
        ),
        {"code": f"CLIENT-{cycle_id[:6].upper()}", "name": f"{cycle_label} client"},
    )
    session.execute(
        text(
            "INSERT INTO ref.commodity (id, client_id, commodity_code, commodity_name) "
            "VALUES (:cid, NULL, :code, :name)"
        ),
        {"cid": commodity_id, "code": f"COMM-{cycle_id[:6].upper()}", "name": commodity_name},
    )
    session.execute(
        text(
            "INSERT INTO ref.subcommodity "
            "(subcommodity_id, commodity_id, subcommodity_code, subcommodity_name, active_flag) "
            "VALUES (:sid, :cid, :code, :name, true)"
        ),
        {
            "sid": subcommodity_id,
            "cid": commodity_id,
            "code": f"SUBCOMM-{cycle_id[:6].upper()}",
            "name": subcommodity_name,
        },
    )

    session.execute(
        text(
            "INSERT INTO cyc.cycle (cycle_id, cycle_code, cycle_name, commodity_id, "
            "subcommodity_id, status, why_now, target_effective_date, round_count, "
            "created_at, created_by) VALUES (:cyc, :code, :name, :cid, :sid, 'OPEN', "
            "'Pilot setup ingest', :ted, :rc, :now, :by)"
        ),
        {
            "cyc": cycle_id,
            "code": cycle_code,
            "name": cycle_label,
            "cid": commodity_id,
            "sid": subcommodity_id,
            "ted": effective,
            "rc": rounds,
            "now": now,
            "by": created_by,
        },
    )

    # ref.dc
    for i, dc in enumerate(dcs.values(), start=1):
        session.execute(
            text(
                "INSERT INTO ref.dc (dc_id, dc_code, dc_name, region, division, active_flag) "
                "VALUES (:id, :code, :name, :region, 'Produce', true)"
            ),
            {"id": dc.id, "code": f"DC{i:02d}", "name": dc.name, "region": "EAST"},
        )

    # ref.supplier
    for sup in suppliers.values():
        session.execute(
            text(
                "INSERT INTO ref.supplier (supplier_id, canonical_name, active_flag, created_at) "
                "VALUES (:id, :name, true, :now)"
            ),
            {"id": sup.id, "name": sup.name, "now": now},
        )

    # ref.item + cyc.cycle_lot + cyc.cycle_item_scope + cyc.cycle_lot_item (one item per lot).
    for i, (lot_name, lot) in enumerate(lots.items(), start=1):
        item = items_by_lot[lot_name]
        desc, pack, _ptype, _category = lot_item_meta[lot_name]
        session.execute(
            text(
                "INSERT INTO ref.item (item_id, item_code, description, pack_desc, commodity_id, "
                "subcommodity_id) VALUES (:id, :code, :desc, :pack, :cid, :sid)"
            ),
            {
                "id": item.id,
                "code": f"ITEM-{i:02d}",
                "desc": desc,
                "pack": pack,
                "cid": commodity_id,
                "sid": subcommodity_id,
            },
        )
        session.execute(
            text(
                "INSERT INTO cyc.cycle_lot (lot_id, cycle_id, lot_code, lot_name, active_flag) "
                "VALUES (:id, :cyc, :code, :name, true)"
            ),
            {"id": lot.id, "cyc": cycle_id, "code": f"LOT-{i:02d}", "name": lot.name},
        )
        session.execute(
            text(
                "INSERT INTO cyc.cycle_item_scope (cycle_id, item_id, commodity_id, "
                "subcommodity_id, inclusion_status, added_at, added_by) "
                "VALUES (:cyc, :item, :cid, :sid, 'IN_SCOPE', :now, :by)"
            ),
            {
                "cyc": cycle_id,
                "item": item.id,
                "cid": commodity_id,
                "sid": subcommodity_id,
                "now": now,
                "by": created_by,
            },
        )
        session.execute(
            text(
                "INSERT INTO cyc.cycle_lot_item (lot_item_id, cycle_id, lot_id, item_id, "
                "required_flag, sort_order) VALUES (:lid, :cyc, :lot, :item, true, :so)"
            ),
            {"lid": _id(), "cyc": cycle_id, "lot": lot.id, "item": item.id, "so": i},
        )

    # cyc.cycle_timeframe
    for i, (tf_name, tf) in enumerate(timeframes.items(), start=1):
        start, end, weeks = tf_meta[tf_name]
        start = start or date(now.year, 1 + (i - 1) * 3, 1)
        end = end or date(now.year, 3 + (i - 1) * 3, 28)
        session.execute(
            text(
                "INSERT INTO cyc.cycle_timeframe (tf_id, cycle_id, tf_code, tf_name, "
                "start_date, end_date, week_count) VALUES (:id, :cyc, :code, :name, :s, :e, :w)"
            ),
            {
                "id": tf.id,
                "cyc": cycle_id,
                "code": f"TF{i:02d}",
                "name": tf.name,
                "s": start,
                "e": end,
                "w": weeks,
            },
        )

    # cyc.cycle_round
    for i in range(1, rounds + 1):
        session.execute(
            text(
                "INSERT INTO cyc.cycle_round (round_id, cycle_id, round_number, status, "
                "round_status, is_final) VALUES (:id, :cyc, :n, 'OPEN', 'OPEN', :final)"
            ),
            {"id": _id(), "cyc": cycle_id, "n": i, "final": i == rounds},
        )

    # cyc.cycle_invited_supplier
    for sup in suppliers.values():
        session.execute(
            text(
                "INSERT INTO cyc.cycle_invited_supplier (cycle_id, supplier_id, invited_at, "
                "invited_by) VALUES (:cyc, :sup, :now, :by)"
            ),
            {"cyc": cycle_id, "sup": sup.id, "now": now, "by": created_by},
        )

    # cyc.cycle_projected_volume — keyed at (dc, item, tf); item via the lot's one item.
    for dc_name, lot_name, tf_name, period in volume_rows:
        item = items_by_lot[lot_name]
        weekly = period / Decimal(max(1, tf_meta[tf_name][2]))
        session.execute(
            text(
                "INSERT INTO cyc.cycle_projected_volume (volume_id, cycle_id, dc_id, "
                "item_id, tf_id, volume_input_method, projected_weekly_cases, "
                "projected_period_cases) VALUES (:id, :cyc, :dc, :item, :tf, "
                "'WEEKLY_X_WEEKS', :wk, :pd)"
            ),
            {
                "id": _id(),
                "cyc": cycle_id,
                "dc": dcs[dc_name].id,
                "item": item.id,
                "tf": timeframes[tf_name].id,
                "wk": weekly,
                "pd": period,
            },
        )

    # Incumbents (perf.historical_award_assignment) + routing baseline price-basis row.
    inc_run_id = _id()
    session.execute(
        text(
            "INSERT INTO norm.normalization_run (normalization_run_id, dataset_type, cycle_id, "
            "status) VALUES (:id, 'HISTORICAL_AWARD', :cyc, 'APPROVED')"
        ),
        {"id": inc_run_id, "cyc": cycle_id},
    )
    for dc_name, lot_name, sup_name, routing in incumbent_rows:
        item = items_by_lot[lot_name]
        assignment_id = _id()
        session.execute(
            text(
                "INSERT INTO perf.historical_award_assignment (assignment_id, cycle_id, "
                "dc_id, item_id, supplier_id, effective_start_date, effective_end_date, "
                "awarded_volume_cases, ingestion_run_id, incumbent_flag, created_at, "
                "created_by) VALUES (:id, :cyc, :dc, :item, :sup, :s, :e, :vol, :run, "
                "true, :now, :by)"
            ),
            {
                "id": assignment_id,
                "cyc": cycle_id,
                "dc": dcs[dc_name].id,
                "item": item.id,
                "sup": suppliers[sup_name].id,
                "s": date(now.year - 1, 1, 1),
                "e": date(now.year - 1, 12, 31),
                "vol": Decimal("0"),
                "run": inc_run_id,
                "now": now,
                "by": created_by,
            },
        )
        session.execute(
            text(
                "INSERT INTO perf.historical_awarded_price_basis (price_basis_id, "
                "assignment_id, routing_basis, awarded_price_per_case, preferred_basis_flag, "
                "preferred_basis_source, created_at) VALUES (:id, :aid, 'DELIVERED', :price, "
                "true, :by, :now)"
            ),
            {
                "id": _id(),
                "aid": assignment_id,
                "price": routing,
                "by": created_by,
                "now": now,
            },
        )

    session.flush()
    return cycle_id
