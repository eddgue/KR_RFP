"""Flexible bid ingest — "take my file as-is, figure it out, parse it" (PILOT_SYSTEM_DESIGN §4).

Beyond the strict key-validated path (`bid_ingester.ingest_template`, which only reads OUR own
key-stamped template), the pilot offers a flexible ingest for the messy spreadsheets a supplier
actually sends — different column words, a different column order, no embedded keys. This module is
the inference + adaptation core:

  * `infer_bid_mapping(data, cycle) -> MappingProposal` — read the file as-is and INFER which
    columns are supplier / DC / lot / all-in / fob / volume, using (a) produce-sourcing header
    SYNONYMS and (b) matching the cells' VALUES against the cycle's KNOWN supplier / DC / lot NAMES
    (a column whose cells resolve to known names locks that role). Returns a plain-language proposal
    the skill SHOWS the buyer for a quick confirm; ambiguity is surfaced, never guessed silently.
  * `apply_mapping(data, mapping, scope) -> bytes` — produce a CLEAN, key-stamped OWNED template:
    start from the owned template for the scope (keys embedded, D21), then overlay the messy file's
    price/volume cells onto the matching scope rows (resolving each messy row's supplier/DC/lot
    NAMES to the scope identity). The result ingests through the strict key-validated path.

CLEAN-ROOM (ADR-0001): synthetic/scope-driven only; no `reference/` import. The inference is a
heuristic over the cycle's own known names + a small synonym table — it never invents an identity.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.domain.bid.template_generator import build_template_workbook
from app.domain.bid.template_schema import (
    BODY_START_ROW,
    HEADER_ROW,
    SHEET_BIDS,
    BidColumn,
    CycleScope,
)
from app.output.types import CycleView

# The flexible-ingest fields we try to locate in a messy file. The IDENTITY fields (supplier/dc/lot)
# place a row in the cycle grain; the PRICE/volume fields carry the numbers the engine scores.
FIELD_SUPPLIER = "supplier"
FIELD_DC = "dc"
FIELD_LOT = "lot"
FIELD_ALL_IN = "all_in"
FIELD_FOB = "fob"
FIELD_VOLUME = "volume"

_IDENTITY_FIELDS = (FIELD_SUPPLIER, FIELD_DC, FIELD_LOT)
_PRICE_FIELDS = (FIELD_ALL_IN, FIELD_FOB, FIELD_VOLUME)
_ALL_FIELDS = (*_IDENTITY_FIELDS, *_PRICE_FIELDS)

# Header SYNONYMS (produce-sourcing lingo) — matched case/space-folded as substrings, longest first
# so "all in" beats a bare "price". Each tuple is (field, ordered synonyms).
_HEADER_SYNONYMS: tuple[tuple[str, tuple[str, ...]], ...] = (
    (FIELD_ALL_IN, ("all in", "all-in", "allin", "landed", "delivered price", "net price")),
    (FIELD_FOB, ("fob", "farm gate", "farm-gate", "ex works", "ex-works")),
    (FIELD_VOLUME, ("volume", "cases", "qty", "quantity", "weekly", "vol offered")),
    (FIELD_SUPPLIER, ("supplier", "vendor", "grower", "shipper", "bidder", "company")),
    (FIELD_DC, ("dc", "distribution center", "distribution centre", "warehouse", "ship to")),
    (FIELD_LOT, ("lot", "item", "product", "sku", "commodity")),
)


@dataclass(frozen=True)
class ColumnMapping:
    """One inferred column → field mapping with how it was decided (plain language)."""

    field: str
    column_index: int  # 1-based column in the messy sheet
    source_header: str
    basis: str  # "header" (matched a synonym) or "values" (cells matched known names)
    confidence: str  # "high" / "medium" / "low"


@dataclass
class MappingProposal:
    """The inferred mapping the skill SHOWS the buyer for a quick confirm before applying.

    `mappings` is field → ColumnMapping for every field we located. `header_row` / `sheet_name`
    record where the table was found. `ambiguities` lists, in plain language, anything we could NOT
    pin down (a missing identity column, a price column we couldn't find) — surfaced, never guessed.
    """

    sheet_name: str
    header_row: int
    mappings: dict[str, ColumnMapping] = field(default_factory=dict)
    ambiguities: list[str] = field(default_factory=list)

    @property
    def is_confident(self) -> bool:
        """True when every identity field + a price field is mapped, with nothing left ambiguous."""

        has_identity = all(f in self.mappings for f in _IDENTITY_FIELDS)
        has_price = any(f in self.mappings for f in (FIELD_ALL_IN, FIELD_FOB))
        return has_identity and has_price and not self.ambiguities

    def describe(self) -> str:
        """A short plain-language summary of the mapping (for the skill to show the buyer)."""

        lines = [f"Reading sheet '{self.sheet_name}' (headers on row {self.header_row}):"]
        for fld in _ALL_FIELDS:
            m = self.mappings.get(fld)
            if m is not None:
                lines.append(
                    f"  - {fld}: column '{m.source_header}' "
                    f"(matched by {m.basis}, {m.confidence} confidence)"
                )
        for a in self.ambiguities:
            lines.append(f"  ! {a}")
        return "\n".join(lines)


def _norm(value: object) -> str:
    return " ".join(str(value).strip().lower().split()) if value is not None else ""


def _find_header_row(ws: Worksheet, max_scan: int = 10) -> int:
    """Find the most likely header row: the first row (<= max_scan) with >= 2 non-empty text cells.

    A pragmatic heuristic for a messy file — most supplier sheets carry a title band then a header
    row of column labels. We pick the earliest text-rich row so the rows below it are the body.
    """

    best_row = 1
    best_score = -1
    for r in range(1, min(ws.max_row, max_scan) + 1):
        text_cells = 0
        for c in range(1, ws.max_column + 1):
            value = ws.cell(row=r, column=c).value
            if isinstance(value, str) and value.strip():
                text_cells += 1
        if text_cells > best_score:
            best_score = text_cells
            best_row = r
    return best_row


def _match_header(header: str) -> str | None:
    """Return the field a header text matches by SYNONYM, or None (first matching field wins)."""

    norm = _norm(header)
    if not norm:
        return None
    for fld, synonyms in _HEADER_SYNONYMS:
        for syn in synonyms:
            if syn in norm:
                return fld
    return None


def infer_bid_mapping(data: bytes, cycle: CycleView) -> MappingProposal:
    """Infer the messy file's column → field mapping against the cycle's known scope (§4).

    Two signals, combined: (1) the column HEADER matched against produce-sourcing synonyms, and
    (2) the column's VALUES matched against the cycle's KNOWN supplier / DC / lot NAMES (a column
    whose cells resolve to known names locks that identity role, even if the header is odd). Header
    matches and value matches that AGREE read "high"; a value-only lock reads "medium". Anything we
    can't pin down lands in `ambiguities` for the buyer to resolve — we never guess silently.
    """

    wb = load_workbook(BytesIO(data), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    header_row = _find_header_row(ws)

    # Known-name sets for the value-matching signal (normalized).
    known = {
        FIELD_SUPPLIER: {_norm(s.name): s for s in cycle.suppliers},
        FIELD_DC: {_norm(d.name): d for d in cycle.dcs},
        FIELD_LOT: {_norm(lot.name): lot for lot in cycle.lots},
    }

    n_cols = ws.max_column
    headers = [
        (c, str(ws.cell(row=header_row, column=c).value or "").strip())
        for c in range(1, n_cols + 1)
    ]

    # Value-match score per (column, identity field): fraction of body cells resolving to a known
    # name. The body is everything below the header row.
    value_hits: dict[tuple[int, str], int] = {}
    value_total: dict[int, int] = {}
    for c, _hdr in headers:
        total = 0
        for r in range(header_row + 1, ws.max_row + 1):
            cell = _norm(ws.cell(row=r, column=c).value)
            if not cell:
                continue
            total += 1
            for fld in _IDENTITY_FIELDS:
                if cell in known[fld]:
                    value_hits[(c, fld)] = value_hits.get((c, fld), 0) + 1
        value_total[c] = total

    wb.close()

    proposal = MappingProposal(sheet_name=ws.title, header_row=header_row)

    # Decide each column. A column is claimed by the field with the strongest combined signal.
    claimed_fields: dict[str, ColumnMapping] = {}
    for c, hdr in headers:
        header_field = _match_header(hdr)
        # Best value-match identity field for this column (>= 50% of non-empty cells resolve).
        total = value_total.get(c, 0)
        best_value_field: str | None = None
        best_frac = 0.0
        if total:
            for fld in _IDENTITY_FIELDS:
                frac = value_hits.get((c, fld), 0) / total
                if frac > best_frac:
                    best_frac = frac
                    best_value_field = fld
        value_field = best_value_field if best_frac >= 0.5 else None

        chosen: str | None = None
        basis = ""
        confidence = "low"
        if value_field is not None and header_field == value_field:
            chosen, basis, confidence = value_field, "header+values", "high"
        elif value_field is not None:
            chosen, basis, confidence = value_field, "values", "medium"
        elif header_field is not None:
            chosen = header_field
            basis = "header"
            confidence = "high" if header_field in _PRICE_FIELDS else "medium"

        if chosen is None:
            continue
        candidate = ColumnMapping(
            field=chosen,
            column_index=c,
            source_header=hdr or f"(column {c})",
            basis=basis,
            confidence=confidence,
        )
        prior = claimed_fields.get(chosen)
        if prior is None or _rank(candidate) > _rank(prior):
            if prior is not None:
                proposal.ambiguities.append(
                    f"Two columns looked like '{chosen}' "
                    f"('{prior.source_header}' and '{hdr}') — kept '{candidate.source_header}'."
                )
            claimed_fields[chosen] = candidate

    proposal.mappings.update(claimed_fields)

    # Surface what's missing, in the buyer's terms.
    for fld in _IDENTITY_FIELDS:
        if fld not in proposal.mappings:
            proposal.ambiguities.append(
                f"Couldn't find the {fld} column — which column holds the {fld}?"
            )
    if FIELD_ALL_IN not in proposal.mappings and FIELD_FOB not in proposal.mappings:
        proposal.ambiguities.append(
            "Couldn't find a price column (All-In or FOB) — which column holds the bid price?"
        )
    return proposal


_CONFIDENCE_RANK = {"high": 3, "medium": 2, "low": 1}


def _rank(m: ColumnMapping) -> int:
    return _CONFIDENCE_RANK.get(m.confidence, 0)


def apply_mapping(data: bytes, mapping: MappingProposal, scope: CycleScope) -> bytes:
    """Produce a CLEAN, key-stamped owned template from the messy file + the confirmed mapping.

    Builds the owned template for `scope` (all scope rows, keys embedded — D21), then overlays the
    messy file's price/volume cells onto the matching scope rows. Each messy row is placed by
    resolving its supplier/DC/lot NAMES (via the mapped columns) to the scope identity; the price is
    written to EVERY matching (supplier, DC, lot) cell in the scope (across the cycle's TFs/items,
    since a messy supplier sheet rarely carries the TF/item grain). Returns owned-template xlsx
    bytes ready for the strict key-validated ingest.
    """

    # Build the clean owned template (scope rows pre-filled with keys + names, price cells blank).
    wb_out = build_template_workbook(scope)
    ws_out = wb_out[SHEET_BIDS]
    header_index = {
        str(ws_out.cell(row=HEADER_ROW, column=c).value).strip(): c
        for c in range(1, ws_out.max_column + 1)
        if ws_out.cell(row=HEADER_ROW, column=c).value is not None
    }

    sup_col = _col(mapping, FIELD_SUPPLIER)
    dc_col = _col(mapping, FIELD_DC)
    lot_col = _col(mapping, FIELD_LOT)
    all_in_col = _col(mapping, FIELD_ALL_IN)
    fob_col = _col(mapping, FIELD_FOB)
    vol_col = _col(mapping, FIELD_VOLUME)

    # Read the messy file's priced rows keyed by (supplier, dc, lot) NORMALIZED names.
    wb_in = load_workbook(BytesIO(data), data_only=True, read_only=True)
    ws_in = wb_in[wb_in.sheetnames[0]]
    priced: dict[tuple[str, str, str], dict[str, float]] = {}
    for r in range(mapping.header_row + 1, ws_in.max_row + 1):
        sup = _norm(ws_in.cell(row=r, column=sup_col).value) if sup_col else ""
        dc = _norm(ws_in.cell(row=r, column=dc_col).value) if dc_col else ""
        lot = _norm(ws_in.cell(row=r, column=lot_col).value) if lot_col else ""
        if not (sup and dc and lot):
            continue
        values: dict[str, float] = {}
        if all_in_col:
            _put_number(values, FIELD_ALL_IN, ws_in.cell(row=r, column=all_in_col).value)
        if fob_col:
            _put_number(values, FIELD_FOB, ws_in.cell(row=r, column=fob_col).value)
        if vol_col:
            _put_number(values, FIELD_VOLUME, ws_in.cell(row=r, column=vol_col).value)
        if values:
            priced[(sup, dc, lot)] = values
    wb_in.close()

    # Overlay onto the owned template by matching each scope row's NAMES to the messy keys.
    all_in_out = header_index[BidColumn.ALL_IN.value]
    fob_out = header_index[BidColumn.FOB.value]
    weekly_out = header_index[BidColumn.WEEKLY_VOL_OFFERED.value]
    total_out = header_index[BidColumn.TOTAL_VOL_OFFERED.value]
    sup_name_col = header_index[BidColumn.SUPPLIER.value]
    dc_name_col = header_index[BidColumn.DC.value]
    lot_name_col = header_index[BidColumn.LOT.value]

    for row in range(BODY_START_ROW, ws_out.max_row + 1):
        sup = _norm(ws_out.cell(row=row, column=sup_name_col).value)
        dc = _norm(ws_out.cell(row=row, column=dc_name_col).value)
        lot = _norm(ws_out.cell(row=row, column=lot_name_col).value)
        cell_values = priced.get((sup, dc, lot))
        if cell_values is None:
            continue
        if FIELD_ALL_IN in cell_values:
            ws_out.cell(row=row, column=all_in_out, value=cell_values[FIELD_ALL_IN])
        if FIELD_FOB in cell_values:
            ws_out.cell(row=row, column=fob_out, value=cell_values[FIELD_FOB])
        if FIELD_VOLUME in cell_values:
            ws_out.cell(row=row, column=weekly_out, value=cell_values[FIELD_VOLUME])
            ws_out.cell(row=row, column=total_out, value=cell_values[FIELD_VOLUME])

    buffer = BytesIO()
    wb_out.save(buffer)
    return buffer.getvalue()


def _col(mapping: MappingProposal, field_name: str) -> int | None:
    m = mapping.mappings.get(field_name)
    return m.column_index if m is not None else None


def _put_number(target: dict[str, float], key: str, value: object) -> None:
    """Coerce a messy cell to a float and store it under `key`; skip blanks / non-numerics."""

    if value is None or isinstance(value, bool):
        return
    if isinstance(value, (int, float, Decimal)):
        target[key] = float(value)
        return
    if isinstance(value, str):
        try:
            target[key] = float(value.strip())
        except ValueError:
            return
