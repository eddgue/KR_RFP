"""The POST-AWARD ADJUSTMENTS workbook generator (D23/D24, ADR-0014) — importable.

Renders the versioned post-award adjustment picture for a frozen award: the full version history
(v0 baseline -> vN), the current effective price per cell at a chosen version, and the per-cell
changes that version introduced. The doc carries a PROMINENT "which version" heading (PILOT step 5):
a banner title `POST-AWARD ADJUSTMENTS — {award_code}` plus a bold subtitle
`Version {N} · as of {effective_date}`.

Freeze-and-layer (ADR-0014): the frozen baseline (`awd.award_line`) is the immutable raw award; the
effective price is the baseline overlaid by the append-only versioned layers (`awd.award_adjustment`
/ `awd.award_adjustment_line`). The raw award is never overwritten; this doc reads, never mutates.

Presentation-quality (D24): reuses the shared `app.output.formatting` styling pass. Names not keys
(D23): DC / supplier / lot / timeframe are resolved to NAMES via the cycle_id on `awd.award` (ref.dc
/ ref.supplier / cyc.cycle_lot / cyc.cycle_timeframe). Deterministic: stable sort, no clock in the
body. The rendered reason/type come from the STORED rows (D28: no hardcoded reasons).
"""

from __future__ import annotations

from collections.abc import Iterable
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select, text
from sqlalchemy.orm import Session

from app.domain.awd.models import Award, AwardAdjustment, AwardAdjustmentLine, AwardLine
from app.domain.awd.service import (
    CellKey,
    VersionRow,
    award_versions,
    effective_award,
)
from app.domain.cyc.models import CycleLot, CycleTimeframe
from app.engine.formulas import price_delta
from app.output.formatting import (
    _SUBTITLE_FONT,
    DECISION_SUPPORT_STRAP,
    NUMFMT_INT,
    NUMFMT_MONEY,
    Col,
    format_table,
)

# The version-heading template (PILOT step 5) — kept as a constant so the test can assert the exact
# emitted text without re-deriving it.
VERSION_SUBTITLE = "Version {n} · as of {effective_date}"
BANNER_TITLE = "POST-AWARD ADJUSTMENTS — {award_code}"


@dataclass(frozen=True)
class _Names:
    """Resolved display names for the cells of one award (D23 — names, not keys)."""

    dc: dict[str, str]
    supplier: dict[str, str]
    lot: dict[str, str]
    tf: dict[str, str]

    def cell_label(self, key: CellKey) -> tuple[str, str, str, str]:
        """(DC name, Lot name, Supplier name, TF name) for a (dc, lot, tf, supplier) cell key."""

        dc_id, lot_id, tf_id, supplier_id = key
        return (
            self.dc.get(dc_id, dc_id[:8]),
            self.lot.get(lot_id, lot_id[:8]),
            self.supplier.get(supplier_id, supplier_id[:8]),
            self.tf.get(tf_id, tf_id[:8]),
        )


def _resolve_names(session: Session, award: Award) -> _Names:
    """Resolve DC / supplier / lot / timeframe NAMES for the award's cells (via its cycle_id).

    Lots + timeframes are cycle-scoped (read by `cycle_id` via the ORM); DCs + suppliers are global
    reference (`ref.dc` / `ref.supplier` are not ORM-mapped, so they are read by key with `text()`,
    the same pattern the scenario-workbook generator uses). All reads are keyed (D21), rendered by
    name (D23). `ref.supplier`'s name column is `canonical_name`.
    """

    dc_names: dict[str, str] = dict(
        session.execute(text("SELECT dc_id, dc_name FROM ref.dc")).all()  # type: ignore[arg-type]
    )
    sup_names: dict[str, str] = dict(
        session.execute(text("SELECT supplier_id, canonical_name FROM ref.supplier")).all()  # type: ignore[arg-type]
    )
    lot_names: dict[str, str] = dict(
        session.execute(
            select(CycleLot.lot_id, CycleLot.lot_name).where(CycleLot.cycle_id == award.cycle_id)
        ).all()  # type: ignore[arg-type]
    )
    tf_names: dict[str, str] = dict(
        session.execute(
            select(CycleTimeframe.tf_id, CycleTimeframe.tf_name).where(
                CycleTimeframe.cycle_id == award.cycle_id
            )
        ).all()  # type: ignore[arg-type]
    )
    return _Names(dc=dc_names, supplier=sup_names, lot=lot_names, tf=tf_names)


def write_post_award_adjustments_xlsx(
    session: Session,
    *,
    award_id: str,
    as_of_version: int | None = None,
    output_path: Path,
) -> Path:
    """Write the versioned post-award adjustments workbook for an award; return the path.

    Three tabs: `Versions` (full history v0->vN), `Current Effective Prices` (per cell: frozen
    baseline, effective price at this version, cumulative delta), `This Version's Changes` (the
    per-cell prior->new changes for version N). N = `as_of_version` or the latest. Every tab carries
    the explicit version heading (PILOT step 5). Names not keys (D23); deterministic.
    """

    award = session.execute(select(Award).where(Award.award_id == award_id)).scalar_one()
    names = _resolve_names(session, award)
    history = award_versions(session, award_id=award_id)

    latest_version = max((h["version_no"] for h in history), default=0)
    version_n = as_of_version if as_of_version is not None else latest_version
    version_n = min(version_n, latest_version)

    # The effective date shown in the heading = the as-of version's effective date (v0 = frozen_at).
    eff_date_by_version = {h["version_no"]: h["effective_date"] for h in history}
    heading_eff_date = eff_date_by_version.get(version_n, award.frozen_at.date())

    subtitle = VERSION_SUBTITLE.format(n=version_n, effective_date=heading_eff_date)
    banner = BANNER_TITLE.format(award_code=award.award_code)

    wb = Workbook()
    _write_versions_tab(wb, banner, subtitle, history)
    _write_effective_tab(wb, banner, subtitle, session, award_id, version_n, names)
    _write_changes_tab(wb, banner, subtitle, session, award_id, version_n, names)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def _version_banner_lines(subtitle: str) -> list[str]:
    """The bold version subtitle (PILOT step 5) above the decision-support strap."""

    return [subtitle, DECISION_SUPPORT_STRAP]


def _write_versions_tab(
    wb: Workbook,
    banner: str,
    subtitle: str,
    history: list[VersionRow],
) -> None:
    """Versions tab — the full history v0 (frozen baseline) -> vN (D24)."""

    ws = wb.active
    assert ws is not None  # noqa: S101
    ws.title = "Versions"

    columns = [
        Col("Version", 9, NUMFMT_INT),
        Col("Type", 22),
        Col("Effective Date", 16),
        Col("Reason", 46),
        Col("Recorded By", 20),
        Col("Recorded At", 20),
        Col("# Cells Changed", 14, NUMFMT_INT),
    ]
    header_row = 5
    row = header_row + 1
    for h in history:
        ws.cell(row=row, column=1, value=h["version_no"])
        ws.cell(row=row, column=2, value=h["adjustment_type"])
        ws.cell(row=row, column=3, value=h["effective_date"].strftime("%Y-%m-%d"))
        ws.cell(row=row, column=4, value=h["reason"])
        ws.cell(row=row, column=5, value=h["created_by"])
        ws.cell(row=row, column=6, value=h["created_at"].strftime("%Y-%m-%d %H:%M"))
        ws.cell(row=row, column=7, value=h["n_lines"])
        row += 1

    format_table(
        ws,
        title=banner,
        subtitle_lines=_version_banner_lines(subtitle),
        columns=columns,
        n_body_rows=len(history),
        header_row=header_row,
        add_total=False,
    )
    _bold_subtitle(ws, len(columns))


def _write_effective_tab(
    wb: Workbook,
    banner: str,
    subtitle: str,
    session: Session,
    award_id: str,
    version_n: int,
    names: _Names,
) -> None:
    """Current Effective Prices tab — per cell: frozen baseline, effective @ N, cumulative delta."""

    ws = wb.create_sheet("Current Effective Prices")

    baseline = {
        (dc_id, lot_id, tf_id, supplier_id): frozen_price
        for dc_id, lot_id, tf_id, supplier_id, frozen_price in session.execute(
            select(
                AwardLine.dc_id,
                AwardLine.lot_id,
                AwardLine.tf_id,
                AwardLine.supplier_id,
                AwardLine.frozen_price,
            ).where(AwardLine.award_id == award_id)
        ).all()
    }
    effective = effective_award(session, award_id=award_id, as_of_version=version_n)

    columns = [
        Col("DC", 22),
        Col("Lot", 22),
        Col("Supplier", 28),
        Col("Timeframe", 16),
        Col("Frozen Baseline $/case", 18, NUMFMT_MONEY),
        Col(f"Effective $/case (v{version_n})", 20, NUMFMT_MONEY),
        Col("Cumulative Δ $/case", 18, NUMFMT_MONEY),
    ]
    header_row = 5
    rows = _sorted_cells(baseline.keys(), names)
    row = header_row + 1
    for key in rows:
        dc_name, lot_name, sup_name, tf_name = names.cell_label(key)
        base = baseline.get(key, Decimal("0"))
        eff = effective.get(key, base)
        ws.cell(row=row, column=1, value=dc_name)
        ws.cell(row=row, column=2, value=lot_name)
        ws.cell(row=row, column=3, value=sup_name)
        ws.cell(row=row, column=4, value=tf_name)
        ws.cell(row=row, column=5, value=float(base))
        ws.cell(row=row, column=6, value=float(eff))
        ws.cell(row=row, column=7, value=float(price_delta(eff, base)))
        row += 1

    format_table(
        ws,
        title=banner,
        subtitle_lines=_version_banner_lines(subtitle),
        columns=columns,
        n_body_rows=len(rows),
        header_row=header_row,
        total_label="TOTAL",
        add_total=False,
    )
    _bold_subtitle(ws, len(columns))


def _write_changes_tab(
    wb: Workbook,
    banner: str,
    subtitle: str,
    session: Session,
    award_id: str,
    version_n: int,
    names: _Names,
) -> None:
    """This Version's Changes tab — the per-cell prior->new changes introduced by version N."""

    ws = wb.create_sheet("This Version's Changes")

    change_rows: list[tuple[CellKey, Decimal, Decimal, Decimal]] = []
    if version_n >= 1:
        rows = session.execute(
            select(
                AwardAdjustmentLine.dc_id,
                AwardAdjustmentLine.lot_id,
                AwardAdjustmentLine.tf_id,
                AwardAdjustmentLine.supplier_id,
                AwardAdjustmentLine.prior_price,
                AwardAdjustmentLine.new_price,
                AwardAdjustmentLine.delta,
            )
            .join(
                AwardAdjustment,
                AwardAdjustment.adjustment_id == AwardAdjustmentLine.adjustment_id,
            )
            .where(
                AwardAdjustment.award_id == award_id,
                AwardAdjustment.version_no == version_n,
            )
        ).all()
        for dc_id, lot_id, tf_id, supplier_id, prior, new, delta in rows:
            change_rows.append(((dc_id, lot_id, tf_id, supplier_id), prior, new, delta))

    by_key = {ck: (p, n, d) for ck, p, n, d in change_rows}
    ordered = _sorted_cells(by_key.keys(), names)

    columns = [
        Col("DC", 22),
        Col("Lot", 22),
        Col("Supplier", 28),
        Col("Timeframe", 16),
        Col("Prior $/case", 16, NUMFMT_MONEY),
        Col("New $/case", 16, NUMFMT_MONEY),
        Col("Δ $/case", 14, NUMFMT_MONEY),
    ]
    header_row = 5
    row = header_row + 1
    for key in ordered:
        dc_name, lot_name, sup_name, tf_name = names.cell_label(key)
        prior, new, delta = by_key[key]
        ws.cell(row=row, column=1, value=dc_name)
        ws.cell(row=row, column=2, value=lot_name)
        ws.cell(row=row, column=3, value=sup_name)
        ws.cell(row=row, column=4, value=tf_name)
        ws.cell(row=row, column=5, value=float(prior))
        ws.cell(row=row, column=6, value=float(new))
        ws.cell(row=row, column=7, value=float(delta))
        row += 1

    if not ordered:
        note_row = header_row + 1
        ws.cell(
            row=note_row,
            column=1,
            value="No per-cell price changes recorded for this version (v0 = frozen baseline).",
        )

    format_table(
        ws,
        title=banner,
        subtitle_lines=_version_banner_lines(subtitle),
        columns=columns,
        n_body_rows=len(ordered),
        header_row=header_row,
        add_total=False,
    )
    _bold_subtitle(ws, len(columns))


def _sorted_cells(keys: Iterable[CellKey], names: _Names) -> list[CellKey]:
    """Deterministic cell order by resolved (DC, Lot, Supplier, TF) names."""

    return sorted(keys, key=lambda k: names.cell_label(k))


def _bold_subtitle(ws: Worksheet, span: int) -> None:
    """Make the version subtitle (row 2 of the banner) BOLD so the 'which version' line stands out.

    `format_table`/`_title_block` write the subtitle lines in the italic subtitle font starting at
    banner row 2; row 2 is the `Version N · as of ...` line (PILOT step 5). Re-style it bold.
    """

    from openpyxl.styles import Font

    cell = ws.cell(row=2, column=1)
    base = _SUBTITLE_FONT
    cell.font = Font(bold=True, italic=True, color=base.color, size=11)
