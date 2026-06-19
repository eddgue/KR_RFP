"""Shared presentation formatting for the xlsx output surfaces (D24).

A reusable styling pass for every client-openable workbook: a titled header block, a
bold white-on-color header row, sensible column widths, $/% number formats, thin
borders, freeze panes under the header, a TOTAL/summary row, and an AutoFilter. NOT a
raw CSV-like dump (D24).

MOVED out of the demo so the scenario-workbook generator (and any other feature that
produces a client workbook) imports the presentation language from ONE place.
"""

from __future__ import annotations

from dataclasses import dataclass

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# D24 — PRESENTATION FORMATTING. A reusable styling pass for every xlsx output:
#   a titled header block, a bold white-on-color header row, sensible column
#   widths, $/% number formats, thin borders, freeze panes under the header, a
#   TOTAL/summary row, and an AutoFilter. NOT a raw CSV-like dump (D24).
# ---------------------------------------------------------------------------
NUMFMT_MONEY = "$#,##0.00"
NUMFMT_PCT = "0.0%"  # applied to a FRACTION (0.05 -> 5.0%)
NUMFMT_PCT_WHOLE = "0.0%"
NUMFMT_INT = "#,##0"

# Brand palette (decision-support neutral; readable on a projector).
_HEADER_FILL = PatternFill("solid", fgColor="1F3864")  # deep navy
_TITLE_FILL = PatternFill("solid", fgColor="2E5496")  # lighter navy band
_TOTAL_FILL = PatternFill("solid", fgColor="D9E1F2")  # pale blue summary band
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_TITLE_FONT = Font(bold=True, color="FFFFFF", size=13)
_SUBTITLE_FONT = Font(italic=True, color="FFFFFF", size=9)
_TOTAL_FONT = Font(bold=True, color="1F3864")
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)
_WRAP_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Comparison-tool accent fills (the alignment surfaces highlight the picture the team debates).
_MIN_FILL = PatternFill("solid", fgColor="C6EFCE")  # best/min price per row (green)
_MIN_FONT = Font(bold=True, color="006100")
_BENCH_FILL = PatternFill("solid", fgColor="FCE4D6")  # Scenario A benchmark row (peach)
_REC_FILL = PatternFill("solid", fgColor="DDEBF7")  # Scenario B recommendation row (blue)
_REC_PICK_FILL = PatternFill("solid", fgColor="BDD7EE")  # the recommended supplier cell (blue)
_INCUMBENT_FILL = PatternFill("solid", fgColor="FFF2CC")  # incumbent marker (amber)
_BREACH_FILL = PatternFill("solid", fgColor="FFC7CE")  # cap-breach (red)
_BREACH_FONT = Font(bold=True, color="9C0006")

# The standard provenance strap every presentation surface carries (ADR-0006).
DECISION_SUPPORT_STRAP = "DECISION-SUPPORT — recommends, does not assert"


@dataclass(frozen=True)
class Col:
    """One column in a formatted table: header text, width, and a number format."""

    header: str
    width: int = 16
    number_format: str | None = None  # None -> text/general
    total: str = ""  # "sum" -> SUM over the body; "" -> no total cell


def _title_block(
    ws: Worksheet,
    *,
    title: str,
    subtitle_lines: list[str],
    span: int,
    start_row: int = 1,
) -> int:
    """Write a titled header block across `span` columns; return the next free row.

    Row 1 = cycle/strategy title (large, white-on-navy). Following rows = subtitle
    lines (date, strategy, the decision-support strap). The whole block is merged
    across the table width so it reads as a banner, not a stray cell (D24).
    """

    last_col = get_column_letter(span)
    row = start_row
    ws.merge_cells(f"A{row}:{last_col}{row}")
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = _TITLE_FONT
    cell.fill = _TITLE_FILL
    cell.alignment = _LEFT
    ws.row_dimensions[row].height = 22
    for c in range(1, span + 1):
        ws.cell(row=row, column=c).fill = _TITLE_FILL
    row += 1
    for line in subtitle_lines:
        ws.merge_cells(f"A{row}:{last_col}{row}")
        cell = ws.cell(row=row, column=1, value=line)
        cell.font = _SUBTITLE_FONT
        cell.fill = _TITLE_FILL
        cell.alignment = _LEFT
        for c in range(1, span + 1):
            ws.cell(row=row, column=c).fill = _TITLE_FILL
        row += 1
    return row + 1  # one blank spacer row below the banner


def format_table(
    ws: Worksheet,
    *,
    title: str,
    subtitle_lines: list[str],
    columns: list[Col],
    n_body_rows: int,
    header_row: int | None = None,
    total_label_col: int = 1,
    total_label: str = "TOTAL",
    add_total: bool = True,
    add_autofilter: bool = True,
) -> dict[str, int]:
    """Apply the full D24 presentation pass to a sheet whose body is ALREADY written.

    Expects the caller to have written the data rows starting at `header_row + 1`.
    Writes the title banner (above the header), styles the header row bold
    white-on-color, sets column widths + number formats, draws thin borders over
    the table, freezes panes under the header, appends a styled TOTAL row that SUMs
    the money/count columns, and turns on an AutoFilter. Returns key row indices.
    """

    span = len(columns)
    # Title banner occupies rows 1..(header_row-1); caller may pass header_row, else
    # we compute it from the banner height.
    if header_row is None:
        next_row = _title_block(
            ws, title=title, subtitle_lines=subtitle_lines, span=span
        )
        header_row = next_row
    else:
        _title_block(
            ws, title=title, subtitle_lines=subtitle_lines, span=span, start_row=1
        )

    # Header row — bold white-on-color, centered, wrapped, bordered.
    for ci, col in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=ci, value=col.header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _WRAP_CENTER
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(ci)].width = col.width
    ws.row_dimensions[header_row].height = 30

    body_start = header_row + 1
    body_end = body_start + n_body_rows - 1

    # Body — number formats + borders + alignment.
    for ci, col in enumerate(columns, start=1):
        for r in range(body_start, body_end + 1):
            cell = ws.cell(row=r, column=ci)
            cell.border = _BORDER
            if col.number_format:
                cell.number_format = col.number_format
                cell.alignment = _CENTER
            else:
                cell.alignment = _LEFT

    # TOTAL / summary row.
    total_row = None
    if add_total and n_body_rows > 0:
        total_row = body_end + 1
        ws.cell(row=total_row, column=total_label_col, value=total_label)
        for ci, col in enumerate(columns, start=1):
            cell = ws.cell(row=total_row, column=ci)
            cell.fill = _TOTAL_FILL
            cell.font = _TOTAL_FONT
            cell.border = _BORDER
            cell.alignment = _CENTER if col.number_format else _LEFT
            if col.total == "sum" and n_body_rows > 0:
                letter = get_column_letter(ci)
                cell.value = f"=SUM({letter}{body_start}:{letter}{body_end})"
                cell.number_format = col.number_format or NUMFMT_INT

    # Freeze panes directly under the header (title + header stay on screen).
    ws.freeze_panes = ws.cell(row=body_start, column=1)

    # AutoFilter across the header + body (not the title banner, not the total).
    if add_autofilter and n_body_rows > 0:
        last_col = get_column_letter(span)
        ws.auto_filter.ref = f"A{header_row}:{last_col}{body_end}"

    return {
        "header_row": header_row,
        "body_start": body_start,
        "body_end": body_end,
        "total_row": total_row if total_row is not None else body_end,
    }
