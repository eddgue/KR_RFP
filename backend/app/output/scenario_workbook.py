"""The ALIGNMENT / COMPARISON Scenario Workbook generator (D26/D27) — importable.

MOVED out of `demo/run_cycle_demo.py` so any feature (not just the demo) can produce the
18-tab scenario workbook from a sealed engine run. The generator reads ONLY the governed
sealed records (`eng.*`, `bid.bid_line`, ...) plus a `CycleView` (the resolved cycle scope);
it never imports from the demo. Build the `CycleView` with the demo's `seed_cycle` (synthetic)
or with `app.cycle.loader.load_cycle` (a REAL persisted cycle).

The `award` it consumes is described structurally (`AwardView` / `AwardCellView` Protocols) so
the demo keeps its own concrete `SelectedAward` / `AwardedCell` award-selection types without
this module depending on them.

SYNTHETIC-friendly: every readable cell renders RESOLVED NAMES (D23); a trailing key column is
kept for traceability. Decision-support only — recommends, never asserts (ADR-0006).
"""

from __future__ import annotations

from collections import defaultdict
from collections.abc import Sequence
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Protocol

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule, FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.domain.eng.models import AnalysisRun, AnalysisScenario
from app.engine.formulas import (
    awarded_cases,
    construct_price_from_parts,
    delta_vs_historical,
    line_spend,
    savings_dollars,
    savings_fraction,
)
from app.engine.interface import EngineConfig
from app.output.capacity_check import evaluate_capacity, load_active_capacity
from app.output.formatting import (
    _BENCH_FILL,
    _BORDER,
    _BREACH_FILL,
    _BREACH_FONT,
    _CENTER,
    _HEADER_FILL,
    _HEADER_FONT,
    _INCUMBENT_FILL,
    _LEFT,
    _MIN_FILL,
    _MIN_FONT,
    _REC_FILL,
    _REC_PICK_FILL,
    _SUBTITLE_FONT,
    _TITLE_FILL,
    _TITLE_FONT,
    _TOTAL_FILL,
    _TOTAL_FONT,
    _WRAP_CENTER,
    DECISION_SUPPORT_STRAP,
    NUMFMT_INT,
    NUMFMT_MONEY,
    NUMFMT_PCT,
    Col,
    _title_block,
    format_table,
)
from app.output.synthetic import (
    FRESHNESS_WATCH_DAYS,
    LOT_PRODUCT_TYPE,
    WEEKS_PER_TF,
    _dc_region,
)
from app.output.types import CycleView

# Default output location (kept for the demo's existing path); callers may override via output_path.
OUTPUT_DIR = Path(__file__).resolve().parents[2] / "demo" / "output"


class AwardCellView(Protocol):
    """Structural view of one frozen award cell the generator reads (keys + economics).

    Read-only properties so a frozen dataclass (e.g. the demo's `AwardedCell`) satisfies it.
    """

    @property
    def dc_id(self) -> str: ...
    @property
    def lot_id(self) -> str: ...
    @property
    def item_id(self) -> str: ...
    @property
    def tf_id(self) -> str: ...
    @property
    def supplier_id(self) -> str: ...
    @property
    def volume_share(self) -> Decimal: ...
    @property
    def awarded_price(self) -> Decimal: ...
    @property
    def period_cases(self) -> Decimal: ...
    @property
    def routing_baseline(self) -> Decimal: ...


class AwardView(Protocol):
    """Structural view of the promoted award the generator reads (selected scenario + cells).

    Read-only properties + a covariant `Sequence` of cells so a frozen dataclass with a
    `tuple[AwardedCell, ...]` (e.g. the demo's `SelectedAward`) satisfies it.
    """

    @property
    def scenario_code(self) -> str: ...
    @property
    def scenario_label(self) -> str: ...
    @property
    def cells(self) -> Sequence[AwardCellView]: ...


# ---------------------------------------------------------------------------
# 8) ALIGNMENT / COMPARISON SCENARIO WORKBOOK (D26/D27) — SCENARIO_WORKBOOK.xlsx
#
# The team-alignment tool — the comparison surfaces buyers/category/sourcing work through to
# DECIDE in the alignment call (D26), with MANIPULABLE data (D27 — pivot/drill/filter, depth on
# demand, not fixed tables). Generated from the sealed records:
#   * Summary             — the headline: cycle/strategy + the A-vs-B alignment call + a guide.
#   * Scenario Comparison — the lenses A-G SIDE BY SIDE (which lens): spend, Δ between lenses,
#                           savings vs baseline / vs STLY, supplier count, cap-breaches, # cells;
#                           A (benchmark) + B (recommendation) highlighted; + a LIVE Custom row
#                           (formulas off the Custom Scenario tab, read side by side vs A-G, D27).
#                           Below it an EXPANDABLE DRILL (outline grouping, summaryBelow=False):
#                           each scenario TOTAL → per-DC → per-supplier, opens COLLAPSED to totals,
#                           expand for depth on demand (D27). Then a PER-DC MATRIX (DCs down,
#                           scenarios across → spend + supplier mix) so the team sees WHERE lenses
#                           differ.
#   * Supplier Comparison — THE CENTERPIECE (the v3 FOB comparison). One row per (DC x lot x item
#                           x TF) cell with demand, baseline + incumbent $/case, then ONE COLUMN
#                           PER SUPPLIER (all-in $/case, blank if no bid). The MIN per row is
#                           highlighted (best price); incumbent + recommended pick flagged. Min
#                           $/case, Recommended supplier + RecScore, then a second block = cost
#                           impact vs baseline % per supplier (the premium each carries). The
#                           competitive picture the team scans to compare suppliers per cell.
#   * Custom Scenario     — THE INTERACTIVE TAB (D25). One row per cell with a data-validation
#                           DROPDOWN of the eligible suppliers (by NAME) and LIVE formulas: the
#                           chosen supplier's $/case (SUMIFS over the hidden `_Prices` grid), vs Min
#                           / vs Incumbent / vs Baseline %, line spend = price x volume; everything
#                           rolls to a TOTAL spend + savings-vs-baseline + a per-DC cap-breach flag.
#                           Pre-filled with Scenario B; changing any dropdown recomputes live (and
#                           drives the live Custom row + column on Scenario Comparison).
#   * Data (pivot me)     — THE FLAT MANIPULABLE DATASET (D27). One row per (scenario × DC × lot ×
#                           item × TF × awarded-supplier) with every metric, as a REAL Excel Table
#                           (ListObject) with AutoFilter — drop a native PivotTable / filter / slice
#                           on it to cut the data any way. Rich but neat: detail lives here.
#   * Scored Bids         — per bid: names + the 5 factors + RecScore + eligible?.
#   * _Prices (hidden)    — the supplier x cell price grid + per-cell Min/Incumbent/Baseline
#                           reference grid the live formulas reference.
# ---------------------------------------------------------------------------
@dataclass(frozen=True)
class CellInfo:
    """One (DC x lot x item x TF) allocation cell, resolved to names + the competitive picture.

    The unit the Supplier Comparison + Custom Scenario tabs are built on: every eligible supplier's
    $/case side by side, the recommended pick + its RecScore, the incumbent, and the baseline — the
    v3 FOB-comparison the team debates per cell (D26).
    """

    dc_id: str
    lot_id: str
    item_id: str
    tf_id: str
    cell_key: str  # a stable text key "DCxx|LOT-xx|TFxx" the _Prices grid is keyed on
    dc_name: str
    lot_name: str
    item_name: str
    tf_name: str
    volume: Decimal  # projected period cases for the cell
    baseline_price: Decimal  # incumbent routing baseline $/case
    incumbent_name: str  # the incumbent supplier NAME for this DC x lot (reference point)
    eligible_suppliers: list[str]  # supplier NAMES eligible for this cell (dropdown list)
    rec_supplier: str  # Scenario B recommended supplier NAME (the pre-filled pick)
    rec_score: Decimal  # the recommended pick's RecScore (0-100)
    price_by_supplier: dict[str, Decimal]  # supplier NAME -> $/case for this cell
    score_by_supplier: dict[str, Decimal]  # supplier NAME -> RecScore for this cell
    transit_by_supplier: dict[str, int | None]  # supplier NAME -> real lane transit (None=absent)
    rec_type: str  # the engine's authoritative B reason for this cell (§5; "" if none)


def _transit_by_lane(session: Session, cycle_id: str) -> dict[tuple[str, str], int]:
    """Real supplier→DC transit days from the bids (a stable lane property, round-independent).

    Reads the `transit_days` column suppliers fill on the bid template. Returns ONLY lanes that
    carry a value — a lane with no submitted transit is absent here, so the analysis shows no
    transit for it (no synthetic proxy; transit is surfaced only when it is real data).
    """

    # ACTIVE rows only (`is_scoreable`): a superseded submission's transit must not win the MAX.
    # Fan-out is a non-issue here — the per-period rows replicate the transit, so MAX is stable.
    rows = session.execute(
        text(
            "SELECT supplier_id, dc_id, MAX(transit_days) FROM bid.bid_line "
            "WHERE cycle_id = :cyc AND transit_days IS NOT NULL AND is_scoreable = true "
            "GROUP BY supplier_id, dc_id"
        ),
        {"cyc": cycle_id},
    ).all()
    return {(s, d): int(t) for s, d, t in rows if t is not None}


def _line_price(
    all_in: object, fob: object, delivery: object, vegcool: object, lot_discount: object
) -> Decimal | None:
    """The canonical §7 price for a persisted bid_line's component columns (E-39 formula).

    Mirrors what the engine actually scored: All-In if present, else FOB + delivery + vegcool −
    lot_discount. Reading raw `submitted_all_in_case` alone (as the workbook used to) DROPPED
    component-basis bids (All-In NULL, FOB+surcharges set) the engine scored — so they vanished
    from the price grids / market stats / coverage / FOB tabs while awards used the constructed
    price. Routing every price read through this keeps the workbook consistent with the engine.
    """

    return construct_price_from_parts(
        Decimal(str(all_in)) if all_in is not None else None,
        Decimal(str(fob)) if fob is not None else None,
        Decimal(str(delivery)) if delivery is not None else Decimal("0"),
        Decimal(str(vegcool)) if vegcool is not None else Decimal("0"),
        Decimal(str(lot_discount)) if lot_discount is not None else Decimal("0"),
    )


def _gather_cells(
    session: Session,
    seeded: CycleView,
    analysis_run_id: str,
    final_round_id: str,
    award: AwardView,
) -> list[CellInfo]:
    """Resolve every allocation cell to names + the competitive picture (the v3 FOB comparison).

    Prices come from the FINAL-round `bid.bid_line` rows (the priced reality the engine scored);
    eligibility + RecScore come from `eng.bid_score`; the B pick comes from the selected award; the
    incumbent comes from the seed. All by KEY (D21); all rendered by NAME (D23). The result is what
    the Supplier Comparison + Custom Scenario tabs are built on.
    """

    dc_name = {dc.id: dc.name for dc in seeded.dcs}
    lot_name = {lot.id: lot.name for lot in seeded.lots}
    sup_name = {sup.id: sup.name for sup in seeded.suppliers}
    tf_name = {tf.id: tf.name for tf in seeded.tfs}
    dc_code = {dc.id: dc.code for dc in seeded.dcs}
    lot_code = {lot.id: lot.code for lot in seeded.lots}
    tf_code = {tf.id: tf.code for tf in seeded.tfs}
    item_for_lot = {seeded.lots[i].id: seeded.items[i] for i in range(len(seeded.lots))}

    # Final-round prices per (supplier, dc, lot, tf) from the persisted bid lines.
    # OPTION B (INTAKE §1a): bids are STORED flat at the 13 fiscal periods (one row per period in a
    # timeframe's span, identical payload). The competitive grid is TIMEFRAME-grain, so collapse the
    # (≤13×) period rows to ONE price per (supplier, dc, lot, tf) with DISTINCT ON — the fanned rows
    # are identical, so any one returns the same price (a pure tf-grain row, period NULL, is its own
    # representative). This keeps the grid byte-identical to the pre-fan-out timeframe-grain read.
    # Only ACTIVE (`is_scoreable`) rows are read — mirrors the engine's `_read_bid_lines` filter,
    # so a superseded re-submission (prior lines flipped non-scoreable) can't leak a stale price.
    price_rows = session.execute(
        text(
            "SELECT DISTINCT ON (supplier_id, dc_id, lot_id, tf_id) "
            "supplier_id, dc_id, lot_id, tf_id, submitted_all_in_case, "
            "fob_case, delivery_surcharge_case, vegcool_surcharge_case, lot_discount_case "
            "FROM bid.bid_line WHERE cycle_id = :cyc AND round_id = :rnd "
            "AND is_scoreable = true "
            "ORDER BY supplier_id, dc_id, lot_id, tf_id, fiscal_period_id NULLS LAST, bid_line_id"
        ),
        {"cyc": seeded.cycle_id, "rnd": final_round_id},
    ).all()
    price_by: dict[tuple[str, str, str, str], Decimal] = {}
    for sup_id, dc_id, lot_id, tf_id, all_in, fob, deliv, veg, disc in price_rows:
        price = _line_price(all_in, fob, deliv, veg, disc)
        if price is not None:
            price_by[(sup_id, dc_id, lot_id, tf_id)] = price

    # Eligibility + RecScore per (supplier, dc, lot, tf) from the sealed scores.
    score_rows = session.execute(
        text(
            "SELECT supplier_id, dc_id, lot_id, tf_id, is_eligible, rec_score "
            "FROM eng.bid_score WHERE analysis_run_id = :run"
        ),
        {"run": analysis_run_id},
    ).all()
    eligible_by: dict[tuple[str, str, str, str], bool] = {}
    recscore_by: dict[tuple[str, str, str, str], Decimal] = {}
    for s, d, lo, t, e, rs in score_rows:
        eligible_by[(s, d, lo, t)] = bool(e)
        recscore_by[(s, d, lo, t)] = Decimal(str(rs))

    # The Scenario B recommended supplier per cell (the pre-filled dropdown pick).
    rec_by_cell = {
        (c.dc_id, c.lot_id, c.tf_id): sup_name.get(c.supplier_id, c.supplier_id[:6])
        for c in award.cells
    }
    # The engine's AUTHORITATIVE per-cell reason (RecType) for the B pick — sealed on the award
    # (D28: outputs render the engine's reason, never a hardcoded clause). Keyed (dc, lot, tf).
    rec_type_by_cell: dict[tuple[str, str, str], str] = {
        (r[0], r[1], r[2]): (r[3] or "")
        for r in session.execute(
            text(
                "SELECT a.dc_id, a.lot_id, a.tf_id, a.rec_type "
                "FROM eng.analysis_scenario_award a "
                "JOIN eng.analysis_scenario s ON s.analysis_scenario_id = a.analysis_scenario_id "
                "WHERE s.analysis_run_id = :run AND s.scenario_code = 'B'"
            ),
            {"run": analysis_run_id},
        ).all()
    }
    # Incumbent supplier name per (dc, lot) — a reference point throughout (D26).
    incumbent_name_by = {
        (dc_id, lot_id): sup_name.get(sup_id, sup_id[:6])
        for (dc_id, lot_id), sup_id in seeded.incumbent_by_dc_lot.items()
    }

    transit_lane = _transit_by_lane(session, seeded.cycle_id)

    cells: list[CellInfo] = []
    for c in award.cells:
        key_t = (c.dc_id, c.lot_id, c.tf_id)
        # Suppliers that BOTH priced this cell AND scored eligible -> the valid dropdown.
        price_map: dict[str, Decimal] = {}
        score_map: dict[str, Decimal] = {}
        transit_map: dict[str, int | None] = {}
        eligible_names: list[str] = []
        for sup in seeded.suppliers:
            p = price_by.get((sup.id, c.dc_id, c.lot_id, c.tf_id))
            if p is None:
                continue
            name = sup_name.get(sup.id, sup.id[:6])
            price_map[name] = p
            transit_map[name] = transit_lane.get((sup.id, c.dc_id))
            sc = recscore_by.get((sup.id, c.dc_id, c.lot_id, c.tf_id))
            if sc is not None:
                score_map[name] = sc
            if eligible_by.get((sup.id, c.dc_id, c.lot_id, c.tf_id), False):
                eligible_names.append(name)
        # Always include the recommended pick even if the gate marked it ineligible elsewhere.
        rec = rec_by_cell.get(key_t, "")
        if rec and rec not in eligible_names and rec in price_map:
            eligible_names.append(rec)
        if not eligible_names:  # fall back to anyone who priced, so the dropdown is never empty
            eligible_names = sorted(price_map)
        eligible_names = sorted(dict.fromkeys(eligible_names))

        item = item_for_lot.get(c.lot_id)
        cell_key = (
            f"{dc_code.get(c.dc_id, c.dc_id[:4])}|{lot_code.get(c.lot_id, c.lot_id[:4])}|"
            f"{tf_code.get(c.tf_id, c.tf_id[:4])}"
        )
        rec_name = rec or (eligible_names[0] if eligible_names else "")
        cells.append(
            CellInfo(
                dc_id=c.dc_id,
                lot_id=c.lot_id,
                item_id=item.id if item else c.lot_id,
                tf_id=c.tf_id,
                cell_key=cell_key,
                dc_name=dc_name.get(c.dc_id, c.dc_id[:6]),
                lot_name=lot_name.get(c.lot_id, c.lot_id[:6]),
                item_name=item.name if item else "",
                tf_name=tf_name.get(c.tf_id, c.tf_id[:6]),
                volume=c.period_cases,
                baseline_price=c.routing_baseline,
                incumbent_name=incumbent_name_by.get((c.dc_id, c.lot_id), ""),
                eligible_suppliers=eligible_names,
                rec_supplier=rec_name,
                rec_score=score_map.get(rec_name, Decimal("0")),
                price_by_supplier=price_map,
                score_by_supplier=score_map,
                transit_by_supplier=transit_map,
                rec_type=rec_type_by_cell.get(key_t, ""),
            )
        )
    # Stable, readable order.
    cells.sort(key=lambda c: (c.dc_name, c.lot_name, c.tf_name))
    return cells


# ---------------------------------------------------------------------------
# Scenario-comparison data — the "which lens" side-by-side, per DC + total (D26).
# ---------------------------------------------------------------------------
@dataclass(frozen=True)
class ScenarioRollup:
    """One lens rolled up for the side-by-side: spend, deltas, savings, counts, breaches (D26)."""

    code: str
    label: str
    total_spend: Decimal
    delta_vs_a: Decimal  # Δ vs A (lowest-cost benchmark)
    savings_vs_baseline_frac: Decimal  # vs incumbent-routing baseline
    savings_vs_stly_frac: Decimal  # vs synthesized prior-year baseline (STLY proxy)
    n_suppliers: int
    n_cap_breaches: int
    n_cells: int
    spend_by_dc: dict[str, Decimal]  # dc_name -> lens spend at that DC
    suppliers_by_dc: dict[str, str]  # dc_name -> recommended supplier names at that DC


# Synthesized STLY (Same-Time-Last-Year) uplift: with no STLY feed we model last year's actual-paid
# as a small uplift over this cycle's incumbent-routing baseline, so "Savings vs STLY" is a clearly-
# labelled SYNTHETIC reference, not a real feed (D11 actual-paid baseline; D26 reference points).
_STLY_UPLIFT = Decimal("1.04")  # prior-year actual-paid modeled ~4% above this year's baseline


def _gather_scenario_rollups(
    session: Session,
    seeded: CycleView,
    scenarios: list[AnalysisScenario],
    analysis_run_id: str,
) -> tuple[list[ScenarioRollup], Decimal, Decimal]:
    """Roll every lens up to the side-by-side comparison + per-DC matrix (D26).

    Returns (rollups A-G ordered, baseline total spend, STLY total spend). Spend, supplier
    count, cap-breaches and per-DC spend/supplier mix all come from the persisted
    `eng.analysis_scenario_award` rows; the baseline + STLY totals are derived from the seeded
    incumbent routing (STLY = a clearly-labelled synthetic prior-year uplift).
    """

    dc_name = {dc.id: dc.name for dc in seeded.dcs}
    sup_name = {sup.id: sup.name for sup in seeded.suppliers}

    # Volume per (dc, lot, tf) cell to weight award spend (price * cases * share).
    vol_by_cell = dict(seeded.period_cases_by_cell)

    # Baseline (incumbent routing) total spend across all cells, and the STLY proxy total.
    baseline_total = Decimal("0")
    for (dc_id, lot_id, _tf_id), cases in vol_by_cell.items():
        baseline_total += line_spend(
            seeded.incumbent_routing.get((dc_id, lot_id), Decimal("0")), cases
        )
    stly_total = (baseline_total * _STLY_UPLIFT).quantize(Decimal("0.01"))

    rows = session.execute(
        text(
            "SELECT s.scenario_code, a.dc_id, a.lot_id, a.tf_id, a.supplier_id, "
            "a.volume_share, a.awarded_price, a.cap_breach_flag "
            "FROM eng.analysis_scenario_award a "
            "JOIN eng.analysis_scenario s ON s.analysis_scenario_id = a.analysis_scenario_id "
            "WHERE s.analysis_run_id = :run"
        ),
        {"run": analysis_run_id},
    ).all()

    # Accumulate per scenario: spend, per-DC spend, per-DC supplier set, suppliers, breaches, cells.
    spend_by: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    dc_spend_by: dict[str, dict[str, Decimal]] = defaultdict(
        lambda: defaultdict(lambda: Decimal("0"))
    )
    dc_sups_by: dict[str, dict[str, set[str]]] = defaultdict(lambda: defaultdict(set))
    sups_by: dict[str, set[str]] = defaultdict(set)
    breaches_by: dict[str, int] = defaultdict(int)
    cells_by: dict[str, set[tuple[str, str, str]]] = defaultdict(set)
    for code, dc_id, lot_id, tf_id, supplier_id, share, price, breach in rows:
        cases = vol_by_cell.get((dc_id, lot_id, tf_id), Decimal("1"))
        line = line_spend(Decimal(str(price)), awarded_cases(cases, Decimal(str(share))))
        dcn = dc_name.get(dc_id, dc_id[:6])
        name = sup_name.get(supplier_id, supplier_id[:6])
        spend_by[code] += line
        dc_spend_by[code][dcn] += line
        dc_sups_by[code][dcn].add(name)
        sups_by[code].add(name)
        cells_by[code].add((dc_id, lot_id, tf_id))
        if breach:
            breaches_by[code] += 1

    a_spend = spend_by.get("A", Decimal("0"))
    rollups: list[ScenarioRollup] = []
    for s in sorted(scenarios, key=lambda x: x.scenario_code):
        code = s.scenario_code
        spend = spend_by.get(code, Decimal("0"))
        sv_base = savings_fraction(baseline_total, spend)
        sv_stly = savings_fraction(stly_total, spend)
        rollups.append(
            ScenarioRollup(
                code=code,
                label=s.label,
                total_spend=spend,
                delta_vs_a=spend - a_spend,
                savings_vs_baseline_frac=sv_base,
                savings_vs_stly_frac=sv_stly,
                n_suppliers=len(sups_by.get(code, set())),
                n_cap_breaches=breaches_by.get(code, 0),
                n_cells=len(cells_by.get(code, set())),
                spend_by_dc=dict(dc_spend_by.get(code, {})),
                suppliers_by_dc={
                    dcn: ", ".join(sorted(s for s in names if s))
                    for dcn, names in dc_sups_by.get(code, {}).items()
                },
            )
        )
    return rollups, baseline_total, stly_total


# ---------------------------------------------------------------------------
# D27 — DRILL-DOWN + FLAT DATASET precompute. The "data is manipulable" core:
# every scenario's awards resolved to one fully-detailed row per
# (scenario × DC × lot × item × TF × awarded-supplier), with all metrics
# (volume, $/case, spend, savings vs baseline, the 5 scores, premium, flags).
# This single sealed-record set feeds BOTH the expandable scenario→DC→supplier
# pivot (rolled up two ways) AND the flat `Data (pivot me)` Excel Table.
# ---------------------------------------------------------------------------
@dataclass(frozen=True)
class AwardDetail:
    """One fully-resolved award line for the drill pivot + the flat pivot-me dataset (D27)."""

    scenario_code: str
    scenario_label: str
    dc_name: str
    lot_name: str
    item_name: str
    tf_name: str
    supplier_name: str
    incumbent_name: str
    is_incumbent: bool
    volume: Decimal  # awarded cases = projected period cases × volume_share
    volume_share: Decimal
    price: Decimal  # awarded $/case
    spend: Decimal  # price × awarded cases
    baseline_price: Decimal
    baseline_spend: Decimal  # baseline $/case × awarded cases
    savings_vs_baseline: Decimal  # baseline_spend − spend ($)
    premium_vs_baseline_frac: Decimal  # (price − baseline) / baseline
    price_score: Decimal
    coverage_score: Decimal
    hist_score: Decimal
    zrisk_score: Decimal
    continuity_score: Decimal
    rec_score: Decimal
    cap_breach: bool
    is_fallback: bool
    transit_days: int | None  # real lane transit supplier→DC (None if not supplied)
    relationship: str  # "Preserve" (incumbent kept) | "Create" (new supplier won)


def _gather_award_details(
    session: Session,
    seeded: CycleView,
    scenarios: list[AnalysisScenario],
    analysis_run_id: str,
) -> list[AwardDetail]:
    """Resolve every A-G award row to a fully-detailed line (names + metrics + 5 scores + flags).

    One row per (scenario × DC × lot × item × TF × awarded-supplier). The award metrics come from
    `eng.analysis_scenario_award`; the five factor scores + RecScore are joined per cell+supplier
    from `eng.bid_score`; volume from the seeded projected cases × the award's volume_share; the
    baseline from the seeded incumbent routing. All by KEY (D21), all rendered by NAME (D23). This
    one set drives the drill pivot AND the flat pivot-me table.
    """

    dc_name = {dc.id: dc.name for dc in seeded.dcs}
    lot_name = {lot.id: lot.name for lot in seeded.lots}
    sup_name = {sup.id: sup.name for sup in seeded.suppliers}
    tf_name = {tf.id: tf.name for tf in seeded.tfs}
    item_for_lot = {seeded.lots[i].id: seeded.items[i] for i in range(len(seeded.lots))}
    incumbent_name_by = {
        (dc_id, lot_id): sup_name.get(sup_id, sup_id[:6])
        for (dc_id, lot_id), sup_id in seeded.incumbent_by_dc_lot.items()
    }
    vol_by_cell = dict(seeded.period_cases_by_cell)
    label_by_code = {s.scenario_code: s.label for s in scenarios}

    # Five-factor scores per (supplier, dc, lot, tf) for the run (joined onto each award).
    score_rows = session.execute(
        text(
            "SELECT supplier_id, dc_id, lot_id, tf_id, price_score, coverage_score, "
            "hist_score, zrisk_score, continuity_score, rec_score "
            "FROM eng.bid_score WHERE analysis_run_id = :run"
        ),
        {"run": analysis_run_id},
    ).all()
    score_by: dict[tuple[str, str, str, str], tuple[Decimal, ...]] = {}
    for s, d, lo, t, ps, cov, hist, z, cont, rec in score_rows:
        score_by[(s, d, lo, t)] = tuple(Decimal(str(v)) for v in (ps, cov, hist, z, cont, rec))

    award_rows = session.execute(
        text(
            "SELECT s.scenario_code, a.dc_id, a.lot_id, a.tf_id, a.supplier_id, "
            "a.volume_share, a.awarded_price, a.cap_breach_flag, a.is_fallback "
            "FROM eng.analysis_scenario_award a "
            "JOIN eng.analysis_scenario s ON s.analysis_scenario_id = a.analysis_scenario_id "
            "WHERE s.analysis_run_id = :run"
        ),
        {"run": analysis_run_id},
    ).all()

    transit_lane = _transit_by_lane(session, seeded.cycle_id)

    details: list[AwardDetail] = []
    for code, dc_id, lot_id, tf_id, sup_id, share, price, breach, fallback in award_rows:
        share_d = Decimal(str(share))
        price_d = Decimal(str(price))
        cases = awarded_cases(vol_by_cell.get((dc_id, lot_id, tf_id), Decimal("0")), share_d)
        baseline = seeded.incumbent_routing.get((dc_id, lot_id), Decimal("0"))
        spend = line_spend(price_d, cases)
        baseline_spend = line_spend(baseline, cases)
        premium = delta_vs_historical(price_d, baseline) or Decimal("0")
        scores = score_by.get((sup_id, dc_id, lot_id, tf_id), tuple(Decimal("0") for _ in range(6)))
        item = item_for_lot.get(lot_id)
        sup_disp = sup_name.get(sup_id, sup_id[:6])
        inc_disp = incumbent_name_by.get((dc_id, lot_id), "")
        is_inc = sup_disp == inc_disp
        transit = transit_lane.get((sup_id, dc_id))
        details.append(
            AwardDetail(
                scenario_code=code,
                scenario_label=label_by_code.get(code, code),
                dc_name=dc_name.get(dc_id, dc_id[:6]),
                lot_name=lot_name.get(lot_id, lot_id[:6]),
                item_name=item.name if item else "",
                tf_name=tf_name.get(tf_id, tf_id[:6]),
                supplier_name=sup_disp,
                incumbent_name=inc_disp,
                is_incumbent=(sup_disp == inc_disp),
                volume=cases,
                volume_share=share_d,
                price=price_d,
                spend=spend,
                baseline_price=baseline,
                baseline_spend=baseline_spend,
                savings_vs_baseline=savings_dollars(baseline_spend, spend),
                premium_vs_baseline_frac=premium,
                price_score=scores[0],
                coverage_score=scores[1],
                hist_score=scores[2],
                zrisk_score=scores[3],
                continuity_score=scores[4],
                rec_score=scores[5],
                cap_breach=bool(breach),
                is_fallback=bool(fallback),
                transit_days=transit,
                relationship="Preserve" if is_inc else "Create",
            )
        )
    details.sort(key=lambda d: (d.scenario_code, d.dc_name, d.lot_name, d.tf_name, d.supplier_name))
    return details


def _write_summary_tab(
    wb: Workbook,
    seeded: CycleView,
    config: EngineConfig,
    rollups: list[ScenarioRollup],
    version: _RunVersion,
) -> None:
    """Summary tab — the headline: cycle/strategy + the A-vs-B alignment call (D26).

    The banner carries the MID-CYCLE ALIGNMENT version heading: the run's 1-based analysis
    version, round number, and sealed timestamp (the sponsor's mid-cycle versioning), above
    the existing decision-support strap.
    """

    ws = wb.active
    assert ws is not None  # noqa: S101
    ws.title = "Summary"

    by_code = {r.code: r for r in rollups}
    rec = by_code.get("B")
    bench = by_code.get("A")
    headline = "—"
    if rec is not None and bench is not None and bench.total_spend > 0:
        delta_pct = (rec.total_spend - bench.total_spend) / bench.total_spend * Decimal("100")
        direction = "above" if delta_pct >= 0 else "below"
        headline = (
            f"Recommendation = Scenario B ({rec.label}): "
            f"${rec.total_spend:,.0f} spend, {rec.savings_vs_baseline_frac * 100:.1f}% vs "
            f"baseline, {abs(delta_pct):.1f}% {direction} the Scenario A lowest-cost benchmark "
            f"(${bench.total_spend:,.0f}) — the risk-adjusted premium for coverage/continuity."
        )

    columns = [
        Col("Item", 34),
        Col("Value", 60),
    ]
    header_row = 6  # banner rows 1-5
    rows_data: list[tuple[str, str]] = [
        ("Cycle", f"{seeded.cycle_code} — {seeded.cycle_name}"),
        (
            "Scope",
            f"{len(seeded.dcs)} DCs × {len(seeded.lots)} lots × {len(seeded.tfs)} timeframes; "
            f"{len(seeded.suppliers)} invited suppliers",
        ),
        (
            "Strategy",
            f"{config.preset.value} preset · weights P{config.weight_price}/"
            f"Cov{config.weight_coverage}/Hist{config.weight_historical}/Z{config.weight_zrisk}/"
            f"Cont{config.weight_continuity} · max {config.max_sup_dc} suppliers/DC",
        ),
        ("Headline (A vs B)", headline),
        (
            "How to use this workbook",
            "Supplier Comparison = every supplier per cell side by side (the price debate). "
            "Scenario Comparison = the lenses side by side (which lens). "
            "Custom Scenario = override per cell, watch spend/savings recompute live.",
        ),
    ]
    row = header_row + 1
    for label, value in rows_data:
        c1 = ws.cell(row=row, column=1, value=label)
        c1.font = _TOTAL_FONT
        c1.alignment = _LEFT
        c1.border = _BORDER
        c2 = ws.cell(row=row, column=2, value=value)
        c2.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c2.border = _BORDER
        ws.row_dimensions[row].height = 30
        row += 1

    format_table(
        ws,
        title=f"MID-CYCLE ALIGNMENT ANALYSIS — {seeded.cycle_name}",
        subtitle_lines=[
            f"Cycle {seeded.cycle_code} · Round {version.round_number} "
            f"(run {version.round_seq}) · Analysis v{version.seq} · "
            f"sealed {version.run_finished_at:%Y-%m-%d %H:%M}",
            "ALIGNMENT / COMPARISON tool (D26) — what the team works through to DECIDE, "
            "not a final summary.",
            f"Generated {date.today():%Y-%m-%d} · SYNTHETIC names & prices · "
            f"{DECISION_SUPPORT_STRAP}",
        ],
        columns=columns,
        n_body_rows=len(rows_data),
        header_row=header_row,
        add_total=False,
        add_autofilter=False,
    )


@dataclass(frozen=True)
class _DcRollup:
    """Per-DC roll-up of one scenario's awards (the level-1 drill row + its supplier children)."""

    dc_name: str
    spend: Decimal
    volume: Decimal
    savings_vs_baseline: Decimal
    n_suppliers: int
    suppliers: list[AwardDetail]  # the per-supplier (level-2) child lines, volume-desc


def _rollup_scenario(rows: list[AwardDetail]) -> tuple[Decimal, Decimal, Decimal, list[_DcRollup]]:
    """Roll one scenario's award rows into (total spend, total volume, total savings, per-DC rows).

    Per-DC rows aggregate a DC's award lines to spend/volume/savings + a distinct-supplier count,
    and carry the per-supplier child lines (each a real award line). Drives the expandable drill.
    """

    by_dc: dict[str, list[AwardDetail]] = defaultdict(list)
    for d in rows:
        by_dc[d.dc_name].append(d)
    dc_rollups: list[_DcRollup] = []
    tot_spend = tot_vol = tot_sav = Decimal("0")
    for dcn in sorted(by_dc):
        lines = by_dc[dcn]
        spend = sum((x.spend for x in lines), Decimal("0"))
        vol = sum((x.volume for x in lines), Decimal("0"))
        sav = sum((x.savings_vs_baseline for x in lines), Decimal("0"))
        tot_spend += spend
        tot_vol += vol
        tot_sav += sav
        dc_rollups.append(
            _DcRollup(
                dc_name=dcn,
                spend=spend,
                volume=vol,
                savings_vs_baseline=sav,
                n_suppliers=len({x.supplier_name for x in lines}),
                suppliers=sorted(lines, key=lambda x: (-float(x.volume), x.supplier_name)),
            )
        )
    return tot_spend, tot_vol, tot_sav, dc_rollups


# Drill columns: a fixed 8-wide grid shared by the total / DC / supplier rows.
_DRILL_HEADERS = (
    ("Scenario / DC / Supplier", None),
    ("Spend", NUMFMT_MONEY),
    ("Volume (cases)", NUMFMT_INT),
    ("Savings vs Baseline ($)", NUMFMT_MONEY),
    ("# Suppliers / $/case", NUMFMT_MONEY),
    ("Volume share", NUMFMT_PCT),
    ("Premium vs Baseline", NUMFMT_PCT),
    ("Flags", None),
)


def _write_scenario_drill(
    ws: Worksheet,
    details: list[AwardDetail],
    span: int,
    start_row: int,
) -> int:
    """Write the EXPANDABLE scenario→DC→supplier drill region; return the next free row (D27).

    Excel outline grouping: each scenario TOTAL row is level 0; its per-DC rows are level 1; each
    DC's per-supplier rows are level 2. With `summaryBelow=False` the +/- buttons collapse the
    detail UNDER each scenario total. Rows are written collapsed (hidden + collapsed flags set) so
    the region OPENS as scenario totals only, expandable on demand.
    """

    # Section banner.
    tcell = ws.cell(
        row=start_row,
        column=1,
        value="DRILL-DOWN — expand a scenario (+) to per-DC, then to per-supplier (D27)",
    )
    tcell.font = _TITLE_FONT
    tcell.fill = _TITLE_FILL
    tcell.alignment = _LEFT
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=span)
    for col in range(1, span + 1):
        ws.cell(row=start_row, column=col).fill = _TITLE_FILL

    # Header row for the drill grid.
    hrow = start_row + 1
    for ci, (htext, _fmt) in enumerate(_DRILL_HEADERS, start=1):
        hc = ws.cell(row=hrow, column=ci, value=htext)
        hc.font = _HEADER_FONT
        hc.fill = _HEADER_FILL
        hc.alignment = _WRAP_CENTER
        hc.border = _BORDER
    ws.row_dimensions[hrow].height = 28

    by_scen: dict[str, list[AwardDetail]] = defaultdict(list)
    label_by: dict[str, str] = {}
    for d in details:
        by_scen[d.scenario_code].append(d)
        label_by[d.scenario_code] = d.scenario_label

    row = hrow + 1
    for code in sorted(by_scen):
        tot_spend, tot_vol, tot_sav, dc_rollups = _rollup_scenario(by_scen[code])
        # --- Scenario TOTAL row (outline level 0; the collapse handle). ---
        scen_label = f"Scenario {code} — {label_by[code]}"
        _drill_cells(
            ws,
            row,
            [
                scen_label,
                tot_spend,
                tot_vol,
                tot_sav,
                len({d.supplier_name for d in by_scen[code]}),
                None,
                None,
                "",
            ],
            level=0,
            bold=True,
            fill=(_BENCH_FILL if code == "A" else _REC_FILL if code == "B" else _TOTAL_FILL),
        )
        # The total row carries the collapse control for the child block below it.
        ws.row_dimensions[row].collapsed = True
        row += 1
        for dcr in dc_rollups:
            # --- per-DC row (level 1; hidden+collapsed so detail opens closed). ---
            _drill_cells(
                ws,
                row,
                [
                    f"   {dcr.dc_name}",
                    dcr.spend,
                    dcr.volume,
                    dcr.savings_vs_baseline,
                    dcr.n_suppliers,
                    None,
                    None,
                    "",
                ],
                level=1,
                bold=False,
                hidden=True,
                collapsed=True,
            )
            row += 1
            for sup in dcr.suppliers:
                # --- per-supplier row (level 2; hidden). $/case sits in the "# Suppliers" col. ---
                flags = []
                if sup.cap_breach:
                    flags.append("CAP-BREACH")
                if sup.is_fallback:
                    flags.append("FALLBACK")
                if sup.is_incumbent:
                    flags.append("INCUMBENT")
                _drill_cells(
                    ws,
                    row,
                    [
                        f"      {sup.supplier_name}",
                        sup.spend,
                        sup.volume,
                        sup.savings_vs_baseline,
                        sup.price,
                        sup.volume_share,
                        sup.premium_vs_baseline_frac,
                        ", ".join(flags) if flags else "—",
                    ],
                    level=2,
                    bold=False,
                    hidden=True,
                )
                row += 1
    return row


def _drill_cells(
    ws: Worksheet,
    row: int,
    values: list[object],
    *,
    level: int,
    bold: bool,
    hidden: bool = False,
    collapsed: bool = False,
    fill: PatternFill | None = None,
) -> None:
    """Write one drill row: 8 values with shared formats; set the outline level + hidden flag."""

    for ci, (val, (_h, fmt)) in enumerate(zip(values, _DRILL_HEADERS, strict=True), start=1):
        cell = ws.cell(row=row, column=ci)
        cell.value = float(val) if isinstance(val, Decimal) else val
        cell.border = _BORDER
        if fmt and val is not None and not isinstance(val, str):
            cell.number_format = fmt
            cell.alignment = _CENTER
        else:
            cell.alignment = _LEFT
        if bold:
            cell.font = _TOTAL_FONT
        if fill is not None:
            cell.fill = fill
    rd = ws.row_dimensions[row]
    rd.outline_level = level
    if hidden:
        rd.hidden = True
    if collapsed:
        rd.collapsed = True


# The Custom Scenario tab's fixed layout (mirrors `_write_custom_scenario_tab`) so the live Custom
# column in Scenario Comparison can reference it by deterministic address. header_row=6 → body 7..,
# columns: DC=A, Supplier=E, $/case=F, Volume=J, Line Spend=K, Cell Key=M; total row = body_end+1.
_CUSTOM_SHEET = "Custom Scenario"
_CUSTOM_HEADER_ROW = 6


def _custom_refs(n_cells: int) -> dict[str, str]:
    """Deterministic A1-style ranges into the Custom Scenario tab for the live Custom column."""

    body_start = _CUSTOM_HEADER_ROW + 1
    body_end = body_start + n_cells - 1
    total_row = body_end + 1
    q = f"'{_CUSTOM_SHEET}'"
    return {
        "total_spend": f"{q}!$K${total_row}",  # live total spend
        "dc": f"{q}!$A${body_start}:$A${body_end}",  # DC name per row
        "sup": f"{q}!$E${body_start}:$E${body_end}",  # chosen supplier per row
        "spend": f"{q}!$K${body_start}:$K${body_end}",  # live line spend per row
        "vol": f"{q}!$J${body_start}:$J${body_end}",  # volume per row
        "price": f"{q}!$F${body_start}:$F${body_end}",  # live $/case per row
        "baseline": f"{q}!$L${body_start}:$L${body_end}",  # baseline $/case per row
    }


def _write_custom_drill(
    ws: Worksheet,
    cells: list[CellInfo],
    refs: dict[str, str],
    start_row: int,
) -> int:
    """Write the LIVE Custom block in the same drill grid: total → per-DC → per-supplier (D27).

    Every value is a SUMIFS/SUMPRODUCT against the Custom Scenario tab's live spend + chosen
    supplier columns, so as the buyer changes picks there, the Custom total / per-DC / supplier rows
    recompute and read side by side against A-G. Per-supplier rows list every supplier that COULD be
    picked in that DC; each row's spend/volume is the live SUMIFS for that supplier (0 when not the
    current pick). Same outline levels as A-G (DC=1, supplier=2), opens collapsed under the total.
    """

    dc_r, sup_r, spend_r, vol_r = refs["dc"], refs["sup"], refs["spend"], refs["vol"]
    # Per-DC supplier universe (who could be picked in that DC) from the cells' eligible lists.
    dc_universe: dict[str, set[str]] = defaultdict(set)
    for c in cells:
        dc_universe[c.dc_name].update(c.eligible_suppliers)

    row = start_row
    # --- Custom TOTAL row (level 0; the collapse handle for the live block). ---
    _drill_cells(
        ws,
        row,
        [
            "Custom (LIVE — off Custom Scenario tab)",
            f"={refs['total_spend']}",
            f"=SUM({vol_r})",
            f"=SUMPRODUCT(({refs['baseline']}-{refs['price']})*{vol_r})",
            f'=SUMPRODUCT(({sup_r}<>"")/COUNTIF({sup_r},{sup_r}&""))',
            None,
            None,
            "live",
        ],
        level=0,
        bold=True,
        fill=_REC_PICK_FILL,
    )
    ws.row_dimensions[row].collapsed = True
    row += 1
    for dcn in sorted(dc_universe):
        # --- per-DC LIVE row (level 1): SUMIFS over the Custom tab restricted to this DC. ---
        _drill_cells(
            ws,
            row,
            [
                f"   {dcn}",
                f'=SUMIFS({spend_r},{dc_r},"{dcn}")',
                f'=SUMIFS({vol_r},{dc_r},"{dcn}")',
                None,
                f'=SUMPRODUCT(({dc_r}="{dcn}")/COUNTIFS({dc_r},{dc_r},{sup_r},{sup_r}))',
                None,
                None,
                "live",
            ],
            level=1,
            bold=False,
            hidden=True,
            collapsed=True,
        )
        row += 1
        for sup in sorted(dc_universe[dcn]):
            # --- per-supplier LIVE row (level 2): this supplier's live spend/volume in this DC. ---
            _drill_cells(
                ws,
                row,
                [
                    f"      {sup}",
                    f'=SUMIFS({spend_r},{dc_r},"{dcn}",{sup_r},"{sup}")',
                    f'=SUMIFS({vol_r},{dc_r},"{dcn}",{sup_r},"{sup}")',
                    None,
                    None,
                    None,
                    None,
                    "live (0 if not picked)",
                ],
                level=2,
                bold=False,
                hidden=True,
            )
            row += 1
    return row


def _write_scenario_comparison_tab(
    wb: Workbook,
    rollups: list[ScenarioRollup],
    baseline_total: Decimal,
    stly_total: Decimal,
    dc_names: list[str],
    details: list[AwardDetail],
    cells: list[CellInfo],
) -> None:
    """Scenario Comparison tab — lenses side by side + LIVE Custom + an EXPANDABLE drill (D26/D27).

    Top block: rows = scenarios A-G + a LIVE Custom row (formulas off the Custom Scenario tab, so it
    reads side by side against A-G and updates as the buyer changes picks). Below it: a DRILL-DOWN
    region using Excel outline grouping — each scenario TOTAL (level 0) expands to per-DC rows
    (level 1) which expand to per-supplier rows (level 2); opens COLLAPSED to the totals. Then the
    per-DC matrix (DCs down, scenarios across → spend) so the team sees WHERE lenses differ (D26).
    """

    n_cells = len(cells)
    ws = wb.create_sheet("Scenario Comparison")
    # summaryBelow=False → the +/- outline buttons sit on the scenario TOTAL row and collapse the
    # DC/supplier detail UNDER it (depth-on-demand, D27). The view opens collapsed to totals.
    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.sheet_properties.outlinePr.applyStyles = True

    custom_refs = _custom_refs(n_cells)

    columns = [
        Col("Lens", 7),
        Col("Scenario (label)", 30),
        Col("Total Spend", 16, NUMFMT_MONEY),
        Col("Δ vs A (lowest-cost)", 18, NUMFMT_MONEY),
        Col("Savings vs Baseline", 16, NUMFMT_PCT),
        Col("Savings vs STLY*", 16, NUMFMT_PCT),
        Col("# Suppliers", 11, NUMFMT_INT),
        Col("# Cap-Breaches", 13, NUMFMT_INT),
        Col("# Cells", 9, NUMFMT_INT),
    ]
    header_row = 6
    body_start = header_row + 1
    row = body_start
    a_spend_cell = f"$C${body_start}"  # Scenario A's Total Spend cell (Δ-vs-A anchor for Custom)
    for r in rollups:
        ws.cell(row=row, column=1, value=r.code)
        ws.cell(row=row, column=2, value=r.label)
        ws.cell(row=row, column=3, value=float(r.total_spend))
        ws.cell(row=row, column=4, value=float(r.delta_vs_a))
        ws.cell(row=row, column=5, value=float(r.savings_vs_baseline_frac))
        ws.cell(row=row, column=6, value=float(r.savings_vs_stly_frac))
        ws.cell(row=row, column=7, value=r.n_suppliers)
        ws.cell(row=row, column=8, value=r.n_cap_breaches)
        ws.cell(row=row, column=9, value=r.n_cells)
        row += 1
    # LIVE Custom row — formulas off the Custom Scenario tab so it reads side by side vs A-G and
    # recomputes as the buyer changes any dropdown there (D27 live custom-vs-scenarios).
    custom_row = row
    ws.cell(row=custom_row, column=1, value="Cust")
    ws.cell(row=custom_row, column=2, value="Custom build (LIVE — off Custom Scenario tab)")
    # Total spend = the Custom tab's live total. Δ vs A and savings recompute off it.
    ws.cell(row=custom_row, column=3, value=f"={custom_refs['total_spend']}")
    ws.cell(row=custom_row, column=4, value=f"={custom_refs['total_spend']}-{a_spend_cell}")
    ws.cell(
        row=custom_row,
        column=5,
        value=f"=IF({baseline_total}=0,0,({baseline_total}-{custom_refs['total_spend']})/{baseline_total})",
    )
    ws.cell(
        row=custom_row,
        column=6,
        value=f"=IF({stly_total}=0,0,({stly_total}-{custom_refs['total_spend']})/{stly_total})",
    )
    # # Suppliers = distinct chosen suppliers across the Custom tab's body (live).
    ws.cell(
        row=custom_row,
        column=7,
        value=(
            f'=SUMPRODUCT(({custom_refs["sup"]}<>"")/'
            f'COUNTIF({custom_refs["sup"]},{custom_refs["sup"]}&""))'
        ),
    )
    ws.cell(row=custom_row, column=8, value="(see Custom tab)")
    ws.cell(row=custom_row, column=9, value=n_cells)
    n_body = (custom_row - body_start) + 1

    format_table(
        ws,
        title="SCENARIO COMPARISON — lenses side by side + LIVE Custom + drill (which lens?)",
        subtitle_lines=[
            "Rows A-G = the lenses; the Cust row is LIVE off the Custom Scenario tab (change a "
            "pick there and Custom spend / Δ vs A / savings update here). A = benchmark (peach), "
            "B = recommendation (blue).",
            "▶ DRILL: below this table each scenario TOTAL expands (+ in the margin) to per-DC, "
            "then to per-supplier — opens collapsed to totals, expand for depth on demand (D27).",
            "*Savings vs STLY uses a MODELED prior-year baseline (no live STLY feed yet: "
            f"prior-year actual-paid modeled ~{(_STLY_UPLIFT - 1) * 100:.0f}% over this year's "
            "incumbent routing). Clearly labelled an estimate.",
            f"SYNTHETIC · baseline ${baseline_total:,.0f} · STLY* ${stly_total:,.0f} · "
            f"{DECISION_SUPPORT_STRAP}",
        ],
        columns=columns,
        n_body_rows=n_body,
        header_row=header_row,
        add_total=False,
    )

    # Highlight the benchmark (A) and recommendation (B) rows across the table.
    span = len(columns)
    for off, r in enumerate(rollups):
        if r.code in ("A", "B"):
            fill = _BENCH_FILL if r.code == "A" else _REC_FILL
            for col in range(1, span + 1):
                ws.cell(row=body_start + off, column=col).fill = fill

    # --- EXPANDABLE DRILL: scenario total → per-DC → per-supplier (outline grouping, D27). ---
    drill_next = _write_scenario_drill(ws, details, span, custom_row + 2)
    # The LIVE Custom block sits in the same drill grid, side by side with A-G (total → per-DC →
    # per-supplier, all SUMIFS off the Custom Scenario tab; recomputes as the buyer changes picks).
    drill_next = _write_custom_drill(ws, cells, custom_refs, drill_next)

    # --- Per-DC matrix below the drill: DCs down, scenarios across → spend (D26). ---
    matrix_title_row = drill_next + 1
    tcell = ws.cell(
        row=matrix_title_row,
        column=1,
        value="PER-DC MATRIX — where the lenses differ (spend per DC × scenario)",
    )
    tcell.font = _TITLE_FONT
    tcell.fill = _TITLE_FILL
    tcell.alignment = _LEFT
    ws.merge_cells(
        start_row=matrix_title_row, start_column=1, end_row=matrix_title_row, end_column=span
    )
    for col in range(1, span + 1):
        ws.cell(row=matrix_title_row, column=col).fill = _TITLE_FILL

    # Matrix header: DC | A | B | ... | G | + supplier mix note row beneath each DC.
    mh_row = matrix_title_row + 1
    ws.cell(row=mh_row, column=1, value="DC").font = _HEADER_FONT
    ws.cell(row=mh_row, column=1).fill = _HEADER_FILL
    ws.cell(row=mh_row, column=1).alignment = _WRAP_CENTER
    ws.cell(row=mh_row, column=1).border = _BORDER
    for ci, r in enumerate(rollups, start=2):
        hc = ws.cell(row=mh_row, column=ci, value=f"{r.code} spend")
        hc.font = _HEADER_FONT
        hc.fill = _HEADER_FILL
        hc.alignment = _WRAP_CENTER
        hc.border = _BORDER
    ws.row_dimensions[mh_row].height = 26

    mrow = mh_row + 1
    for dcn in dc_names:
        dcell = ws.cell(row=mrow, column=1, value=dcn)
        dcell.alignment = _LEFT
        dcell.border = _BORDER
        dcell.font = _TOTAL_FONT
        for ci, r in enumerate(rollups, start=2):
            spend = r.spend_by_dc.get(dcn, Decimal("0"))
            sc = ws.cell(row=mrow, column=ci, value=float(spend))
            sc.number_format = NUMFMT_MONEY
            sc.alignment = _CENTER
            sc.border = _BORDER
            if r.code == "A":
                sc.fill = _BENCH_FILL
            elif r.code == "B":
                sc.fill = _REC_FILL
        mrow += 1
        # Supplier-mix sub-row (which suppliers each lens uses at this DC).
        mixcell = ws.cell(row=mrow, column=1, value=f"   ↳ {dcn} supplier mix")
        mixcell.alignment = _LEFT
        mixcell.font = _SUBTITLE_FONT
        for ci, r in enumerate(rollups, start=2):
            mix = r.suppliers_by_dc.get(dcn, "")
            mc = ws.cell(row=mrow, column=ci, value=mix)
            mc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            mc.font = _SUBTITLE_FONT
            mc.border = _BORDER
        ws.row_dimensions[mrow].height = 24
        mrow += 1


def _write_supplier_comparison_tab(
    wb: Workbook,
    config: EngineConfig,
    cells: list[CellInfo],
    seeded: CycleView,
) -> None:
    """Supplier Comparison tab — THE CENTERPIECE: every supplier per cell side by side (D26).

    Rows = each (DC × lot × item × TF) cell with demand, Baseline $/case, Incumbent $/case. Then one
    $/case column PER SUPPLIER (blank if no bid). Conditional formatting highlights the MIN per row
    (best price). Then Min $/case, Recommended supplier, RecScore, and a SECOND block = cost-impact-
    vs-baseline % per supplier (the premium each supplier carries). Incumbent + recommended flagged.
    This is the v3 FOB comparison the team scans to compare suppliers per cell.
    """

    ws = wb.create_sheet("Supplier Comparison")

    # Supplier order = the seeded order (stable, readable). One $/case column per supplier.
    sup_order = [sup.name for sup in seeded.suppliers]
    n_sup = len(sup_order)

    # Fixed lead columns, then n_sup price columns, then Min / Rec / RecScore,
    # then n_sup cost-impact columns.
    lead = [
        Col("DC", 16),
        Col("Lot", 20),
        Col("Item", 22),
        Col("Timeframe", 16),
        Col("Demand (cases)", 13, NUMFMT_INT, total="sum"),
        Col("Baseline $/case", 14, NUMFMT_MONEY),
        Col("Incumbent", 22),
        Col("Incumbent $/case", 14, NUMFMT_MONEY),
    ]
    price_cols = [Col(name, 14, NUMFMT_MONEY) for name in sup_order]
    mid = [
        Col("Min $/case", 13, NUMFMT_MONEY),
        Col("Recommended", 22),
        Col("RecScore", 10, NUMFMT_INT),
    ]
    impact_cols = [Col(f"{name} (impact)", 13, NUMFMT_PCT) for name in sup_order]
    columns = lead + price_cols + mid + impact_cols

    n_lead = len(lead)
    price_start_col = n_lead + 1  # 1-based col of the first supplier price column
    price_end_col = price_start_col + n_sup - 1
    min_col = price_end_col + 1
    rec_col = min_col + 1
    recscore_col = rec_col + 1
    impact_start_col = recscore_col + 1

    header_row = 7
    body_start = header_row + 1
    row = body_start
    for c in cells:
        prices = [c.price_by_supplier.get(name) for name in sup_order]
        present = [p for p in prices if p is not None]
        min_price = min(present) if present else None
        inc_price = c.price_by_supplier.get(c.incumbent_name)

        ws.cell(row=row, column=1, value=c.dc_name)
        ws.cell(row=row, column=2, value=c.lot_name)
        ws.cell(row=row, column=3, value=c.item_name)
        ws.cell(row=row, column=4, value=c.tf_name)
        ws.cell(row=row, column=5, value=float(c.volume))
        ws.cell(row=row, column=6, value=float(c.baseline_price))
        ws.cell(row=row, column=7, value=c.incumbent_name)
        ws.cell(row=row, column=8, value=float(inc_price) if inc_price is not None else None)
        for i, p in enumerate(prices):
            cell = ws.cell(row=row, column=price_start_col + i)
            cell.value = float(p) if p is not None else None
            # Mark the recommended pick's price cell + the incumbent's price cell.
            if sup_order[i] == c.rec_supplier and p is not None:
                cell.fill = _REC_PICK_FILL
                cell.font = Font(bold=True, color="1F3864")
            elif sup_order[i] == c.incumbent_name and p is not None:
                cell.fill = _INCUMBENT_FILL
        ws.cell(row=row, column=min_col, value=float(min_price) if min_price is not None else None)
        ws.cell(row=row, column=rec_col, value=c.rec_supplier)
        ws.cell(row=row, column=recscore_col, value=float(c.rec_score))
        # Cost-impact vs baseline % per supplier (positive = premium over baseline).
        for i, p in enumerate(prices):
            cell = ws.cell(row=row, column=impact_start_col + i)
            if p is not None and c.baseline_price > 0:
                cell.value = float((p - c.baseline_price) / c.baseline_price)
        row += 1

    format_table(
        ws,
        title="SUPPLIER COMPARISON — every supplier per cell, side by side (the FOB comparison)",
        subtitle_lines=[
            "Each row = one DC × lot × item × TF cell. Left block = each supplier's all-in "
            "$/case (blank = no bid). GREEN = lowest (best) price per row. Recommended pick = "
            "bold blue; incumbent = amber.",
            "Right block = cost impact vs baseline % per supplier (the premium each carries). "
            "Min $/case, Recommended supplier + RecScore in the middle.",
            f"SYNTHETIC · {n_sup} suppliers · {DECISION_SUPPORT_STRAP}",
        ],
        columns=columns,
        n_body_rows=len(cells),
        header_row=header_row,
    )

    body_end = body_start + len(cells) - 1
    # Conditional formatting: highlight the MIN price per row across the supplier price columns.
    if cells:
        first_letter = get_column_letter(price_start_col)
        last_letter = get_column_letter(price_end_col)
        price_range = f"{first_letter}{body_start}:{last_letter}{body_end}"
        # Per-row MIN: the cell equals the row's min over the supplier price block (ignores blanks
        # via MIN). Anchored so the formula slides per row but the row's range is absolute-column.
        min_formula = (
            f'AND({first_letter}{body_start}<>"",'
            f"{first_letter}{body_start}=MIN(${first_letter}{body_start}:${last_letter}{body_start}))"
        )
        ws.conditional_formatting.add(
            price_range,
            FormulaRule(formula=[min_formula], fill=_MIN_FILL, font=_MIN_FONT, stopIfTrue=False),
        )
        # Conditional formatting on the impact block: premiums >12% (the eligibility ceiling) red.
        imp_first = get_column_letter(impact_start_col)
        imp_last = get_column_letter(impact_start_col + n_sup - 1)
        impact_range = f"{imp_first}{body_start}:{imp_last}{body_end}"
        ws.conditional_formatting.add(
            impact_range,
            CellIsRule(
                operator="greaterThan",
                formula=[str(float(config.global_premium_threshold))],
                fill=_BREACH_FILL,
                font=_BREACH_FONT,
            ),
        )


def _write_prices_helper(
    wb: Workbook,
    cells: list[CellInfo],
) -> str:
    """Hidden `_Prices` sheet: the supplier × cell price grid + a per-cell reference grid.

    Left grid (cols A-D): Match Key | Cell Key | Supplier | $/case, where Match Key = cell_key &
    "@" & supplier (one SUMIFS on Match Key resolves the chosen supplier's price for a cell). Right
    grid (cols F-I): Cell Key | Min $/case | Incumbent $/case | Baseline $/case — so the Custom
    Scenario tab can show the picked price vs Min / vs Incumbent / vs Baseline LIVE. Returns the
    sheet name.
    """

    sheet_name = "_Prices"
    ws = wb.create_sheet(sheet_name)
    ws["A1"] = "Match Key"
    ws["B1"] = "Cell Key"
    ws["C1"] = "Supplier"
    ws["D1"] = "$/case"
    ws["E1"] = "Transit (days)"
    r = 2
    for c in cells:
        for supplier, price in c.price_by_supplier.items():
            ws.cell(row=r, column=1, value=f"{c.cell_key}@{supplier}")
            ws.cell(row=r, column=2, value=c.cell_key)
            ws.cell(row=r, column=3, value=supplier)
            ws.cell(row=r, column=4, value=float(price))
            ws.cell(row=r, column=4).number_format = NUMFMT_MONEY
            ws.cell(row=r, column=5, value=c.transit_by_supplier.get(supplier, 0))
            r += 1

    # Right reference grid: per-cell Min / Incumbent / Baseline (the comparison anchors).
    ws["F1"] = "Cell Key"
    ws["G1"] = "Min $/case"
    ws["H1"] = "Incumbent $/case"
    ws["I1"] = "Baseline $/case"
    rr = 2
    for c in cells:
        present = list(c.price_by_supplier.values())
        min_price = min(present) if present else Decimal("0")
        inc_price = c.price_by_supplier.get(c.incumbent_name, Decimal("0"))
        ws.cell(row=rr, column=6, value=c.cell_key)
        ws.cell(row=rr, column=7, value=float(min_price)).number_format = NUMFMT_MONEY
        ws.cell(row=rr, column=8, value=float(inc_price)).number_format = NUMFMT_MONEY
        ws.cell(row=rr, column=9, value=float(c.baseline_price)).number_format = NUMFMT_MONEY
        rr += 1

    for col, width in (
        ("A", 36),
        ("B", 18),
        ("C", 26),
        ("D", 12),
        ("F", 18),
        ("G", 12),
        ("H", 16),
        ("I", 14),
    ):
        ws.column_dimensions[col].width = width
    ws.sheet_state = "hidden"
    return sheet_name


def _write_custom_scenario_tab(
    wb: Workbook,
    config: EngineConfig,
    cells: list[CellInfo],
    prices_sheet: str,
) -> None:
    """THE INTERACTIVE TAB — per-cell supplier dropdowns + live spend/savings/cap-breach.

    Layout: an instruction banner, then one row per cell with a Supplier dropdown
    (data-validation list = the cell's eligible suppliers, by NAME) and LIVE formulas:
      $/case  = SUMIFS(_Prices!$D, _Prices!$A, CellKey & "@" & ChosenSupplier)
      Volume  = a literal (projected period cases)
      Line Spend = $/case × Volume
    Below the table: a summary block with TOTAL spend (=SUM), baseline total, savings
    vs baseline (live), and per-DC distinct-supplier counts with a cap-breach flag
    (distinct suppliers in a DC > max_sup_dc). Pre-filled with Scenario B's picks.
    """

    ws = wb.create_sheet("Custom Scenario")

    columns = [
        Col("DC", 16),
        Col("Lot", 20),
        Col("Item", 22),
        Col("Timeframe", 16),
        Col("Supplier (change me ▼)", 24),
        Col("$/case (live)", 13, NUMFMT_MONEY),
        Col("vs Min (live)", 12, NUMFMT_PCT),
        Col("vs Incumbent (live)", 14, NUMFMT_PCT),
        Col("vs Baseline (live)", 14, NUMFMT_PCT),
        Col("Volume (cases)", 13, NUMFMT_INT, total="sum"),
        Col("Line Spend (live)", 16, NUMFMT_MONEY, total="sum"),
        Col("Baseline $/case", 14, NUMFMT_MONEY),
        Col("Cell Key", 16),  # the SUMIFS key (kept visible-but-narrow for traceability)
        Col("Transit days (live)", 14, NUMFMT_INT),  # hidden-cost data point in the builder
    ]
    header_row = 6  # banner rows 1-5 (instruction is prominent)
    body_start = header_row + 1
    body_end = body_start + len(cells) - 1

    # Column letters used inside the live formulas (recompute as the dropdown changes).
    col_supplier = get_column_letter(5)
    col_price = get_column_letter(6)
    col_vol = get_column_letter(10)
    col_spend = get_column_letter(11)
    col_baseline = get_column_letter(12)
    col_cellkey = get_column_letter(13)
    p_key = f"'{prices_sheet}'!$A:$A"
    p_val = f"'{prices_sheet}'!$D:$D"
    p_transit = f"'{prices_sheet}'!$E:$E"
    # The per-cell reference grid (Cell Key | Min | Incumbent | Baseline) in _Prices F:I.
    ref_key = f"'{prices_sheet}'!$F:$F"
    ref_min = f"'{prices_sheet}'!$G:$G"
    ref_inc = f"'{prices_sheet}'!$H:$H"

    row = body_start
    for c in cells:
        ws.cell(row=row, column=1, value=c.dc_name)
        ws.cell(row=row, column=2, value=c.lot_name)
        ws.cell(row=row, column=3, value=c.item_name)
        ws.cell(row=row, column=4, value=c.tf_name)
        # The interactive pick — pre-filled with Scenario B's recommended supplier.
        ws.cell(row=row, column=5, value=c.rec_supplier)
        # LIVE: chosen supplier's $/case via SUMIFS on the hidden _Prices match key.
        ws.cell(
            row=row,
            column=6,
            value=(f'=SUMIFS({p_val},{p_key},{col_cellkey}{row}&"@"&{col_supplier}{row})'),
        )
        # LIVE: picked price vs the cell's Min / Incumbent / Baseline (grounded comparisons, D26).
        ws.cell(
            row=row,
            column=7,
            value=(
                f"=IF(SUMIFS({ref_min},{ref_key},{col_cellkey}{row})=0,0,"
                f"({col_price}{row}-SUMIFS({ref_min},{ref_key},{col_cellkey}{row}))"
                f"/SUMIFS({ref_min},{ref_key},{col_cellkey}{row}))"
            ),
        )
        ws.cell(
            row=row,
            column=8,
            value=(
                f"=IF(SUMIFS({ref_inc},{ref_key},{col_cellkey}{row})=0,0,"
                f"({col_price}{row}-SUMIFS({ref_inc},{ref_key},{col_cellkey}{row}))"
                f"/SUMIFS({ref_inc},{ref_key},{col_cellkey}{row}))"
            ),
        )
        ws.cell(
            row=row,
            column=9,
            value=(
                f"=IF({col_baseline}{row}=0,0,"
                f"({col_price}{row}-{col_baseline}{row})/{col_baseline}{row})"
            ),
        )
        # Volume is a literal projected demand for the cell.
        ws.cell(row=row, column=10, value=float(c.volume))
        # LIVE: line spend = price × volume.
        ws.cell(row=row, column=11, value=f"={col_price}{row}*{col_vol}{row}")
        ws.cell(row=row, column=12, value=float(c.baseline_price))
        ws.cell(row=row, column=13, value=c.cell_key)
        # LIVE: chosen supplier's lane transit (hidden cost) — updates with the dropdown.
        ws.cell(
            row=row,
            column=14,
            value=f'=SUMIFS({p_transit},{p_key},{col_cellkey}{row}&"@"&{col_supplier}{row})',
        )
        row += 1

    fmt = format_table(
        ws,
        title="CUSTOM SCENARIO — override the supplier per cell, watch it recompute LIVE",
        subtitle_lines=[
            "▶ Change the Supplier dropdown on any row — $/case, vs Min/Incumbent/Baseline, "
            "Total Spend, Savings + Cap-Breach all update live.",
            "Opens on Scenario B's recommendation. Dropdown lists only the eligible "
            "suppliers for that cell (by name). Grounded in the Supplier Comparison.",
            f"SYNTHETIC · max {config.max_sup_dc} suppliers/DC · {DECISION_SUPPORT_STRAP}",
        ],
        columns=columns,
        n_body_rows=len(cells),
        header_row=header_row,
        total_label_col=1,
        total_label="TOTAL (live)",
    )
    total_row = fmt["total_row"]

    # Per-cell data-validation dropdowns — list = the cell's eligible suppliers (by NAME).
    # openpyxl inline list formula must be wrapped in double-quotes; commas separate items.
    for offset, c in enumerate(cells):
        r = body_start + offset
        items = ",".join(c.eligible_suppliers)
        # Excel inline lists are capped ~255 chars; our short demo names fit comfortably.
        dv = DataValidation(
            type="list",
            formula1=f'"{items}"',
            allow_blank=False,
            showDropDown=False,  # False => the dropdown arrow IS shown (Excel quirk)
        )
        dv.error = "Pick a supplier from the eligible list for this cell."
        dv.errorTitle = "Not an eligible supplier"
        dv.prompt = "Choose the supplier for this DC × lot × timeframe cell."
        dv.promptTitle = "Override supplier"
        ws.add_data_validation(dv)
        dv.add(ws.cell(row=r, column=5))

    # --- LIVE summary block below the table (spend / savings / cap-breach). ---
    s_label_col = 4  # put labels in col D, values in col E for readability
    s_val_col = 5
    val_letter = get_column_letter(s_val_col)
    base_total = sum((c.baseline_price * c.volume for c in cells), Decimal("0"))

    summary_top = total_row + 2
    sr = summary_top  # row of the first block; baseline is the next row
    blocks: list[tuple[str, object, str | None]] = [
        # Total custom spend = the table's live total spend cell.
        ("Total Custom Spend (live):", f"={col_spend}{total_row}", NUMFMT_MONEY),
        ("Baseline Spend (incumbent routing):", float(base_total), NUMFMT_MONEY),
        # Savings $ = baseline - custom; live off the two rows above.
        (
            "Savings vs Baseline ($, live):",
            f"={val_letter}{sr + 1}-{val_letter}{sr}",
            NUMFMT_MONEY,
        ),
        (
            "Savings vs Baseline (%, live):",
            (
                f"=IF({val_letter}{sr + 1}=0,0,"
                f"({val_letter}{sr + 1}-{val_letter}{sr})/{val_letter}{sr + 1})"
            ),
            NUMFMT_PCT,
        ),
    ]

    r = summary_top
    for label, value, numfmt in blocks:
        lcell = ws.cell(row=r, column=s_label_col, value=label)
        lcell.font = _TOTAL_FONT
        lcell.alignment = _LEFT
        vcell = ws.cell(row=r, column=s_val_col, value=value)
        vcell.font = _TOTAL_FONT
        vcell.fill = _TOTAL_FILL
        vcell.border = _BORDER
        vcell.alignment = _CENTER
        if numfmt:
            vcell.number_format = numfmt
        r += 1

    # --- Per-DC cap-breach block: distinct chosen suppliers per DC vs max_sup_dc. ---
    dc_names = sorted({c.dc_name for c in cells})
    cap_title_row = r + 1
    tcell = ws.cell(
        row=cap_title_row,
        column=s_label_col,
        value=f"Cap-breach check — distinct suppliers per DC (cap = {config.max_sup_dc}):",
    )
    tcell.font = _TOTAL_FONT
    tcell.alignment = _LEFT

    dc_col = get_column_letter(1)
    cap_row = cap_title_row + 1
    for dc in dc_names:
        # Distinct chosen suppliers in this DC across the body rows: a SUMPRODUCT over
        # 1/COUNTIFS restricted to the DC's rows. Built as an array-ish SUMPRODUCT that
        # Excel evaluates live as the dropdowns change.
        # distinct = SUM over rows in DC of 1/(count of same supplier within the DC rows).
        rng_dc = f"${dc_col}${body_start}:${dc_col}${body_end}"
        rng_sup = f"${col_supplier}${body_start}:${col_supplier}${body_end}"
        # Distinct chosen suppliers within this DC's rows: SUMPRODUCT of (row is in DC) /
        # (count of rows sharing this row's DC AND supplier). COUNTIFS is >=1 for every
        # row (each row matches itself), so there is no div/0; the (dc=…) numerator zeroes
        # non-DC rows, leaving the distinct count for the DC. Recomputes live.
        distinct_formula = (
            f'=SUMPRODUCT(({rng_dc}="{dc}")/COUNTIFS({rng_dc},{rng_dc},{rng_sup},{rng_sup}))'
        )
        lcell = ws.cell(row=cap_row, column=s_label_col, value=dc)
        lcell.alignment = _LEFT
        dcell = ws.cell(row=cap_row, column=s_val_col, value=distinct_formula)
        dcell.alignment = _CENTER
        dcell.border = _BORDER
        # Flag column right of the count.
        flag_formula = f'=IF({val_letter}{cap_row}>{config.max_sup_dc},"⚠ CAP BREACH","OK")'
        fcell = ws.cell(row=cap_row, column=s_val_col + 1, value=flag_formula)
        fcell.font = _TOTAL_FONT
        fcell.alignment = _LEFT
        cap_row += 1

    # Header labels for the cap block columns.
    ws.cell(row=cap_title_row + 0, column=s_val_col, value="# suppliers").font = _SUBTITLE_FONT


# ---------------------------------------------------------------------------
# D27 — FLAT MANIPULABLE DATASET. `Data (pivot me)` — one row per
# (scenario × DC × lot × item × TF × awarded-supplier), every metric, as a real
# Excel Table (ListObject) with AutoFilter. The buyer drops their OWN native
# PivotTable / filter / slicer on it to slice any way they like. Rich but neat:
# the detail lives here for self-serve pivoting, not dumped on the headline tabs.
# ---------------------------------------------------------------------------
_DATA_COLUMNS: tuple[tuple[str, str | None], ...] = (
    ("Scenario", None),
    ("Scenario Label", None),
    ("DC", None),
    ("Lot", None),
    ("Item", None),
    ("Timeframe", None),
    ("Awarded Supplier", None),
    ("Incumbent", None),
    ("Is Incumbent", None),
    ("Volume (cases)", NUMFMT_INT),
    ("Volume Share", NUMFMT_PCT),
    ("$/case", NUMFMT_MONEY),
    ("Spend", NUMFMT_MONEY),
    ("Baseline $/case", NUMFMT_MONEY),
    ("Baseline Spend", NUMFMT_MONEY),
    ("Savings vs Baseline ($)", NUMFMT_MONEY),
    ("Premium vs Baseline", NUMFMT_PCT),
    ("Price Score", NUMFMT_INT),
    ("Coverage Score", NUMFMT_INT),
    ("Historical Score", NUMFMT_INT),
    ("Z-Risk Score", NUMFMT_INT),
    ("Continuity Score", NUMFMT_INT),
    ("RecScore", NUMFMT_INT),
    ("Cap-Breach", None),
    ("Fallback", None),
    ("Transit (days)", NUMFMT_INT),
    ("Relationship", None),
)


def _write_data_pivot_tab(wb: Workbook, details: list[AwardDetail]) -> None:
    """`Data (pivot me)` — the flat manipulable dataset as a real Excel Table (ListObject, D27).

    One row per (scenario × DC × lot × item × TF × awarded-supplier) with all metrics. Registered as
    an Excel Table with AutoFilter so the buyer can drop a native PivotTable / filter / slice on it
    (we do NOT build a programmatic PivotTable — openpyxl support is unreliable; the Table is the
    robust path). A one-line note tells the buyer how to pivot it.
    """

    ws = wb.create_sheet("Data (pivot me)")

    # A one-line note above the table (row 1) — how to pivot. Table starts at row 2.
    note = ws.cell(
        row=1,
        column=1,
        value=(
            "Insert > PivotTable on this table to slice any way you like "
            "(by scenario, DC, supplier, lot…). SYNTHETIC data — names & prices invented."
        ),
    )
    note.font = Font(italic=True, color="1F3864", size=10)

    header_row = 2
    for ci, (htext, _fmt) in enumerate(_DATA_COLUMNS, start=1):
        hc = ws.cell(row=header_row, column=ci, value=htext)
        hc.font = _HEADER_FONT
        hc.fill = _HEADER_FILL
        hc.alignment = _WRAP_CENTER

    row = header_row + 1
    for d in details:
        vals: list[object] = [
            d.scenario_code,
            d.scenario_label,
            d.dc_name,
            d.lot_name,
            d.item_name,
            d.tf_name,
            d.supplier_name,
            d.incumbent_name,
            "Yes" if d.is_incumbent else "No",
            float(d.volume),
            float(d.volume_share),
            float(d.price),
            float(d.spend),
            float(d.baseline_price),
            float(d.baseline_spend),
            float(d.savings_vs_baseline),
            float(d.premium_vs_baseline_frac),
            float(d.price_score),
            float(d.coverage_score),
            float(d.hist_score),
            float(d.zrisk_score),
            float(d.continuity_score),
            float(d.rec_score),
            "Yes" if d.cap_breach else "No",
            "Yes" if d.is_fallback else "No",
            d.transit_days,
            d.relationship,
        ]
        for ci, (val, (_h, fmt)) in enumerate(zip(vals, _DATA_COLUMNS, strict=True), start=1):
            cell = ws.cell(row=row, column=ci, value=val)
            if fmt:
                cell.number_format = fmt
        row += 1

    n_rows = len(details)
    last_col = get_column_letter(len(_DATA_COLUMNS))
    body_end = header_row + n_rows
    # Column widths for legibility.
    for ci, (htext, _fmt) in enumerate(_DATA_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = max(11, min(26, len(htext) + 2))

    # Register the REAL Excel Table (ListObject) — gives the AutoFilter + the pivot-ready source.
    if n_rows > 0:
        ref = f"A{header_row}:{last_col}{body_end}"
        table = Table(displayName="AwardData", ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showRowStripes=True,
            showColumnStripes=False,
            showFirstColumn=False,
            showLastColumn=False,
        )
        ws.add_table(table)
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


# ---------------------------------------------------------------------------
# DESIGN-STUDY ADDITIONS (SCENARIO_TOOL_DESIGN_STUDY.md §4) — the high-value v3
# views we were light on, added clean (depth-on-demand, D27; single-purpose
# tabs, named for the question), grounded in our sealed records (§5):
#   * Lowest-Cost Check  — why the B recommendation ≠ the cheapest (governance).
#   * Coverage           — offered vs required volume + cover ratio + band.
#   * Detailed Scoring   — the 5 factors PLUS the market stats that explain them.
#   * TF Comparison      — TF1 vs TF2 per DC×lot (only when >1 TF).
#   * Round Evolution    — R1→…→Rn price movement per cell×supplier (only >1 round).
#   * Data Quality       — no-bids / missing coverage / advisory gate flags (never fatal).
# Visual design-language (color/type) is DEFERRED to the downstream design
# review; these add STRUCTURE / VIEWS / interaction only.
# ---------------------------------------------------------------------------
@dataclass(frozen=True)
class ScoreDetail:
    """One scored (supplier × cell) line + the per-group market stats that explain its scores."""

    dc_name: str
    lot_name: str
    item_name: str
    tf_name: str
    supplier_name: str
    price: Decimal
    mkt_min: Decimal
    mkt_avg: Decimal
    prem_vs_low_frac: Decimal  # (price − mkt_min) / mkt_min
    z_score: Decimal  # (price − mkt_avg) / mkt_std
    bidder_count: int
    price_score: Decimal
    coverage_score: Decimal
    hist_score: Decimal
    zrisk_score: Decimal
    continuity_score: Decimal
    rec_score: Decimal
    is_eligible: bool
    gate_flags: str
    is_recommended: bool  # the Scenario B pick for this cell
    is_incumbent: bool


def _gather_score_detail(
    session: Session,
    seeded: CycleView,
    analysis_run_id: str,
    final_round_id: str,
    award: AwardView,
) -> list[ScoreDetail]:
    """Resolve every scored bid to a detail line + the market stats (MktMin/Avg, Z, PremVsLow).

    The 5 factor scores + RecScore + eligibility come from `eng.bid_score`; the market stats are
    recomputed per group key [dc, lot, tf] from the persisted FINAL-round `bid.bid_line` prices (the
    same prices the engine scored). This is the auditability the proven v3 `Detailed Scoring` tab
    gives — the recommendation explained, not asserted (ADR-0006). All by KEY (D21), by NAME (D23).
    """

    dc_name = {dc.id: dc.name for dc in seeded.dcs}
    lot_name = {lot.id: lot.name for lot in seeded.lots}
    sup_name = {sup.id: sup.name for sup in seeded.suppliers}
    tf_name = {tf.id: tf.name for tf in seeded.tfs}
    item_for_lot = {seeded.lots[i].id: seeded.items[i].name for i in range(len(seeded.lots))}
    rec_keys = {(c.dc_id, c.lot_id, c.tf_id, c.supplier_id) for c in award.cells}
    inc_by = dict(seeded.incumbent_by_dc_lot)

    # Final-round prices per (supplier, dc, lot, tf) → per-group market stats (min/avg/std/count).
    # OPTION B (INTAKE §1a): bids are STORED flat at the 13 fiscal periods (≤13 identical rows per
    # cell). The market stats (esp. the bid COUNT `n`) are TIMEFRAME-grain, so collapse the period
    # rows to ONE price per (supplier, dc, lot, tf) with DISTINCT ON — otherwise every supplier's
    # price would be counted ≤13× and the group count/std would inflate. The fanned rows are
    # identical, so the deduped read is byte-identical to the pre-fan-out timeframe grain. Only
    # ACTIVE (`is_scoreable`) rows are read — mirrors the engine, so a superseded re-submission
    # can't inflate the group stats with a stale price.
    price_rows = session.execute(
        text(
            "SELECT DISTINCT ON (supplier_id, dc_id, lot_id, tf_id) "
            "supplier_id, dc_id, lot_id, tf_id, submitted_all_in_case, "
            "fob_case, delivery_surcharge_case, vegcool_surcharge_case, lot_discount_case "
            "FROM bid.bid_line WHERE cycle_id = :cyc AND round_id = :rnd "
            "AND is_scoreable = true "
            "ORDER BY supplier_id, dc_id, lot_id, tf_id, fiscal_period_id NULLS LAST, bid_line_id"
        ),
        {"cyc": seeded.cycle_id, "rnd": final_round_id},
    ).all()
    prices_by_group: dict[tuple[str, str, str], list[Decimal]] = defaultdict(list)
    price_by: dict[tuple[str, str, str, str], Decimal] = {}
    for sup_id, dc_id, lot_id, tf_id, all_in, fob, deliv, veg, disc in price_rows:
        p = _line_price(all_in, fob, deliv, veg, disc)
        if p is not None:
            prices_by_group[(dc_id, lot_id, tf_id)].append(p)
            price_by[(sup_id, dc_id, lot_id, tf_id)] = p

    def _stats(group: tuple[str, str, str]) -> tuple[Decimal, Decimal, Decimal, int]:
        vals = prices_by_group.get(group, [])
        n = len(vals)
        if n == 0:
            return Decimal("0"), Decimal("0"), Decimal("0"), 0
        mn = min(vals)
        avg = sum(vals, Decimal("0")) / Decimal(n)
        var = sum(((v - avg) ** 2 for v in vals), Decimal("0")) / Decimal(n)
        std = var.sqrt() if var > 0 else Decimal("0")
        return mn, avg, std, n

    score_rows = session.execute(
        text(
            "SELECT supplier_id, dc_id, lot_id, tf_id, price_score, coverage_score, hist_score, "
            "zrisk_score, continuity_score, rec_score, is_eligible, gate_flags "
            "FROM eng.bid_score WHERE analysis_run_id = :run"
        ),
        {"run": analysis_run_id},
    ).all()

    details: list[ScoreDetail] = []
    for sup_id, dc_id, lot_id, tf_id, ps, cov, hist, z, cont, rec, elig, flags in score_rows:
        mn, avg, std, n = _stats((dc_id, lot_id, tf_id))
        price = price_by.get((sup_id, dc_id, lot_id, tf_id), Decimal("0"))
        prem = (price - mn) / mn if mn > 0 else Decimal("0")
        zsc = (price - avg) / std if std > 0 else Decimal("0")
        sup_disp = sup_name.get(sup_id, sup_id[:6])
        details.append(
            ScoreDetail(
                dc_name=dc_name.get(dc_id, dc_id[:6]),
                lot_name=lot_name.get(lot_id, lot_id[:6]),
                item_name=item_for_lot.get(lot_id, ""),
                tf_name=tf_name.get(tf_id, tf_id[:6]),
                supplier_name=sup_disp,
                price=price,
                mkt_min=mn,
                mkt_avg=avg,
                prem_vs_low_frac=prem,
                z_score=zsc,
                bidder_count=n,
                price_score=Decimal(str(ps)),
                coverage_score=Decimal(str(cov)),
                hist_score=Decimal(str(hist)),
                zrisk_score=Decimal(str(z)),
                continuity_score=Decimal(str(cont)),
                rec_score=Decimal(str(rec)),
                is_eligible=bool(elig),
                gate_flags=flags or "",
                is_recommended=(dc_id, lot_id, tf_id, sup_id) in rec_keys,
                is_incumbent=inc_by.get((dc_id, lot_id)) == sup_id,
            )
        )
    details.sort(key=lambda d: (d.dc_name, d.lot_name, d.tf_name, -float(d.rec_score)))
    return details


def _rec_type_reason(rec_type: str, prem: Decimal, *, is_lowest: bool) -> str:
    """Render the engine's AUTHORITATIVE RecType label as a specific reason (D28).

    The category (Lowest cost / Coverage advantage / Comparable / Defensible / Risk-adjusted) is the
    engine's sealed `rec_type` — we RENDER it, we do not re-derive it. The premium % is from the
    sealed prices. No generic catch-all: each cell states the reason that actually applies.
    """

    pct = f"{prem * 100:.1f}% over the market low"
    if is_lowest:
        return "Recommended pick IS the market-low bid (the floor)."
    rendered = {
        "Lowest cost": f"Lowest cost — {pct}, effectively at the floor (within the 2% band).",
        "Coverage advantage": (
            f"Coverage advantage — clears >120% of required volume; {pct} buys supply security."
        ),
        "Comparable premium": f"Comparable — {pct}, inside the 3% comparable band.",
        "Defensible premium": f"Defensible — {pct}, inside the 7% defensible band.",
        "Risk-adjusted": (
            f"Risk-adjusted — {pct}, earned on RecScore (coverage/continuity), not price alone."
        ),
    }
    if rec_type in rendered:
        return rendered[rec_type]
    if rec_type:
        return f"{rec_type} — {pct}."
    return f"{pct} — risk-adjusted on RecScore. Benchmark = Scenario A."


def _write_lowest_cost_check_tab(
    wb: Workbook, cells: list[CellInfo], details: list[ScoreDetail]
) -> None:
    """Lowest-Cost Check — why the B recommendation is (or isn't) the cheapest bid (from v3).

    Per cell: the recommended supplier's $/case vs the market-low, the premium it carries, whether
    the rec IS the lowest, and a plain-language reason. This is the governance reconciliation the
    proven v3 tool gives and we lacked — recommends-not-asserts made legible (ADR-0006).
    """

    ws = wb.create_sheet("Lowest-Cost Check")
    # Per-cell recommended detail line, keyed by (dc,lot,tf) names via the rec flag.
    rec_by_cell = {(d.dc_name, d.lot_name, d.tf_name): d for d in details if d.is_recommended}
    columns = [
        Col("DC", 16),
        Col("Lot", 20),
        Col("Item", 22),
        Col("Timeframe", 16),
        Col("Recommended", 22),
        Col("Rec $/case", 12, NUMFMT_MONEY),
        Col("Market-low $/case", 14, NUMFMT_MONEY),
        Col("Premium vs Low", 13, NUMFMT_PCT),
        Col("Is Lowest?", 11),
        Col("Why not lowest", 42),
    ]
    header_row = 6
    row = header_row + 1
    n = 0
    for c in cells:
        d = rec_by_cell.get((c.dc_name, c.lot_name, c.tf_name))
        present = [p for p in c.price_by_supplier.values() if p is not None]
        min_price = min(present) if present else Decimal("0")
        rec_price = d.price if d else c.price_by_supplier.get(c.rec_supplier, Decimal("0"))
        prem = (rec_price - min_price) / min_price if min_price > 0 else Decimal("0")
        is_lowest = abs(rec_price - min_price) < Decimal("0.005")
        reason = _rec_type_reason(c.rec_type, prem, is_lowest=is_lowest)
        ws.cell(row=row, column=1, value=c.dc_name)
        ws.cell(row=row, column=2, value=c.lot_name)
        ws.cell(row=row, column=3, value=c.item_name)
        ws.cell(row=row, column=4, value=c.tf_name)
        ws.cell(row=row, column=5, value=c.rec_supplier)
        ws.cell(row=row, column=6, value=float(rec_price))
        ws.cell(row=row, column=7, value=float(min_price))
        ws.cell(row=row, column=8, value=float(prem))
        ws.cell(row=row, column=9, value="Yes" if is_lowest else "No")
        ws.cell(row=row, column=10, value=reason)
        row += 1
        n += 1

    fmt = format_table(
        ws,
        title="LOWEST-COST CHECK — is the recommendation the cheapest? if not, why?",
        subtitle_lines=[
            "Per cell: the Scenario B pick vs the market-low bid. Premium = the risk-adjusted "
            "cost traded for coverage/continuity. A = the lowest-cost benchmark.",
            f"SYNTHETIC · {DECISION_SUPPORT_STRAP}",
        ],
        columns=columns,
        n_body_rows=n,
        header_row=header_row,
        add_total=False,
    )
    body_start, body_end = fmt["body_start"], fmt["body_end"]
    if n > 0:
        # Premium-vs-low > eligibility ceiling → amber-flag the premium cell (advisory).
        ws.conditional_formatting.add(
            f"H{body_start}:H{body_end}",
            CellIsRule(
                operator="greaterThan",
                formula=["0.07"],
                fill=_INCUMBENT_FILL,
                font=Font(bold=True, color="9C6500"),
            ),
        )


def _write_detailed_scoring_tab(wb: Workbook, details: list[ScoreDetail]) -> None:
    """Detailed Scoring — the 5 factors PLUS the market stats that explain them (upgraded, from v3).

    The proven v3 `Detailed Scoring` exposes MktMin/Avg, Z-score, PremVsLow and bidder count next to
    the factor scores, so the recommendation is auditable. We were showing only the factors; this
    adds the explaining stats. Real Excel Table (filterable). Recommended + incumbent rows flagged.
    """

    ws = wb.create_sheet("Detailed Scoring")
    columns = [
        Col("DC", 15),
        Col("Lot", 18),
        Col("Timeframe", 15),
        Col("Supplier", 22),
        Col("$/case", 10, NUMFMT_MONEY),
        Col("Mkt Min", 10, NUMFMT_MONEY),
        Col("Mkt Avg", 10, NUMFMT_MONEY),
        Col("Prem vs Low", 11, NUMFMT_PCT),
        Col("Z-Score", 9, NUMFMT_MONEY),
        Col("# Bidders", 9, NUMFMT_INT),
        Col("Price", 8, NUMFMT_INT),
        Col("Cov", 7, NUMFMT_INT),
        Col("Hist", 7, NUMFMT_INT),
        Col("Z-Risk", 8, NUMFMT_INT),
        Col("Cont", 7, NUMFMT_INT),
        Col("RecScore", 10, NUMFMT_INT),
        Col("Eligible?", 10),
        Col("Rec/Inc", 12),
        Col("Gate flags", 26),
    ]
    header_row = 5
    row = header_row + 1
    for d in details:
        tag = []
        if d.is_recommended:
            tag.append("REC")
        if d.is_incumbent:
            tag.append("INC")
        ws.cell(row=row, column=1, value=d.dc_name)
        ws.cell(row=row, column=2, value=d.lot_name)
        ws.cell(row=row, column=3, value=d.tf_name)
        ws.cell(row=row, column=4, value=d.supplier_name)
        ws.cell(row=row, column=5, value=float(d.price))
        ws.cell(row=row, column=6, value=float(d.mkt_min))
        ws.cell(row=row, column=7, value=float(d.mkt_avg))
        ws.cell(row=row, column=8, value=float(d.prem_vs_low_frac))
        ws.cell(row=row, column=9, value=float(d.z_score))
        ws.cell(row=row, column=10, value=d.bidder_count)
        ws.cell(row=row, column=11, value=float(d.price_score))
        ws.cell(row=row, column=12, value=float(d.coverage_score))
        ws.cell(row=row, column=13, value=float(d.hist_score))
        ws.cell(row=row, column=14, value=float(d.zrisk_score))
        ws.cell(row=row, column=15, value=float(d.continuity_score))
        ws.cell(row=row, column=16, value=float(d.rec_score))
        ws.cell(row=row, column=17, value="Yes" if d.is_eligible else "No")
        ws.cell(row=row, column=18, value="·".join(tag) if tag else "—")
        ws.cell(row=row, column=19, value=d.gate_flags or "—")
        row += 1

    format_table(
        ws,
        title="DETAILED SCORING — the five factors + the market stats that explain them",
        subtitle_lines=[
            "Per scored bid (final round). Market stats (Mkt Min/Avg, Prem vs Low, Z-Score, "
            "# bidders) explain the factor scores → RecScore. Eligible? = passed the gates. "
            "REC = Scenario B pick; INC = incumbent.",
            f"SYNTHETIC · {DECISION_SUPPORT_STRAP}",
        ],
        columns=columns,
        n_body_rows=len(details),
        header_row=header_row,
        add_total=False,
    )


@dataclass(frozen=True)
class CoverageRow:
    """One (supplier × cell) coverage line: offered vs required volume + cover ratio + band."""

    dc_name: str
    lot_name: str
    tf_name: str
    supplier_name: str
    price: Decimal
    req_cases: Decimal
    offered_cases: Decimal
    cover_ratio: Decimal
    band: str
    is_eligible: bool
    is_recommended: bool


def _gather_coverage(
    session: Session,
    seeded: CycleView,
    analysis_run_id: str,
    final_round_id: str,
    award: AwardView,
) -> list[CoverageRow]:
    """Per (supplier × cell): offered vs required volume + cover ratio + band (from v3 Coverage).

    Required cases come from the seeded projected volume; offered cases from the persisted
    `bid.bid_line.volume_minimum_cases` (the vol-offered captured at intake); eligibility from
    `eng.bid_score`. The coverage band mirrors v3's floors (the 0.80 eligibility floor). The view we
    lacked — capacity reality next to price, not folded into a single score.
    """

    dc_name = {dc.id: dc.name for dc in seeded.dcs}
    lot_name = {lot.id: lot.name for lot in seeded.lots}
    sup_name = {sup.id: sup.name for sup in seeded.suppliers}
    tf_name = {tf.id: tf.name for tf in seeded.tfs}
    req_by = dict(seeded.period_cases_by_cell)  # (dc, lot, tf) -> required period cases
    rec_keys = {(c.dc_id, c.lot_id, c.tf_id, c.supplier_id) for c in award.cells}

    elig_rows = session.execute(
        text(
            "SELECT supplier_id, dc_id, lot_id, tf_id, is_eligible "
            "FROM eng.bid_score WHERE analysis_run_id = :run"
        ),
        {"run": analysis_run_id},
    ).all()
    elig_by = {(s, d, lo, t): bool(e) for s, d, lo, t, e in elig_rows}

    # OPTION B (INTAKE §1a): bids are STORED flat at the 13 fiscal periods (≤13 identical rows per
    # cell). Coverage is per (supplier × cell), so collapse the period rows to ONE per (supplier,
    # dc, lot, tf) with DISTINCT ON — otherwise every cell would emit ≤13 duplicate coverage rows.
    # The fanned rows are identical, so the deduped read matches the pre-fan-out timeframe grain.
    # Only ACTIVE (`is_scoreable`) rows are read — mirrors the engine, so a superseded re-submission
    # can't report stale coverage.
    bid_rows = session.execute(
        text(
            "SELECT DISTINCT ON (supplier_id, dc_id, lot_id, tf_id) "
            "supplier_id, dc_id, lot_id, tf_id, submitted_all_in_case, "
            "fob_case, delivery_surcharge_case, vegcool_surcharge_case, lot_discount_case, "
            "volume_minimum_cases "
            "FROM bid.bid_line WHERE cycle_id = :cyc AND round_id = :rnd "
            "AND is_scoreable = true "
            "ORDER BY supplier_id, dc_id, lot_id, tf_id, fiscal_period_id NULLS LAST, bid_line_id"
        ),
        {"cyc": seeded.cycle_id, "rnd": final_round_id},
    ).all()

    rows: list[CoverageRow] = []
    for sup_id, dc_id, lot_id, tf_id, all_in, fob, deliv, veg, disc, offered in bid_rows:
        price = _line_price(all_in, fob, deliv, veg, disc)
        if price is None:
            continue
        req = req_by.get((dc_id, lot_id, tf_id), Decimal("0"))
        off = Decimal(str(offered)) if offered is not None else Decimal("0")
        ratio = off / req if req > 0 else Decimal("0")
        if req <= 0:
            band = "As-Needed"
        elif ratio < Decimal("0.50"):
            band = "Critical (<50%)"
        elif ratio < Decimal("0.80"):
            band = "Short (<80%)"
        elif ratio < Decimal("1.00"):
            band = "Partial (<100%)"
        elif ratio <= Decimal("1.20"):
            band = "Full"
        else:
            band = "Surplus (>120%)"
        rows.append(
            CoverageRow(
                dc_name=dc_name.get(dc_id, dc_id[:6]),
                lot_name=lot_name.get(lot_id, lot_id[:6]),
                tf_name=tf_name.get(tf_id, tf_id[:6]),
                supplier_name=sup_name.get(sup_id, sup_id[:6]),
                price=Decimal(str(price)),
                req_cases=req,
                offered_cases=off,
                cover_ratio=ratio,
                band=band,
                is_eligible=elig_by.get((sup_id, dc_id, lot_id, tf_id), False),
                is_recommended=(dc_id, lot_id, tf_id, sup_id) in rec_keys,
            )
        )
    rows.sort(key=lambda r: (r.dc_name, r.lot_name, r.tf_name, -float(r.cover_ratio)))
    return rows


def _write_coverage_tab(wb: Workbook, rows: list[CoverageRow], config: EngineConfig) -> None:
    """Coverage — offered vs required volume + cover ratio + band per (supplier × cell)."""

    ws = wb.create_sheet("Coverage")
    columns = [
        Col("DC", 16),
        Col("Lot", 20),
        Col("Timeframe", 16),
        Col("Supplier", 22),
        Col("$/case", 11, NUMFMT_MONEY),
        Col("Req (cases)", 12, NUMFMT_INT),
        Col("Offered (cases)", 13, NUMFMT_INT),
        Col("Cover Ratio", 12, NUMFMT_PCT),
        Col("Band", 16),
        Col("Eligible?", 10),
        Col("Rec?", 8),
    ]
    header_row = 6
    row = header_row + 1
    for r in rows:
        ws.cell(row=row, column=1, value=r.dc_name)
        ws.cell(row=row, column=2, value=r.lot_name)
        ws.cell(row=row, column=3, value=r.tf_name)
        ws.cell(row=row, column=4, value=r.supplier_name)
        ws.cell(row=row, column=5, value=float(r.price))
        ws.cell(row=row, column=6, value=float(r.req_cases))
        ws.cell(row=row, column=7, value=float(r.offered_cases))
        ws.cell(row=row, column=8, value=float(r.cover_ratio))
        ws.cell(row=row, column=9, value=r.band)
        ws.cell(row=row, column=10, value="Yes" if r.is_eligible else "No")
        ws.cell(row=row, column=11, value="Yes" if r.is_recommended else "—")
        row += 1

    fmt = format_table(
        ws,
        title="COVERAGE — offered vs required volume (capacity reality next to price)",
        subtitle_lines=[
            f"Offered vs required cases → cover ratio + band. The "
            f"{config.coverage_floor * 100:.0f}% floor is the eligibility gate; "
            "Short/Critical bands flag capacity risk (advisory).",
            f"SYNTHETIC · {DECISION_SUPPORT_STRAP}",
        ],
        columns=columns,
        n_body_rows=len(rows),
        header_row=header_row,
        add_total=False,
    )
    body_start, body_end = fmt["body_start"], fmt["body_end"]
    if rows:
        # Cover ratio below the floor → red (the capacity risk the team must see).
        ws.conditional_formatting.add(
            f"H{body_start}:H{body_end}",
            CellIsRule(
                operator="lessThan",
                formula=[str(float(config.coverage_floor))],
                fill=_BREACH_FILL,
                font=_BREACH_FONT,
            ),
        )


@dataclass(frozen=True)
class CapacityCheckDisplayRow:
    """One awarded cell vs the supplier's stated capacity, resolved to names (E-38b)."""

    dc_name: str
    lot_name: str
    tf_name: str
    supplier_name: str
    allocated_cases: Decimal
    allocated_weekly_cases: Decimal
    max_period_cases: Decimal | None
    max_weekly_cases: Decimal | None
    status: str
    over_capacity: bool
    has_statement: bool


def _gather_capacity_check(
    session: Session, seeded: CycleView, award: AwardView
) -> list[CapacityCheckDisplayRow]:
    """Allocation-vs-stated-capacity for the recommended award, resolved to names (E-38b / G-G).

    Reads the cycle's ACTIVE stated ceilings (`load_active_capacity`) and compares each awarded cell
    (supplier × dc × lot × tf, allocation = period_cases × volume_share) against them via the pure
    `evaluate_capacity`. Decision-support only — surfaces over-capacity; never changes the award.
    Over-capacity rows sort to the top so the buyer sees the risk first.
    """

    dc_name = {dc.id: dc.name for dc in seeded.dcs}
    lot_name = {lot.id: lot.name for lot in seeded.lots}
    tf_name = {tf.id: tf.name for tf in seeded.tfs}
    sup_name = {sup.id: sup.name for sup in seeded.suppliers}

    capacity = load_active_capacity(session, seeded.cycle_id)
    rows = evaluate_capacity(award.cells, capacity, weeks_per_tf=WEEKS_PER_TF)

    out = [
        CapacityCheckDisplayRow(
            dc_name=dc_name.get(r.dc_id, r.dc_id[:6]),
            lot_name=lot_name.get(r.lot_id, r.lot_id[:6]),
            tf_name=tf_name.get(r.tf_id, r.tf_id[:6]),
            supplier_name=sup_name.get(r.supplier_id, r.supplier_id[:6]),
            allocated_cases=r.allocated_cases,
            allocated_weekly_cases=r.allocated_weekly_cases,
            max_period_cases=r.max_period_cases,
            max_weekly_cases=r.max_weekly_cases,
            status=r.status,
            over_capacity=r.over_capacity,
            has_statement=r.has_statement,
        )
        for r in rows
    ]
    # Over-capacity first, then by cell — the risk rows lead.
    out.sort(key=lambda x: (not x.over_capacity, x.dc_name, x.lot_name, x.tf_name, x.supplier_name))
    return out


def _write_capacity_check_tab(wb: Workbook, rows: list[CapacityCheckDisplayRow]) -> None:
    """Capacity Check — recommended allocation vs each supplier's STATED capacity (E-38b / G-G).

    The operator-facing safety surface: "are we recommending beyond what a supplier said they can
    supply?" Over-capacity cells are flagged red. Decision-support only (ADR-0006) — it never
    changes the award; the buyer decides.
    """

    ws = wb.create_sheet("Capacity Check")
    columns = [
        Col("DC", 16),
        Col("Lot", 20),
        Col("Timeframe", 16),
        Col("Supplier", 22),
        Col("Allocated (cases)", 14, NUMFMT_INT),
        Col("Allocated / wk", 13, NUMFMT_INT),
        Col("Stated Max (period)", 16, NUMFMT_INT),
        Col("Stated Max (weekly)", 16, NUMFMT_INT),
        Col("Status", 18),
    ]
    header_row = 6
    row = header_row + 1
    for r in rows:
        ws.cell(row=row, column=1, value=r.dc_name)
        ws.cell(row=row, column=2, value=r.lot_name)
        ws.cell(row=row, column=3, value=r.tf_name)
        ws.cell(row=row, column=4, value=r.supplier_name)
        ws.cell(row=row, column=5, value=float(r.allocated_cases))
        ws.cell(row=row, column=6, value=float(r.allocated_weekly_cases))
        ws.cell(
            row=row,
            column=7,
            value=float(r.max_period_cases) if r.max_period_cases is not None else "—",
        )
        ws.cell(
            row=row,
            column=8,
            value=float(r.max_weekly_cases) if r.max_weekly_cases is not None else "—",
        )
        ws.cell(row=row, column=9, value=r.status)
        row += 1

    over = sum(1 for r in rows if r.over_capacity)
    no_stmt = sum(1 for r in rows if not r.has_statement)
    within = len(rows) - over - no_stmt
    fmt = format_table(
        ws,
        title="CAPACITY CHECK — recommended allocation vs each supplier's STATED capacity",
        subtitle_lines=[
            f"{over} cell(s) OVER stated capacity · {within} within · {no_stmt} with no stated "
            f"ceiling. Weekly = allocated ÷ {WEEKS_PER_TF} wks/TF; a cell flags if EITHER the "
            "period or weekly ceiling is exceeded.",
            "Decision-support: flags where the recommendation books a supplier beyond what they "
            "stated they can supply — it never changes the award (E-38b / G-G).",
            f"SYNTHETIC · {DECISION_SUPPORT_STRAP}",
        ],
        columns=columns,
        n_body_rows=len(rows),
        header_row=header_row,
        add_total=False,
    )
    body_start, body_end = fmt["body_start"], fmt["body_end"]
    if rows:
        # OVER CAPACITY → red (the supply risk the buyer must see before booking).
        ws.conditional_formatting.add(
            f"I{body_start}:I{body_end}",
            CellIsRule(
                operator="equal",
                formula=['"OVER CAPACITY"'],
                fill=_BREACH_FILL,
                font=_BREACH_FONT,
            ),
        )


def _write_tf_comparison_tab(wb: Workbook, cells: list[CellInfo]) -> None:
    """TF Comparison — TF1 vs TF2 rec supplier/price per DC×lot, with a split flag (from v3).

    Only written when the cycle has >1 timeframe. Per DC×lot: the recommended supplier + price in
    each TF side by side, whether it's the SAME supplier across TFs, and a SplitFlag when the TFs
    award differently — the seasonal-split story the proven v3 tool surfaces.
    """

    ws = wb.create_sheet("TF Comparison")
    tf_names = sorted({c.tf_name for c in cells})
    # Group cells by (dc, lot, item) → row; one column pair per TF (rec supplier + rec $/case).
    by_dclot: dict[tuple[str, str, str], dict[str, tuple[str, Decimal]]] = defaultdict(dict)
    for c in cells:
        by_dclot[(c.dc_name, c.lot_name, c.item_name)][c.tf_name] = (
            c.rec_supplier,
            c.price_by_supplier.get(c.rec_supplier, Decimal("0")),
        )

    columns: list[Col] = [Col("DC", 16), Col("Lot", 20), Col("Item", 22)]
    for tfn in tf_names:
        columns.append(Col(f"{tfn} — Supplier", 22))
        columns.append(Col(f"{tfn} — $/case", 12, NUMFMT_MONEY))
    columns.append(Col("Same supplier?", 13))
    columns.append(Col("Split flag", 12))

    header_row = 6
    row = header_row + 1
    for dcn, lotn, itemn in sorted(by_dclot):
        per_tf = by_dclot[(dcn, lotn, itemn)]
        ws.cell(row=row, column=1, value=dcn)
        ws.cell(row=row, column=2, value=lotn)
        ws.cell(row=row, column=3, value=itemn)
        col = 4
        sups_seen: set[str] = set()
        for tfn in tf_names:
            sup, price = per_tf.get(tfn, ("—", Decimal("0")))
            ws.cell(row=row, column=col, value=sup)
            ws.cell(row=row, column=col + 1, value=float(price) if price else None)
            if sup and sup != "—":
                sups_seen.add(sup)
            col += 2
        same = len(sups_seen) <= 1
        ws.cell(row=row, column=col, value="Yes" if same else "No")
        ws.cell(row=row, column=col + 1, value="—" if same else "⚠ SPLIT")
        row += 1

    format_table(
        ws,
        title="TF COMPARISON — the recommendation across timeframes (seasonal split?)",
        subtitle_lines=[
            "Per DC × lot: the recommended supplier + $/case in each timeframe, side by side. "
            "SPLIT = the timeframes award different suppliers (a seasonal split to align on).",
            f"SYNTHETIC · {DECISION_SUPPORT_STRAP}",
        ],
        columns=columns,
        n_body_rows=row - (header_row + 1),
        header_row=header_row,
        add_total=False,
    )


@dataclass(frozen=True)
class RoundEvoRow:
    """One (supplier × cell) round-evolution line: the per-round price seq + first→last Δ."""

    dc_name: str
    lot_name: str
    tf_name: str
    supplier_name: str
    prices: list[Decimal | None]  # one entry per round (None = no priced bid that round)
    delta: Decimal  # last priced − first priced
    pct: Decimal
    direction: str


def _gather_round_evolution(
    session: Session, seeded: CycleView, round_ids: list[str]
) -> list[RoundEvoRow]:
    """Per (supplier × cell): the priced $/case across each round R1..Rn + first→last Δ + direction.

    Reads the priced `bid.bid_line` rows for EVERY round (we persist all rounds; the engine scores
    only the final). Returns rows only where >=2 rounds have a price, so the movement is real. The
    negotiation story the proven v3 `Round Evolution` tool surfaces and we lacked.
    """

    dc_name = {dc.id: dc.name for dc in seeded.dcs}
    lot_name = {lot.id: lot.name for lot in seeded.lots}
    sup_name = {sup.id: sup.name for sup in seeded.suppliers}
    tf_name = {tf.id: tf.name for tf in seeded.tfs}

    # ACTIVE rows only (`is_scoreable`): for a re-submitted round the superseded and current prices
    # share the same (cell, round) key, so without this filter the per-round price would be the
    # arbitrary last-written row. Fan-out is a non-issue — the per-period rows replicate one price
    # per (cell, round), so the dict assignment is idempotent.
    rows = session.execute(
        text(
            "SELECT supplier_id, dc_id, lot_id, tf_id, round_id, submitted_all_in_case, "
            "fob_case, delivery_surcharge_case, vegcool_surcharge_case, lot_discount_case "
            "FROM bid.bid_line WHERE cycle_id = :cyc AND is_scoreable = true"
        ),
        {"cyc": seeded.cycle_id},
    ).all()
    # (sup,dc,lot,tf) -> {round_id: price}
    by_cell: dict[tuple[str, str, str, str], dict[str, Decimal]] = defaultdict(dict)
    for sup_id, dc_id, lot_id, tf_id, rnd_id, all_in, fob, deliv, veg, disc in rows:
        price = _line_price(all_in, fob, deliv, veg, disc)
        if price is not None:
            by_cell[(sup_id, dc_id, lot_id, tf_id)][rnd_id] = price

    out: list[RoundEvoRow] = []
    for (sup_id, dc_id, lot_id, tf_id), prices in by_cell.items():
        seq: list[Decimal | None] = [prices.get(rid) for rid in round_ids]
        present = [p for p in seq if p is not None]
        if len(present) < 2:
            continue
        first, last = present[0], present[-1]
        delta = last - first
        pct = delta / first if first > 0 else Decimal("0")
        direction = "↓ down" if delta < 0 else ("↑ up" if delta > 0 else "→ flat")
        out.append(
            RoundEvoRow(
                dc_name=dc_name.get(dc_id, dc_id[:6]),
                lot_name=lot_name.get(lot_id, lot_id[:6]),
                tf_name=tf_name.get(tf_id, tf_id[:6]),
                supplier_name=sup_name.get(sup_id, sup_id[:6]),
                prices=seq,
                delta=delta,
                pct=pct,
                direction=direction,
            )
        )
    out.sort(key=lambda r: (r.dc_name, r.lot_name, r.tf_name, r.supplier_name))
    return out


def _write_round_evolution_tab(
    wb: Workbook, rows: list[RoundEvoRow], round_labels: list[str]
) -> None:
    """Round Evolution — $/case per round + first→last Δ + direction per (supplier × cell)."""

    ws = wb.create_sheet("Round Evolution")
    columns: list[Col] = [Col("DC", 16), Col("Lot", 20), Col("Timeframe", 16), Col("Supplier", 22)]
    for lbl in round_labels:
        columns.append(Col(lbl, 13, NUMFMT_MONEY))
    columns += [
        Col("Δ first→last", 13, NUMFMT_MONEY),
        Col("Δ %", 10, NUMFMT_PCT),
        Col("Direction", 11),
    ]
    header_row = 6
    row = header_row + 1
    for r in rows:
        ws.cell(row=row, column=1, value=r.dc_name)
        ws.cell(row=row, column=2, value=r.lot_name)
        ws.cell(row=row, column=3, value=r.tf_name)
        ws.cell(row=row, column=4, value=r.supplier_name)
        col = 5
        for p in r.prices:
            ws.cell(row=row, column=col, value=float(p) if p is not None else None)
            col += 1
        ws.cell(row=row, column=col, value=float(r.delta))
        ws.cell(row=row, column=col + 1, value=float(r.pct))
        ws.cell(row=row, column=col + 2, value=r.direction)
        row += 1

    format_table(
        ws,
        title="ROUND EVOLUTION — how each bid moved across the negotiation rounds",
        subtitle_lines=[
            "Per supplier × cell: the priced $/case in each round, then first→last movement. "
            "Down = competitive tension working. Only cells priced in ≥2 rounds shown.",
            f"SYNTHETIC · {DECISION_SUPPORT_STRAP}",
        ],
        columns=columns,
        n_body_rows=len(rows),
        header_row=header_row,
        add_total=False,
    )


def _write_data_quality_tab(
    wb: Workbook,
    seeded: CycleView,
    cells: list[CellInfo],
    details: list[ScoreDetail],
) -> None:
    """Data Quality — no-bid cells / advisory gate flags surfaced, never fatal (v3 Missing Data).

    Mirrors the proven v3 `Missing Data` tab: a flag / count / detail register of the data-quality
    conditions the team should see — suppliers that didn't bid a cell, cells with thin competition,
    and advisory gate flags — surfaced (transparency) but non-blocking (decision-support, ADR-0006).
    """

    ws = wb.create_sheet("Data Quality")
    n_sup = len(seeded.suppliers)

    # No-bid count: per cell, suppliers invited minus suppliers who priced.
    no_bid_total = sum(max(0, n_sup - len(c.price_by_supplier)) for c in cells)
    thin_cells = sum(1 for c in cells if len(c.price_by_supplier) < 3)
    gate_flagged = sum(1 for d in details if d.gate_flags)
    ineligible = sum(1 for d in details if not d.is_eligible)

    issues: list[tuple[str, int, str]] = [
        (
            "No-bid cells (supplier × cell with no priced bid)",
            no_bid_total,
            "Suppliers that did not price a cell. Expected (suppliers decline cells); the cell is "
            "still scored from those who bid. Non-blocking.",
        ),
        (
            "Thin-competition cells (<3 bidders)",
            thin_cells,
            "Cells with fewer than 3 priced bids — the Z-score is less reliable; advisory "
            "flag, the bid stays eligible.",
        ),
        (
            "Bids carrying an advisory gate flag",
            gate_flagged,
            "Outlier / low-bidder-count notes accumulated in gate_flags. Advisory — does not by "
            "itself make a bid ineligible.",
        ),
        (
            "Bids gated ineligible",
            ineligible,
            "Failed a hard gate (no valid price / premium over ceiling / coverage < floor). Scored "
            "but not awardable. Recommends-not-asserts: surfaced for review.",
        ),
    ]
    columns = [Col("Flag", 46), Col("Count", 9, NUMFMT_INT), Col("Detail / Action", 70)]
    header_row = 6
    row = header_row + 1
    for flag, count, detail in issues:
        ws.cell(row=row, column=1, value=flag)
        ws.cell(row=row, column=2, value=count)
        c3 = ws.cell(row=row, column=3, value=detail)
        c3.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 42
        row += 1

    format_table(
        ws,
        title="DATA QUALITY — surfaced, never hidden, never fatal",
        subtitle_lines=[
            "The data-quality conditions the team should see before aligning. All non-blocking "
            "(decision-support) — the run completes and recommends; the human reviews these.",
            f"SYNTHETIC · {DECISION_SUPPORT_STRAP}",
        ],
        columns=columns,
        n_body_rows=len(issues),
        header_row=header_row,
        add_total=False,
        add_autofilter=False,
    )


def _augment_summary_index(wb: Workbook, tab_index: list[tuple[str, str]]) -> None:
    """Add a 'What's in this workbook' tab index to the Summary tab — the front door (study §3.7).

    The proven v3 tool proved 20 flat tabs need navigation. With ~12 tabs we add a compact index on
    the Overview/Summary tab: each tab + the one question it answers, so the buyer knows where to.
    """

    ws = wb["Summary"]
    start = (ws.max_row or 1) + 2
    title = ws.cell(row=start, column=1, value="What's in this workbook (where to go)")
    title.font = _TITLE_FONT
    title.fill = _TITLE_FILL
    title.alignment = _LEFT
    ws.merge_cells(start_row=start, start_column=1, end_row=start, end_column=2)
    ws.cell(row=start, column=2).fill = _TITLE_FILL
    r = start + 1
    hc1 = ws.cell(row=r, column=1, value="Tab")
    hc2 = ws.cell(row=r, column=2, value="Answers the question")
    for hc in (hc1, hc2):
        hc.font = _HEADER_FONT
        hc.fill = _HEADER_FILL
        hc.border = _BORDER
        hc.alignment = _CENTER
    r += 1
    for tab, question in tab_index:
        # A band separator (empty question) groups the tabs into the decision flow
        # (Decide -> Compare -> Diligence -> Build & slice) the real allocation models use.
        if question == "":
            band = ws.cell(row=r, column=1, value=tab)
            band.font = Font(bold=True, italic=True, color="1F3864", size=10)
            band.fill = _TOTAL_FILL
            band.alignment = _LEFT
            band.border = _BORDER
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
            ws.cell(row=r, column=2).fill = _TOTAL_FILL
            ws.cell(row=r, column=2).border = _BORDER
            ws.row_dimensions[r].height = 18
            r += 1
            continue
        c1 = ws.cell(row=r, column=1, value=tab)
        c1.font = _TOTAL_FONT
        c1.alignment = _LEFT
        c1.border = _BORDER
        c2 = ws.cell(row=r, column=2, value=question)
        c2.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c2.border = _BORDER
        ws.row_dimensions[r].height = 22
        r += 1


def _write_controls_tab(
    wb: Workbook,
    seeded: CycleView,
    config: EngineConfig,
    baseline_total: Decimal,
    stly_total: Decimal,
    rec_spend: Decimal,
    negotiation_savings: Decimal,
) -> None:
    """Controls / Assumptions — the cockpit (how this cycle was run + the frozen engine config).

    The real allocation models open with a `Controls` tab (commodity, horizon, weeks/periods,
    volume) that drives the whole model. We surface the run's assumptions, the savings baselines,
    and the frozen `EngineConfig` weights/thresholds in ONE place so every downstream number is
    traceable to how the engine was parameterised. Key/value, banded by section.
    """

    ws = wb.create_sheet("Controls")
    next_row = _title_block(
        ws,
        title="Controls & Assumptions — how this cycle was run",
        subtitle_lines=[
            seeded.cycle_name,
            DECISION_SUPPORT_STRAP,
        ],
        span=2,
    )

    total_cases = sum(seeded.period_cases_by_cell.values(), Decimal("0"))
    # Real horizon = sum of the cycle's timeframe week counts (falls back to the legacy estimate
    # only if the loader could not resolve week counts).
    total_weeks = seeded.horizon_weeks or (len(seeded.tfs) * WEEKS_PER_TF)
    rec_savings = baseline_total - rec_spend

    # (section | label | value | numfmt)  — numfmt None => text.
    rows: list[tuple[str, str, object, str | None]] = [
        ("Cycle", "Commodity", seeded.commodity_name or seeded.cycle_name, None),
        ("Cycle", "Cycle code", seeded.cycle_code, None),
        ("Cycle", "Horizon (weeks)", total_weeks, NUMFMT_INT),
        ("Cycle", "Timeframes (seasons)", len(seeded.tfs), NUMFMT_INT),
        ("Cycle", "Rounds run", len(seeded.rounds), NUMFMT_INT),
        ("Scope", "DCs", len(seeded.dcs), NUMFMT_INT),
        ("Scope", "Lots (items)", len(seeded.lots), NUMFMT_INT),
        ("Scope", "Suppliers invited", len(seeded.suppliers), NUMFMT_INT),
        ("Scope", "Total projected cases (period)", total_cases, NUMFMT_INT),
        ("Baselines", "Incumbent baseline spend (iTrade routing)", baseline_total, NUMFMT_MONEY),
        ("Baselines", "STLY baseline (modeled — iTrade pending)", stly_total, NUMFMT_MONEY),
        ("Baselines", "Recommended (Scenario B) spend", rec_spend, NUMFMT_MONEY),
        ("Baselines", "Savings vs incumbent (period $)", rec_savings, NUMFMT_MONEY),
        ("Baselines", "Negotiation savings R1→Final (period $)", negotiation_savings, NUMFMT_MONEY),
        ("Engine weights", "Price weight", config.weight_price, NUMFMT_PCT),
        ("Engine weights", "Coverage weight", config.weight_coverage, NUMFMT_PCT),
        ("Engine weights", "Historical weight", config.weight_historical, NUMFMT_PCT),
        ("Engine weights", "Z-Risk weight", config.weight_zrisk, NUMFMT_PCT),
        ("Engine weights", "Continuity weight", config.weight_continuity, NUMFMT_PCT),
        ("Engine rules", "Max suppliers per DC (split cap)", config.max_sup_dc, NUMFMT_INT),
        ("Engine rules", "Concentration flag threshold", config.conc_thresh, NUMFMT_PCT),
        ("Engine rules", "Global premium ceiling", config.global_premium_threshold, NUMFMT_PCT),
        ("Engine rules", "Coverage eligibility floor", config.coverage_floor, NUMFMT_PCT),
    ]

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 30
    r = next_row
    current_section = ""
    for section, label, value, numfmt in rows:
        if section != current_section:
            band = ws.cell(row=r, column=1, value=section)
            band.font = _HEADER_FONT
            band.fill = _HEADER_FILL
            band.alignment = _LEFT
            band.border = _BORDER
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
            ws.cell(row=r, column=2).fill = _HEADER_FILL
            ws.cell(row=r, column=2).border = _BORDER
            current_section = section
            r += 1
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = _TOTAL_FONT
        lc.alignment = _LEFT
        lc.border = _BORDER
        vc = ws.cell(row=r, column=2, value=value)
        vc.alignment = _CENTER if numfmt else _LEFT
        vc.border = _BORDER
        if numfmt:
            vc.number_format = numfmt
        r += 1

    note = ws.cell(
        row=r + 1,
        column=1,
        value="Schema-backed: incumbent baseline, savings, weights, rules, rounds. "
        "Modeled (pending real feeds): STLY uplift (iTrade), product type — labelled where shown.",
    )
    note.font = Font(italic=True, color="808080", size=9)
    note.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=2)
    ws.row_dimensions[r + 1].height = 30
    ws.sheet_view.showGridLines = False


@dataclass
class _DcSignoff:
    """One DC rolled up for the savings-first sign-off table (the headline output)."""

    dc_name: str
    incumbent: str
    recommended: str
    n_cells: int
    rec_spend: Decimal
    baseline_spend: Decimal
    stly_spend: Decimal
    negotiation: Decimal


def _write_award_summary_tab(
    wb: Workbook,
    details: list[AwardDetail],
    award_scenario_code: str,
    award_scenario_label: str,
    negotiation_by_dc: dict[str, Decimal],
    lot_product_type: dict[str, str],
) -> None:
    """Award Summary (Sign-off) — THE headline: per DC, Incumbent → Recommended + savings.

    The savings-first table the real allocation models put up front and buyers sign off on:
    per DC the incumbent vs the recommended supplier(s), the recommended period spend, savings
    in DOLLARS against TWO baselines (incumbent + STLY), and the round-over-round negotiation
    capture. Plus a Conventional/Organic split (the product-type segmentation the real models
    carry). Decision-support — recommends, does not assert (ADR-0006).
    """

    ws = wb.create_sheet("Award Summary")
    rec = [d for d in details if d.scenario_code == award_scenario_code]

    by_dc: dict[str, _DcSignoff] = {}
    for d in sorted(rec, key=lambda x: x.dc_name):
        s = by_dc.get(d.dc_name)
        if s is None:
            s = _DcSignoff(
                dc_name=d.dc_name,
                incumbent=d.incumbent_name,
                recommended="",
                n_cells=0,
                rec_spend=Decimal("0"),
                baseline_spend=Decimal("0"),
                stly_spend=Decimal("0"),
                negotiation=negotiation_by_dc.get(d.dc_name, Decimal("0")),
            )
            by_dc[d.dc_name] = s
        s.n_cells += 1
        s.rec_spend += d.spend
        s.baseline_spend += d.baseline_spend
        s.stly_spend += d.baseline_spend * _STLY_UPLIFT
    # recommended supplier list per DC (distinct, readable)
    sups_by_dc: dict[str, set[str]] = defaultdict(set)
    for d in rec:
        sups_by_dc[d.dc_name].add(d.supplier_name)
    for dcn, s in by_dc.items():
        s.recommended = ", ".join(sorted(sups_by_dc[dcn]))

    columns: list[Col] = [
        Col("DC", 16),
        Col("Incumbent", 22),
        Col("Recommended", 26),
        Col("Cells", 8, NUMFMT_INT, total="sum"),
        Col("Rec spend (period)", 18, NUMFMT_MONEY, total="sum"),
        Col("Incumbent baseline", 18, NUMFMT_MONEY, total="sum"),
        Col("Savings vs incumbent $", 18, NUMFMT_MONEY, total="sum"),
        Col("Savings vs incumbent %", 16, NUMFMT_PCT),
        Col("Savings vs STLY $", 16, NUMFMT_MONEY, total="sum"),
        Col("Negotiation R1→Final $", 18, NUMFMT_MONEY, total="sum"),
    ]
    header_row = 5
    body_start = header_row + 1
    ordered = sorted(by_dc.values(), key=lambda x: x.dc_name)
    for i, s in enumerate(ordered):
        r = body_start + i
        sav_inc = s.baseline_spend - s.rec_spend
        sav_inc_pct = sav_inc / s.baseline_spend if s.baseline_spend > 0 else Decimal("0")
        sav_stly = s.stly_spend - s.rec_spend
        ws.cell(row=r, column=1, value=s.dc_name)
        ws.cell(row=r, column=2, value=s.incumbent)
        ws.cell(row=r, column=3, value=s.recommended)
        ws.cell(row=r, column=4, value=s.n_cells)
        ws.cell(row=r, column=5, value=float(s.rec_spend))
        ws.cell(row=r, column=6, value=float(s.baseline_spend))
        ws.cell(row=r, column=7, value=float(sav_inc))
        ws.cell(row=r, column=8, value=float(sav_inc_pct))
        ws.cell(row=r, column=9, value=float(sav_stly))
        ws.cell(row=r, column=10, value=float(s.negotiation))

    rows_meta = format_table(
        ws,
        title=f"Award Summary (Sign-off) — Recommended: Scenario {award_scenario_code} "
        f"({award_scenario_label})",
        subtitle_lines=[
            "Per DC: incumbent → recommended, savings in $ vs TWO baselines + negotiation capture",
            DECISION_SUPPORT_STRAP,
        ],
        columns=columns,
        n_body_rows=len(ordered),
        header_row=header_row,
    )
    # Blended savings-% in the TOTAL row = total savings / total baseline (a sum is meaningless).
    total_row = rows_meta["total_row"]
    ws.cell(
        row=total_row,
        column=8,
        value=f"=G{total_row}/F{total_row}",
    ).number_format = NUMFMT_PCT
    ws.cell(row=total_row, column=8).font = _TOTAL_FONT
    ws.cell(row=total_row, column=8).fill = _TOTAL_FILL
    ws.cell(row=total_row, column=8).border = _BORDER
    ws.cell(row=total_row, column=8).alignment = _CENTER

    # Visual cue: savings $ green when positive, red when the deal costs more (columns G, I, J).
    body_last = body_start + len(ordered) - 1
    for col in ("G", "I", "J"):
        rng = f"{col}{body_start}:{col}{body_last}"
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="greaterThan", formula=["0"], fill=_MIN_FILL, font=_MIN_FONT)
        )
        ws.conditional_formatting.add(
            rng,
            CellIsRule(operator="lessThan", formula=["0"], fill=_BREACH_FILL, font=_BREACH_FONT),
        )

    # Product-type split (Conventional vs Organic) — the segmentation the real models carry.
    pt_spend: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    pt_base: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    for d in rec:
        pt = lot_product_type.get(d.lot_name, "Conventional")
        pt_spend[pt] += d.spend
        pt_base[pt] += d.baseline_spend
    sec = total_row + 3
    title = ws.cell(row=sec, column=1, value="By product type (Conventional / Organic)")
    title.font = _TITLE_FONT
    title.fill = _TITLE_FILL
    title.alignment = _LEFT
    for c in range(1, 6):
        ws.cell(row=sec, column=c).fill = _TITLE_FILL
    ws.merge_cells(start_row=sec, start_column=1, end_row=sec, end_column=5)
    hdr = sec + 1
    for ci, h in enumerate(
        ["Product type", "Rec spend", "Baseline", "Savings $", "Savings %"], start=1
    ):
        c = ws.cell(row=hdr, column=ci, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = _CENTER
        c.border = _BORDER
    for i, pt in enumerate(sorted(pt_spend)):
        r = hdr + 1 + i
        sav = pt_base[pt] - pt_spend[pt]
        pct = sav / pt_base[pt] if pt_base[pt] > 0 else Decimal("0")
        vals: list[tuple[object, str | None]] = [
            (pt, None),
            (float(pt_spend[pt]), NUMFMT_MONEY),
            (float(pt_base[pt]), NUMFMT_MONEY),
            (float(sav), NUMFMT_MONEY),
            (float(pct), NUMFMT_PCT),
        ]
        for ci, (v, nf) in enumerate(vals, start=1):
            c = ws.cell(row=r, column=ci, value=v)
            c.border = _BORDER
            c.alignment = _CENTER if nf else _LEFT
            if nf:
                c.number_format = nf
    ws.sheet_view.showGridLines = False


@dataclass(frozen=True)
class FobRow:
    """One (lot × DC × supplier) FOB decomposition line for the freight-transparency view."""

    lot_name: str
    dc_name: str
    region: str
    supplier_name: str
    fob: Decimal
    delivery: Decimal
    vegcool: Decimal
    all_in: Decimal
    transit_days: int | None  # real lane transit (None if not supplied)


def _gather_fob(session: Session, seeded: CycleView, final_round_id: str) -> list[FobRow]:
    """Per (lot × DC × supplier): the FOB / Delivery / VegCool / All-In decomposition (final round).

    Reads the persisted `bid.bid_line` components (`fob_case`, `delivery_surcharge_case`,
    `vegcool_surcharge_case`, `submitted_all_in_case`) — real schema columns (migration 0007).
    One row per (lot, DC, supplier); the freight does not vary by TF in the demo, so we take the
    first priced line per group. Names by D23.
    """

    lot_name = {lot.id: lot.name for lot in seeded.lots}
    sup_name = {sup.id: sup.name for sup in seeded.suppliers}
    dc_name = {dc.id: dc.name for dc in seeded.dcs}
    dc_region = {dc.id: _dc_region(i) for i, dc in enumerate(seeded.dcs)}
    transit_lane = _transit_by_lane(session, seeded.cycle_id)

    # ACTIVE rows only (`is_scoreable`) so a superseded re-submission can't surface a stale price;
    # the ORDER BY makes the "first priced line per (dc, lot, supplier)" DETERMINISTIC and collapses
    # Option-B fan-out (prefer a real period row, NULL period last, then by id) to one stable pick.
    rows = session.execute(
        text(
            "SELECT dc_id, lot_id, supplier_id, fob_case, delivery_surcharge_case, "
            "vegcool_surcharge_case, lot_discount_case, submitted_all_in_case FROM bid.bid_line "
            "WHERE cycle_id = :cyc AND round_id = :rnd "
            "AND (submitted_all_in_case IS NOT NULL OR fob_case IS NOT NULL) "
            "AND is_scoreable = true "
            "ORDER BY dc_id, lot_id, supplier_id, tf_id, fiscal_period_id NULLS LAST, bid_line_id"
        ),
        {"cyc": seeded.cycle_id, "rnd": final_round_id},
    ).all()

    seen: set[tuple[str, str, str]] = set()
    out: list[FobRow] = []
    for dc_id, lot_id, sup_id, fob, deliv, veg, disc, allin in rows:
        key = (dc_id, lot_id, sup_id)
        if key in seen:
            continue
        # Constructed landed = All-In if present, else FOB + surcharges − discount (E-39); a
        # component-basis bid (All-In NULL) now appears instead of being filtered out.
        constructed = _line_price(allin, fob, deliv, veg, disc)
        if constructed is None:
            continue
        seen.add(key)
        out.append(
            FobRow(
                lot_name=lot_name.get(lot_id, lot_id[:6]),
                dc_name=dc_name.get(dc_id, dc_id[:6]),
                region=dc_region.get(dc_id, ""),
                supplier_name=sup_name.get(sup_id, sup_id[:6]),
                fob=Decimal(str(fob)) if fob is not None else Decimal("0"),
                delivery=Decimal(str(deliv)) if deliv is not None else Decimal("0"),
                vegcool=Decimal(str(veg)) if veg is not None else Decimal("0"),
                all_in=constructed,
                transit_days=transit_lane.get((sup_id, dc_id)),
            )
        )
    out.sort(key=lambda x: (x.lot_name, x.dc_name, x.all_in, x.supplier_name))
    return out


def _write_fob_analysis_tab(wb: Workbook, rows: list[FobRow]) -> None:
    """Landed & Hidden Costs — FOB + freight = landed, PLUS transit time (freshness/lead-time).

    The real models devote a whole `FOB analysis` tab (+ a separate Delivery Charge tab) to
    stripping freight off the landed price. We surface the same AND the non-price hidden costs:
    each bid decomposed FOB → +Delivery (lane freight, by region) → +VegCool → = All-In, the
    cheapest landed bid per (lot, DC) highlighted, the lane **transit days** + a **freshness
    watch** (perishable produce — a hidden cost the headline price hides), and a regional freight
    summary. Freight from real `bid.bid_line` component columns; transit is a labelled lane proxy.
    """

    ws = wb.create_sheet("Landed & Hidden Costs")
    columns: list[Col] = [
        Col("Lot", 20),
        Col("DC", 16),
        Col("Region", 10),
        Col("Supplier", 22),
        Col("FOB $/case", 14, NUMFMT_MONEY),
        Col("+ Delivery", 12, NUMFMT_MONEY),
        Col("+ VegCool", 12, NUMFMT_MONEY),
        Col("= All-In $/case", 14, NUMFMT_MONEY),
        Col("Freight % of All-In", 14, NUMFMT_PCT),
        Col("Transit (days)", 12, NUMFMT_INT),
        Col("Freshness", 14),
    ]
    header_row = 5
    body_start = header_row + 1
    # cheapest landed All-In per (lot, DC) for the min-highlight.
    min_by_group: dict[tuple[str, str], Decimal] = {}
    for x in rows:
        k = (x.lot_name, x.dc_name)
        if k not in min_by_group or x.all_in < min_by_group[k]:
            min_by_group[k] = x.all_in

    for i, x in enumerate(rows):
        r = body_start + i
        freight = x.delivery + x.vegcool
        freight_pct = (freight / x.all_in) if x.all_in > 0 else Decimal("0")
        ws.cell(row=r, column=1, value=x.lot_name)
        ws.cell(row=r, column=2, value=x.dc_name)
        ws.cell(row=r, column=3, value=x.region)
        ws.cell(row=r, column=4, value=x.supplier_name)
        ws.cell(row=r, column=5, value=float(x.fob))
        ws.cell(row=r, column=6, value=float(x.delivery))
        ws.cell(row=r, column=7, value=float(x.vegcool))
        ws.cell(row=r, column=8, value=float(x.all_in))
        ws.cell(row=r, column=9, value=float(freight_pct))
        ws.cell(row=r, column=10, value=x.transit_days if x.transit_days is not None else "—")
        watch = x.transit_days is not None and x.transit_days > FRESHNESS_WATCH_DAYS
        fresh = ws.cell(
            row=r,
            column=11,
            value="⚠ watch" if watch else ("ok" if x.transit_days is not None else "—"),
        )
        if watch:
            fresh.fill = _INCUMBENT_FILL
            fresh.font = Font(bold=True, color="7F6000")
        if x.all_in == min_by_group[(x.lot_name, x.dc_name)]:
            for c in (4, 8):
                ws.cell(row=r, column=c).fill = _MIN_FILL
                ws.cell(row=r, column=c).font = _MIN_FONT

    format_table(
        ws,
        title="Landed & Hidden Costs — farm-gate + freight = landed, plus transit/freshness",
        subtitle_lines=[
            "Cheapest LANDED bid per (lot, DC) green · freight = Delivery + VegCool · "
            f"transit > {FRESHNESS_WATCH_DAYS}d = freshness watch (hidden cost)",
            DECISION_SUPPORT_STRAP,
        ],
        columns=columns,
        n_body_rows=len(rows),
        header_row=header_row,
        add_total=False,
    )

    # Regional freight summary — avg delivery per region (why landed cost differs by lane).
    by_region: dict[str, list[Decimal]] = defaultdict(list)
    for x in rows:
        by_region[x.region].append(x.delivery)
    sec = body_start + len(rows) + 2
    title = ws.cell(row=sec, column=1, value="Regional freight (avg Delivery $/case by lane)")
    title.font = _TITLE_FONT
    title.fill = _TITLE_FILL
    title.alignment = _LEFT
    for c in range(1, 4):
        ws.cell(row=sec, column=c).fill = _TITLE_FILL
    ws.merge_cells(start_row=sec, start_column=1, end_row=sec, end_column=3)
    hdr = sec + 1
    for ci, h in enumerate(["Region", "Avg Delivery $/case", "Lanes"], start=1):
        c = ws.cell(row=hdr, column=ci, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = _CENTER
        c.border = _BORDER
    for i, region in enumerate(sorted(by_region)):
        vals = by_region[region]
        avg = sum(vals, Decimal("0")) / Decimal(len(vals)) if vals else Decimal("0")
        r = hdr + 1 + i
        ws.cell(row=r, column=1, value=region).border = _BORDER
        cc = ws.cell(row=r, column=2, value=float(avg))
        cc.number_format = NUMFMT_MONEY
        cc.alignment = _CENTER
        cc.border = _BORDER
        cn = ws.cell(row=r, column=3, value=len(vals))
        cn.number_format = NUMFMT_INT
        cn.alignment = _CENTER
        cn.border = _BORDER
    ws.sheet_view.showGridLines = False


# ---------------------------------------------------------------------------
# Incumbent Retention — SURFACE THE PREMIUM, the human decides (relationship pillar).
# The engine recommends on economics and keeps a price-eligible incumbent on a near-tie (continuity
# weight). An incumbent over the premium ceiling is GATED — never silently paid. This tab puts a
# dollar figure on keeping each incumbent the engine did not pick, so the buyer overrides with eyes
# open (a recorded Scenario-F override), not by feel.
# ---------------------------------------------------------------------------
_RETAIN_RECOMMENDED = "Retained — recommended"
_RETAIN_ELIGIBLE = "Eligible — not the pick"
_RETAIN_GATED = "Gated by premium"
_RETAIN_NOBID = "No bid"


@dataclass(frozen=True)
class IncumbentRetentionRow:
    """One incumbent cell: keep cost vs. the recommendation (D26 relationship surface)."""

    dc_name: str
    lot_name: str
    tf_name: str
    incumbent: str
    incumbent_price: Decimal | None  # incumbent's bid $/case (None -> did not bid)
    status: str
    rec_supplier: str
    rec_price: Decimal | None
    volume: Decimal
    premium_per_case: Decimal | None  # incumbent - recommendation (None if no incumbent bid)
    premium_period: Decimal | None  # premium_per_case x period volume
    premium_pct: Decimal | None  # premium_per_case / recommendation price


def _gather_incumbent_retention(cells: list[CellInfo]) -> list[IncumbentRetentionRow]:
    """Per incumbent cell: the cost to keep them vs the recommendation, sorted by $ at stake.

    Reads ONLY the resolved `CellInfo` (incumbent name, the per-supplier price map, the eligible
    list, the B pick, the volume) — the same governed records the comparison tabs use; no new
    queries, no boilerplate. Status is derived, not asserted: recommended / eligible-but-not-picked
    / gated-by-premium / no-bid.
    """

    rows: list[IncumbentRetentionRow] = []
    for c in cells:
        inc = c.incumbent_name
        if not inc:
            continue
        inc_price = c.price_by_supplier.get(inc)
        rec_price = c.price_by_supplier.get(c.rec_supplier)
        if c.rec_supplier and inc == c.rec_supplier:
            status = _RETAIN_RECOMMENDED
        elif inc in c.eligible_suppliers:
            status = _RETAIN_ELIGIBLE
        elif inc_price is not None:
            status = _RETAIN_GATED
        else:
            status = _RETAIN_NOBID
        prem_case = (
            inc_price - rec_price if inc_price is not None and rec_price is not None else None
        )
        prem_period = prem_case * c.volume if prem_case is not None else None
        has_rec = rec_price is not None and rec_price != Decimal("0")
        prem_pct = (
            prem_case / rec_price if prem_case is not None and has_rec and rec_price else None
        )
        rows.append(
            IncumbentRetentionRow(
                dc_name=c.dc_name,
                lot_name=c.lot_name,
                tf_name=c.tf_name,
                incumbent=inc,
                incumbent_price=inc_price,
                status=status,
                rec_supplier=c.rec_supplier,
                rec_price=rec_price,
                volume=c.volume,
                premium_per_case=prem_case,
                premium_period=prem_period,
                premium_pct=prem_pct,
            )
        )
    # Biggest retention bill first (the decisions that cost the most); no-bid / retained settle low.
    rows.sort(
        key=lambda r: r.premium_period if r.premium_period is not None else Decimal("-1"),
        reverse=True,
    )
    return rows


def _write_incumbent_retention_tab(wb: Workbook, rows: list[IncumbentRetentionRow]) -> None:
    """Incumbent Retention — the dollar cost of keeping each incumbent the engine did not pick.

    The engine recommends on economics; relationships are the buyer's call. Each row prices that
    call: keep the incumbent (Δ to retain $/case × period volume) or take the recommendation. Rows
    where the incumbent is over the premium ceiling are flagged GATED — retaining one is a recorded
    override, never a silent payment. Sorted by $ at stake; a relationship-budget total at the foot.
    """

    ws = wb.create_sheet("Incumbent Retention")
    columns: list[Col] = [
        Col("DC", 16),
        Col("Lot", 24),
        Col("TF", 8),
        Col("Incumbent", 18),
        Col("Incumbent $/case", 14, NUMFMT_MONEY),
        Col("Status", 20),
        Col("Recommended", 18),
        Col("Rec $/case", 12, NUMFMT_MONEY),
        Col("Δ to retain $/case", 14, NUMFMT_MONEY),
        Col("Period cases", 12, NUMFMT_INT),
        Col("Premium to retain $", 16, NUMFMT_MONEY, total="sum"),
        Col("Retain? (your call)", 16),
    ]
    header_row = 5
    body_start = header_row + 1
    for i, r in enumerate(rows):
        rr = body_start + i

        def _m(v: Decimal | None) -> float | str:
            return float(v) if v is not None else "—"

        ws.cell(row=rr, column=1, value=r.dc_name)
        ws.cell(row=rr, column=2, value=r.lot_name)
        ws.cell(row=rr, column=3, value=r.tf_name)
        ws.cell(row=rr, column=4, value=r.incumbent)
        ws.cell(row=rr, column=5, value=_m(r.incumbent_price))
        scell = ws.cell(row=rr, column=6, value=r.status)
        ws.cell(row=rr, column=7, value=r.rec_supplier)
        ws.cell(row=rr, column=8, value=_m(r.rec_price))
        ws.cell(row=rr, column=9, value=_m(r.premium_per_case))
        ws.cell(row=rr, column=10, value=float(r.volume))
        ws.cell(row=rr, column=11, value=_m(r.premium_period))
        # colour the status: gated = amber (a real override decision), retained = green.
        if r.status == _RETAIN_GATED:
            scell.fill = _INCUMBENT_FILL
            scell.font = Font(bold=True, color="7F6000")
        elif r.status == _RETAIN_RECOMMENDED:
            scell.fill = _MIN_FILL
            scell.font = _MIN_FONT

    format_table(
        ws,
        title="Incumbent Retention — what keeping each incumbent costs vs. the recommendation",
        subtitle_lines=[
            "The engine recommends on economics and already keeps a price-eligible incumbent on a "
            "near-tie. 'Gated by premium' = the incumbent's bid is over the eligibility ceiling, "
            "so the engine excludes it — retaining is a recorded override, never silently paid.",
            "Δ to retain = incumbent − recommendation; Premium to retain $ = Δ × period cases. "
            "Sort/keep what the relationship is worth. " + DECISION_SUPPORT_STRAP,
        ],
        columns=columns,
        n_body_rows=len(rows),
        header_row=header_row,
        total_label="RELATIONSHIP BUDGET — premium to retain every non-recommended incumbent",
        total_label_col=1,
    )

    # Status tally — the picture at a glance (counts + the $ already saved by the gate).
    by_status: dict[str, int] = defaultdict(int)
    gated_premium = Decimal("0")
    for r in rows:
        by_status[r.status] += 1
        if r.status in (_RETAIN_GATED, _RETAIN_ELIGIBLE) and r.premium_period is not None:
            gated_premium += r.premium_period
    sec = body_start + len(rows) + 3
    title = ws.cell(
        row=sec,
        column=1,
        value="Incumbent picture — counts by status + the relationship budget at stake",
    )
    title.font = _TITLE_FONT
    title.fill = _TITLE_FILL
    title.alignment = _LEFT
    ws.merge_cells(start_row=sec, start_column=1, end_row=sec, end_column=6)
    for c in range(1, 7):
        ws.cell(row=sec, column=c).fill = _TITLE_FILL
    bs = by_status.get
    lines = [
        (f"{_RETAIN_RECOMMENDED} (engine kept the incumbent)", bs(_RETAIN_RECOMMENDED, 0)),
        (f"{_RETAIN_ELIGIBLE} (in range, lost on score)", bs(_RETAIN_ELIGIBLE, 0)),
        (f"{_RETAIN_GATED} (over the ceiling — override to keep)", bs(_RETAIN_GATED, 0)),
        (f"{_RETAIN_NOBID} (incumbent did not bid)", bs(_RETAIN_NOBID, 0)),
    ]
    rr = sec + 1
    for label, n in lines:
        ws.cell(row=rr, column=1, value=label).alignment = _LEFT
        nc = ws.cell(row=rr, column=4, value=n)
        nc.number_format = NUMFMT_INT
        nc.alignment = _CENTER
        rr += 1
    bc = ws.cell(
        row=rr,
        column=1,
        value="Relationship budget — premium to retain every non-recommended incumbent that bid",
    )
    bc.font = _TOTAL_FONT
    bc.alignment = _LEFT
    vc = ws.cell(row=rr, column=4, value=float(gated_premium))
    vc.number_format = NUMFMT_MONEY
    vc.font = _TOTAL_FONT
    vc.fill = _TOTAL_FILL
    ws.sheet_view.showGridLines = False


def _write_share_relationships_tab(
    wb: Workbook,
    details: list[AwardDetail],
    scenarios: list[AnalysisScenario],
    seeded: CycleView,
    config: EngineConfig,
    rec_code: str,
) -> None:
    """Share & Relationships — RELATIONSHIP CAPITAL in the repeated game (pillar 3).

    Each supplier's % share of total spend in EVERY scenario (a heatmap), their relationship
    type — **Preserve** (incumbent kept) vs **Create** (new supplier earned in) — and a
    **dependency** flag where share ≥ the concentration threshold (over-giving weakens your
    next-round bargaining position). Below: a relationship ledger for the recommendation —
    incumbents retained, new suppliers introduced, incumbents dropped. The structural rule —
    reward competitiveness with business, don't build a dependency you can be pressured by.
    """

    ws = wb.create_sheet("Share & Relationships")
    incumbent_names = {
        seeded_sup_name(seeded, sup_id) for sup_id in set(seeded.incumbent_by_dc_lot.values())
    }
    codes = [s.scenario_code for s in sorted(scenarios, key=lambda x: x.scenario_code)]
    # spend per (scenario, supplier) and per scenario total.
    spend: dict[tuple[str, str], Decimal] = defaultdict(lambda: Decimal("0"))
    scen_total: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    for d in details:
        spend[(d.scenario_code, d.supplier_name)] += d.spend
        scen_total[d.scenario_code] += d.spend
    suppliers = [s.name for s in seeded.suppliers]

    columns: list[Col] = [Col("Supplier", 24), Col("Relationship", 13)]
    for code in codes:
        columns.append(Col(f"{code} share", 10, NUMFMT_PCT))
    columns.append(Col("Max share", 11, NUMFMT_PCT))
    columns.append(Col("Dependency?", 13))
    header_row = 6
    body_start = header_row + 1
    first_scen_col = 3
    for i, sup in enumerate(suppliers):
        r = body_start + i
        ws.cell(row=r, column=1, value=sup)
        ws.cell(row=r, column=2, value="Preserve" if sup in incumbent_names else "Create")
        shares: list[float] = []
        for j, code in enumerate(codes):
            tot = scen_total.get(code, Decimal("0"))
            sh = float(spend[(code, sup)] / tot) if tot > 0 else 0.0
            shares.append(sh)
            ws.cell(row=r, column=first_scen_col + j, value=sh)
        mx = max(shares) if shares else 0.0
        ws.cell(row=r, column=first_scen_col + len(codes), value=mx)
        dep = ws.cell(
            row=r,
            column=first_scen_col + len(codes) + 1,
            value="DEPENDENCY" if mx >= float(config.conc_thresh) else "",
        )
        if mx >= float(config.conc_thresh):
            dep.font = _BREACH_FONT
            dep.fill = _BREACH_FILL

    format_table(
        ws,
        title="Share & Relationships — who carries the business, and the relationship behind it",
        subtitle_lines=[
            "Preserve = incumbent kept · Create = new supplier earned in · "
            f"Dependency = share ≥ {config.conc_thresh:.0%} (concentration risk)",
            DECISION_SUPPORT_STRAP,
        ],
        columns=columns,
        n_body_rows=len(suppliers),
        header_row=header_row,
        add_total=False,
    )
    # Heatmap across the per-scenario share columns (low=green → high=red concentration).
    last_scen_col = get_column_letter(first_scen_col + len(codes) - 1)
    share_end = body_start + len(suppliers) - 1
    ws.conditional_formatting.add(
        f"{get_column_letter(first_scen_col)}{body_start}:{last_scen_col}{share_end}",
        ColorScaleRule(
            start_type="num",
            start_value=0,
            start_color="FFFFFF",
            mid_type="num",
            mid_value=float(config.conc_thresh),
            mid_color="FFEB9C",
            end_type="num",
            end_value=0.6,
            end_color="F8696B",
        ),
    )

    # Relationship ledger for the recommendation (Scenario B).
    rec = [d for d in details if d.scenario_code == rec_code]
    won = {d.supplier_name for d in rec}
    preserved = sorted(won & incumbent_names)
    created = sorted(won - incumbent_names)
    dropped = sorted(incumbent_names - won)
    sec = body_start + len(suppliers) + 2
    title = ws.cell(
        row=sec, column=1, value=f"Relationship ledger — Scenario {rec_code} (recommended)"
    )
    title.font = _TITLE_FONT
    title.fill = _TITLE_FILL
    title.alignment = _LEFT
    for c in range(1, 5):
        ws.cell(row=sec, column=c).fill = _TITLE_FILL
    ws.merge_cells(start_row=sec, start_column=1, end_row=sec, end_column=4)
    ledger = [
        ("Relationships PRESERVED (incumbent retained)", ", ".join(preserved) or "—"),
        ("Relationships CREATED (new supplier earned in)", ", ".join(created) or "—"),
        ("Incumbents NOT awarded (relationship at risk)", ", ".join(dropped) or "—"),
    ]
    for i, (label, val) in enumerate(ledger):
        r = sec + 1 + i
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = _TOTAL_FONT
        lc.alignment = _LEFT
        lc.border = _BORDER
        vc = ws.cell(row=r, column=2, value=val)
        vc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        vc.border = _BORDER
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        ws.row_dimensions[r].height = 22
    ws.sheet_view.showGridLines = False


@dataclass
class _SupplierPlay:
    """One supplier's negotiation behaviour across the rounds (reading their incentives)."""

    name: str
    role: str  # Incumbent | Challenger
    cells: int
    avg_move_frac: Decimal  # mean R1→final price move (negative = conceded)
    avg_prem_frac: Decimal  # mean premium vs market-low (final round)
    sustainability_flags: int  # below-market outliers (Z < −2) — real risk to validate
    read: str  # the game-theoretic behavioural label


def _gather_negotiation(
    evo_rows: list[RoundEvoRow],
    score_detail: list[ScoreDetail],
    incumbent_names: set[str],
) -> tuple[list[_SupplierPlay], Decimal, Decimal]:
    """Read each supplier's negotiation behaviour from observable moves (asymmetric-info frame).

    Concession = the R1→final price move (from `RoundEvoRow`); market position = premium-vs-low +
    Z-score (from `ScoreDetail`, final round). Returns (per-supplier plays, incumbent avg move,
    challenger avg move) — the asymmetry that says whether the incumbent is competing or leaning
    on the installed base. Distinguishes a below-market **real risk** (Z<−2, validate) from a
    priced-high-and-firm **leverage/theater** play.
    """

    moves: dict[str, list[Decimal]] = defaultdict(list)
    for e in evo_rows:
        moves[e.supplier_name].append(e.pct)
    prem: dict[str, list[Decimal]] = defaultdict(list)
    zlow: dict[str, int] = defaultdict(int)
    for s in score_detail:
        prem[s.supplier_name].append(s.prem_vs_low_frac)
        if s.z_score < Decimal("-2"):
            zlow[s.supplier_name] += 1

    def _avg(xs: list[Decimal]) -> Decimal:
        return sum(xs, Decimal("0")) / Decimal(len(xs)) if xs else Decimal("0")

    names = sorted(set(moves) | set(prem))
    plays: list[_SupplierPlay] = []
    inc_moves: list[Decimal] = []
    chal_moves: list[Decimal] = []
    for name in names:
        role = "Incumbent" if name in incumbent_names else "Challenger"
        amove = _avg(moves[name])
        aprem = _avg(prem[name])
        flags = zlow[name]
        if role == "Incumbent":
            inc_moves.append(amove)
        else:
            chal_moves.append(amove)
        # Game-theoretic read from concession + market position.
        if flags > 0 and amove <= Decimal("-0.04"):
            read = "Below-market & conceding — validate sustainability (real risk)"
        elif role == "Incumbent" and amove > Decimal("-0.01"):
            read = "Holding the installed base — test the leverage"
        elif aprem > Decimal("0.07") and amove > Decimal("-0.01"):
            read = "Priced high & firm — likely negotiation theater"
        elif amove <= Decimal("-0.04"):
            read = "Conceding hard — hungry for the volume"
        else:
            read = "Competing in line with the field"
        plays.append(
            _SupplierPlay(
                name=name,
                role=role,
                cells=len(moves[name]),
                avg_move_frac=amove,
                avg_prem_frac=aprem,
                sustainability_flags=flags,
                read=read,
            )
        )
    return plays, _avg(inc_moves), _avg(chal_moves)


def _write_negotiation_dynamics_tab(
    wb: Workbook,
    plays: list[_SupplierPlay],
    inc_move: Decimal,
    chal_move: Decimal,
) -> None:
    """Negotiation Dynamics — fairness & leverage in a repeated game (pillar 4).

    Are you being treated fairly? Reads each supplier's concession behaviour (how far they moved
    R1→final), the incumbent's move vs the field's (do they lean on tenure?), and separates a
    below-market **real risk** from a priced-high-and-firm **leverage/theater** play. The
    structural rule: be predictable in process, flexible only where the economics justify it.
    """

    ws = wb.create_sheet("Negotiation Dynamics")
    columns: list[Col] = [
        Col("Supplier", 24),
        Col("Role", 12),
        Col("Cells moved", 11, NUMFMT_INT),
        Col("Avg move R1→Final", 16, NUMFMT_PCT),
        Col("Avg premium vs low", 16, NUMFMT_PCT),
        Col("Sustainability flags", 16, NUMFMT_INT),
        Col("Negotiation read", 46),
    ]
    header_row = 6
    body_start = header_row + 1
    ordered = sorted(plays, key=lambda p: (p.role != "Incumbent", p.avg_move_frac))
    for i, p in enumerate(ordered):
        r = body_start + i
        ws.cell(row=r, column=1, value=p.name)
        ws.cell(row=r, column=2, value=p.role)
        ws.cell(row=r, column=3, value=p.cells)
        ws.cell(row=r, column=4, value=float(p.avg_move_frac))
        ws.cell(row=r, column=5, value=float(p.avg_prem_frac))
        ws.cell(row=r, column=6, value=p.sustainability_flags)
        ws.cell(row=r, column=7, value=p.read)
        if p.role == "Incumbent":
            ws.cell(row=r, column=2).fill = _INCUMBENT_FILL

    format_table(
        ws,
        title="Negotiation Dynamics — are you being treated fairly? (repeated game, asym. info)",
        subtitle_lines=[
            "Concession = R1→Final move (more negative = conceded more) · "
            "predictable in process, flexible only where economics justify",
            DECISION_SUPPORT_STRAP,
        ],
        columns=columns,
        n_body_rows=len(ordered),
        header_row=header_row,
        add_total=False,
    )
    # Data bars on the concession column make "who moved" scannable (negative axis).
    ws.conditional_formatting.add(
        f"D{body_start}:D{body_start + len(ordered) - 1}",
        DataBarRule(start_type="min", end_type="max", color="5B9BD5", showValue=True),
    )

    # The fairness headline — incumbent move vs the field.
    gap = inc_move - chal_move  # >0 means incumbent conceded LESS than the field
    if gap > Decimal("0.005"):
        verdict = (
            f"The incumbent conceded {inc_move:.1%} vs the field's {chal_move:.1%} — "
            "LESS than challengers. You are paying partly for tenure, not competitiveness: "
            "press the installed-base lots or signal credible switching."
        )
    elif gap < Decimal("-0.005"):
        verdict = (
            f"The incumbent conceded {inc_move:.1%} vs the field's {chal_move:.1%} — "
            "MORE than challengers. They are defending the relationship competitively (fair)."
        )
    else:
        verdict = (
            f"The incumbent ({inc_move:.1%}) moved in line with the field ({chal_move:.1%}) — "
            "competitive, no obvious leverage play."
        )
    sec = body_start + len(ordered) + 2
    title = ws.cell(row=sec, column=1, value="Fairness read — incumbent concession vs the field")
    title.font = _TITLE_FONT
    title.fill = _TITLE_FILL
    title.alignment = _LEFT
    for c in range(1, 8):
        ws.cell(row=sec, column=c).fill = _TITLE_FILL
    ws.merge_cells(start_row=sec, start_column=1, end_row=sec, end_column=7)
    vc = ws.cell(row=sec + 1, column=1, value=verdict)
    vc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    vc.border = _BORDER
    ws.merge_cells(start_row=sec + 1, start_column=1, end_row=sec + 1, end_column=7)
    ws.row_dimensions[sec + 1].height = 46
    ws.sheet_view.showGridLines = False


def seeded_sup_name(seeded: CycleView, sup_id: str) -> str:
    """Readable supplier name for an id (D23)."""

    for s in seeded.suppliers:
        if s.id == sup_id:
            return s.name
    return sup_id[:6]


@dataclass(frozen=True)
class _Kpi:
    """One headline KPI card, tagged to one of the four decision lenses (with a lens colour)."""

    lens: str
    label: str
    value: str  # pre-formatted
    color: str  # hex fill for the lens


def _compute_kpis(
    rec_details: list[AwardDetail],
    baseline_total: Decimal,
    rec_spend: Decimal,
    incumbent_names: set[str],
    inc_move: Decimal,
    chal_move: Decimal,
) -> list[_Kpi]:
    """The four-lens headline: cost & savings · hidden costs · relationships · negotiation."""

    savings = baseline_total - rec_spend
    savings_pct = savings / baseline_total if baseline_total > 0 else Decimal("0")
    t_vals = [d.transit_days for d in rec_details if d.transit_days is not None]
    avg_transit = (sum(t_vals) / len(t_vals)) if t_vals else 0.0
    fresh = sum(1 for t in t_vals if t > FRESHNESS_WATCH_DAYS)
    preserved = len({d.supplier_name for d in rec_details if d.relationship == "Preserve"})
    created = len({d.supplier_name for d in rec_details if d.relationship == "Create"})
    gap = inc_move - chal_move
    fair = (
        "leaning on tenure"
        if gap > Decimal("0.005")
        else ("competing" if gap < Decimal("-0.005") else "in line")
    )
    cost, hidden, rel, neg = "548235", "BF8F00", "2E5496", "1F3864"
    return [
        _Kpi("Cost & savings", "Savings vs incumbent", f"${savings:,.0f}", cost),
        _Kpi("Cost & savings", "Savings %", f"{savings_pct:.1%}", cost),
        _Kpi("Hidden costs", "Avg transit (days)", f"{avg_transit:.1f}", hidden),
        _Kpi("Hidden costs", "Freshness watches", f"{fresh}", hidden),
        _Kpi("Relationships", "Preserved / Created", f"{preserved} / {created}", rel),
        _Kpi(
            "Negotiation",
            "Incumbent vs field move",
            f"{inc_move:.1%} / {chal_move:.1%} ({fair})",
            neg,
        ),
    ]


def _write_kpi_band(wb: Workbook, kpis: list[_Kpi]) -> None:
    """A KPI scorecard band on the Overview — the four lenses, big values, lens-colour coded.

    Visual readability only (not final brand): a colour-coded Lens | Metric | Value scorecard so
    the headline reads at a glance. The downstream design review owns the final visual language.
    """

    ws = wb["Summary"]
    start = (ws.max_row or 1) + 2
    title = ws.cell(
        row=start,
        column=1,
        value="Headline KPIs — the four lenses (cost · hidden cost · relationships · negotiation)",
    )
    title.font = _TITLE_FONT
    title.fill = _TITLE_FILL
    title.alignment = _LEFT
    ws.merge_cells(start_row=start, start_column=1, end_row=start, end_column=2)
    ws.cell(row=start, column=2).fill = _TITLE_FILL
    r = start + 1
    for k in kpis:
        lens = ws.cell(row=r, column=1, value=f"{k.lens} — {k.label}")
        lens.fill = PatternFill("solid", fgColor=k.color)
        lens.font = Font(bold=True, color="FFFFFF", size=10)
        lens.alignment = _LEFT
        lens.border = _BORDER
        val = ws.cell(row=r, column=2, value=k.value)
        val.font = Font(bold=True, color=k.color, size=14)
        val.alignment = _LEFT
        val.border = _BORDER
        ws.row_dimensions[r].height = 24
        r += 1


def _write_custom_dashboard_tab(
    wb: Workbook,
    n_cells: int,
    seeded: CycleView,
    config: EngineConfig,
    baseline_total: Decimal,
    rec_spend: Decimal,
    rec_avg_transit: float,
    rec_fresh: int,
    rec_n_suppliers: int,
    incumbent_names: set[str],
) -> None:
    """Custom Dashboard — the lenses LIVE off the Custom Scenario build (not the automated B).

    Every value is an Excel formula reading the Custom Scenario rows, so as the buyer changes a
    supplier dropdown the whole dashboard recomputes: total spend & savings (cost), avg transit &
    freshness (hidden cost), per-supplier share & dependency (relationships) — each shown beside
    the recommended Scenario B so 'your build vs the recommendation' is one glance. The 'see all
    the dashboards move when I build custom, not just the automated one' ask (D25/D27).
    """

    ws = wb.create_sheet("Custom Dashboard")
    s, e = 7, 6 + n_cells
    cs = "'Custom Scenario'"
    sup_r = f"{cs}!$E${s}:$E${e}"
    spend_r = f"{cs}!$K${s}:$K${e}"
    vol_r = f"{cs}!$J${s}:$J${e}"
    base_r = f"{cs}!$L${s}:$L${e}"
    tr_r = f"{cs}!$N${s}:$N${e}"
    dc_r = f"{cs}!$A${s}:$A${e}"
    base_spend = f"SUMPRODUCT({vol_r},{base_r})"
    cust_spend = f"SUM({spend_r})"

    next_row = _title_block(
        ws,
        title="Custom Dashboard — LIVE off your Custom Scenario build",
        subtitle_lines=[
            "Change a supplier on the Custom Scenario tab and every number here recomputes.",
            DECISION_SUPPORT_STRAP,
        ],
        span=4,
    )
    for col, w in (("A", 32), ("B", 18), ("C", 18), ("D", 16)):
        ws.column_dimensions[col].width = w

    def section(row: int, label: str, span: int = 4) -> int:
        c = ws.cell(row=row, column=1, value=label)
        c.font = _TITLE_FONT
        c.fill = _TITLE_FILL
        c.alignment = _LEFT
        for cc in range(1, span + 1):
            ws.cell(row=row, column=cc).fill = _TITLE_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
        return row + 1

    def header(row: int, labels: list[str]) -> int:
        for ci, h in enumerate(labels, start=1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font = _HEADER_FONT
            c.fill = _HEADER_FILL
            c.alignment = _CENTER
            c.border = _BORDER
        return row + 1

    # --- Section 1: your build vs the recommendation (live KPIs) ---
    r = section(next_row, "Your build vs the recommendation")
    r = header(r, ["Metric", "Custom (LIVE)", "Recommended (B)", "Δ vs rec"])
    base_rec = baseline_total - rec_spend
    pct_rec = float(base_rec / baseline_total) if baseline_total > 0 else 0.0
    kpi_rows: list[tuple[str, str, float, str | None]] = [
        ("Total spend", f"={cust_spend}", float(rec_spend), NUMFMT_MONEY),
        (
            "Savings vs incumbent $",
            f"={base_spend}-{cust_spend}",
            float(base_rec),
            NUMFMT_MONEY,
        ),
        (
            "Savings vs incumbent %",
            f"=IFERROR(({base_spend}-{cust_spend})/{base_spend},0)",
            pct_rec,
            NUMFMT_PCT,
        ),
        (
            "Avg transit (days, vol-wtd)",
            f"=IFERROR(SUMPRODUCT({vol_r},{tr_r})/SUM({vol_r}),0)",
            rec_avg_transit,
            "0.0",
        ),
        (
            "Freshness watches (cells)",
            f'=COUNTIF({tr_r},">{FRESHNESS_WATCH_DAYS}")',
            float(rec_fresh),
            NUMFMT_INT,
        ),
        (
            "Suppliers used",
            f"=SUMPRODUCT(1/COUNTIF({sup_r},{sup_r}))",
            float(rec_n_suppliers),
            NUMFMT_INT,
        ),
    ]
    for label, cust_f, rec_v, nf in kpi_rows:
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = _TOTAL_FONT
        lc.border = _BORDER
        lc.alignment = _LEFT
        cc = ws.cell(row=r, column=2, value=cust_f)
        rc = ws.cell(row=r, column=3, value=rec_v)
        dc = ws.cell(row=r, column=4, value=f"=B{r}-C{r}")
        for cell in (cc, rc, dc):
            cell.border = _BORDER
            cell.alignment = _CENTER
            if nf:
                cell.number_format = nf
        cc.font = Font(bold=True, color="1F3864")
        r += 1
    r += 1

    # --- Section 2: supplier share — LIVE ---
    r = section(r, "Supplier share — LIVE (relationships & dependency)")
    r = header(r, ["Supplier", "Role", "Custom share", "Dependency"])
    share_start = r
    for sup in seeded.suppliers:
        nm = sup.name.replace('"', '""')
        ws.cell(row=r, column=1, value=sup.name).border = _BORDER
        role = ws.cell(
            row=r, column=2, value="Preserve" if sup.name in incumbent_names else "Create"
        )
        role.border = _BORDER
        role.alignment = _CENTER
        sh = ws.cell(
            row=r,
            column=3,
            value=f'=IFERROR(SUMIF({sup_r},"{nm}",{spend_r})/{cust_spend},0)',
        )
        sh.number_format = NUMFMT_PCT
        sh.alignment = _CENTER
        sh.border = _BORDER
        dep = ws.cell(
            row=r,
            column=4,
            value=f'=IF(C{r}>={float(config.conc_thresh)},"DEPENDENCY","")',
        )
        dep.alignment = _CENTER
        dep.border = _BORDER
        r += 1
    share_end = r - 1
    ws.conditional_formatting.add(
        f"C{share_start}:C{share_end}",
        ColorScaleRule(
            start_type="num",
            start_value=0,
            start_color="FFFFFF",
            mid_type="num",
            mid_value=float(config.conc_thresh),
            mid_color="FFEB9C",
            end_type="num",
            end_value=0.6,
            end_color="F8696B",
        ),
    )
    ws.conditional_formatting.add(
        f"D{share_start}:D{share_end}",
        CellIsRule(
            operator="equal", formula=['"DEPENDENCY"'], fill=_BREACH_FILL, font=_BREACH_FONT
        ),
    )
    r += 1

    # --- Section 3: spend by DC — LIVE ---
    r = section(r, "Spend by DC — LIVE", span=3)
    r = header(r, ["DC", "Custom spend", "Savings vs incumbent"])
    for dc in seeded.dcs:
        nm = dc.name.replace('"', '""')
        base_dc = f'SUMPRODUCT(({dc_r}="{nm}")*{vol_r}*{base_r})'
        ws.cell(row=r, column=1, value=dc.name).border = _BORDER
        cspend = ws.cell(row=r, column=2, value=f'=SUMIF({dc_r},"{nm}",{spend_r})')
        cspend.number_format = NUMFMT_MONEY
        cspend.alignment = _CENTER
        cspend.border = _BORDER
        sav = ws.cell(row=r, column=3, value=f'={base_dc}-SUMIF({dc_r},"{nm}",{spend_r})')
        sav.number_format = NUMFMT_MONEY
        sav.alignment = _CENTER
        sav.border = _BORDER
        r += 1
    ws.freeze_panes = ws.cell(row=next_row, column=1)
    ws.sheet_view.showGridLines = False


@dataclass(frozen=True)
class _RunVersion:
    """The mid-cycle alignment version surfaced on the Summary banner (from sealed records)."""

    seq: int  # 1-based ordinal of THIS run among the cycle's sealed runs (by start time)
    round_seq: int  # 1-based ordinal of THIS run among the cycle's runs OF THIS ROUND
    round_number: int  # cyc.cycle_round.round_number for the final round
    run_finished_at: datetime  # eng.analysis_run.run_finished_at for THIS run


def _run_version(
    session: Session, cycle_id: str, analysis_run_id: str, final_round_id: str
) -> _RunVersion:
    """Compute the run version = the 1-based ordinal of THIS sealed analysis run for the cycle.

    The cycle seals one `eng.analysis_run` per alignment pass; we order the cycle's runs by
    `run_started_at` and take the count of runs that started at or before THIS run (so the first
    pass reads `v1`, the next `v2`, …). The round number comes from `cyc.cycle_round` for the
    final round; the sealed timestamp from this run's `run_finished_at`. History is already sealed
    per run — this only SURFACES the version (D26).
    """

    this_run = (
        session.query(AnalysisRun).filter(AnalysisRun.analysis_run_id == analysis_run_id).one()
    )
    seq = (
        session.query(AnalysisRun)
        .filter(
            AnalysisRun.cycle_id == cycle_id,
            AnalysisRun.run_started_at <= this_run.run_started_at,
        )
        .count()
    )
    # Per-round ordinal too: a buyer re-running Round 2 sees "round run 1, 2…" instead of being
    # confused by the cycle-global v3/v4 (which counts every round's runs).
    round_seq = (
        session.query(AnalysisRun)
        .filter(
            AnalysisRun.cycle_id == cycle_id,
            AnalysisRun.round_id == final_round_id,
            AnalysisRun.run_started_at <= this_run.run_started_at,
        )
        .count()
    )
    round_number = session.execute(
        text("SELECT round_number FROM cyc.cycle_round WHERE round_id = :rid"),
        {"rid": final_round_id},
    ).scalar_one()
    return _RunVersion(
        seq=seq,
        round_seq=round_seq,
        round_number=int(round_number),
        run_finished_at=this_run.run_finished_at,
    )


def _stamp_real_provenance(wb: Workbook) -> None:
    """Restamp the demo's SYNTHETIC provenance tokens to LIVE/real for a real cycle run.

    The tab straps are written with the literal 'SYNTHETIC' / 'SYNTHESIZED' (the generator's demo
    lineage). On a real run those mislabel real supplier names and prices, so replace the provenance
    tokens in every string cell. 'SYNTHETIC' appears ONLY in these straps, so the replacement is
    safe; the specific phrases are ordered before the bare token.
    """

    repls = [
        ("SYNTHETIC data — names & prices invented.", "LIVE CYCLE DATA — real names & prices."),
        ("SYNTHETIC names & prices", "LIVE CYCLE DATA — real names & prices"),
        ("SYNTHETIC", "LIVE CYCLE DATA"),
        ("SYNTHESIZED", "MODELED"),
    ]
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and ("SYNTHETIC" in v or "SYNTHESIZED" in v):
                    for a, b in repls:
                        v = v.replace(a, b)
                    cell.value = v


def write_scenario_workbook_xlsx(
    session: Session,
    cycle: CycleView,
    config: EngineConfig,
    analysis_run_id: str,
    final_round_id: str,
    award: AwardView,
    *,
    output_path: Path | None = None,
    synthetic: bool = False,
) -> Path:
    """Generate the ALIGNMENT / COMPARISON Scenario Workbook (D26/D27) from the sealed records.

    Redesigned per SCENARIO_TOOL_DESIGN_STUDY.md §4 — single-purpose tabs named for the
    question, read left-to-right as a decision flow, depth via drill/filter/live (D27),
    not 20-tab sprawl. Tabs: Summary/Overview (headline + tab index = the front door) ·
    Scenario Comparison (lenses side by side + LIVE Custom + drill + per-DC matrix) ·
    Supplier Comparison (CENTERPIECE — every supplier per cell) · Lowest-Cost Check (why
    the rec ≠ cheapest) · Coverage (offered vs required) · Detailed Scoring (5 factors +
    the market stats behind them) · TF Comparison (if >1 TF) · Round Evolution (if >1
    round) · Data Quality (surfaced, never fatal) · Custom Scenario (interactive, D25) ·
    Data (pivot me) (flat Excel Table, D27) · _Prices (hidden live-formula grid). The
    high-value v3 views we were light on are ADDED clean; Bidder Detail / Top-5 /
    Share-of-Business / Regional / Vol-Util are reachable by drilling Scenario Comparison +
    pivoting Data (depth-on-demand), not as separate tabs. Visual design-language
    (color/type) is DEFERRED to the downstream design review.
    """

    seeded = cycle  # internal alias: every gather/write step reads the resolved CycleView.

    # --- Version heading (sponsor: mid-cycle alignment versioning). The run version is the 1-based
    #     ordinal of THIS analysis run among the cycle's sealed runs ordered by start time; round
    #     number + sealed timestamp come straight off the governed records (history is sealed per
    #     run — we are SURFACING the version, not minting it). ---
    version = _run_version(session, cycle.cycle_id, analysis_run_id, final_round_id)

    out_path = output_path if output_path is not None else (OUTPUT_DIR / "SCENARIO_WORKBOOK.xlsx")
    out_path.parent.mkdir(parents=True, exist_ok=True)

    scenarios = (
        session.query(AnalysisScenario)
        .filter(AnalysisScenario.analysis_run_id == analysis_run_id)
        .order_by(AnalysisScenario.scenario_code)
        .all()
    )
    cells = _gather_cells(session, seeded, analysis_run_id, final_round_id, award)
    rollups, baseline_total, stly_total = _gather_scenario_rollups(
        session, seeded, scenarios, analysis_run_id
    )
    details = _gather_award_details(session, seeded, scenarios, analysis_run_id)
    score_detail = _gather_score_detail(session, seeded, analysis_run_id, final_round_id, award)
    coverage_rows = _gather_coverage(session, seeded, analysis_run_id, final_round_id, award)
    dc_names = [dc.name for dc in seeded.dcs]

    # Product type per lot (DEMO-illustrative segmentation, by lot order).
    lot_product_type = {
        seeded.lots[i].name: LOT_PRODUCT_TYPE[i % len(LOT_PRODUCT_TYPE)]
        for i in range(len(seeded.lots))
    }

    # Round-evolution rows (negotiation story) — computed once, reused by the sign-off negotiation
    # column AND the Round Evolution tab. Empty when single-round.
    evo_rows: list[RoundEvoRow] = []
    round_labels: list[str] = []
    if len(seeded.rounds) > 1:
        round_ids = [r.id for r in seeded.rounds]
        round_labels = [r.name for r in seeded.rounds]
        evo_rows = _gather_round_evolution(session, seeded, round_ids)

    # Per-DC negotiation capture (R1→Final) on the RECOMMENDED award cells, + recommended spend.
    rec_code = award.scenario_code
    evo_by_key = {(e.dc_name, e.lot_name, e.tf_name, e.supplier_name): e for e in evo_rows}
    negotiation_by_dc: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    rec_spend = Decimal("0")
    for d in details:
        if d.scenario_code != rec_code:
            continue
        rec_spend += d.spend
        e = evo_by_key.get((d.dc_name, d.lot_name, d.tf_name, d.supplier_name))
        if e is not None:
            negotiation_by_dc[d.dc_name] += (-e.delta) * d.volume
    negotiation_total = sum(negotiation_by_dc.values(), Decimal("0"))

    # Negotiation behaviour (fairness/leverage, pillar 4) — concession vs the field, real-risk vs
    # theater. Reads observable moves (round evo) + market position (scores).
    incumbent_names = {
        seeded_sup_name(seeded, sup_id) for sup_id in set(seeded.incumbent_by_dc_lot.values())
    }
    plays, inc_move, chal_move = _gather_negotiation(evo_rows, score_detail, incumbent_names)

    # The four-pillar KPI band on the Overview (cost · hidden cost · relationships · fairness).
    rec_details = [d for d in details if d.scenario_code == rec_code]
    kpis = _compute_kpis(
        rec_details, baseline_total, rec_spend, incumbent_names, inc_move, chal_move
    )

    # The tab index for the Overview front door — banded into the FOUR decision lenses the
    # repeated-game frame needs (Decide/cost · Relationships · Negotiation · Diligence · Build).
    tab_index: list[tuple[str, str]] = [
        ("Controls", "How was this run? Horizon, scope, baselines, engine weights & rules."),
        ("§ DECIDE — cost & savings", ""),
        ("Award Summary", "The recommendation: per-DC incumbent→recommended + savings $ to sign."),
        ("Scenario Comparison", "Which lens? A-G + LIVE Custom side by side, drill to DC."),
        ("Lowest-Cost Check", "Is the rec the cheapest? If not, the premium it trades."),
        ("§ COMPARE SUPPLIERS", ""),
        ("Supplier Comparison", "Which supplier per cell? Every eligible supplier's $/case."),
        ("Landed & Hidden Costs", "Beyond price: freight + transit/freshness — the true cost."),
        ("§ RELATIONSHIPS & NEGOTIATION", ""),
        ("Share & Relationships", "Who carries the business — preserve vs create + dependency."),
        ("Incumbent Retention", "What keeping each incumbent costs vs the rec."),
        ("Negotiation Dynamics", "Treated fairly? Concession vs the field, real risk vs theater."),
        ("§ DILIGENCE", ""),
        ("Coverage", "Can they supply it? Offered vs required volume + cover band."),
        ("Capacity Check", "Beyond stated capacity? Allocation vs each supplier's ceiling."),
        ("Detailed Scoring", "Why these scores? The 5 factors + the market stats behind them."),
    ]

    wb = Workbook()
    # fullCalcOnLoad → Excel recomputes every live formula (the Custom column + Custom tab) on open,
    # so the workbook opens with the live cross-tab values already resolved (D27).
    wb.calculation.fullCalcOnLoad = True
    _write_summary_tab(wb, seeded, config, rollups, version)
    _write_kpi_band(wb, kpis)
    # --- Cockpit + headline sign-off (practitioner layer from the real allocation models). ---
    _write_controls_tab(
        wb, seeded, config, baseline_total, stly_total, rec_spend, negotiation_total
    )
    _write_award_summary_tab(
        wb, details, rec_code, award.scenario_label, negotiation_by_dc, lot_product_type
    )
    # --- DECIDE / COMPARE (alignment surfaces, D26). ---
    _write_scenario_comparison_tab(
        wb, rollups, baseline_total, stly_total, dc_names, details, cells
    )
    _write_lowest_cost_check_tab(wb, cells, score_detail)
    _write_supplier_comparison_tab(wb, config, cells, seeded)
    _write_fob_analysis_tab(wb, _gather_fob(session, seeded, final_round_id))
    # --- RELATIONSHIPS & NEGOTIATION (the repeated-game pillars 3 & 4). ---
    _write_share_relationships_tab(wb, details, scenarios, seeded, config, rec_code)
    _write_incumbent_retention_tab(wb, _gather_incumbent_retention(cells))
    _write_negotiation_dynamics_tab(wb, plays, inc_move, chal_move)
    # --- DILIGENCE (the receipts behind the recommendation). ---
    _write_coverage_tab(wb, coverage_rows, config)
    _write_capacity_check_tab(wb, _gather_capacity_check(session, seeded, award))
    _write_detailed_scoring_tab(wb, score_detail)
    # TF Comparison only when the cycle has >1 timeframe (else the view is degenerate).
    if len(seeded.tfs) > 1:
        _write_tf_comparison_tab(wb, cells)
        tab_index.append(
            ("TF Comparison", "Same supplier across timeframes? Seasonal split to align on.")
        )
    # Round Evolution only when the cycle has >1 round (the negotiation story needs movement).
    if evo_rows:
        _write_round_evolution_tab(wb, evo_rows, round_labels)
        tab_index.append(
            ("Round Evolution", "How did bids move across rounds? First→last price + direction.")
        )
    _write_data_quality_tab(wb, seeded, cells, score_detail)
    tab_index.append(
        ("Data Quality", "What's missing/thin? No-bids, thin competition, gate flags.")
    )
    # --- BUILD & SLICE — interactive + self-serve (KEEP, D25/D27). ---
    tab_index.append(("§ BUILD & SLICE", ""))
    prices_sheet = _write_prices_helper(wb, cells)
    _write_custom_scenario_tab(wb, config, cells, prices_sheet)
    # The Custom Dashboard recomputes the lenses LIVE off the builder (not the automated B).
    # Volume-weighted transit (matches the live SUMPRODUCT) so Δ-vs-rec reads 0 on the B pre-fill.
    t_pairs = [(d.transit_days, d.volume) for d in rec_details if d.transit_days is not None]
    t_vol = sum((vol for _t, vol in t_pairs), Decimal("0"))
    rec_avg_transit = (
        float(sum((t * vol for t, vol in t_pairs), Decimal("0")) / t_vol) if t_vol > 0 else 0.0
    )
    rec_fresh = sum(1 for t, _vol in t_pairs if t > FRESHNESS_WATCH_DAYS)
    rec_n_suppliers = len({d.supplier_name for d in rec_details})
    _write_custom_dashboard_tab(
        wb,
        len(cells),
        seeded,
        config,
        baseline_total,
        rec_spend,
        rec_avg_transit,
        rec_fresh,
        rec_n_suppliers,
        incumbent_names,
    )
    _write_data_pivot_tab(wb, details)
    tab_index.append(
        ("Custom Scenario", "Build your own: override the supplier per cell, recomputes live.")
    )
    tab_index.append(
        ("Custom Dashboard", "Your build's lenses LIVE — spend, savings, share, transit move too.")
    )
    tab_index.append(
        ("Data (pivot me)", "Slice it yourself: a flat Excel Table to drop a native PivotTable on.")
    )

    # Front door: the banded tab index on the Overview/Summary tab.
    _augment_summary_index(wb, tab_index)

    # Provenance: the generator writes the SYNTHETIC placeholder strap (demo lineage). On a REAL
    # cycle run, restamp the provenance tokens so the file never mislabels real names & prices.
    if not synthetic:
        _stamp_real_provenance(wb)

    wb.save(out_path)
    return out_path
