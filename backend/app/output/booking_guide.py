"""The post-award BOOKING GUIDE generators (D22/D23/D24) — importable.

The booking guide is the LAST step of the loop (after a scenario is SELECTED, promoted to an award,
frozen + signed off — D22): it is generated FROM THE FROZEN AWARD, never straight off a scenario.
Two audiences, two artifacts:

  * `write_booking_guide_internal_xlsx` — the buyers/pricing MASTER: one row per awarded
    DC × lot × item × TF — awarded supplier (NAME, D23), FOB/landed $/case, awarded volume,
    routing baseline + savings — what pricing uses to update the system (D9).
  * `write_supplier_award_guides_xlsx`  — one SHEET per awarded supplier in a single internal
    workbook: "here is what you've been awarded" (that supplier's lots/DCs/volumes/prices only).
  * `write_supplier_award_guide_files`  — the same, but as one separate FILE per supplier (the
    sendable artifact for E-37: a supplier's award notification attaches its own file, never the
    combined workbook — which would expose every other supplier's awards).

MOVED out of `demo/run_cycle_demo.py` so any feature (the pilot, not just the demo) can produce the
booking guides from a `CycleView` (the resolved cycle scope — demo seed OR `load_cycle`) plus a
frozen award. The award is described STRUCTURALLY (`BookingAwardView` / `BookingCellView` Protocols)
so the demo's concrete `SelectedAward` / `AwardedCell` satisfy it without this module depending on
them — and the pilot can pass an award assembled from the governed `awd.*` records the same way.

Names not keys (D23): every readable cell renders the resolved supplier/DC/lot/item/TF NAME; a
trailing key reference column trails for traceability. Presentation-quality (D24): reuses the
shared `app.output.formatting` styling pass. SYNTHETIC-friendly; decision-support provenance strap.
"""

from __future__ import annotations

from collections import defaultdict
from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import Protocol

from openpyxl import Workbook

from app.engine.formulas import awarded_cases, line_spend, savings_fraction
from app.output.formatting import (
    DECISION_SUPPORT_STRAP,
    NUMFMT_INT,
    NUMFMT_MONEY,
    NUMFMT_PCT,
    Col,
    format_table,
)
from app.output.types import CycleView


class BookingCellView(Protocol):
    """Structural view of one frozen award cell the booking guide reads (keys + economics).

    Read-only properties so a frozen dataclass (the demo's `AwardedCell`, or the pilot's award cell
    assembled from `awd.award_line`) satisfies it.
    """

    @property
    def dc_id(self) -> str: ...
    @property
    def lot_id(self) -> str: ...
    @property
    def item_id(self) -> str: ...
    @property
    def tf_id(self) -> str: ...
    @property
    def supplier_id(self) -> str: ...
    @property
    def volume_share(self) -> Decimal: ...
    @property
    def awarded_price(self) -> Decimal: ...
    @property
    def period_cases(self) -> Decimal: ...
    @property
    def routing_baseline(self) -> Decimal: ...


class BookingAwardView(Protocol):
    """Structural view of the promoted award the booking guide reads (selected scenario + cells)."""

    @property
    def scenario_code(self) -> str: ...
    @property
    def scenario_label(self) -> str: ...
    @property
    def cells(self) -> Sequence[BookingCellView]: ...


def _provenance_line(synthetic: bool) -> str:
    """The provenance token for a booking-guide strap — SYNTHETIC for a rehearsal, else LIVE."""

    return (
        "SYNTHETIC — names & prices invented"
        if synthetic
        else "LIVE CYCLE DATA — real names & prices"
    )


def write_booking_guide_internal_xlsx(
    cycle: CycleView,
    award: BookingAwardView,
    *,
    output_path: Path,
    synthetic: bool = False,
) -> Path:
    """The buyers/pricing master booking guide (D22 internal version) — FROM THE AWARD.

    One row per awarded DC × lot × item × TF: awarded supplier (NAME, D23), FOB/landed $/case,
    awarded volume, routing baseline + savings — what pricing uses to update the system (D9).
    """

    dc_name = {dc.id: dc.name for dc in cycle.dcs}
    lot_name = {lot.id: lot.name for lot in cycle.lots}
    item_name = {item.id: item.name for item in cycle.items}
    sup_name = {sup.id: sup.name for sup in cycle.suppliers}
    tf_name = {tf.id: tf.name for tf in cycle.tfs}
    dc_code = {dc.id: dc.code for dc in cycle.dcs}
    lot_code = {lot.id: lot.code for lot in cycle.lots}
    sup_code = {sup.id: sup.code for sup in cycle.suppliers}

    wb = Workbook()
    ws = wb.active
    assert ws is not None  # noqa: S101
    ws.title = "Internal Booking Guide"

    columns = [
        Col("DC", 18),
        Col("Lot", 22),
        Col("Item", 26),
        Col("Timeframe", 20),
        Col("Awarded Supplier", 26),
        Col("Volume Share", 13, NUMFMT_PCT),
        Col("FOB $/case", 14, NUMFMT_MONEY),
        Col("Landed $/case", 14, NUMFMT_MONEY),
        Col("Awarded Period Cases", 18, NUMFMT_INT, total="sum"),
        Col("Line Spend", 20, NUMFMT_MONEY, total="sum"),
        Col("Routing Baseline $/case", 20, NUMFMT_MONEY),
        Col("Savings vs Baseline", 16, NUMFMT_PCT),
        Col("Key ref (DC·lot·sup)", 22),  # traceability — names lead, keys trail (D23)
    ]
    header_row = 5  # title banner occupies rows 1-4
    row = header_row + 1
    n_rows = 0
    for c in sorted(
        award.cells, key=lambda c: (dc_name.get(c.dc_id, ""), lot_name.get(c.lot_id, ""))
    ):
        savings_frac = savings_fraction(c.routing_baseline, c.awarded_price)
        cases = awarded_cases(c.period_cases, c.volume_share)
        key_ref = (
            f"{dc_code.get(c.dc_id, c.dc_id[:6])}·{lot_code.get(c.lot_id, c.lot_id[:6])}·"
            f"{sup_code.get(c.supplier_id, c.supplier_id[:6])}"
        )
        ws.cell(row=row, column=1, value=dc_name.get(c.dc_id, c.dc_id[:6]))
        ws.cell(row=row, column=2, value=lot_name.get(c.lot_id, c.lot_id[:6]))
        ws.cell(row=row, column=3, value=item_name.get(c.item_id, c.item_id[:6]))
        ws.cell(row=row, column=4, value=tf_name.get(c.tf_id, c.tf_id[:6]))
        ws.cell(row=row, column=5, value=sup_name.get(c.supplier_id, c.supplier_id[:6]))
        ws.cell(row=row, column=6, value=float(c.volume_share))  # fraction -> 0.0% fmt
        # Demo economics use All-In as both the FOB and the landed basis (placeholders only).
        ws.cell(row=row, column=7, value=float(c.awarded_price))
        ws.cell(row=row, column=8, value=float(c.awarded_price))
        ws.cell(row=row, column=9, value=float(cases))
        ws.cell(row=row, column=10, value=float(line_spend(c.awarded_price, cases)))
        ws.cell(row=row, column=11, value=float(c.routing_baseline))
        ws.cell(row=row, column=12, value=float(savings_frac))  # fraction -> 0.0% fmt
        ws.cell(row=row, column=13, value=key_ref)
        row += 1
        n_rows += 1

    format_table(
        ws,
        title=f"INTERNAL BOOKING GUIDE — {cycle.cycle_name}",
        subtitle_lines=[
            f"Awarded from Scenario {award.scenario_code} ({award.scenario_label}) · "
            f"post-award: selected → awarded → frozen → signed off (D22)",
            f"Generated {date.today():%Y-%m-%d} · {_provenance_line(synthetic)} · "
            f"{DECISION_SUPPORT_STRAP}",
        ],
        columns=columns,
        n_body_rows=n_rows,
        header_row=header_row,
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def write_supplier_award_guides_xlsx(
    cycle: CycleView,
    award: BookingAwardView,
    *,
    output_path: Path,
    synthetic: bool = False,
) -> Path:
    """The per-supplier award guides (D22 per-supplier version) — one SHEET per awarded supplier.

    Each sheet shows ONLY that supplier's awarded lots/DCs/volumes/prices — "here is what you've
    been awarded." All NAMES (D23); no other supplier's data appears on a supplier's sheet.
    """

    dc_name = {dc.id: dc.name for dc in cycle.dcs}
    lot_name = {lot.id: lot.name for lot in cycle.lots}
    item_name = {item.id: item.name for item in cycle.items}
    sup_name = {sup.id: sup.name for sup in cycle.suppliers}
    tf_name = {tf.id: tf.name for tf in cycle.tfs}

    cells_by_sup: dict[str, list[BookingCellView]] = defaultdict(list)
    for c in award.cells:
        cells_by_sup[c.supplier_id].append(c)

    wb = Workbook()
    # Drop the default empty sheet once we add the first real one.
    default_ws = wb.active

    columns = [
        Col("DC", 18),
        Col("Lot", 22),
        Col("Item", 26),
        Col("Timeframe", 20),
        Col("Volume Share", 13, NUMFMT_PCT),
        Col("Awarded Period Cases", 18, NUMFMT_INT, total="sum"),
        Col("Awarded $/case", 16, NUMFMT_MONEY),
        Col("Line Spend", 20, NUMFMT_MONEY, total="sum"),
    ]
    header_row = 5  # title banner occupies rows 1-4
    # Stable, readable order: awarded suppliers by name.
    for sup_id in sorted(cells_by_sup, key=lambda s: sup_name.get(s, s)):
        sup_disp = sup_name.get(sup_id, sup_id[:6])
        title = _sheet_title(sup_disp)
        ws = wb.create_sheet(title=title)
        row = header_row + 1
        n_rows = 0
        for c in sorted(
            cells_by_sup[sup_id],
            key=lambda c: (dc_name.get(c.dc_id, ""), lot_name.get(c.lot_id, "")),
        ):
            cases = awarded_cases(c.period_cases, c.volume_share)
            ws.cell(row=row, column=1, value=dc_name.get(c.dc_id, c.dc_id[:6]))
            ws.cell(row=row, column=2, value=lot_name.get(c.lot_id, c.lot_id[:6]))
            ws.cell(row=row, column=3, value=item_name.get(c.item_id, c.item_id[:6]))
            ws.cell(row=row, column=4, value=tf_name.get(c.tf_id, c.tf_id[:6]))
            ws.cell(row=row, column=5, value=float(c.volume_share))
            ws.cell(row=row, column=6, value=float(cases))
            ws.cell(row=row, column=7, value=float(c.awarded_price))
            ws.cell(row=row, column=8, value=float(line_spend(c.awarded_price, cases)))
            row += 1
            n_rows += 1
        format_table(
            ws,
            title=f"AWARD GUIDE — {sup_disp}",
            subtitle_lines=[
                f"{cycle.cycle_name}: here is what you've been awarded",
                f"Generated {date.today():%Y-%m-%d} · {_provenance_line(synthetic)} · "
                f"{DECISION_SUPPORT_STRAP}",
            ],
            columns=columns,
            n_body_rows=n_rows,
            header_row=header_row,
        )

    if default_ws is not None and len(wb.sheetnames) > 1:
        wb.remove(default_ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


@dataclass(frozen=True)
class _SingleSupplierAward:
    """A `BookingAwardView` carrying ONE supplier's cells — to render that supplier's own file."""

    scenario_code: str
    scenario_label: str
    cells: Sequence[BookingCellView]


def supplier_guide_label(
    award_id: str, award_code: str, supplier_name: str, supplier_id: str
) -> str:
    """The `stage_filename` LABEL for one supplier's individual award guide.

    `stage_filename` slugifies this. The **award_id** (the unique award PK) makes the filename
    unique even if two awards in a run reuse the same `award_code` (codes aren't enforced unique),
    so a later freeze never overwrites an earlier award's per-supplier file and a draft never
    points at another award's guide; the award code + supplier name are kept for human readability,
    and the short supplier-id suffix keeps two like-named suppliers distinct.
    """

    return f"award_guide_{award_code}_{supplier_name}_{supplier_id[:6]}_{award_id}"


def write_supplier_award_guide_files(
    cycle: CycleView,
    award: BookingAwardView,
    *,
    paths_by_supplier: Mapping[str, Path],
    synthetic: bool = False,
) -> dict[str, Path]:
    """One INDIVIDUAL award-guide FILE per awarded supplier — only that supplier's data.

    The combined `write_supplier_award_guides_xlsx` packs every supplier on its own SHEET in one
    workbook: fine internally, but NOT safe to attach to a single supplier's email — it would expose
    every other supplier's awarded volumes/prices. This writes a separate one-sheet workbook per
    supplier (reusing the same renderer over a single-supplier award view) so each can be the
    attachment on that supplier's award notification (E-37). Returns the written path per supplier
    id (suppliers in `paths_by_supplier` that actually have awarded cells).
    """

    written: dict[str, Path] = {}
    for sup_id, path in paths_by_supplier.items():
        sup_cells = [c for c in award.cells if c.supplier_id == sup_id]
        if not sup_cells:
            continue
        one = _SingleSupplierAward(award.scenario_code, award.scenario_label, sup_cells)
        write_supplier_award_guides_xlsx(cycle, one, output_path=path, synthetic=synthetic)
        written[sup_id] = path
    return written


def _sheet_title(name: str) -> str:
    """A safe (<=31 char, no forbidden chars) Excel sheet title from a supplier name."""

    cleaned = "".join(ch for ch in name if ch not in "[]:*?/\\")
    return cleaned[:31] or "Supplier"
