"""D20 round-trip proof: generate a template -> fill (synthetic) -> ingest -> grain round-trips.

The system owns BOTH ends (generate + ingest) of ONE owned schema (D20). This asserts that a
template generated for a synthetic cycle scope, filled with synthetic bids, ingests back to the
SAME supplier x DC x lot x item x TF x round grain, with the cost COMPONENTS preserved exactly.
"""

from __future__ import annotations

from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from app.domain.bid.bid_ingester import (
    PRICE_BASIS_ALL_IN,
    PRICE_BASIS_FALLBACK,
    Completeness,
    QuarantineReason,
    ingest_template,
)
from app.domain.bid.template_generator import (
    build_template_workbook,
    generate_template_bytes,
)
from app.domain.bid.template_schema import (
    BID_HEADERS,
    HEADER_ROW,
    SHEET_BIDS,
    SHEET_CAPACITY,
    SHEET_INSTRUCTIONS,
    BidColumn,
)
from tests.bid.synthetic import build_scope

_D = Decimal


def _fill_bids(template_bytes: bytes) -> bytes:
    """Fill the generated template's blank price cells with synthetic bids (the supplier step).

    Row 0 -> All-In bid. Row 1 -> component-fallback bid. The rest -> All-In with volume.
    """

    wb = load_workbook(BytesIO(template_bytes))
    ws = wb[SHEET_BIDS]
    col = {h: i for i, h in enumerate(BID_HEADERS, start=1)}

    def put(row: int, column: BidColumn, value: object) -> None:
        ws.cell(row=row, column=col[column.value], value=value)

    body_rows = list(range(HEADER_ROW + 1, ws.max_row + 1))
    for offset, row in enumerate(body_rows):
        if offset == 0:
            # All-In bid (price basis ALL_IN), with volume.
            put(row, BidColumn.ALL_IN, "100.00")
            put(row, BidColumn.TOTAL_VOL_OFFERED, "500")
            put(row, BidColumn.WEEKLY_VOL_OFFERED, "50")
            put(row, BidColumn.PRICING_COMMENTS, "firm")
        elif offset == 1:
            # Component-fallback bid: FOB 90 + Delivery 5 + VegCool 3 - LotDisc 2 = 96.00.
            put(row, BidColumn.FOB, "90.00")
            put(row, BidColumn.DELIVERY_SURCHARGE, "5.00")
            put(row, BidColumn.VEGCOOL_SURCHARGE, "3.00")
            put(row, BidColumn.LOT_DISCOUNT, "2.00")
            put(row, BidColumn.TOTAL_VOL_OFFERED, "300")
        else:
            put(row, BidColumn.ALL_IN, "101.50")
            put(row, BidColumn.TOTAL_VOL_OFFERED, "200")

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def test_generated_template_has_owned_sheets_and_grain() -> None:
    scope = build_scope()
    wb = build_template_workbook(scope)
    assert wb.sheetnames == [SHEET_INSTRUCTIONS, SHEET_BIDS, SHEET_CAPACITY]
    ws = wb[SHEET_BIDS]
    headers = [ws.cell(row=HEADER_ROW, column=c).value for c in range(1, len(BID_HEADERS) + 1)]
    assert headers == list(BID_HEADERS)
    # One Bids body row per scope cell.
    assert ws.max_row - HEADER_ROW == len(scope.rows)


def test_round_trip_grain_and_components_exact() -> None:
    scope = build_scope()

    template_bytes = generate_template_bytes(scope)
    filled_bytes = _fill_bids(template_bytes)
    # D21: ingest OUR template by KEY VALIDATION against the scope (no name resolver in sight).
    result = ingest_template(filled_bytes, scope)

    # Nothing quarantined — a clean self-generated round-trip.
    assert result.quarantined == []
    # No name warnings — the generator wrote the scope's own names against its own keys.
    assert result.name_warnings == []
    # Same number of lines as scope rows (grain count round-trips).
    assert len(result.lines) == len(scope.rows)

    # The ingested identity grain matches the scope's identity grain EXACTLY (the D20 proof).
    ingested_grain = {
        (
            line.round_code,
            line.identity.supplier_id,
            line.identity.dc_id,
            line.identity.lot_id,
            line.identity.item_id,
            line.identity.tf_code,
        )
        for line in result.lines
    }
    scope_grain = {
        (r.round_code, r.supplier_id, r.dc_id, r.lot_id, r.item_id, r.tf_code) for r in scope.rows
    }
    assert ingested_grain == scope_grain

    # D21: the EMBEDDED KEY tuples round-trip EXACTLY — generated scope keys == ingested keys.
    ingested_keys = {
        (
            line.identity.cycle_id,
            line.identity.round_id,
            line.identity.tf_id,
            line.identity.lot_id,
            line.identity.item_id,
            line.identity.dc_id,
            line.identity.supplier_id,
        )
        for line in result.lines
    }
    assert ingested_keys == scope.key_set()

    by_row = {line.source_row_number: line for line in result.lines}
    first_row = min(by_row)

    # Row 0 — All-In basis, price verbatim, components carried.
    all_in_line = by_row[first_row]
    assert all_in_line.price_basis == PRICE_BASIS_ALL_IN
    assert all_in_line.landed_cost_per_case == _D("100.00")
    assert all_in_line.components.all_in == _D("100.00")
    assert all_in_line.completeness is Completeness.BID
    assert all_in_line.total_vol_offered == _D("500")
    assert all_in_line.weekly_vol_offered == _D("50")
    assert all_in_line.pricing_comments == "firm"

    # Row 1 — component fallback, §7 sum, components preserved exactly.
    fb_line = by_row[first_row + 1]
    assert fb_line.price_basis == PRICE_BASIS_FALLBACK
    assert fb_line.landed_cost_per_case == _D("96.00")  # 90 + 5 + 3 - 2
    assert fb_line.components.fob == _D("90.00")
    assert fb_line.components.delivery_surcharge == _D("5.00")
    assert fb_line.components.vegcool_surcharge == _D("3.00")
    assert fb_line.components.lot_discount == _D("2.00")
    assert fb_line.completeness is Completeness.BID

    # All lines flagged BID; counts reconcile.
    assert result.bid_count == len(scope.rows)
    assert result.no_bid_count == 0
    assert result.incomplete_count == 0


def _fill_all_in(template_bytes: bytes) -> tuple[bytes, list[int]]:
    """Fill every body row with a simple All-In bid; return (bytes, body row numbers)."""

    wb = load_workbook(BytesIO(template_bytes))
    ws = wb[SHEET_BIDS]
    col = {h: i for i, h in enumerate(BID_HEADERS, start=1)}
    body_rows = list(range(HEADER_ROW + 1, ws.max_row + 1))
    for row in body_rows:
        ws.cell(row=row, column=col[BidColumn.ALL_IN.value], value="100.00")
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue(), body_rows


def test_tampered_embedded_key_quarantines_not_resolved() -> None:
    """D21: a row whose embedded LOT_ID is tampered to an unknown key quarantines (UNKNOWN_KEY).

    The display names are LEFT INTACT and valid — proving we do NOT silently re-resolve the row
    from its names; an out-of-scope embedded key is fatal for that row.
    """

    scope = build_scope()
    template_bytes = generate_template_bytes(scope)

    wb = load_workbook(BytesIO(template_bytes))
    ws = wb[SHEET_BIDS]
    col = {h: i for i, h in enumerate(BID_HEADERS, start=1)}
    body_rows = list(range(HEADER_ROW + 1, ws.max_row + 1))
    for row in body_rows:
        ws.cell(row=row, column=col[BidColumn.ALL_IN.value], value="100.00")
    # Tamper ONLY the embedded LOT_ID key on the first row (names stay valid + in-scope).
    ws.cell(row=body_rows[0], column=col[BidColumn.LOT_ID.value], value="LOT-TAMPERED")
    buffer = BytesIO()
    wb.save(buffer)
    tampered = buffer.getvalue()

    result = ingest_template(tampered, scope)

    assert len(result.quarantined) == 1
    assert result.quarantined[0].reason is QuarantineReason.UNKNOWN_KEY
    # The tampered row did NOT become a parsed line (no name-based re-resolve).
    assert len(result.lines) == len(scope.rows) - 1


def test_blank_embedded_key_quarantines_missing_key() -> None:
    """D21: clearing a locked KEY-ID cell quarantines MISSING_KEY (never name-resolved)."""

    scope = build_scope()
    template_bytes = generate_template_bytes(scope)
    filled, body_rows = _fill_all_in(template_bytes)

    wb = load_workbook(BytesIO(filled))
    ws = wb[SHEET_BIDS]
    col = {h: i for i, h in enumerate(BID_HEADERS, start=1)}
    # Clear the locked SUPPLIER_ID cell (openpyxl ignores value=None on a populated cell, so "").
    ws.cell(row=body_rows[0], column=col[BidColumn.SUPPLIER_ID.value], value="")
    buffer = BytesIO()
    wb.save(buffer)

    result = ingest_template(buffer.getvalue(), scope)

    assert len(result.quarantined) == 1
    assert result.quarantined[0].reason is QuarantineReason.MISSING_KEY
    assert len(result.lines) == len(scope.rows) - 1


def test_name_mismatch_warns_but_keys_still_resolve() -> None:
    """D21: a display NAME that disagrees with the keyed identity WARNS — it does not re-resolve.

    We overwrite a row's SUPPLIER display name (an attribute) but leave its embedded keys intact.
    The row STILL loads on its keys (correct supplier_id), and a name-mismatch warning is raised.
    """

    scope = build_scope()
    template_bytes = generate_template_bytes(scope)
    filled, body_rows = _fill_all_in(template_bytes)

    wb = load_workbook(BytesIO(filled))
    ws = wb[SHEET_BIDS]
    col = {h: i for i, h in enumerate(BID_HEADERS, start=1)}
    # Capture the embedded supplier_id we expect the row to keep loading under.
    kept_supplier_id = ws.cell(row=body_rows[0], column=col[BidColumn.SUPPLIER_ID.value]).value
    ws.cell(row=body_rows[0], column=col[BidColumn.SUPPLIER.value], value="Totally Wrong Name")
    buffer = BytesIO()
    wb.save(buffer)

    result = ingest_template(buffer.getvalue(), scope)

    # No quarantine — the row loads on its KEYS (the name is an attribute, not the join key).
    assert result.quarantined == []
    assert len(result.lines) == len(scope.rows)
    # Exactly one name-mismatch WARNING, on the Supplier column.
    assert len(result.name_warnings) == 1
    warning = result.name_warnings[0]
    assert warning.column == BidColumn.SUPPLIER.value
    assert warning.found_name == "Totally Wrong Name"
    # The row that warned still carries the CORRECT keyed supplier_id (no re-resolve to the name).
    warned_line = next(
        line for line in result.lines if line.source_row_number == body_rows[0]
    )
    assert warned_line.identity.supplier_id == kept_supplier_id
