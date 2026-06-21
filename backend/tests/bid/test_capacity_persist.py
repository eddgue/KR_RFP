"""E-38 capacity persistence: `ingest_bids` writes bid.capacity_statement + bid.capacity_constraint.

Integration (real Postgres): stand up a real cycle, generate its template, fill BOTH the Bids and
the Capacity sheets, ingest, and assert the stated ceilings landed as CELL-scoped constraints under
a per-supplier statement that RIDES the supplier's own bid submission/artifact (shared FK chain),
and that a re-submission SUPERSEDES the prior statement (so the cap check never reads stale data).
"""

from __future__ import annotations

from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import text

from app.domain.bid.template_schema import (
    BODY_START_ROW,
    CAPACITY_KEY_ID_COLUMNS,
    HEADER_ROW,
    SHEET_BIDS,
    SHEET_CAPACITY,
    BidColumn,
    CapacityColumn,
)
from app.pilot.service import PilotService
from app.pilot.vault import stage_filename
from tests.pilot.test_pilot_cycle_e2e import _build_filled_setup

CapKeys = tuple[str, str, str, str, str, str]  # cycle, supplier, dc, lot, item, tf


def _headers(ws: Worksheet) -> dict[str, int]:
    return {
        str(ws.cell(row=HEADER_ROW, column=c).value).strip(): c
        for c in range(1, ws.max_column + 1)
        if ws.cell(row=HEADER_ROW, column=c).value is not None
    }


def _fill(template_bytes: bytes, cap_fills: dict[int, tuple[float | None, float | None]]) -> bytes:
    """Fill every Bids row with an All-In bid, and the Capacity rows in `cap_fills` with ceilings.

    `cap_fills` maps a Capacity body-row offset (0-based) -> (max_weekly, max_total). Returns the
    saved bytes. Bids are filled on every cell so each supplier has a real submission/artifact.
    """

    wb = load_workbook(BytesIO(template_bytes))
    bids = wb[SHEET_BIDS]
    bh = _headers(bids)
    for row in range(BODY_START_ROW, bids.max_row + 1):
        sup = str(bids.cell(row=row, column=bh[BidColumn.SUPPLIER.value]).value or "").strip()
        if not sup:
            continue
        bids.cell(row=row, column=bh[BidColumn.ALL_IN.value], value=20.0)
        bids.cell(row=row, column=bh[BidColumn.WEEKLY_VOL_OFFERED.value], value=600)
        bids.cell(row=row, column=bh[BidColumn.TOTAL_VOL_OFFERED.value], value=7800)

    cap = wb[SHEET_CAPACITY]
    ch = _headers(cap)
    for offset, (weekly, total) in cap_fills.items():
        row = BODY_START_ROW + offset
        if weekly is not None:
            cap.cell(row=row, column=ch[CapacityColumn.MAX_WEEKLY_CASES.value], value=weekly)
        if total is not None:
            cap.cell(row=row, column=ch[CapacityColumn.MAX_TOTAL_CASES.value], value=total)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _capacity_keys(template_bytes: bytes, offset: int) -> CapKeys:
    """The embedded key identity of a Capacity body row (cycle, supplier, dc, lot, item, tf)."""

    wb = load_workbook(BytesIO(template_bytes))
    cap = wb[SHEET_CAPACITY]
    ch = _headers(cap)
    row = BODY_START_ROW + offset
    return tuple(  # type: ignore[return-value]
        str(cap.cell(row=row, column=ch[c.value]).value) for c in CAPACITY_KEY_ID_COLUMNS
    )


def _distinct_cell_offsets(template_bytes: bytes, n: int) -> list[tuple[int, CapKeys]]:
    """Pick `n` Capacity body rows with DISTINCT (supplier, dc, lot, tf) cells.

    A CELL constraint is per supplier x dc x lot x tf — two suppliers (or two items under one lot)
    can share a dc/lot/tf, so the test must fill rows that are unambiguous at that grain.
    """

    wb = load_workbook(BytesIO(template_bytes))
    cap = wb[SHEET_CAPACITY]
    seen: set[tuple[str, str, str, str]] = set()
    chosen: list[tuple[int, CapKeys]] = []
    for offset in range(cap.max_row - HEADER_ROW):
        keys = _capacity_keys(template_bytes, offset)
        cell = (keys[1], keys[2], keys[3], keys[5])  # supplier, dc, lot, tf
        if cell in seen:
            continue
        seen.add(cell)
        chosen.append((offset, keys))
        if len(chosen) == n:
            break
    return chosen


def _prepare_cycle(tmp_path: Path, db_session) -> tuple[PilotService, object, str, Path]:  # type: ignore[no-untyped-def]
    service = PilotService(tmp_path, isolate_db=False)
    paths = service.start_run(commodity="Tomatoes", label="Capacity Persist")
    setup_path = paths.inputs / stage_filename(1, "setup_kickoff")
    setup_path.write_bytes(_build_filled_setup())
    cycle_id = service.ingest_setup(db_session, paths, setup_path)
    template_path = service.generate_bid_template(db_session, paths, 1)
    return service, paths, cycle_id, template_path


@pytest.mark.integration
def test_capacity_persisted_as_cell_constraints(tmp_path: Path, db_session) -> None:  # type: ignore[no-untyped-def]
    service, paths, cycle_id, template_path = _prepare_cycle(tmp_path, db_session)

    # Two DISTINCT cells: cell A (weekly + total), cell B (total only).
    (off_a, keys_a), (off_b, keys_b) = _distinct_cell_offsets(template_path.read_bytes(), 2)
    template_path.write_bytes(
        _fill(template_path.read_bytes(), {off_a: (500.0, 6500.0), off_b: (None, 7000.0)})
    )

    service.ingest_bids(db_session, paths, 1, template_path)

    cons = db_session.execute(
        text(
            "SELECT cs.supplier_id, cc.scope_type, cc.dc_id, cc.lot_id, cc.tf_id, "
            "cc.max_weekly_cases, cc.max_period_cases, cc.capacity_statement_id "
            "FROM bid.capacity_constraint cc JOIN bid.capacity_statement cs "
            "  ON cs.capacity_statement_id = cc.capacity_statement_id "
            "WHERE cc.cycle_id = :c"
        ),
        {"c": cycle_id},
    ).all()
    assert len(cons) == 2
    by_cell = {(r[0], r[2], r[3], r[4]): r for r in cons}  # (supplier, dc, lot, tf)

    # cell A -> supplier/dc/lot/tf from keys_a (indices 1,2,3,5), weekly 500 + total 6500.
    c_a = by_cell[(keys_a[1], keys_a[2], keys_a[3], keys_a[5])]
    assert c_a[1] == "CELL"
    assert float(c_a[5]) == 500.0
    assert float(c_a[6]) == 6500.0
    # cell B -> total only (weekly NULL).
    c_b = by_cell[(keys_b[1], keys_b[2], keys_b[3], keys_b[5])]
    assert c_b[5] is None
    assert float(c_b[6]) == 7000.0

    # Each constraint hangs off a SUBMITTED statement that rides a real submission + artifact.
    stmt_ids = {c_a[7], c_b[7]}
    stmts = db_session.execute(
        text(
            "SELECT supplier_id, status, submission_id, source_artifact_id, round_id "
            "FROM bid.capacity_statement WHERE capacity_statement_id = ANY(:ids)"
        ),
        {"ids": list(stmt_ids)},
    ).all()
    assert {s[1] for s in stmts} == {"SUBMITTED"}
    for supplier_id, _status, submission_id, artifact_id, _round in stmts:
        assert submission_id is not None and artifact_id is not None
        # The statement rides the SAME submission the supplier's bids came in on (shared chain).
        rode = db_session.execute(
            text(
                "SELECT count(*) FROM bid.bid_submission WHERE submission_id = :s "
                "AND supplier_id = :sup AND cycle_id = :c"
            ),
            {"s": submission_id, "sup": supplier_id, "c": cycle_id},
        ).scalar_one()
        assert rode == 1


@pytest.mark.integration
def test_capacity_resubmit_supersedes_prior_statement(tmp_path: Path, db_session) -> None:  # type: ignore[no-untyped-def]
    service, paths, cycle_id, template_path = _prepare_cycle(tmp_path, db_session)
    keys0 = _capacity_keys(template_path.read_bytes(), 0)

    # First submission: weekly 500 on cell 0.
    first = paths.inputs / "first_bids.xlsx"
    first.write_bytes(_fill(template_path.read_bytes(), {0: (500.0, None)}))
    service.ingest_bids(db_session, paths, 1, first)

    # Re-submission for the same round: weekly 900 on the same cell.
    second = paths.inputs / "second_bids.xlsx"
    second.write_bytes(_fill(template_path.read_bytes(), {0: (900.0, None)}))
    service.ingest_bids(db_session, paths, 1, second)

    statuses = [
        r[0]
        for r in db_session.execute(
            text(
                "SELECT cs.status FROM bid.capacity_statement cs WHERE cs.cycle_id = :c "
                "AND cs.supplier_id = :sup"
            ),
            {"c": cycle_id, "sup": keys0[1]},
        ).all()
    ]
    # Exactly one active (SUBMITTED) statement; the prior one is SUPERSEDED (append-only).
    assert sorted(statuses) == ["SUBMITTED", "SUPERSEDED"]

    # The ACTIVE statement's ceiling is the new 900 — the cap check filters out the superseded one.
    active_weekly = db_session.execute(
        text(
            "SELECT cc.max_weekly_cases FROM bid.capacity_constraint cc "
            "JOIN bid.capacity_statement cs "
            "  ON cs.capacity_statement_id = cc.capacity_statement_id "
            "WHERE cs.cycle_id = :c AND cs.supplier_id = :sup AND cs.status = 'SUBMITTED'"
        ),
        {"c": cycle_id, "sup": keys0[1]},
    ).scalar_one()
    assert float(active_weekly) == 900.0


@pytest.mark.integration
def test_load_active_capacity_reads_persisted_ceilings(tmp_path: Path, db_session) -> None:  # type: ignore[no-untyped-def]
    """The reader the E-38b capacity tab depends on returns the persisted CELL ceiling by cell."""

    from app.output.capacity_check import load_active_capacity

    service, paths, cycle_id, template_path = _prepare_cycle(tmp_path, db_session)
    (off_a, keys_a) = _distinct_cell_offsets(template_path.read_bytes(), 1)[0]
    template_path.write_bytes(_fill(template_path.read_bytes(), {off_a: (500.0, 6500.0)}))
    service.ingest_bids(db_session, paths, 1, template_path)

    cap = load_active_capacity(db_session, cycle_id)
    cell = (keys_a[1], keys_a[2], keys_a[3], keys_a[5])  # supplier, dc, lot, tf
    assert cell in cap
    assert float(cap[cell].max_period_cases) == 6500.0
    assert float(cap[cell].max_weekly_cases) == 500.0
