"""Legacy-resilience: a SYNTHETIC legacy-shaped fixture maps to the same grain or quarantines.

D20: our owned template is the live contract; messy legacy shapes are migration-resilience proof
only. This builds OUR OWN synthetic legacy-shaped workbook (NOT the quarantined real files) that
mimics the reference reality (header NOT on our row, reference column wording, a `No Bid` cell, an
unmapped trailing column) and asserts the adapter+ingester map it to the SAME bid_line grain, with
no-bid cells flagged and unresolvable rows quarantined — never guessed.
"""

from __future__ import annotations

from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook

from app.domain.bid.bid_ingester import Completeness, QuarantineReason
from app.domain.bid.legacy_adapter import ingest_legacy
from tests.bid.synthetic import build_resolver

_D = Decimal

_LEGACY_SHEET = "6. Vol and Pricing Capability"  # echoes the reference's pricing tab name
_LEGACY_HEADER_ROW = 4  # NOT our row 2 — the reference puts headers at row 4 (§1.2)

# Legacy headers (reference wording), incl. a trailing column we don't map (inflation, §2).
_LEGACY_HEADERS = [
    "Round ID",
    "Bid Type",
    "Supplier",
    "DC Name",
    "Lot_ID",
    "Product Description",
    "TF",
    "All-In $/Case",
    "FOB $/Case (Corrugate)",
    "Delivery Surcharge",
    "VegCool Surcharge",
    "Lot Discount",
    "Weekly Vol Cap",
    "Total Vol Cap",
    "Comments",
    "Distance (mi) to DC",  # unmapped trailing field -> dropped, not guessed
]


def _build_synthetic_legacy_workbook() -> bytes:
    """A synthetic legacy-shaped workbook (our own — not real data)."""

    wb = Workbook()
    ws = wb.active
    assert ws is not None  # noqa: S101
    ws.title = _LEGACY_SHEET
    ws.cell(row=1, column=1, value="Kroger Bid Document — Synthetic Legacy")  # title band

    for i, header in enumerate(_LEGACY_HEADERS, start=1):
        ws.cell(row=_LEGACY_HEADER_ROW, column=i, value=header)

    def write(row: int, values: dict[str, object]) -> None:
        for header, value in values.items():
            ws.cell(row=row, column=_LEGACY_HEADERS.index(header) + 1, value=value)

    base = _LEGACY_HEADER_ROW + 1
    # An All-In row.
    write(
        base,
        {
            "Round ID": "R1",
            "Bid Type": "Initial FOB",
            "Supplier": "Acme Produce",
            "DC Name": "Atlanta DC",
            "Lot_ID": "Grape Lot",
            "Product Description": "Grape Tomato 1lb",
            "TF": "TF1",
            "All-In $/Case": "100.00",
            "Total Vol Cap": "500",
            "Distance (mi) to DC": "412",
        },
    )
    # A component-fallback row.
    write(
        base + 1,
        {
            "Round ID": "R1",
            "Bid Type": "Initial FOB",
            "Supplier": "Bravo Farms",
            "DC Name": "Memphis DC",
            "Lot_ID": "Roma Lot",
            "Product Description": "Roma XL 25lb",
            "TF": "TF1",
            "FOB $/Case (Corrugate)": "90.00",
            "Delivery Surcharge": "5.00",
            "VegCool Surcharge": "3.00",
            "Lot Discount": "2.00",
            "Total Vol Cap": "300",
        },
    )
    # A `No Bid` declined cell -> must become a no_bid line (not a zero price), §4.
    write(
        base + 2,
        {
            "Round ID": "R1",
            "Bid Type": "Initial FOB",
            "Supplier": "Acme Produce",
            "DC Name": "Memphis DC",
            "Lot_ID": "Roma Lot",
            "Product Description": "Roma XL 25lb",
            "TF": "TF1",
            "All-In $/Case": "No Bid",
        },
    )
    # An unresolvable supplier -> quarantined (not guessed).
    write(
        base + 3,
        {
            "Round ID": "R1",
            "Bid Type": "Initial FOB",
            "Supplier": "Ghost Vendor Inc",
            "DC Name": "Atlanta DC",
            "Lot_ID": "Grape Lot",
            "Product Description": "Grape Tomato 1lb",
            "TF": "TF1",
            "All-In $/Case": "99.00",
        },
    )

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def test_legacy_shape_maps_to_same_grain_or_quarantines() -> None:
    legacy_bytes = _build_synthetic_legacy_workbook()
    resolver = build_resolver()

    result = ingest_legacy(
        legacy_bytes,
        resolver,
        sheet_name=_LEGACY_SHEET,
        legacy_header_row=_LEGACY_HEADER_ROW,
    )

    # The unresolvable supplier row is quarantined (resilience: clean failure, no guess).
    assert len(result.quarantined) == 1
    assert result.quarantined[0].reason is QuarantineReason.UNRESOLVED_SUPPLIER

    # Three resolvable rows mapped to the SAME bid_line grain as our owned template.
    assert len(result.lines) == 3
    by_id = {
        (
            line.identity.supplier_id,
            line.identity.dc_id,
            line.identity.lot_id,
            line.identity.item_id,
            line.identity.tf_code,
        ): line
        for line in result.lines
    }
    # All-In row resolved + costed.
    all_in = by_id[("SUP-1", "DC-1", "LOT-1", "ITEM-1", "TF1")]
    assert all_in.landed_cost_per_case == _D("100.00")
    assert all_in.completeness is Completeness.BID

    # Component-fallback row reconstructs the §7 sum from the legacy columns.
    fb = by_id[("SUP-2", "DC-2", "LOT-2", "ITEM-2", "TF1")]
    assert fb.landed_cost_per_case == _D("96.00")
    assert fb.completeness is Completeness.BID

    # The `No Bid` cell became a no_bid line (NOT a zero price).
    no_bid = by_id[("SUP-1", "DC-2", "LOT-2", "ITEM-2", "TF1")]
    assert no_bid.completeness is Completeness.NO_BID
    assert no_bid.landed_cost_per_case is None
