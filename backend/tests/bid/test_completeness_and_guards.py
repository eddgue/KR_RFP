"""Completeness flags + the double-subtract guard + quarantine (don't guess).

Asserts: no_bid (all price cells blank), incomplete (partial price intent), the §7 double-subtract
case is BLOCKED (quarantined, not silently recomputed), and an unresolvable supplier is quarantined
rather than guessed.
"""

from __future__ import annotations

from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from app.domain.bid.bid_ingester import (
    Completeness,
    ParsedComponents,
    QuarantineReason,
    construct_price,
    ingest_template,
)
from app.domain.bid.template_generator import generate_template_bytes
from app.domain.bid.template_schema import BID_HEADERS, HEADER_ROW, SHEET_BIDS, BidColumn
from tests.bid.synthetic import build_scope

_D = Decimal


def _fill(template_bytes: bytes, filler) -> bytes:  # type: ignore[no-untyped-def]
    wb = load_workbook(BytesIO(template_bytes))
    ws = wb[SHEET_BIDS]
    col = {h: i for i, h in enumerate(BID_HEADERS, start=1)}
    body_rows = list(range(HEADER_ROW + 1, ws.max_row + 1))
    filler(ws, col, body_rows)
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def test_no_bid_and_incomplete_and_bid_flags() -> None:
    scope = build_scope()
    template = generate_template_bytes(scope)

    def filler(ws, col, body_rows):  # type: ignore[no-untyped-def]
        # Row 0: a real bid. Row 1: leave price/volume blank but KEEP the embedded keys (no_bid).
        # Row 2: only a volume, no price (incomplete). Remaining rows: real bids.
        ws.cell(row=body_rows[0], column=col[BidColumn.ALL_IN.value], value="100.00")
        # body_rows[1] price/volume left blank -> no_bid (keys are intact from the generator)
        ws.cell(
            row=body_rows[2], column=col[BidColumn.TOTAL_VOL_OFFERED.value], value="120"
        )  # vol but no price -> incomplete
        for row in body_rows[3:]:
            ws.cell(row=row, column=col[BidColumn.ALL_IN.value], value="100.00")

    result = ingest_template(_fill(template, filler), scope)

    assert result.quarantined == []
    assert result.no_bid_count == 1
    assert result.incomplete_count == 1
    assert result.bid_count == len(scope.rows) - 2
    # The blank row is parsed (placed in the grain) but flagged NO_BID, not dropped.
    no_bids = [line for line in result.lines if line.completeness is Completeness.NO_BID]
    assert len(no_bids) == 1
    assert no_bids[0].landed_cost_per_case is None


def test_double_subtract_blocked_and_quarantined() -> None:
    scope = build_scope()
    template = generate_template_bytes(scope)

    def filler(ws, col, body_rows):  # type: ignore[no-untyped-def]
        # Row 0: All-In AND a Lot Discount both populated -> double-subtract footgun.
        ws.cell(row=body_rows[0], column=col[BidColumn.ALL_IN.value], value="95.00")
        ws.cell(row=body_rows[0], column=col[BidColumn.LOT_DISCOUNT.value], value="2.00")
        for row in body_rows[1:]:
            ws.cell(row=row, column=col[BidColumn.ALL_IN.value], value="100.00")

    result = ingest_template(_fill(template, filler), scope)

    # The ambiguous row is quarantined (not silently recomputed), the rest ingest cleanly.
    assert len(result.quarantined) == 1
    assert result.quarantined[0].reason is QuarantineReason.DOUBLE_SUBTRACT
    assert result.bid_count == len(scope.rows) - 1


def test_construct_price_guard_unit() -> None:
    # All-In verbatim when no discount.
    price, basis, err = construct_price(
        ParsedComponents(_D("95.00"), None, _D("0"), _D("0"), _D("0"))
    )
    assert price == _D("95.00") and basis == "ALL_IN" and err is None
    # All-In + Lot Discount -> blocked (no price, error surfaced).
    price, basis, err = construct_price(
        ParsedComponents(_D("95.00"), None, _D("0"), _D("0"), _D("2.00"))
    )
    assert price is None and basis is None and err is not None
    # Fallback sum.
    price, basis, err = construct_price(
        ParsedComponents(None, _D("90.00"), _D("5.00"), _D("3.00"), _D("2.00"))
    )
    assert price == _D("96.00") and basis == "COMPONENT_FALLBACK" and err is None


def test_unknown_supplier_name_warns_not_quarantined_d21() -> None:
    """D21: for OUR template, an unknown supplier NAME is an attribute mismatch — WARN, not fail.

    The pre-D21 behavior (quarantine UNRESOLVED_SUPPLIER on an unknown name) is gone for our own
    template: identity comes from the embedded KEY, not the name. An unknown name therefore loads
    on its keys and raises a name-mismatch warning — it is NEVER re-resolved from the name. The
    name-resolver quarantine survives ONLY on the legacy path (see test_legacy_resilience).
    """

    scope = build_scope()
    template = generate_template_bytes(scope)

    def filler(ws, col, body_rows):  # type: ignore[no-untyped-def]
        # Corrupt one row's supplier display NAME to an unknown label (keys left intact), price all.
        ws.cell(
            row=body_rows[0], column=col[BidColumn.SUPPLIER.value], value="Unknown Vendor LLC"
        )
        for row in body_rows:
            ws.cell(row=row, column=col[BidColumn.ALL_IN.value], value="100.00")

    result = ingest_template(_fill(template, filler), scope)

    # No quarantine — the row loads on its embedded keys.
    assert result.quarantined == []
    assert len(result.lines) == len(scope.rows)
    # The mismatching display name produced a WARNING (not a re-resolve).
    assert len(result.name_warnings) == 1
    assert result.name_warnings[0].column == BidColumn.SUPPLIER.value
    assert result.name_warnings[0].found_name == "Unknown Vendor LLC"
