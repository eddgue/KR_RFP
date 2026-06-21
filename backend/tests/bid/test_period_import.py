"""Intake stores bids FLAT at the 13 fiscal periods; engine/awards stay timeframe-grain (§1a).

End-to-end proof of the flat-13 storage path on a real Postgres, now that the fan-out is wired
NATIVELY into intake (`PilotService._persist_bid_lines`) rather than demonstrated by the test:

  1. stand up a real cycle (valid FK parents) whose single timeframe's calendar dates DO map onto
     the seeded FY2026 calendar, and INGEST a bid template filled across the full column set (All-In
     / FOB / surcharges / discount / volume);
  2. assert intake STORED the bids per-period: each logical priced cell is fanned out to one
     `bid.bid_line` per fiscal period in its timeframe's span (`fiscal_period_id` populated), every
     period in the span is covered exactly once, and every column survives the fan-out verbatim;
  3. assert the API/contract count semantics: `ingest_bids` returns the LOGICAL line count (one per
     identity cell), NOT the fanned storage row count;
  4. assert the ENGINE / SCENARIO / AWARD output is UNCHANGED by the storage-grain change — the
     runner collapses the period rows to one representative row per (dc, lot, tf, supplier) before
     building `BidInput`, so a period-grain run and a control tf-grain run yield byte-identical
     engine inputs, scores, and scenario awards (the safety invariant);
  5. the GRACEFUL FALLBACK: a synthetic cycle whose timeframe dates fall OUTSIDE the seeded calendar
     still ingests at tf-grain (`fiscal_period_id` NULL), one row per cell, exactly as before.

Uses the potato sample when present (real prices); falls back to synthetic prices when it is absent
(the sample is git-ignored), so the test is CI-safe. Integration (needs a live Postgres at head).
"""

from __future__ import annotations

import json
from decimal import Decimal
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import text

from app.domain.bid.models import BidLine
from app.domain.bid.template_schema import BODY_START_ROW, HEADER_ROW, SHEET_BIDS, BidColumn
from app.domain.eng.runner import EngineRunner
from app.fiscal.calendar import period_for_date
from app.pilot.service import PilotService
from app.pilot.vault import stage_filename
from tests.pilot.test_pilot_cycle_e2e import _build_filled_setup

_POTATO_SAMPLE = (
    Path(__file__).resolve().parents[2].parent
    / "reference"
    / "samples"
    / ("potato_2026_rfp_input.xlsx")
)

# The price/volume columns we assert survive the fan-out verbatim onto every period row.
_CARRIED_COLUMNS = (
    "submitted_all_in_case",
    "fob_case",
    "delivery_surcharge_case",
    "vegcool_surcharge_case",
    "lot_discount_case",
    "volume_minimum_cases",
    "transit_days",
    "currency_code",
    "price_basis",
)

# The identity octuple cell key — what intake fans out PER, and what the engine collapses BACK to.
_CELL_KEY_COLUMNS = ("supplier_id", "dc_id", "lot_id", "item_id", "tf_id")


def _potato_price_pool() -> list[tuple[float, float]]:
    """Real (all-in, FOB) pairs from the potato sample; empty when that file is absent (CI)."""

    if not _POTATO_SAMPLE.is_file():
        return []
    wb = load_workbook(_POTATO_SAMPLE, read_only=True, data_only=True)
    pool: list[tuple[float, float]] = []
    for row in wb["IN_Bids"].iter_rows(min_row=5, values_only=True):
        if not row or row[0] is None:
            continue
        all_in, fob = row[7], row[8]
        if isinstance(all_in, int | float) and isinstance(fob, int | float):
            pool.append((float(all_in), float(fob)))
        if len(pool) >= 64:
            break
    return pool


def _header_map(ws: Worksheet) -> dict[str, int]:
    return {
        str(ws.cell(row=HEADER_ROW, column=c).value).strip(): c
        for c in range(1, ws.max_column + 1)
        if ws.cell(row=HEADER_ROW, column=c).value is not None
    }


def _fill_template_full_columns(template_bytes: bytes, pool: list[tuple[float, float]]) -> bytes:
    """Fill every scope row across the full column set, ALTERNATING the two valid price bases.

    A bid is either All-In basis OR component basis (FOB + surcharges) — never both on one row. To
    exercise the WHOLE column set with valid bids: even rows take the potato All-In/FOB (+ transit),
    odd rows take FOB + the three surcharge/discount components. Volume is set on every row.
    """

    wb = load_workbook(BytesIO(template_bytes))
    ws = wb[SHEET_BIDS]
    headers = _header_map(ws)

    def put(row: int, bid_col: BidColumn, value: object) -> None:
        c = headers.get(bid_col.value)
        if c is not None:
            ws.cell(row=row, column=c, value=value)

    fallback = [(18.40, 17.40), (24.75, 23.75), (31.41, 30.41), (26.50, 25.50)]
    prices = pool or fallback
    idx = 0
    for row in range(BODY_START_ROW, ws.max_row + 1):
        sup = str(ws.cell(row=row, column=headers[BidColumn.SUPPLIER.value]).value or "").strip()
        lot = str(ws.cell(row=row, column=headers[BidColumn.LOT.value]).value or "").strip()
        if not sup or not lot:
            continue
        all_in, fob = prices[idx % len(prices)]
        if idx % 2 == 0:  # All-In basis (real potato price) + transit
            put(row, BidColumn.ALL_IN, all_in)
            put(row, BidColumn.FOB, fob)
            put(row, BidColumn.TRANSIT_DAYS, 3)
        else:  # component basis: FOB + surcharges/discount (no All-In)
            put(row, BidColumn.FOB, fob)
            put(row, BidColumn.DELIVERY_SURCHARGE, 0.85)
            put(row, BidColumn.VEGCOOL_SURCHARGE, 0.40)
            put(row, BidColumn.LOT_DISCOUNT, 0.25)
        put(row, BidColumn.WEEKLY_VOL_OFFERED, 600)
        put(row, BidColumn.TOTAL_VOL_OFFERED, 7800)
        idx += 1

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _fill_all_in(template_bytes: bytes, all_in: float) -> bytes:
    """Fill EVERY scope row with one All-In price — a clean, unique price marker per submission.

    Used by the supersession regression test: a constant All-In price on every cell means the whole
    submission is identifiable by one number, so a stale (superseded) price leaking into any read is
    detectable. FOB is set just below All-In (a valid All-In-basis bid also records FOB).
    """

    wb = load_workbook(BytesIO(template_bytes))
    ws = wb[SHEET_BIDS]
    headers = _header_map(ws)
    for row in range(BODY_START_ROW, ws.max_row + 1):
        sup = str(ws.cell(row=row, column=headers[BidColumn.SUPPLIER.value]).value or "").strip()
        lot = str(ws.cell(row=row, column=headers[BidColumn.LOT.value]).value or "").strip()
        if not sup or not lot:
            continue
        ws.cell(row=row, column=headers[BidColumn.ALL_IN.value], value=all_in)
        ws.cell(row=row, column=headers[BidColumn.FOB.value], value=round(all_in - 1.0, 2))
        ws.cell(row=row, column=headers[BidColumn.TRANSIT_DAYS.value], value=3)
        ws.cell(row=row, column=headers[BidColumn.WEEKLY_VOL_OFFERED.value], value=600)
        ws.cell(row=row, column=headers[BidColumn.TOTAL_VOL_OFFERED.value], value=7800)
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _expected_span(db_session, cycle_id: str) -> dict[str, list[int]]:  # type: ignore[no-untyped-def]
    """Per timeframe, the fiscal-period numbers its stored dates cover (via `period_for_date`)."""

    span_by_tf: dict[str, list[int]] = {}
    for tf_id, start, end in db_session.execute(
        text("SELECT tf_id, start_date, end_date FROM cyc.cycle_timeframe WHERE cycle_id = :c"),
        {"c": cycle_id},
    ).all():
        first, last = period_for_date(start), period_for_date(end)
        assert first.fiscal_year == last.fiscal_year == 2026
        span_by_tf[tf_id] = list(range(first.period, last.period + 1))
    return span_by_tf


@pytest.mark.integration
def test_intake_stores_bids_flat_at_fiscal_periods(tmp_path: Path, db_session) -> None:  # type: ignore[no-untyped-def]
    """Intake fans each priced cell out to one bid.bid_line PER period in its tf's span (§1a)."""

    pool = _potato_price_pool()
    service = PilotService(tmp_path, isolate_db=False)
    paths = service.start_run(commodity="Colored Potatoes", label="Period Import")
    setup_path = paths.inputs / stage_filename(1, "setup_kickoff")
    setup_path.write_bytes(_build_filled_setup())
    cycle_id = service.ingest_setup(db_session, paths, setup_path)

    template_path = service.generate_bid_template(db_session, paths, 1)
    template_path.write_bytes(_fill_template_full_columns(template_path.read_bytes(), pool))

    # (c) `ingest_bids` returns the LOGICAL line count (one per identity cell), NOT the fanned rows.
    n_lines = service.ingest_bids(db_session, paths, 1, template_path)
    assert n_lines > 0

    span_by_tf = _expected_span(db_session, cycle_id)
    # The synthetic "Spring 2026" timeframe must cover SEVERAL periods for a real fan-out.
    assert len(span_by_tf) == 1, "this fixture has exactly one timeframe"
    span = next(iter(span_by_tf.values()))
    n_periods = len(span)
    assert n_periods > 1, "the timeframe must cover several periods for a real fan-out"

    all_rows = db_session.query(BidLine).filter(BidLine.cycle_id == cycle_id).all()

    # (a) intake stored bids at the PERIOD grain — every row carries a fiscal_period_id, none NULL.
    assert all_rows, "bids were persisted"
    assert all(r.fiscal_period_id is not None for r in all_rows), "every row is period-grain"
    # one row per (logical cell × period in the span): n_lines logical cells × n_periods.
    assert len(all_rows) == n_lines * n_periods

    # the LOGICAL line count (distinct identity cells) is exactly `ingested` — fan-out is hidden.
    distinct_cells = {tuple(getattr(r, c) for c in _CELL_KEY_COLUMNS) for r in all_rows}
    assert len(distinct_cells) == n_lines

    # (b) every fiscal period in the span is covered exactly once per cell, each linked to a real
    # ref.fiscal_period; the whole span and nothing outside it.
    linked = db_session.execute(
        text(
            "SELECT fp.fiscal_year, fp.period, count(*) "
            "FROM bid.bid_line bl JOIN ref.fiscal_period fp ON bl.fiscal_period_id = fp.id::text "
            "WHERE bl.cycle_id = :c GROUP BY fp.fiscal_year, fp.period ORDER BY fp.period"
        ),
        {"c": cycle_id},
    ).all()
    assert [(fy, p) for fy, p, _ in linked] == [(2026, p) for p in span]
    assert all(cnt == n_lines for _, _, cnt in linked)  # each period got every cell's price

    # every COLUMN survives the fan-out verbatim — every period row of a cell is identical.
    by_cell: dict[tuple[str, ...], list[BidLine]] = {}
    for r in all_rows:
        by_cell.setdefault(tuple(getattr(r, c) for c in _CELL_KEY_COLUMNS), []).append(r)
    for rows in by_cell.values():
        assert len(rows) == n_periods
        head, *tail = rows
        for r in tail:
            for column in _CARRIED_COLUMNS:
                assert getattr(r, column) == getattr(head, column), f"{column} not preserved"

    # the full column set was genuinely exercised: some cells priced All-In (a Decimal), others
    # carried the surcharge components — both survived into the by-period rows.
    all_in_vals = [r.submitted_all_in_case for r in all_rows if r.submitted_all_in_case is not None]
    surcharge_vals = [r.delivery_surcharge_case for r in all_rows if r.delivery_surcharge_case]
    assert all_in_vals and all(isinstance(v, Decimal) for v in all_in_vals)
    assert surcharge_vals, "the surcharge/component columns were exercised and carried through"


@pytest.mark.integration
def test_period_grain_storage_leaves_engine_output_unchanged(tmp_path: Path, db_session) -> None:  # type: ignore[no-untyped-def]
    """The HARD invariant: the engine sees the SAME inputs whether bids are stored per-period or tf.

    The runner's representative-row collapse must reduce the period-grain rows to exactly one bid
    per (dc, lot, tf, supplier) — identical to the pre-fan-out timeframe grain. We prove it by
    running the input assembly on the REAL period-grain rows and on a synthetic tf-grain control
    from the same logical cells, and asserting the resulting `BidInput`s are identical. We then run
    the full engine and confirm it scored exactly the LOGICAL cell count (no doubling) and sealed
    the seven scenarios.
    """

    service = PilotService(tmp_path, isolate_db=False)
    paths = service.start_run(commodity="Colored Potatoes", label="Period Invariant")
    setup_path = paths.inputs / stage_filename(1, "setup_kickoff")
    setup_path.write_bytes(_build_filled_setup())
    cycle_id = service.ingest_setup(db_session, paths, setup_path)
    cycle = service._load_cycle(db_session, paths)
    round_id = cycle.rounds[0].id

    template_path = service.generate_bid_template(db_session, paths, 1)
    template_path.write_bytes(_fill_template_full_columns(template_path.read_bytes(), []))
    n_lines = service.ingest_bids(db_session, paths, 1, template_path)

    runner = EngineRunner(db_session)
    tf_code_by_id, _ = runner._timeframe_maps(cycle_id)
    lot_by_item = runner._lot_by_item(cycle_id)

    # The REAL period-grain rows the runner reads, then collapses.
    period_rows = runner._read_bid_lines(cycle_id, round_id)
    assert len(period_rows) > n_lines, "storage is fanned out to >1 row per cell"

    # A tf-grain CONTROL: one representative row per logical cell (what storage looked like before).
    representative = runner._representative_lines(period_rows)
    assert len(representative) == n_lines, "collapse yields exactly one row per logical cell"

    period_bids = runner._assemble_bids(period_rows, tf_code_by_id, lot_by_item)
    control_bids = runner._assemble_bids(representative, tf_code_by_id, lot_by_item)

    def _engine_facing(bids):  # type: ignore[no-untyped-def]
        # Everything the engine scores on EXCEPT the surrogate bid_id (a fresh uuid per row): the
        # cell key, the landed cost, eligibility, components, volume. These must be identical.
        return sorted(
            (
                b.supplier_id,
                b.dc_no,
                b.lot_id,
                b.tf_code,
                str(b.landed_cost_per_case),
                b.eligible,
                b.is_incumbent,
                str(b.components.all_in),
                str(b.components.fob),
                str(b.total_vol_offered),
            )
            for b in bids
        )

    assert _engine_facing(period_bids) == _engine_facing(control_bids), (
        "the representative-row collapse must yield the SAME engine inputs as the tf grain"
    )
    assert len(period_bids) == n_lines, "the engine sees one bid per cell, not the fanned rows"

    # The collapse is DETERMINISTIC (stable input-hash manifest): re-collapsing yields the same ids.
    again = runner._representative_lines(period_rows)
    assert [r.bid_line_id for r in representative] == [r.bid_line_id for r in again]

    # End to end: the sealed run scores exactly the LOGICAL cells (no doubling) + seals 7 scenarios.
    alignment_path = service.run_round(db_session, paths, 1)
    n_scores = db_session.execute(
        text(
            "SELECT count(*) FROM eng.bid_score s JOIN eng.analysis_run r "
            "ON r.analysis_run_id = s.analysis_run_id WHERE r.cycle_id = :c"
        ),
        {"c": cycle_id},
    ).scalar_one()
    assert n_scores == n_lines, "the engine scored the logical cells, never the fanned storage rows"
    n_scenarios = db_session.execute(
        text(
            "SELECT count(*) FROM eng.analysis_scenario s JOIN eng.analysis_run r "
            "ON r.analysis_run_id = s.analysis_run_id WHERE r.cycle_id = :c"
        ),
        {"c": cycle_id},
    ).scalar_one()
    assert n_scenarios == 7  # the A–G lenses

    # The ALIGNMENT WORKBOOK is timeframe-grain too: its per-(supplier × cell) tabs must carry one
    # row per LOGICAL cell — NOT one per fanned period row. A dedupe miss in those gathers (the
    # Detailed Scoring market stats, the Coverage rows) would surface here as `× n_periods` rows.
    #   * Detailed Scoring lists EVERY scored cell -> exactly n_lines rows.
    #   * Coverage lists cells with a CONSTRUCTIBLE price (All-In or component FOB) -> n_lines.
    # The `# Bidders` stat on Detailed Scoring is the per-group count that would inflate by periods
    # if the stats gather double-counted the period rows; assert it stays the true bidder count (2).
    wb = load_workbook(alignment_path)

    def _body_rows(tab: str) -> list[int]:
        ws = wb[tab]
        # The header row is the one whose column 4 == "Supplier"; data rows follow, Supplier-filled.
        header = next(
            r for r in range(1, ws.max_row + 1) if ws.cell(row=r, column=4).value == "Supplier"
        )
        return [
            r
            for r in range(header + 1, ws.max_row + 1)
            if ws.cell(row=r, column=4).value not in (None, "")
        ]

    detail_rows = _body_rows("Detailed Scoring")
    assert len(detail_rows) == n_lines, (
        f"Detailed Scoring has {len(detail_rows)} rows, expected {n_lines} (per cell, not fanned)"
    )
    coverage_rows = _body_rows("Coverage")
    # #2 regression: Coverage now lists EVERY priced cell — including component-basis (FOB-only)
    # bids, which were previously DROPPED because the gather read raw `submitted_all_in_case`.
    # With the canonical constructed price (E-39), the ~half of cells filled component-basis now
    # appear, so the count is the full logical-cell count (never the fanned period rows).
    assert len(coverage_rows) == n_lines, (
        f"Coverage has {len(coverage_rows)} rows, expected {n_lines} (every priced cell, incl. "
        "component-basis FOB-only bids; not fanned)"
    )
    # # Bidders (col 10) is the per-group count of distinct All-In prices — never inflated by the
    # fan-out. There are 2 suppliers per cell, so the true count is ≤ 2; a stats-dedupe miss would
    # multiply it by n_periods (≥ 4). Assert it stays at the real grain (≤ the supplier count).
    ds = wb["Detailed Scoring"]
    bidder_counts = [ds.cell(row=r, column=10).value for r in detail_rows]
    fanout_factor = len(period_rows) // n_lines  # ≥ 2 (the span has several periods)
    assert max(bidder_counts) <= 2, (
        f"# Bidders must be <= 2 (the true supplier count), got {max(bidder_counts)}; "
        f"a value >= {fanout_factor} would mean the period rows were double-counted"
    )


@pytest.mark.integration
def test_unmappable_timeframe_falls_back_to_tf_grain(tmp_path: Path, db_session) -> None:  # type: ignore[no-untyped-def]
    """A timeframe whose dates fall OUTSIDE the seeded calendar ingests at tf-grain (NULL period).

    Graceful fallback: the seeded fiscal calendar is FY16–FY36. A cycle whose timeframe dates land
    outside that window can't be resolved to fiscal periods, so intake writes the single tf-grain
    row (`fiscal_period_id` NULL) exactly as before — nothing breaks, the engine path is unchanged.
    """

    service = PilotService(tmp_path, isolate_db=False)
    paths = service.start_run(commodity="Field Tomatoes", label="OOR Fallback")
    setup_path = paths.inputs / stage_filename(1, "setup_kickoff")
    setup_path.write_bytes(_build_filled_setup())
    cycle_id = service.ingest_setup(db_session, paths, setup_path)

    # Shove the cycle's timeframe dates WAY outside the seeded FY16–36 calendar (year 2099).
    db_session.execute(
        text(
            "UPDATE cyc.cycle_timeframe SET start_date = DATE '2099-01-01', "
            "end_date = DATE '2099-03-31' WHERE cycle_id = :c"
        ),
        {"c": cycle_id},
    )
    db_session.flush()

    template_path = service.generate_bid_template(db_session, paths, 1)
    template_path.write_bytes(_fill_template_full_columns(template_path.read_bytes(), []))
    n_lines = service.ingest_bids(db_session, paths, 1, template_path)
    assert n_lines > 0

    rows = db_session.query(BidLine).filter(BidLine.cycle_id == cycle_id).all()
    # Fallback: one tf-grain row per logical cell, ALL with NULL fiscal_period_id (no fan-out).
    assert len(rows) == n_lines
    assert all(r.fiscal_period_id is None for r in rows), "unmappable tf stays at tf-grain (NULL)"


@pytest.mark.integration
def test_resubmission_supersedes_prior_period_rows_in_every_read(
    tmp_path: Path, db_session
) -> None:  # type: ignore[no-untyped-def]
    """A re-submitted round file supersedes prior period rows; NO deduped read shows a stale price.

    Under Option B a re-send supersedes the prior submission — its (fanned) period rows are flipped
    `is_scoreable=false`, never hard-deleted (ADR-0006). The Option-B dedupe reads collapse the
    period grain with `DISTINCT ON (supplier, dc, lot, tf)`, so without an active-row filter they
    could pick a SUPERSEDED period row as a cell's representative and surface a stale price. Each
    deduped read must therefore filter to ACTIVE rows, like the engine's `_read_bid_lines` does.

    The proof is leak-detection by construction: superseded rows never enter legitimate engine math
    (the engine reads active rows only), so the OLD price can appear in the sealed scores or the
    alignment workbook ONLY if a read pulled a superseded row — the bug. We assert it never does,
    and that the NEW (active) price is present. Distinct, price-implausible markers (far above score
    or percentage) make the scan unambiguous. Covers the engine input, all three workbook gathers
    (price grid, Detailed Scoring stats, Coverage) end to end via `run_round`, and the `run_data`
    count.
    """

    old_price, new_price = 1311.07, 8742.93
    service = PilotService(tmp_path, isolate_db=False)
    paths = service.start_run(commodity="Colored Potatoes", label="Resubmit Supersede")
    setup_path = paths.inputs / stage_filename(1, "setup_kickoff")
    setup_path.write_bytes(_build_filled_setup())
    cycle_id = service.ingest_setup(db_session, paths, setup_path)
    cycle = service._load_cycle(db_session, paths)
    round_id = cycle.rounds[0].id

    tpl_bytes = service.generate_bid_template(db_session, paths, 1).read_bytes()

    # Submission 1: the OLD price on every cell.
    first = paths.inputs / "round1_first.xlsx"
    first.write_bytes(_fill_all_in(tpl_bytes, old_price))
    n_lines = service.ingest_bids(db_session, paths, 1, first)
    assert n_lines > 0

    # Submission 2 (same round): the NEW price — supersedes submission 1's rows.
    second = paths.inputs / "round1_second.xlsx"
    second.write_bytes(_fill_all_in(tpl_bytes, new_price))
    n_lines2 = service.ingest_bids(db_session, paths, 1, second)
    assert n_lines2 == n_lines, "the second submission covers the same logical cells"

    def _count(price: float, scoreable: bool) -> int:
        return int(
            db_session.execute(
                text(
                    "SELECT count(*) FROM bid.bid_line WHERE cycle_id = :c "
                    "AND submitted_all_in_case = :p AND is_scoreable = :s"
                ),
                {"c": cycle_id, "p": price, "s": scoreable},
            ).scalar_one()
        )

    # Supersede, never delete: the OLD rows are RETAINED but non-scoreable; the NEW rows are active.
    assert _count(old_price, scoreable=True) == 0, "no superseded (OLD) row stays active"
    assert _count(old_price, scoreable=False) > 0, "prior submission's rows retained (superseded)"
    assert _count(new_price, scoreable=True) > 0, "the current submission's rows are active"

    # (1) ENGINE input: `_read_bid_lines` returns ONLY active rows -> only the NEW price.
    runner = EngineRunner(db_session)
    engine_rows = runner._read_bid_lines(cycle_id, round_id)
    assert engine_rows
    assert all(r.submitted_all_in_case == Decimal(str(new_price)) for r in engine_rows), (
        "the engine scores only the active (current) submission"
    )

    # (2) WORKBOOK end to end: run the round, then scan every tab. The superseded OLD price can only
    # appear if a workbook gather pulled a non-scoreable row; the NEW price must be present.
    alignment_path = service.run_round(db_session, paths, 1)
    wb = load_workbook(alignment_path)
    saw_new = False
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if isinstance(v, int | float) and not isinstance(v, bool):
                    assert abs(float(v) - old_price) > 0.01, (
                        f"superseded OLD price {old_price} leaked into tab '{ws.title}'"
                    )
                    if abs(float(v) - new_price) <= 0.01:
                        saw_new = True
    assert saw_new, "the alignment workbook renders the active NEW price"

    # (3) run_data snapshot: the per-round count stays the LOGICAL active cell count (not doubled).
    run_data_path = service.export_run_data(db_session, paths)
    snapshot = json.loads(run_data_path.read_text(encoding="utf-8"))
    by_round = {int(r["round"]): int(r["bid_lines"]) for r in snapshot["bid_lines_by_round"]}
    assert by_round.get(1) == n_lines, "run_data counts the active submission only"
