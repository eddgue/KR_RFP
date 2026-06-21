"""E-38 capacity round-trip: generate a template -> fill the Capacity sheet -> ingest by KEY.

The Capacity sheet follows the same D21 discipline as the Bids sheet: it embeds system-owned KEY
IDs (cycle/supplier/dc/lot/item/tf), and `ingest_capacity` VALIDATES those embedded keys against the
cycle scope — it never name-resolves. This proves the contract end to end (pure, no DB):
  * a generated template carries the capacity key + display + entry columns;
  * filled cells ingest back to the SAME identity, with the stated ceilings preserved;
  * a blank max cell is NO statement (not a zero); and
  * a tampered / missing / negative cell quarantines, never silently loads.
"""

from __future__ import annotations

from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.domain.bid.bid_ingester import QuarantineReason, ingest_capacity
from app.domain.bid.template_generator import generate_template_bytes
from app.domain.bid.template_schema import (
    CAPACITY_HEADERS,
    CAPACITY_KEY_ID_COLUMNS,
    HEADER_ROW,
    SHEET_CAPACITY,
    CapacityColumn,
)
from tests.bid.synthetic import build_scope

_D = Decimal


def _capacity_sheet(template_bytes: bytes) -> tuple[object, Worksheet, dict[str, int]]:
    wb = load_workbook(BytesIO(template_bytes))
    ws = wb[SHEET_CAPACITY]
    headers = {
        str(ws.cell(row=HEADER_ROW, column=c).value).strip(): c
        for c in range(1, ws.max_column + 1)
        if ws.cell(row=HEADER_ROW, column=c).value is not None
    }
    return wb, ws, headers


def _save(wb: object) -> bytes:
    buffer = BytesIO()
    wb.save(buffer)  # type: ignore[attr-defined]
    return buffer.getvalue()


def test_generated_capacity_sheet_has_keys_and_entry_columns() -> None:
    scope = build_scope()
    _wb, ws, headers = _capacity_sheet(generate_template_bytes(scope))

    emitted = [
        str(ws.cell(row=HEADER_ROW, column=c).value) for c in range(1, len(CAPACITY_HEADERS) + 1)
    ]
    assert emitted == list(CAPACITY_HEADERS)
    # Every key-ID column is present (the validated join identity, D21).
    for key_col in CAPACITY_KEY_ID_COLUMNS:
        assert key_col.value in headers
    assert CapacityColumn.MAX_WEEKLY_CASES.value in headers
    assert CapacityColumn.MAX_TOTAL_CASES.value in headers
    # One capacity body row per distinct supplier x dc x lot x item x tf cell (round-independent).
    distinct = {(r.supplier_id, r.dc_id, r.lot_id, r.item_id, r.tf_id) for r in scope.rows}
    assert ws.max_row - HEADER_ROW == len(distinct)


def test_capacity_round_trips_by_key_blank_is_not_zero() -> None:
    scope = build_scope()
    wb, ws, headers = _capacity_sheet(generate_template_bytes(scope))
    first, second = HEADER_ROW + 1, HEADER_ROW + 2

    # Capture the embedded identity of the two rows we fill (so we can assert it round-trips).
    def keys_of(row: int) -> tuple[str, str, str, str, str, str]:
        return tuple(  # type: ignore[return-value]
            str(ws.cell(row=row, column=headers[c.value]).value) for c in CAPACITY_KEY_ID_COLUMNS
        )

    first_keys, second_keys = keys_of(first), keys_of(second)

    ws.cell(row=first, column=headers[CapacityColumn.MAX_WEEKLY_CASES.value], value=500)
    ws.cell(row=first, column=headers[CapacityColumn.MAX_TOTAL_CASES.value], value=6500)
    ws.cell(row=first, column=headers[CapacityColumn.CAPACITY_NOTES.value], value="firm")
    # Second row: only a total ceiling (weekly left blank — a partial but valid statement).
    ws.cell(row=second, column=headers[CapacityColumn.MAX_TOTAL_CASES.value], value=7000)
    # The remaining rows are left entirely blank — NO capacity statement (must not become a line).

    result = ingest_capacity(_save(wb), scope)

    assert result.quarantined == []
    assert len(result.lines) == 2  # only the two filled cells; the blanks are NOT zero ceilings
    by_row = {line.source_row_number: line for line in result.lines}

    line1 = by_row[first]
    assert (
        line1.cycle_id,
        line1.supplier_id,
        line1.dc_id,
        line1.lot_id,
        line1.item_id,
        line1.tf_id,
    ) == first_keys
    assert line1.cycle_id == scope.cycle_id
    assert (
        line1.cycle_id,
        line1.supplier_id,
        line1.dc_id,
        line1.lot_id,
        line1.item_id,
        line1.tf_id,
    ) in scope.capacity_key_set()
    assert line1.max_weekly_cases == _D("500")
    assert line1.max_period_cases == _D("6500")
    assert line1.notes == "firm"

    line2 = by_row[second]
    assert line2.max_weekly_cases is None  # weekly blank -> no weekly ceiling (not zero)
    assert line2.max_period_cases == _D("7000")
    assert (line2.supplier_id, line2.dc_id, line2.lot_id) == (
        second_keys[1],
        second_keys[2],
        second_keys[3],
    )


def test_tampered_capacity_key_quarantines_not_resolved() -> None:
    scope = build_scope()
    wb, ws, headers = _capacity_sheet(generate_template_bytes(scope))
    row = HEADER_ROW + 1
    ws.cell(row=row, column=headers[CapacityColumn.MAX_TOTAL_CASES.value], value=6500)
    # Tamper ONLY the embedded DC_ID (the display DC Name stays valid + in-scope).
    ws.cell(row=row, column=headers[CapacityColumn.DC_ID.value], value="DC-TAMPERED")

    result = ingest_capacity(_save(wb), scope)

    assert result.lines == []
    assert len(result.quarantined) == 1
    assert result.quarantined[0].reason is QuarantineReason.UNKNOWN_KEY


def test_blank_capacity_key_quarantines_missing_key() -> None:
    scope = build_scope()
    wb, ws, headers = _capacity_sheet(generate_template_bytes(scope))
    row = HEADER_ROW + 1
    ws.cell(row=row, column=headers[CapacityColumn.MAX_TOTAL_CASES.value], value=6500)
    ws.cell(row=row, column=headers[CapacityColumn.SUPPLIER_ID.value], value="")

    result = ingest_capacity(_save(wb), scope)

    assert result.lines == []
    assert len(result.quarantined) == 1
    assert result.quarantined[0].reason is QuarantineReason.MISSING_KEY


def test_negative_capacity_quarantines_bad_numeric() -> None:
    scope = build_scope()
    wb, ws, headers = _capacity_sheet(generate_template_bytes(scope))
    row = HEADER_ROW + 1
    ws.cell(row=row, column=headers[CapacityColumn.MAX_WEEKLY_CASES.value], value=-5)

    result = ingest_capacity(_save(wb), scope)

    assert result.lines == []
    assert len(result.quarantined) == 1
    assert result.quarantined[0].reason is QuarantineReason.BAD_NUMERIC
