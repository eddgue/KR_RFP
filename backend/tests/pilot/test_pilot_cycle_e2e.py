"""Pilot core (PART B) — the WHOLE cycle loop, end to end, on a real Postgres (PLAN §7).

Drives the full decision-support loop through `PilotService` with SYNTHETIC data:
start_run → ingest a synthetic setup (2 DCs, 2 lots, 2 suppliers, 1 TF, 2 rounds) →
generate_bid_template(1) → fill it in-memory + upload → ingest_bids(1) → run_round(1) →
freeze_award → record_adjustment → history → close_run. It asserts the NORMALIZED output
filenames and the in-file VERSION HEADINGS (Analysis v1 / Version 1) the docs carry, and that the
close-out zip holds the full normalized history (inputs/, outputs/, NOTES.md, memory/).

Plus a PURE unit test for `flex_ingest.infer_bid_mapping` on a synthetic "messy" sheet (different
header words + a shuffled column order) → the correct column → field mapping.

Synthetic only (clean-room, ADR-0001). DB-touching parts are marked integration (skip when no DB).
"""

from __future__ import annotations

import json
import zipfile
from datetime import date
from decimal import Decimal
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.cycle.loader import load_cycle
from app.cycle.scope import build_scope_from_cycle
from app.domain.bid.template_schema import BODY_START_ROW, HEADER_ROW, SHEET_BIDS, BidColumn
from app.output.types import CycleView, Entity
from app.pilot.flex_ingest import (
    FIELD_ALL_IN,
    FIELD_DC,
    FIELD_LOT,
    FIELD_SUPPLIER,
    FIELD_VOLUME,
    infer_bid_mapping,
)
from app.pilot.service import PilotService
from app.pilot.setup_template import (
    EXAMPLE_START_ROW,
    HEADER_ROW as SETUP_HEADER_ROW,
    TAB_CYCLE,
    TAB_DCS,
    TAB_INCUMBENTS,
    TAB_LOTS,
    TAB_SUPPLIERS,
    TAB_TIMEFRAMES,
    TAB_VOLUMES,
    build_setup_workbook,
)
from app.pilot.vault import stage_filename


# ---------------------------------------------------------------------------
# helpers — build a synthetic filled SETUP workbook in-memory
# ---------------------------------------------------------------------------
def _setup_header_col(ws: Worksheet, header: str) -> int:
    for col in range(1, ws.max_column + 1):
        if str(ws.cell(row=SETUP_HEADER_ROW, column=col).value or "").strip() == header:
            return col
    raise AssertionError(f"header {header!r} not found on tab {ws.title!r}")


def _write_setup_rows(ws: Worksheet, rows: list[dict[str, object]]) -> None:
    for ri, row in enumerate(rows):
        excel_row = EXAMPLE_START_ROW + ri
        for header, value in row.items():
            ws.cell(row=excel_row, column=_setup_header_col(ws, header), value=value)
    for extra in range(len(rows), len(rows) + 3):
        excel_row = EXAMPLE_START_ROW + extra
        for col in range(1, ws.max_column + 1):
            ws.cell(row=excel_row, column=col, value=None)


def _build_filled_setup() -> bytes:
    """A synthetic setup: 2 DCs, 2 lots, 2 suppliers, 1 TF, 2 rounds, volumes, incumbents."""

    wb = load_workbook(BytesIO(build_setup_workbook()))
    _write_setup_rows(
        wb[TAB_CYCLE],
        [
            {
                "Cycle Label": "E2E Tomatoes Cycle",
                "Commodity": "Field Tomatoes",
                "Sub-commodity": "Round/Vine",
                "Horizon (weeks)": 13,
                "Rounds": 2,
                "Target Effective Date": "2026-12-31",
                "Weight Preset": "balanced",
                "Max Suppliers / DC": 2,
                "Premium Ceiling": 0.12,
                "Concentration Threshold": 0.40,
                "Coverage Floor": 0.80,
            }
        ],
    )
    _write_setup_rows(
        wb[TAB_DCS],
        [
            {"DC Name": "Atlanta DC", "Region": "East", "State": "GA"},
            {"DC Name": "Dallas DC", "Region": "South", "State": "TX"},
        ],
    )
    _write_setup_rows(
        wb[TAB_LOTS],
        [
            {
                "Lot Name": "Lot 1 - Grape",
                "Item Description": "Premium Grape Tomato 10oz",
                "Pack Size / UOM": "10oz clamshell",
                "Product Type": "Conventional",
                "Category": "Tomatoes",
            },
            {
                "Lot Name": "Lot 2 - Roma",
                "Item Description": "Roma Tomato Bulk 25lb",
                "Pack Size / UOM": "25lb carton",
                "Product Type": "Organic",
                "Category": "Tomatoes",
            },
        ],
    )
    _write_setup_rows(
        wb[TAB_SUPPLIERS],
        [
            {"Supplier Name": "Green Valley Farms", "Region / Origin": "East", "Notes": "inc"},
            {"Supplier Name": "Sunbelt Produce", "Region / Origin": "South", "Notes": "new"},
        ],
    )
    _write_setup_rows(
        wb[TAB_TIMEFRAMES],
        [
            {
                "Timeframe Label": "Spring 2026",
                "Start Date": "2026-04-01",
                "End Date": "2026-06-30",
                "Week Count": 13,
            }
        ],
    )
    vol_rows: list[dict[str, object]] = [
        {
            "DC Name": dc,
            "Lot Name": lot,
            "Timeframe": "Spring 2026",
            "Method": "WEEKLY_X_WEEKS",
            "Weekly Cases": 400,
            "Weeks": 13,
        }
        for dc in ("Atlanta DC", "Dallas DC")
        for lot in ("Lot 1 - Grape", "Lot 2 - Roma")
    ]
    _write_setup_rows(wb[TAB_VOLUMES], vol_rows)
    inc_rows: list[dict[str, object]] = [
        {
            "DC Name": dc,
            "Lot Name": lot,
            "Incumbent Supplier": "Green Valley Farms",
            "Routing Baseline $/case": 13.50,
            "Contract Notes": "auto",
        }
        for dc in ("Atlanta DC", "Dallas DC")
        for lot in ("Lot 1 - Grape", "Lot 2 - Roma")
    ]
    _write_setup_rows(wb[TAB_INCUMBENTS], inc_rows)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# helpers — fill the generated OWNED bid template in-memory with synthetic bids
# ---------------------------------------------------------------------------
def _header_map(ws: Worksheet) -> dict[str, int]:
    out: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=HEADER_ROW, column=col).value
        if value is not None:
            out[str(value).strip()] = col
    return out


def _fill_bid_template(template_bytes: bytes) -> bytes:
    """Write synthetic All-In + volume into every scope row so the engine has priced bids.

    Two suppliers, varied so Scenario B has a real split. Green Valley (incumbent) is keener on Lot
    1; Sunbelt is keener on Lot 2 — so each DC's two lots split across two suppliers.
    """

    wb = load_workbook(BytesIO(template_bytes))
    ws = wb[SHEET_BIDS]
    headers = _header_map(ws)
    all_in_col = headers[BidColumn.ALL_IN.value]
    fob_col = headers[BidColumn.FOB.value]
    weekly_col = headers[BidColumn.WEEKLY_VOL_OFFERED.value]
    total_col = headers[BidColumn.TOTAL_VOL_OFFERED.value]
    sup_col = headers[BidColumn.SUPPLIER.value]
    lot_col = headers[BidColumn.LOT.value]

    for row in range(BODY_START_ROW, ws.max_row + 1):
        sup = str(ws.cell(row=row, column=sup_col).value or "").strip()
        lot = str(ws.cell(row=row, column=lot_col).value or "").strip()
        if not sup or not lot:
            continue
        base = Decimal("12.00")
        # Each supplier is keenest on a different lot (a real 2-supplier DC split).
        if "Grape" in lot:
            price = base - (Decimal("1.50") if "Green Valley" in sup else Decimal("0.20"))
        else:
            price = base - (Decimal("1.50") if "Sunbelt" in sup else Decimal("0.20"))
        ws.cell(row=row, column=all_in_col, value=float(price))
        ws.cell(row=row, column=fob_col, value=float(price - Decimal("1.00")))
        ws.cell(row=row, column=weekly_col, value=600)
        ws.cell(row=row, column=total_col, value=600 * 13)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# PURE unit test — flexible-ingest mapping inference on a MESSY sheet
# ---------------------------------------------------------------------------
def _fake_cycle_view() -> CycleView:
    """A tiny in-memory CycleView with known names for the inference test (no DB)."""

    return CycleView(
        cycle_id="cyc-1",
        cycle_code="CYC-TEST",
        cycle_name="Flex Test Cycle",
        client_id="",
        commodity_id="comm-1",
        dcs=[Entity("dc-a", "DC01", "Atlanta DC"), Entity("dc-b", "DC02", "Dallas DC")],
        lots=[Entity("lot-1", "LOT-01", "Lot 1 - Grape"), Entity("lot-2", "LOT-02", "Lot 2 - Roma")],
        items=[Entity("i-1", "ITEM-01", "Grape Tomato"), Entity("i-2", "ITEM-02", "Roma Tomato")],
        tfs=[Entity("tf-1", "TF01", "Spring 2026")],
        rounds=[Entity("r-1", "R1", "Round 1")],
        suppliers=[
            Entity("sup-1", "SUP-01", "Green Valley Farms"),
            Entity("sup-2", "SUP-02", "Sunbelt Produce"),
        ],
        incumbent_by_dc_lot={},
        incumbent_routing={},
        period_cases_by_cell={},
    )


def test_infer_bid_mapping_on_messy_sheet() -> None:
    """A supplier's own sheet — odd header words, shuffled order — maps to the right fields."""

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["Some Supplier's Q2 Quote — please ignore this banner row"])  # a title band
    # Headers in DIFFERENT words and a DIFFERENT order than our template:
    #   warehouse=DC · vendor=supplier · product=lot · delivered price=all-in · cases=volume
    ws.append(["Warehouse", "Vendor", "Product", "Delivered Price", "Cases / Week"])
    ws.append(["Atlanta DC", "Green Valley Farms", "Lot 1 - Grape", 11.40, 600])
    ws.append(["Dallas DC", "Sunbelt Produce", "Lot 2 - Roma", 10.95, 550])
    buffer = BytesIO()
    wb.save(buffer)

    proposal = infer_bid_mapping(buffer.getvalue(), _fake_cycle_view())

    assert proposal.header_row == 2
    assert proposal.mappings[FIELD_DC].source_header == "Warehouse"
    assert proposal.mappings[FIELD_SUPPLIER].source_header == "Vendor"
    assert proposal.mappings[FIELD_LOT].source_header == "Product"
    assert proposal.mappings[FIELD_ALL_IN].source_header == "Delivered Price"
    assert proposal.mappings[FIELD_VOLUME].source_header == "Cases / Week"
    # Identity columns were locked by VALUE matches against the cycle's known names (high conf).
    assert proposal.mappings[FIELD_DC].confidence == "high"
    assert proposal.mappings[FIELD_SUPPLIER].confidence == "high"
    assert proposal.is_confident


# ---------------------------------------------------------------------------
# INTEGRATION — the WHOLE loop, end to end
# ---------------------------------------------------------------------------
@pytest.mark.integration
def test_full_cycle_loop_e2e(tmp_path: Path, db_session) -> None:  # type: ignore[no-untyped-def]
    service = PilotService(tmp_path)

    # 0) start the run + ingest the synthetic setup → a governed cycle.
    paths = service.start_run(commodity="Field Tomatoes", label="E2E Loop")
    setup_path = paths.inputs / stage_filename(1, "setup_kickoff")
    setup_path.write_bytes(_build_filled_setup())
    cycle_id = service.ingest_setup(db_session, paths, setup_path)
    assert cycle_id

    view = load_cycle(db_session, cycle_id)
    assert len(view.dcs) == 2 and len(view.lots) == 2 and len(view.suppliers) == 2
    assert len(view.rounds) == 2

    # 1) generate the Round 1 bid template → inputs/02_round1_bid_template.xlsx
    template_path = service.generate_bid_template(db_session, paths, 1)
    assert template_path.name == "02_round1_bid_template.xlsx"
    assert template_path.is_file()
    assert "upload the round 1 bids" in paths.run_md.read_text().lower()

    # 1b) fill the template synthetically in-memory and "upload" it to inputs/.
    filled = _fill_bid_template(template_path.read_bytes())
    template_path.write_bytes(filled)

    # 1c) ingest the returned bids (key-validated) → bid.bid_line rows.
    n_lines = service.ingest_bids(db_session, paths, 1, template_path)
    # 2 DCs × 2 lots × 1 TF × 2 suppliers = 8 priced cells.
    assert n_lines == 8

    # 2) run the alignment analysis on Round 1 → outputs/04_round1_alignment_v1.xlsx
    alignment_path = service.run_round(db_session, paths, 1)
    assert alignment_path.name == "04_round1_alignment_v1.xlsx"
    assert alignment_path.is_file()
    # The Summary banner carries the mid-cycle "Analysis v1" version heading.
    wb_align = load_workbook(alignment_path)
    summary = wb_align["Summary"]
    banner = "\n".join(
        str(summary.cell(row=r, column=1).value or "") for r in range(1, 6)
    )
    assert "Analysis v1" in banner
    assert "Round 1" in banner

    # 3) freeze the award (human selects Scenario B) → booking guides.
    analysis_run_id = _latest_run_id(db_session, cycle_id)
    award_id = service.freeze_award(
        db_session,
        paths,
        analysis_run_id=analysis_run_id,
        scenario_code="B",
        award_code="AWD-E2E-1",
    )
    assert award_id
    booking_path = paths.outputs / "08_award_booking_guide.xlsx"
    supplier_guides = paths.outputs / "08_award_supplier_guides.xlsx"
    assert booking_path.is_file()
    assert supplier_guides.is_file()

    # 4) record a post-award adjustment (1 cell reprice) → 09_post_award_v1.xlsx
    cell = _first_award_cell(db_session, award_id)
    dc_id, lot_id, tf_id, sup_id, frozen_price = cell
    post_path = service.record_adjustment(
        db_session,
        paths,
        award_id=award_id,
        adjustment_type="NEGOTIATED_REPRICE",
        effective_date=date(2026, 7, 1),
        reason="Synthetic e2e reprice on the first awarded cell",
        line_changes=[(dc_id, lot_id, tf_id, sup_id, frozen_price - Decimal("0.25"))],
    )
    assert post_path.name == "09_post_award_v1.xlsx"
    assert post_path.is_file()
    wb_post = load_workbook(post_path)
    post_banner = "\n".join(
        str(wb_post["Versions"].cell(row=r, column=1).value or "") for r in range(1, 6)
    )
    assert "Version 1" in post_banner

    # 5) history() surfaces the versions: 1 sealed analysis run + the frozen award + its versions.
    hist = service.history(db_session, paths)
    assert len(hist["analysis_runs"]) == 1  # type: ignore[arg-type]
    assert hist["analysis_runs"][0]["version"] == 1  # type: ignore[index]
    assert len(hist["awards"]) == 1  # type: ignore[arg-type]
    award_versions = hist["awards"][0]["versions"]  # type: ignore[index]
    assert {v["version_no"] for v in award_versions} == {0, 1}
    assert "04_round1_alignment_v1.xlsx" in hist["output_files"]  # type: ignore[operator]
    assert "09_post_award_v1.xlsx" in hist["output_files"]  # type: ignore[operator]

    # 5b) the per-run governed-DATA snapshot is written to run_data.json (data-in-git per run),
    # carries names not keys (D23), and reflects the sealed analysis version + the frozen award.
    assert paths.run_data_file.is_file()
    run_data = json.loads(paths.run_data_file.read_text())
    assert run_data["cycle"]["name"] == "E2E Tomatoes Cycle"
    assert "Atlanta DC" in run_data["scope"]["dcs"]  # names, not keys
    assert run_data["bid_lines_by_round"] == [{"round": 1, "bid_lines": 8}]
    assert run_data["analysis_versions"][0]["version"] == 1
    assert run_data["awards"][0]["award_code"] == "AWD-E2E-1"
    award_version_nos = {v["version"] for v in run_data["awards"][0]["versions"]}
    assert award_version_nos == {0, 1}
    # the award lines name their supplier (a known synthetic supplier), never a raw key.
    award_suppliers = {line["supplier"] for line in run_data["awards"][0]["lines"]}
    assert award_suppliers <= {"Green Valley Farms", "Sunbelt Produce"}

    # leave a memory note so the archive's memory/ is exercised.
    service.add_memory(paths, "buyer_note.txt", b"prioritize Dallas", "Buyer ask captured")

    # 6) close_run() → a zip holding the full normalized history.
    zip_path = service.close_run(paths)
    assert zip_path.is_file()
    with zipfile.ZipFile(zip_path) as zf:
        names = zf.namelist()
    slug = paths.slug
    assert any(n.startswith(f"{slug}/inputs/") for n in names)
    assert any(n.startswith(f"{slug}/outputs/") for n in names)
    assert any(n.startswith(f"{slug}/memory/") for n in names)
    assert f"{slug}/NOTES.md" in names
    assert f"{slug}/run_data.json" in names  # the governed-data snapshot rides the archive
    assert any("alignment_v1" in n for n in names)
    assert any("post_award_v1" in n for n in names)

    # purge leaves the archive intact but removes the run folder (records remain in Postgres).
    service.purge_run(slug)
    assert not paths.root.exists()
    assert zip_path.is_file()


@pytest.mark.integration
def test_ingest_any_flexible_roundtrip(tmp_path: Path, db_session) -> None:  # type: ignore[no-untyped-def]
    """ingest_any: propose-then-confirm — a messy file becomes bid_line rows via the strict path."""

    service = PilotService(tmp_path)
    paths = service.start_run(commodity="Field Tomatoes", label="Flex Loop")
    setup_path = paths.inputs / stage_filename(1, "setup_kickoff")
    setup_path.write_bytes(_build_filled_setup())
    cycle_id = service.ingest_setup(db_session, paths, setup_path)
    view = load_cycle(db_session, cycle_id)

    # The buyer drops a supplier's OWN messy sheet (odd headers, no keys, per DC×lot rows).
    messy = _build_messy_supplier_file(view)
    messy_path = paths.inputs / "raw_supplier_drop.xlsx"
    messy_path.write_bytes(messy)

    # confirm=False → a proposal is returned, nothing is ingested yet.
    proposal = service.ingest_any(db_session, paths, 1, messy_path, confirm=False)
    from app.pilot.flex_ingest import MappingProposal  # local import keeps the module surface clear

    assert isinstance(proposal, MappingProposal)
    assert proposal.is_confident

    # confirm=True → normalized file written + ingested via the strict key-validated path.
    count = service.ingest_any(db_session, paths, 1, messy_path, confirm=True)
    assert isinstance(count, int)
    assert count > 0
    assert (paths.inputs / "03_round1_bids_normalized.xlsx").is_file()


def _build_messy_supplier_file(view: CycleView) -> bytes:
    """A messy per-(DC, lot, supplier) sheet with odd headers and a shuffled order (no keys)."""

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["Quarterly Produce Quote (supplier's own format)"])
    ws.append(["Vendor", "Warehouse", "Product Line", "All-In Price", "Weekly Cases"])
    for dc in view.dcs:
        for lot in view.lots:
            for sup in view.suppliers:
                price = 11.50 if "Green Valley" in sup.name else 11.20
                ws.append([sup.name, dc.name, lot.name, price, 600])
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# small DB helpers
# ---------------------------------------------------------------------------
def _latest_run_id(db_session, cycle_id: str) -> str:  # type: ignore[no-untyped-def]
    from sqlalchemy import text

    return db_session.execute(
        text(
            "SELECT analysis_run_id FROM eng.analysis_run WHERE cycle_id = :cyc "
            "ORDER BY run_started_at DESC LIMIT 1"
        ),
        {"cyc": cycle_id},
    ).scalar_one()


def _first_award_cell(db_session, award_id: str):  # type: ignore[no-untyped-def]
    from sqlalchemy import text

    row = db_session.execute(
        text(
            "SELECT dc_id, lot_id, tf_id, supplier_id, frozen_price FROM awd.award_line "
            "WHERE award_id = :aid ORDER BY dc_id, lot_id, supplier_id LIMIT 1"
        ),
        {"aid": award_id},
    ).one()
    return row[0], row[1], row[2], row[3], Decimal(str(row[4]))
