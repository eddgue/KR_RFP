"""Pilot core (PART A) — run vault scaffold, setup template/ingest round-trip, notes/memory, kanban.

Synthetic data only. The DB-touching round-trip (ingest -> load_cycle) runs against the real
Postgres (PLAN §7) and is marked integration (skips when no DB). The pure file/vault assertions
(scaffold identity, normalized filenames, notes/memory, files-only kanban) need no DB.
"""

from __future__ import annotations

from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.pilot.service import PilotService
from app.pilot.setup_template import (
    EXAMPLE_START_ROW,
    HEADER_ROW,
    TAB_CYCLE,
    TAB_DCS,
    TAB_INCUMBENTS,
    TAB_LOTS,
    TAB_SUPPLIERS,
    TAB_TIMEFRAMES,
    TAB_VOLUMES,
    build_setup_workbook,
)
from app.pilot.vault import stage_filename


# ---------------------------------------------------------------------------
# helpers — fill the generated template in-memory with synthetic rows
# ---------------------------------------------------------------------------
def _header_col(ws: Worksheet, header: str) -> int:
    for col in range(1, ws.max_column + 1):
        if str(ws.cell(row=HEADER_ROW, column=col).value or "").strip() == header:
            return col
    raise AssertionError(f"header {header!r} not found on tab {ws.title!r}")


def _write_rows(ws: Worksheet, rows: list[dict[str, object]]) -> None:
    """Overwrite the example rows (and below) with synthetic data — no '(EXAMPLE)' markers."""

    for ri, row in enumerate(rows):
        excel_row = EXAMPLE_START_ROW + ri
        for header, value in row.items():
            ws.cell(row=excel_row, column=_header_col(ws, header), value=value)
    # Clear any leftover greyed example cells below the synthetic rows so they aren't mis-parsed.
    for extra in range(len(rows), len(rows) + 3):
        excel_row = EXAMPLE_START_ROW + extra
        for col in range(1, ws.max_column + 1):
            ws.cell(row=excel_row, column=col, value=None)


def _build_filled_setup() -> bytes:
    """A synthetic filled setup workbook: 2 DCs, 2 lots, 2 suppliers, 1 TF, 2 rounds, vols, incs.

    (The platform requires 2..6 bidding rounds — the cyc.cycle round_count check constraint — so the
    synthetic scope uses the minimum of 2 rounds.)
    """

    wb = load_workbook(BytesIO(build_setup_workbook()))

    _write_rows(
        wb[TAB_CYCLE],
        [
            {
                "Cycle Label": "Test Tomatoes Cycle",
                "Commodity": "Field Tomatoes",
                "Sub-commodity": "Round/Vine",
                "Horizon (weeks)": 13,
                "Rounds": 2,
                "Target Effective Date": "2026-12-31",
                "Weight Preset": "balanced",
                "Max Suppliers / DC": 2,
                "Premium Ceiling": 0.12,
                "Concentration Threshold": 0.40,
                "Coverage Floor": 0.80,
            }
        ],
    )
    _write_rows(
        wb[TAB_DCS],
        [
            {"DC Name": "Atlanta DC", "Region": "East", "State": "GA"},
            {"DC Name": "Dallas DC", "Region": "South", "State": "TX"},
        ],
    )
    _write_rows(
        wb[TAB_LOTS],
        [
            {
                "Lot Name": "Lot 1 - Grape",
                "Item Description": "Premium Grape Tomato 10oz",
                "Pack Size / UOM": "10oz clamshell",
                "Product Type": "Conventional",
                "Category": "Tomatoes",
            },
            {
                "Lot Name": "Lot 2 - Roma",
                "Item Description": "Roma Tomato Bulk 25lb",
                "Pack Size / UOM": "25lb carton",
                "Product Type": "Organic",
                "Category": "Tomatoes",
            },
        ],
    )
    _write_rows(
        wb[TAB_SUPPLIERS],
        [
            {"Supplier Name": "Green Valley Farms", "Region / Origin": "East", "Notes": "inc"},
            {"Supplier Name": "Sunbelt Produce", "Region / Origin": "South", "Notes": "new"},
        ],
    )
    _write_rows(
        wb[TAB_TIMEFRAMES],
        [
            {
                "Timeframe Label": "Spring 2026",
                "Start Date": "2026-04-01",
                "End Date": "2026-06-30",
                "Week Count": 13,
            }
        ],
    )
    # Volumes: 2 DCs x 2 lots x 1 TF = 4 rows.
    vol_rows: list[dict[str, object]] = []
    for dc in ("Atlanta DC", "Dallas DC"):
        for lot in ("Lot 1 - Grape", "Lot 2 - Roma"):
            vol_rows.append(
                {
                    "DC Name": dc,
                    "Lot Name": lot,
                    "Timeframe": "Spring 2026",
                    "Method": "WEEKLY_X_WEEKS",
                    "Weekly Cases": 400,
                    "Weeks": 13,
                }
            )
    _write_rows(wb[TAB_VOLUMES], vol_rows)
    # Incumbents: 2 DCs x 2 lots = 4 rows.
    inc_rows: list[dict[str, object]] = []
    for dc in ("Atlanta DC", "Dallas DC"):
        for lot in ("Lot 1 - Grape", "Lot 2 - Roma"):
            inc_rows.append(
                {
                    "DC Name": dc,
                    "Lot Name": lot,
                    "Incumbent Supplier": "Green Valley Farms",
                    "Routing Baseline $/case": 11.20,
                    "Contract Notes": "auto",
                }
            )
    _write_rows(wb[TAB_INCUMBENTS], inc_rows)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# PURE tests (no DB)
# ---------------------------------------------------------------------------
def test_stage_filename_normalized() -> None:
    assert stage_filename(1, "setup_kickoff") == "01_setup_kickoff.xlsx"
    assert stage_filename(4, "round1_alignment", version=1) == "04_round1_alignment_v1.xlsx"
    assert stage_filename(9, "post_award", version=2) == "09_post_award_v2.xlsx"
    assert stage_filename(2, "round1_bid_template", ext="csv") == "02_round1_bid_template.csv"


def test_start_run_creates_identical_scaffold(tmp_path: Path) -> None:
    service = PilotService(tmp_path, isolate_db=False)
    paths = service.start_run(commodity="Field Tomatoes", label="Spring Tomatoes")

    # The identical scaffold — every directory + manifest exists.
    assert paths.inputs.is_dir()
    assert paths.outputs.is_dir()
    assert paths.memory.is_dir()
    assert paths.notes_md.is_file()
    assert paths.run_md.is_file()
    assert paths.cycle_id_file.is_file()
    # The Setup/Kickoff workbook was written under the normalized stage name.
    setup = paths.inputs / stage_filename(1, "setup_kickoff")
    assert setup.is_file()
    assert setup.read_bytes()[:2] == b"PK"  # a real xlsx (zip) container

    # slug shape: <commodity-slug>-<YYYYMMDD>-<short>
    parts = paths.slug.split("-")
    assert parts[0] == "field" and "tomatoes" in paths.slug
    assert any(len(p) == 8 and p.isdigit() for p in parts)  # the date stamp

    # The vault is a git repo with the run committed.
    assert (tmp_path / ".git").exists()

    # RUN.md carries the kanban skeleton; files-only kanban points at the setup doc.
    run_text = paths.run_md.read_text()
    for bucket in ("## Done", "## Doing", "## Next", "## Waiting on you"):
        assert bucket in run_text


def test_two_runs_are_structurally_identical(tmp_path: Path) -> None:
    service = PilotService(tmp_path, isolate_db=False)
    a = service.start_run(commodity="Tomatoes", label="A")
    b = service.start_run(commodity="Peppers", label="B")

    def shape(p: Path) -> set[str]:
        return {
            child.relative_to(p).as_posix()
            for child in p.rglob("*")
            if ".git" not in child.parts and child.name != ".gitkeep"
        }

    # Same set of files/dirs in each run folder (modulo the generated content).
    assert shape(a.root) == shape(b.root)
    assert len(service.list_runs()) == 2


def test_remember_and_add_memory_append_notes(tmp_path: Path) -> None:
    service = PilotService(tmp_path, isolate_db=False)
    paths = service.start_run(commodity="Tomatoes", label="Notes Run")

    service.remember(paths, "Buyer wants Dallas prioritized")
    notes = paths.notes_md.read_text()
    assert "Buyer wants Dallas prioritized" in notes

    service.add_memory(paths, "supplier_brief.txt", b"hello brief", "Supplier brief from buyer")
    mem_file = paths.memory / "supplier_brief.txt"
    assert mem_file.is_file()
    notes = paths.notes_md.read_text()
    assert "Supplier brief from buyer" in notes
    assert "supplier_brief.txt" in notes  # the file name is linked in the note


def test_setup_template_has_all_tabs_and_dropdowns() -> None:
    wb = load_workbook(BytesIO(build_setup_workbook()))
    for tab in (
        TAB_CYCLE,
        TAB_DCS,
        TAB_LOTS,
        TAB_SUPPLIERS,
        TAB_VOLUMES,
        TAB_INCUMBENTS,
        TAB_TIMEFRAMES,
    ):
        assert tab in wb.sheetnames
    # The Lots tab carries the Product Type closed-domain dropdown.
    lots = wb[TAB_LOTS]
    assert any("Conventional" in (dv.formula1 or "") for dv in lots.data_validations.dataValidation)


# ---------------------------------------------------------------------------
# INTEGRATION — the setup round-trip (template -> fill -> ingest -> load_cycle)
# ---------------------------------------------------------------------------
@pytest.mark.integration
def test_setup_roundtrips_to_cycle(tmp_path: Path, db_session) -> None:  # type: ignore[no-untyped-def]
    from app.cycle.loader import load_cycle

    service = PilotService(tmp_path, isolate_db=False)
    paths = service.start_run(commodity="Field Tomatoes", label="Roundtrip Run")

    # The sponsor "uploads" the filled setup into inputs/ (formal upload step).
    filled = _build_filled_setup()
    uploaded = paths.inputs / stage_filename(1, "setup_kickoff")
    uploaded.write_bytes(filled)

    cycle_id = service.ingest_setup(db_session, paths, uploaded)
    assert cycle_id

    # cycle_id.txt + RUN.md header now carry the link.
    assert paths.cycle_id_file.read_text().strip() == cycle_id
    assert cycle_id in paths.run_md.read_text()

    # load_cycle reconstructs the scope with the right counts.
    view = load_cycle(db_session, cycle_id)
    assert view.cycle_name == "Test Tomatoes Cycle"
    assert len(view.dcs) == 2
    assert len(view.lots) == 2
    assert len(view.items) == 2
    assert len(view.tfs) == 1
    assert len(view.rounds) == 2
    assert len(view.suppliers) == 2
    # 2 DCs x 2 lots x 1 TF = 4 projected-volume cells.
    assert len(view.period_cases_by_cell) == 4
    # 2 DCs x 2 lots = 4 incumbent routing baselines (the synthetic 11.20 $/case).
    assert len(view.incumbent_routing) == 4
    assert all(float(v) == 11.20 for v in view.incumbent_routing.values())

    # The kanban reflects the created cycle in plain language.
    board = service.status(db_session, paths)
    assert any("Cycle created" in entry for entry in board["Done"])
    assert any(
        "bid" in entry.lower() for entry in board["Next"]
    )  # next step phrased in buyer terms


@pytest.mark.integration
def test_setup_ingest_reports_unresolved_rows(tmp_path: Path, db_session) -> None:  # type: ignore[no-untyped-def]
    from app.pilot.setup_ingest import SetupIngestError, ingest_setup_workbook

    # Build a filled workbook then break a cross-reference: a volume row points at an unknown DC.
    wb = load_workbook(BytesIO(_build_filled_setup()))
    vol = wb[TAB_VOLUMES]
    vol.cell(row=EXAMPLE_START_ROW, column=_header_col(vol, "DC Name"), value="Nonexistent DC")
    buffer = BytesIO()
    wb.save(buffer)

    with pytest.raises(SetupIngestError) as exc:
        ingest_setup_workbook(db_session, buffer.getvalue())
    assert any("Nonexistent DC" in p for p in exc.value.problems)
