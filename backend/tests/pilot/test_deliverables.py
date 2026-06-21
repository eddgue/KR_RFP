"""DB-backed deliverable registry (NFS Slice 1) — names match the harness's on-disk filenames.

The web console renders every deliverable on request from the governed DB (ADR-0018). This proves
`enumerate_deliverables` is a faithful projection of what the MCP-harness write path produces today:
it drives a full synthetic run through `PilotService` (which writes the workbooks to disk), then
asserts the registry's enumerated NAMES equal exactly the files the run wrote to inputs/ + outputs/,
and that a sampled render is DATA-identical to the stored workbook (date-stamped provenance lines
differ across days, so timestamped outputs are compared on cell DATA, not raw bytes — E-39).

Also covers the no-cycle case: a run with no ingested setup enumerates exactly the Setup workbook.

Synthetic only (clean-room, ADR-0001); integration (skips when no DB).
"""

from __future__ import annotations

from datetime import date
from decimal import Decimal
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.pilot.deliverables import enumerate_deliverables
from app.pilot.service import PilotService
from app.pilot.vault import is_rehearsal, stage_filename
from tests.pilot.test_pilot_cycle_e2e import (
    _build_filled_setup,
    _fill_bid_template,
    _first_award_cell,
    _latest_run_id,
)

pytestmark = pytest.mark.integration


def _disk_filenames(paths) -> set[str]:  # type: ignore[no-untyped-def]
    """Every workbook the harness wrote to inputs/ + outputs/ (skipping scaffold markers)."""

    names: set[str] = set()
    for directory in (paths.inputs, paths.outputs):
        for entry in directory.iterdir():
            if entry.is_file() and entry.name != ".gitkeep":
                names.add(entry.name)
    return names


def _sheet_data(data: bytes) -> dict[str, list[list[object]]]:
    """The cell VALUES of every sheet (data-only) — the data-identity comparison basis (E-39)."""

    wb = load_workbook(BytesIO(data), data_only=True)
    return {
        name: [[cell.value for cell in row] for row in wb[name].iter_rows()]
        for name in wb.sheetnames
    }


def _cycle_id(paths) -> str | None:  # type: ignore[no-untyped-def]
    value = paths.cycle_id_file.read_text(encoding="utf-8").strip()
    return value or None


def test_enumerate_matches_harness_filenames(tmp_path: Path, db_session) -> None:  # type: ignore[no-untyped-def]
    """The registry's names == the files run_round/freeze_award/record_adjustment write today.

    Drives a full synthetic run through the harness (which persists the workbooks), then asserts the
    DB-derived enumeration reproduces exactly that set of filenames — setup, both round templates,
    the sealed-analysis alignment workbook, the internal + combined supplier guides, every
    per-supplier guide, and the post-award adjustment doc.
    """

    service = PilotService(tmp_path, isolate_db=False)
    paths = service.start_run(commodity="Field Tomatoes", label="Deliverables")
    setup_path = paths.inputs / stage_filename(1, "setup_kickoff")
    setup_path.write_bytes(_build_filled_setup())
    cycle_id = service.ingest_setup(db_session, paths, setup_path)

    # Both rounds' templates (the cycle has 2 rounds), so the registry must enumerate both.
    for round_no in (1, 2):
        tmpl = service.generate_bid_template(db_session, paths, round_no)
        if round_no == 1:
            tmpl.write_bytes(_fill_bid_template(tmpl.read_bytes()))
            service.ingest_bids(db_session, paths, 1, tmpl)

    service.run_round(db_session, paths, 1)
    analysis_run_id = _latest_run_id(db_session, cycle_id)
    award_id = service.freeze_award(
        db_session,
        paths,
        analysis_run_id=analysis_run_id,
        scenario_code="B",
        award_code="AWD-DELIV-1",
    )
    dc_id, lot_id, tf_id, sup_id, frozen_price = _first_award_cell(db_session, award_id)
    service.record_adjustment(
        db_session,
        paths,
        award_id=award_id,
        adjustment_type="NEGOTIATED_REPRICE",
        effective_date=date(2026, 7, 1),
        reason="Registry parity reprice",
        line_changes=[(dc_id, lot_id, tf_id, sup_id, frozen_price - Decimal("0.25"))],
    )

    disk = _disk_filenames(paths)
    enumerated = {
        d.name
        for d in enumerate_deliverables(
            db_session, cycle_id=_cycle_id(paths), slug=paths.slug, rehearsal=is_rehearsal(paths)
        )
    }

    # Exact set parity: the registry produces every file the harness wrote and nothing extra.
    assert enumerated == disk

    # Sanity: the expected normalized names are present (not just an accidental empty match).
    assert "01_setup_kickoff.xlsx" in enumerated
    assert "02_round1_bid_template.xlsx" in enumerated
    assert "05_round2_bid_template.xlsx" in enumerated
    assert "04_round1_alignment_v1.xlsx" in enumerated
    assert "08_award_booking_guide.xlsx" in enumerated
    assert "08_award_supplier_guides.xlsx" in enumerated
    assert "09_post_award_v1.xlsx" in enumerated
    # One per-supplier guide per awarded supplier (award-id stamped, slugified).
    per_supplier = {n for n in enumerated if "award_guide_awd_deliv_1" in n}
    assert per_supplier, enumerated


def test_render_is_data_identical_to_stored(tmp_path: Path, db_session) -> None:  # type: ignore[no-untyped-def]
    """A sampled on-request render reproduces the stored workbook's DATA (E-39, preserving).

    The alignment workbook carries `date.today()` provenance lines, so raw bytes can drift across
    days; this compares the SHEET CELL DATA, which the DB-renders reproduce exactly. The booking
    guide is checked the same way.
    """

    service = PilotService(tmp_path, isolate_db=False)
    paths = service.start_run(commodity="Field Tomatoes", label="DataIdentity")
    setup_path = paths.inputs / stage_filename(1, "setup_kickoff")
    setup_path.write_bytes(_build_filled_setup())
    cycle_id = service.ingest_setup(db_session, paths, setup_path)
    tmpl = service.generate_bid_template(db_session, paths, 1)
    tmpl.write_bytes(_fill_bid_template(tmpl.read_bytes()))
    service.ingest_bids(db_session, paths, 1, tmpl)
    align_path = service.run_round(db_session, paths, 1)
    analysis_run_id = _latest_run_id(db_session, cycle_id)
    service.freeze_award(
        db_session,
        paths,
        analysis_run_id=analysis_run_id,
        scenario_code="B",
        award_code="AWD-DI-1",
    )
    booking_path = paths.outputs / "08_award_booking_guide.xlsx"

    by_name = {
        d.name: d
        for d in enumerate_deliverables(
            db_session, cycle_id=cycle_id, slug=paths.slug, rehearsal=False
        )
    }

    for stored_path in (align_path, booking_path):
        rendered = by_name[stored_path.name].render(db_session)
        assert _sheet_data(rendered) == _sheet_data(stored_path.read_bytes()), stored_path.name


def test_no_cycle_enumerates_only_setup(tmp_path: Path, db_session) -> None:  # type: ignore[no-untyped-def]
    """A run with no ingested setup (no cycle) enumerates exactly the Setup/Kickoff workbook."""

    service = PilotService(tmp_path, isolate_db=False)
    paths = service.start_run(commodity="Field Tomatoes", label="NoCycle")

    items = enumerate_deliverables(db_session, cycle_id=None, slug=paths.slug, rehearsal=False)
    assert [d.name for d in items] == ["01_setup_kickoff.xlsx"]
    # The setup render is a real xlsx (zip) payload, produced with no DB read.
    assert items[0].render(db_session)[:2] == b"PK"
