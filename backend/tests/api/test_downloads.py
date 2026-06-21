"""Download endpoints render on request from the DB, with an EMPTY vault outputs/ (NFS Slice 5).

The web console persists no generated files (ADR-0018): run analysis / freeze / record-adjustment do
the governed DB writes only, and `/files`, `/files/{name}`, `/archive` are projections of
`enumerate_deliverables` rendered on request. This drives a full run through the HTTP surface and
asserts:

  * the vault `outputs/` folder stays EMPTY (no workbook/guide/post-award file is ever written);
  * `/files` lists the DB deliverables (the sealed-analysis alignment workbook, the frozen-award
    guides, the post-award doc) by their normalized names;
  * `/files/{name}` and `/archive` render those bytes, DATA-identical to a direct registry render
    (date-stamped provenance lines differ across days, so timestamped outputs compare on cell DATA).

Integration (real Postgres + a temp vault); reuses the alignment e2e HTTP seed helpers.
"""

from __future__ import annotations

import io
import zipfile
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.pilot.deliverables import enumerate_deliverables
from tests.api.test_alignment import _create_run, _login, _seed_sealed_run

RUNS = "/api/v1/runs"


def _sheet_data(data: bytes) -> dict[str, list[list[object]]]:
    """The cell VALUES of every sheet (data-only) — the data-identity comparison basis (E-39)."""

    wb = load_workbook(BytesIO(data), data_only=True)
    return {
        name: [[cell.value for cell in row] for row in wb[name].iter_rows()]
        for name in wb.sheetnames
    }


def _outputs_files(vault_root, slug: str) -> set[str]:  # type: ignore[no-untyped-def]
    """The files actually on disk in a run's outputs/ dir (skipping scaffold markers)."""

    outputs = vault_root / "runs" / slug / "outputs"
    if not outputs.is_dir():
        return set()
    return {p.name for p in outputs.iterdir() if p.is_file() and p.name != ".gitkeep"}


@pytest.mark.integration
def test_downloads_render_from_db_with_empty_outputs(  # type: ignore[no-untyped-def]
    client, seed_user, vault_root, db_session
) -> None:
    """Analysis → freeze → adjust write NO outputs/ files; downloads render the deliverables."""

    _login(client, seed_user)
    slug = _create_run(client)
    analysis_run_id = _seed_sealed_run(client, slug)  # create→setup→template→import→analysis

    # Freeze B + record one post-award layer — all governed DB writes, no files.
    award_id = client.post(
        f"{RUNS}/{slug}/awards/freeze",
        json={"analysis_run_id": analysis_run_id, "scenario_code": "B", "award_code": "AWD-DL-1"},
    ).json()["award_id"]
    line = client.get(f"{RUNS}/{slug}/awards/{award_id}").json()["lines"][0]
    rec = client.post(
        f"{RUNS}/{slug}/awards/{award_id}/adjustments",
        json={
            "adjustment_type": "MARKET_HIKE",
            "effective_date": "2026-04-01",
            "reason": "download-path reprice",
            "changes": [
                {
                    "dc_id": line["dc_id"],
                    "lot_id": line["lot_id"],
                    "tf_id": line["tf_id"],
                    "supplier_id": line["supplier_id"],
                    "new_price": round(line["frozen_price"] + 3.0, 2),
                }
            ],
        },
    )
    assert rec.status_code == 200, rec.text

    # The console persisted NO generated outputs — outputs/ is empty (render-on-request).
    assert _outputs_files(vault_root, slug) == set()

    # /files lists the DB deliverables by their normalized names.
    listed = client.get(f"{RUNS}/{slug}/files")
    assert listed.status_code == 200
    names = {f["name"] for f in listed.json()}
    assert "01_setup_kickoff.xlsx" in names
    assert "04_round1_alignment_v1.xlsx" in names
    assert "08_award_booking_guide.xlsx" in names
    assert "08_award_supplier_guides.xlsx" in names
    assert "09_post_award_v1.xlsx" in names
    assert all(f["size_bytes"] > 0 for f in listed.json())

    # A direct registry render is the ground truth; /files/{name} must match it on DATA.
    by_name = {
        d.name: d
        for d in enumerate_deliverables(
            db_session, cycle_id=_cycle(db_session, slug), slug=slug, rehearsal=False
        )
    }
    for name in (
        "04_round1_alignment_v1.xlsx",
        "08_award_booking_guide.xlsx",
        "09_post_award_v1.xlsx",
    ):
        resp = client.get(f"{RUNS}/{slug}/files/{name}")
        assert resp.status_code == 200, name
        assert resp.headers["content-type"].startswith("application/vnd.openxmlformats")
        assert name in resp.headers["content-disposition"]
        assert _sheet_data(resp.content) == _sheet_data(by_name[name].render(db_session)), name

    # An unknown deliverable name is a clean 404 (no path-traversal surface — only exact matches).
    assert client.get(f"{RUNS}/{slug}/files/nope.xlsx").status_code == 404

    # /archive zips every deliverable under the slug, each a real xlsx; outputs/ stays empty.
    archive = client.get(f"{RUNS}/{slug}/archive")
    assert archive.status_code == 200
    zf = zipfile.ZipFile(io.BytesIO(archive.content))
    arc_names = set(zf.namelist())
    assert f"{slug}/04_round1_alignment_v1.xlsx" in arc_names
    assert f"{slug}/09_post_award_v1.xlsx" in arc_names
    assert all(zf.read(n)[:2] == b"PK" for n in arc_names)
    assert _outputs_files(vault_root, slug) == set()


def _cycle(db_session, slug: str) -> str:  # type: ignore[no-untyped-def]
    from app.pilot.run_repo import get_run

    run = get_run(db_session, slug)
    assert run is not None and run.cycle_id
    return run.cycle_id
