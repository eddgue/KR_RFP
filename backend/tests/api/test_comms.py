"""Supplier comms (E-37): award-notification email drafts over the HTTP surface.

Integration (real Postgres + temp vault): reuses the alignment HTTP seed helpers (create → setup →
template → import → analysis) + the freeze helper, then exercises the award comms-draft endpoint —
one template-merge draft per awarded supplier, draft-only.
"""

from __future__ import annotations

from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.domain.bid.template_schema import BODY_START_ROW, HEADER_ROW, SHEET_BIDS, BidColumn
from tests.api.test_alignment import _XLSX, _create_run, _login, _seed_sealed_run
from tests.api.test_post_award import _freeze_b

RUNS = "/api/v1/runs"


def _fill_bid_template_bumped(template_bytes: bytes, bump: Decimal) -> bytes:
    """Fill the round template like the e2e seed but add `bump` to every All-In/FOB price.

    Used to RESUBMIT a round with materially different prices so the supersede path fires (the prior
    scored rows flip to non-scoreable), exercising the sealed-run sourcing fix.
    """

    wb = load_workbook(BytesIO(template_bytes))
    ws = wb[SHEET_BIDS]
    hdr = {ws.cell(HEADER_ROW, c).value: c for c in range(1, ws.max_column + 1)}
    all_in_col = hdr[BidColumn.ALL_IN.value]
    fob_col = hdr[BidColumn.FOB.value]
    weekly_col = hdr[BidColumn.WEEKLY_VOL_OFFERED.value]
    total_col = hdr[BidColumn.TOTAL_VOL_OFFERED.value]
    sup_col = hdr[BidColumn.SUPPLIER.value]
    lot_col = hdr[BidColumn.LOT.value]
    for r in range(BODY_START_ROW, ws.max_row + 1):
        sup = str(ws.cell(r, sup_col).value or "").strip()
        lot = str(ws.cell(r, lot_col).value or "").strip()
        if not sup or not lot:
            continue
        base = Decimal("12.00")
        if "Grape" in lot:
            price = base - (Decimal("1.50") if "Green Valley" in sup else Decimal("0.20"))
        else:
            price = base - (Decimal("1.50") if "Sunbelt" in sup else Decimal("0.20"))
        price += bump
        ws.cell(r, all_in_col, value=float(price))
        ws.cell(r, fob_col, value=float(price - Decimal("1.00")))
        ws.cell(r, weekly_col, value=600)
        ws.cell(r, total_col, value=600 * 13)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.mark.integration
def test_award_comms_requires_auth(client, vault_root) -> None:  # type: ignore[no-untyped-def]
    assert client.get(f"{RUNS}/x/awards/a-1/comms/award").status_code == 401


@pytest.mark.integration
def test_award_comms_drafts_one_per_awarded_supplier(client, seed_user, vault_root) -> None:  # type: ignore[no-untyped-def]
    """freeze B → an award draft per supplier; machine-tag subject + data-filled body."""

    _login(client, seed_user)
    slug = _create_run(client)
    analysis_run_id = _seed_sealed_run(client, slug)
    award_id = _freeze_b(client, slug, analysis_run_id, code="AWD-COMMS-1")

    resp = client.get(f"{RUNS}/{slug}/awards/{award_id}/comms/award")
    assert resp.status_code == 200, resp.text
    drafts = resp.json()
    assert len(drafts) >= 1

    for draft in drafts:
        assert draft["email_type"] == "Award Notification"
        # Subject: machine-readable routing tags first, then the human-readable type + cycle.
        assert draft["subject"].startswith("[RFP:")
        assert "[SUP:" in draft["subject"]
        assert "Award Notification –" in draft["subject"]
        # Body merged from governed data.
        assert f"Dear {draft['supplier_name']}," in draft["body"]
        assert "selected for award" in draft["body"]
        assert "Awarded Locations:" in draft["body"]
        # Counts are filled (not visible holes).
        assert "[#AwardedDCCount]" not in draft["body"]
        assert "[#AwardedLotCount]" not in draft["body"]
        # Each draft attaches the supplier's OWN individual guide (generated at freeze) — never the
        # combined all-suppliers workbook (which would leak every supplier's awards).
        assert "[#AwardFileName]" not in draft["body"]
        assert "AwardFileName" not in draft["missing"]
        assert "award_guide" in draft["body"]
        assert "supplier_guides" not in draft["body"]
        # The authenticated user is the draft's buyer; the title is left for the buyer to complete.
        assert seed_user["username"] in draft["body"]
        assert "BuyerTitle" in draft["missing"]


@pytest.mark.integration
def test_award_comms_guide_is_not_stale_across_awards(client, seed_user, vault_root) -> None:  # type: ignore[no-untyped-def]
    """Two awards on one run → an earlier award's drafts attach ITS guide, never the later one's.

    The per-supplier guide is award-code-stamped, so freezing a second scenario can't shadow the
    first award's files (the regression Codex flagged: a fixed-name guide overwritten per freeze).
    """

    _login(client, seed_user)
    slug = _create_run(client)
    analysis_run_id = _seed_sealed_run(client, slug)
    award_b = _freeze_b(client, slug, analysis_run_id, code="AWD-B")
    resp_a = client.post(
        f"{RUNS}/{slug}/awards/freeze",
        json={"analysis_run_id": analysis_run_id, "scenario_code": "A", "award_code": "AWD-A"},
    )
    assert resp_a.status_code == 200, resp_a.text

    drafts = client.get(f"{RUNS}/{slug}/awards/{award_b}/comms/award").json()
    assert len(drafts) >= 1
    for draft in drafts:
        body = draft["body"].lower()
        assert "award file: " in body  # a guide is attached
        assert "awd_b" in body  # the EARLIER award's stamped guide
        assert "awd_a" not in body  # never the later award's guide


@pytest.mark.integration
def test_award_comms_guide_unique_when_award_codes_collide(client, seed_user, vault_root) -> None:  # type: ignore[no-untyped-def]
    """Two awards on one run sharing the SAME award_code still attach distinct, own guides.

    Award codes aren't enforced unique, so the per-supplier guide filename carries the award_id (the
    unique PK) — a later same-coded freeze can't overwrite the earlier award's file, and the earlier
    award's drafts reference its own guide (by id), never the later award's.
    """

    _login(client, seed_user)
    slug = _create_run(client)
    analysis_run_id = _seed_sealed_run(client, slug)
    award_b = _freeze_b(client, slug, analysis_run_id, code="DUP-CODE")
    resp_a = client.post(
        f"{RUNS}/{slug}/awards/freeze",
        json={"analysis_run_id": analysis_run_id, "scenario_code": "A", "award_code": "DUP-CODE"},
    )
    assert resp_a.status_code == 200, resp_a.text
    award_a = resp_a.json()["award_id"]
    assert award_a != award_b

    drafts = client.get(f"{RUNS}/{slug}/awards/{award_b}/comms/award").json()
    assert len(drafts) >= 1
    slug_b = award_b.replace("-", "_")  # the id as it appears slugified in the filename
    slug_a = award_a.replace("-", "_")
    for draft in drafts:
        assert "[#AwardFileName]" not in draft["body"]  # a guide is attached
        assert slug_b in draft["body"]  # THIS award's own guide (by unique id)
        assert slug_a not in draft["body"]  # never the same-coded other award's guide


@pytest.mark.integration
def test_award_comms_unknown_award_404(client, seed_user, vault_root) -> None:  # type: ignore[no-untyped-def]
    _login(client, seed_user)
    slug = _create_run(client)
    _seed_sealed_run(client, slug)
    resp = client.get(f"{RUNS}/{slug}/awards/no-such-award/comms/award")
    assert resp.status_code == 404


@pytest.mark.integration
def test_feedback_comms_requires_auth(client, vault_root) -> None:  # type: ignore[no-untyped-def]
    assert client.get(f"{RUNS}/x/analysis/run-1/comms/feedback").status_code == 401


@pytest.mark.integration
def test_feedback_comms_drafts_above_benchmark_suppliers(client, seed_user, vault_root) -> None:  # type: ignore[no-untyped-def]
    """A sealed round → a round-feedback draft per supplier above the market-low benchmark.

    The synthetic seed splits each DC's two lots across the two suppliers, so on its weaker lot each
    supplier sits above the other's market low — both get a draft with a machine-tag subject and a
    data-filled body (DC summary + hard/soft ask sections, tables expanded, no visible holes).
    """

    _login(client, seed_user)
    slug = _create_run(client)
    analysis_run_id = _seed_sealed_run(client, slug)

    resp = client.get(f"{RUNS}/{slug}/analysis/{analysis_run_id}/comms/feedback")
    assert resp.status_code == 200, resp.text
    drafts = resp.json()
    assert len(drafts) >= 1

    for draft in drafts:
        assert draft["email_type"] == "Round Feedback"
        # Subject: machine-readable routing tags first, then the round-feedback infix + cycle.
        assert draft["subject"].startswith("[RFP:")
        assert "[SUP:" in draft["subject"]
        assert "Feedback –" in draft["subject"]
        assert "[#RoundNumber]" not in draft["subject"]
        # Body merged from governed data: greeting + the three authored sections.
        assert f"Dear {draft['supplier_name']}," in draft["body"]
        assert "DC Summary" in draft["body"]
        assert "Items Requiring Action" in draft["body"]
        assert "Additional Improvement Opportunities" in draft["body"]
        # Every table block expanded (even an empty one renders a header, never a leftover token).
        assert "[#DCSummaryTable]" not in draft["body"]
        assert "[#HardAskTable]" not in draft["body"]
        assert "[#SoftAskTable]" not in draft["body"]
        # The authenticated user is the draft's buyer; the title is left for the buyer to complete.
        assert seed_user["username"] in draft["body"]
        assert "BuyerTitle" in draft["missing"]


@pytest.mark.integration
def test_feedback_comms_unknown_run_404(client, seed_user, vault_root) -> None:  # type: ignore[no-untyped-def]
    _login(client, seed_user)
    slug = _create_run(client)
    _seed_sealed_run(client, slug)
    resp = client.get(f"{RUNS}/{slug}/analysis/no-such-run/comms/feedback")
    assert resp.status_code == 404


@pytest.mark.integration
def test_feedback_comms_sealed_after_resubmit(client, seed_user, vault_root) -> None:  # type: ignore[no-untyped-def]
    """A resubmission after sealing must NOT move the sealed run's feedback prices/market-lows.

    The draft sources from the sealed run's own `bid_score` → `bid_line` rows, so superseding the
    scored rows with a new (higher-priced) submission leaves THAT run's draft byte-for-byte
    unchanged. Without the sealed-run sourcing fix, the draft would re-read the now-current bumped
    rows and the benchmark/premium values would shift.
    """

    _login(client, seed_user)
    slug = _create_run(client)
    analysis_run_id = _seed_sealed_run(client, slug)

    before = client.get(f"{RUNS}/{slug}/analysis/{analysis_run_id}/comms/feedback").json()
    assert len(before) >= 1

    # Resubmit round 1 at materially higher prices — supersedes the scored rows (now non-scoreable).
    tmpl = client.post(f"{RUNS}/{slug}/rounds/1/template")
    template_name = tmpl.json()["filename"]
    template_bytes = client.get(f"{RUNS}/{slug}/files/{template_name}").content
    bumped = _fill_bid_template_bumped(template_bytes, Decimal("5.00"))
    imported = client.post(
        "/api/v1/bids/import",
        data={"run": slug, "round": 1, "mode": "strict"},
        files={"file": (template_name, bumped, _XLSX)},
    )
    assert imported.status_code == 200, imported.text

    after = client.get(f"{RUNS}/{slug}/analysis/{analysis_run_id}/comms/feedback").json()
    assert after == before  # the sealed run's draft is unchanged by a later resubmission


@pytest.mark.integration
def test_rejection_comms_requires_auth(client, vault_root) -> None:  # type: ignore[no-untyped-def]
    assert client.get(f"{RUNS}/x/awards/a-1/comms/rejection").status_code == 401


@pytest.mark.integration
def test_rejection_comms_drafts_per_lost_lot(client, seed_user, vault_root) -> None:  # type: ignore[no-untyped-def]
    """freeze B → a non-selection draft per supplier with a lot they did not win.

    The synthetic seed splits each DC's two lots across the two suppliers, so each supplier loses
    its weaker lot — both get a "RFP Results" draft itemizing the lost lots (their price, the
    market-low benchmark, the % gap, a reason), with a machine-tag subject and no visible holes.
    """

    _login(client, seed_user)
    slug = _create_run(client)
    analysis_run_id = _seed_sealed_run(client, slug)
    award_id = _freeze_b(client, slug, analysis_run_id, code="AWD-COMMS-REJ-1")

    resp = client.get(f"{RUNS}/{slug}/awards/{award_id}/comms/rejection")
    assert resp.status_code == 200, resp.text
    drafts = resp.json()
    assert len(drafts) >= 1

    for draft in drafts:
        assert draft["email_type"] == "RFP Results"
        # Subject: machine-readable routing tags first, then the results infix + cycle.
        assert draft["subject"].startswith("[RFP:")
        assert "[SUP:" in draft["subject"]
        assert "RFP Results –" in draft["subject"]
        # Body merged from governed data: greeting + the evaluation summary section.
        assert f"Dear {draft['supplier_name']}," in draft["body"]
        assert "not selected for award" in draft["body"]
        assert "Evaluation Summary" in draft["body"]
        # The reason table expanded (header present, placeholder gone).
        assert "[#RejectionReasonTable]" not in draft["body"]
        assert "Benchmark Price" in draft["body"]
        # The authenticated user is the draft's buyer; the title is left for the buyer to complete.
        assert seed_user["username"] in draft["body"]
        assert "BuyerTitle" in draft["missing"]


@pytest.mark.integration
def test_rejection_comms_unknown_award_404(client, seed_user, vault_root) -> None:  # type: ignore[no-untyped-def]
    _login(client, seed_user)
    slug = _create_run(client)
    _seed_sealed_run(client, slug)
    resp = client.get(f"{RUNS}/{slug}/awards/no-such-award/comms/rejection")
    assert resp.status_code == 404
