"""E-38b capacity-check workbook tab — renders the allocation-vs-stated-capacity surface.

Pure test of `_write_capacity_check_tab` (the operator-facing safety surface in the alignment
workbook): given resolved display rows, the sheet is written with the right statuses and the
OVER CAPACITY rows are present. No DB — the evaluator + reader are tested separately.
"""

from __future__ import annotations

from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook, load_workbook

from app.output.scenario_workbook import CapacityCheckDisplayRow, _write_capacity_check_tab

_D = Decimal


def _row(
    dc: str, sup: str, alloc: str, maxp: str | None, over: bool, status: str
) -> CapacityCheckDisplayRow:
    return CapacityCheckDisplayRow(
        dc_name=dc,
        lot_name="Roma Lot",
        tf_name="TF1",
        supplier_name=sup,
        allocated_cases=_D(alloc),
        allocated_weekly_cases=_D(alloc) / 13,
        max_period_cases=_D(maxp) if maxp is not None else None,
        max_weekly_cases=None,
        status=status,
        over_capacity=over,
        has_statement=maxp is not None,
    )


def test_capacity_check_tab_renders_and_flags_over() -> None:
    rows = [
        _row("Atlanta DC", "Acme", "9000", "6000", True, "OVER CAPACITY"),
        _row("Memphis DC", "Bravo", "5000", "6000", False, "Within capacity"),
        _row("Dallas DC", "Cresta", "1000", None, False, "No stated capacity"),
    ]
    wb = Workbook()
    _write_capacity_check_tab(wb, rows)

    buf = BytesIO()
    wb.save(buf)
    wb2 = load_workbook(buf)

    assert "Capacity Check" in wb2.sheetnames
    ws = wb2["Capacity Check"]
    statuses = {ws.cell(row=r, column=9).value for r in range(1, ws.max_row + 1)}
    assert "OVER CAPACITY" in statuses
    assert "Within capacity" in statuses
    assert "No stated capacity" in statuses
    # The "—" placeholder is used where a supplier stated no ceiling (period col = 7).
    period_cells = {ws.cell(row=r, column=7).value for r in range(1, ws.max_row + 1)}
    assert "—" in period_cells


def test_capacity_check_tab_empty_is_safe() -> None:
    """No award cells (or none with capacity context) → the sheet still writes, no crash."""

    wb = Workbook()
    _write_capacity_check_tab(wb, [])

    buf = BytesIO()
    wb.save(buf)
    wb2 = load_workbook(buf)
    assert "Capacity Check" in wb2.sheetnames
