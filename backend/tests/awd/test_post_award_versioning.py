"""Post-award freeze-and-layer versioning (ADR-0014) + the versioned doc — integration.

Synthetic only: seeds a minimal cycle (+ round, lots, timeframes, DCs, suppliers), a sealed
`eng.analysis_run` and a scenario 'B' with a couple of `eng.analysis_scenario_award` cells, then:
  * freezes the award (the immutable baseline);
  * records two append-only adjustment layers (v1, v2);
  * asserts version_no increments, effective_award reflects v2 (and as_of_version=1 reflects v1),
    award_versions returns the v0->vN history, and freeze is idempotent;
  * generates the workbook and asserts the explicit Version heading text for the right version and
    that resolved NAMES (not UUID keys) appear (D23).

Runs against the real Postgres (PLAN §7); skips when no DB is reachable (conftest fixtures).
"""

from __future__ import annotations

import uuid
from datetime import UTC, date, datetime
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.domain.awd.service import (
    add_adjustment,
    award_versions,
    effective_award,
    freeze_award,
)
from app.output.post_award_doc import write_post_award_adjustments_xlsx

pytestmark = pytest.mark.integration

_D = Decimal


def _id() -> str:
    return str(uuid.uuid4())


def _seed_cycle_and_run(session: Session) -> dict[str, str]:
    """Seed a minimal cycle + round + lots/timeframes/DCs/suppliers + a sealed B-scenario run.

    Returns the keys the test asserts against (ids + the two cells' supplier/dc/lot/tf + names).
    All synthetic; written by key (D21). Two cells under scenario 'B'.
    """

    cycle_id = _id()
    round_id = _id()
    run_id = _id()
    scenario_id = _id()
    commodity_id = _id()  # text commodity id; subcommodity NULL so the composite FK is skipped

    dc1, dc2 = _id(), _id()
    sup1, sup2 = _id(), _id()
    lot1 = _id()
    tf1 = _id()

    now = datetime.now(UTC).replace(tzinfo=None)

    # ref.dc / ref.supplier (global reference; names resolved in the doc).
    session.execute(
        text(
            "INSERT INTO ref.dc (dc_id, dc_code, dc_name, active_flag) VALUES "
            "(:i1, :c1, :n1, true), (:i2, :c2, :n2, true)"
        ),
        {
            "i1": dc1,
            "c1": "DC01",
            "n1": "Atlanta DC",
            "i2": dc2,
            "c2": "DC02",
            "n2": "Denver DC",
        },
    )
    session.execute(
        text(
            "INSERT INTO ref.supplier (supplier_id, canonical_name, active_flag, created_at) "
            "VALUES (:i1, :n1, true, :ts), (:i2, :n2, true, :ts)"
        ),
        {"i1": sup1, "n1": "Fresh Valley Farms", "i2": sup2, "n2": "Sunrise Produce Co", "ts": now},
    )

    # cyc.cycle (subcommodity NULL -> the composite subcommodity FK is not enforced).
    session.execute(
        text(
            "INSERT INTO cyc.cycle (cycle_id, cycle_code, cycle_name, commodity_id, status, "
            "why_now, target_effective_date, round_count, created_at, created_by) VALUES "
            "(:cid, :code, :name, :com, 'OPEN', 'synthetic test', :ted, 3, :ts, 'tester')"
        ),
        {
            "cid": cycle_id,
            "code": f"CYC-{cycle_id[:8]}",
            "name": "Test Berries Cycle",
            "com": commodity_id,
            "ted": date(2026, 9, 1),
            "ts": now,
        },
    )
    session.execute(
        text(
            "INSERT INTO cyc.cycle_round (round_id, cycle_id, round_number, status, is_final) "
            "VALUES (:rid, :cid, 1, 'OPEN', true)"
        ),
        {"rid": round_id, "cid": cycle_id},
    )
    session.execute(
        text(
            "INSERT INTO cyc.cycle_lot (lot_id, cycle_id, lot_code, lot_name, active_flag) "
            "VALUES (:lid, :cid, 'LOT-A', 'Strawberries Lot', true)"
        ),
        {"lid": lot1, "cid": cycle_id},
    )
    session.execute(
        text(
            "INSERT INTO cyc.cycle_timeframe (tf_id, cycle_id, tf_code, tf_name, start_date, "
            "end_date, week_count) VALUES "
            "(:tid, :cid, 'P1', 'Spring Period', :sd, :ed, 13)"
        ),
        {"tid": tf1, "cid": cycle_id, "sd": date(2026, 3, 1), "ed": date(2026, 6, 1)},
    )

    # eng.analysis_run (sealed) + eng.analysis_scenario 'B' + two award cells.
    session.execute(
        text(
            "INSERT INTO eng.analysis_run (analysis_run_id, cycle_id, round_id, engine_version, "
            "config_preset, status, is_sealed, input_hash_manifest, output_hash_manifest, "
            "run_started_at, run_finished_at, run_by) VALUES "
            "(:rid, :cid, :rnd, 'v3-test', 'balanced', 'SEALED', true, :ih, :oh, :ts, :ts, "
            "'tester')"
        ),
        {
            "rid": run_id,
            "cid": cycle_id,
            "rnd": round_id,
            "ih": "hash" + "0" * 8,
            "oh": "hash" + "1" * 8,
            "ts": now,
        },
    )
    session.execute(
        text(
            "INSERT INTO eng.analysis_scenario (analysis_scenario_id, analysis_run_id, "
            "scenario_code, label) VALUES (:sid, :rid, 'B', 'Recommended')"
        ),
        {"sid": scenario_id, "rid": run_id},
    )
    session.execute(
        text(
            "INSERT INTO eng.analysis_scenario_award (award_id, analysis_scenario_id, dc_id, "
            "lot_id, tf_id, supplier_id, volume_share, awarded_price, is_recommended) VALUES "
            "(:a1, :sid, :dc1, :lot, :tf, :s1, 0.6, 100.00, true), "
            "(:a2, :sid, :dc2, :lot, :tf, :s2, 0.4, 120.00, true)"
        ),
        {
            "a1": _id(),
            "a2": _id(),
            "sid": scenario_id,
            "lot": lot1,
            "tf": tf1,
            "dc1": dc1,
            "dc2": dc2,
            "s1": sup1,
            "s2": sup2,
        },
    )
    session.flush()

    return {
        "cycle_id": cycle_id,
        "run_id": run_id,
        "dc1": dc1,
        "dc2": dc2,
        "sup1": sup1,
        "sup2": sup2,
        "lot1": lot1,
        "tf1": tf1,
        "dc1_name": "Atlanta DC",
        "dc2_name": "Denver DC",
        "sup1_name": "Fresh Valley Farms",
        "sup2_name": "Sunrise Produce Co",
        "lot_name": "Strawberries Lot",
        "tf_name": "Spring Period",
    }


def test_freeze_then_versioned_adjustments(db_session: Session, tmp_path: Path) -> None:
    seed = _seed_cycle_and_run(db_session)

    # FREEZE — promote scenario B to a frozen award (the immutable baseline).
    award_id = freeze_award(
        db_session,
        cycle_id=seed["cycle_id"],
        analysis_run_id=seed["run_id"],
        scenario_code="B",
        award_code="AWD-TEST-01",
        frozen_by="tester",
    )
    assert isinstance(award_id, str)

    # Idempotent: re-freezing the same (cycle, run, scenario) returns the SAME award_id.
    award_id_again = freeze_award(
        db_session,
        cycle_id=seed["cycle_id"],
        analysis_run_id=seed["run_id"],
        scenario_code="B",
        award_code="AWD-TEST-01",
        frozen_by="tester",
    )
    assert award_id_again == award_id

    cell1 = (seed["dc1"], seed["lot1"], seed["tf1"], seed["sup1"])
    cell2 = (seed["dc2"], seed["lot1"], seed["tf1"], seed["sup2"])

    # Baseline effective prices = the frozen scenario-B awarded prices.
    base = effective_award(db_session, award_id=award_id)
    assert base[cell1] == _D("100.000000")
    assert base[cell2] == _D("120.000000")

    # v1 — a market-hike layer on cell1 (100 -> 108).
    v1 = add_adjustment(
        db_session,
        award_id=award_id,
        adjustment_type="MARKET_HIKE",
        effective_date=date(2026, 4, 1),
        reason="Trailing-4wk midpoint reset (rolling-midpoint safety)",
        created_by="tester",
        line_changes=[(seed["dc1"], seed["lot1"], seed["tf1"], seed["sup1"], _D("108.00"))],
    )
    assert v1 == 1

    # v2 — a second layer: cell1 again (108 -> 112) + cell2 (120 -> 118).
    v2 = add_adjustment(
        db_session,
        award_id=award_id,
        adjustment_type="TOLERANCE_BAND",
        effective_date=date(2026, 5, 1),
        reason="Sustained anomaly reprice within collar cap",
        created_by="tester",
        line_changes=[
            (seed["dc1"], seed["lot1"], seed["tf1"], seed["sup1"], _D("112.00")),
            (seed["dc2"], seed["lot1"], seed["tf1"], seed["sup2"], _D("118.00")),
        ],
    )
    assert v2 == 2  # version_no increments

    # effective_award at the latest (v2) overlays both layers.
    eff_v2 = effective_award(db_session, award_id=award_id)
    assert eff_v2[cell1] == _D("112.000000")
    assert eff_v2[cell2] == _D("118.000000")

    # as_of_version=1 reflects ONLY v1 (cell1 at 108; cell2 still baseline 120).
    eff_v1 = effective_award(db_session, award_id=award_id, as_of_version=1)
    assert eff_v1[cell1] == _D("108.000000")
    assert eff_v1[cell2] == _D("120.000000")

    # The v2 cell1 layer's prior_price is the v1 effective (108), not the frozen baseline (100).
    prior = db_session.execute(
        text(
            "SELECT l.prior_price, l.new_price, l.delta FROM awd.award_adjustment_line l "
            "JOIN awd.award_adjustment a ON a.adjustment_id = l.adjustment_id "
            "WHERE a.award_id = :aid AND a.version_no = 2 AND l.dc_id = :dc"
        ),
        {"aid": award_id, "dc": seed["dc1"]},
    ).one()
    assert Decimal(str(prior[0])) == _D("108.000000")
    assert Decimal(str(prior[1])) == _D("112.000000")
    assert Decimal(str(prior[2])) == _D("4.000000")

    # award_versions returns the full history v0 (baseline) -> v2.
    history = award_versions(db_session, award_id=award_id)
    versions = [h["version_no"] for h in history]
    assert versions == [0, 1, 2]
    assert history[0]["adjustment_type"] == "FROZEN"
    assert history[0]["n_lines"] == 2  # two baseline cells
    assert history[1]["adjustment_type"] == "MARKET_HIKE"
    assert history[1]["n_lines"] == 1
    assert history[2]["adjustment_type"] == "TOLERANCE_BAND"
    assert history[2]["n_lines"] == 2

    # ---- The versioned doc: explicit Version heading + names not keys. ----
    out_v2 = tmp_path / "post_award_v2.xlsx"
    write_post_award_adjustments_xlsx(db_session, award_id=award_id, output_path=out_v2)
    wb = load_workbook(out_v2)

    # The Version heading: banner title + the bold "Version N · as of <date>" subtitle (v2).
    versions_ws = wb["Versions"]
    title_cell = versions_ws.cell(row=1, column=1).value
    subtitle_cell = versions_ws.cell(row=2, column=1).value
    assert title_cell == "POST-AWARD ADJUSTMENTS — AWD-TEST-01"
    assert subtitle_cell == "Version 2 · as of 2026-05-01"

    # Names, not UUID keys, in the effective-prices tab.
    eff_ws = wb["Current Effective Prices"]
    all_text = {
        eff_ws.cell(row=r, column=c).value
        for r in range(1, eff_ws.max_row + 1)
        for c in range(1, eff_ws.max_column + 1)
        if isinstance(eff_ws.cell(row=r, column=c).value, str)
    }
    assert seed["dc1_name"] in all_text
    assert seed["sup1_name"] in all_text
    assert seed["lot_name"] in all_text
    assert seed["tf_name"] in all_text
    # No raw UUID key leaked into the readable cells.
    assert seed["dc1"] not in all_text
    assert seed["sup1"] not in all_text

    # ---- The doc at as_of_version=1 carries the v1 heading (a different historical version). ----
    out_v1 = tmp_path / "post_award_v1.xlsx"
    write_post_award_adjustments_xlsx(
        db_session, award_id=award_id, as_of_version=1, output_path=out_v1
    )
    wb1 = load_workbook(out_v1)
    assert wb1["Versions"].cell(row=2, column=1).value == "Version 1 · as of 2026-04-01"
    # This-version's-changes for v1 shows the single cell1 change (108 prior 100).
    chg_ws = wb1["This Version's Changes"]
    chg_text = {
        chg_ws.cell(row=r, column=c).value
        for r in range(1, chg_ws.max_row + 1)
        for c in range(1, chg_ws.max_column + 1)
    }
    assert seed["dc1_name"] in chg_text
    assert 108.0 in chg_text  # the new price for the v1 cell


def test_deterministic_doc(db_session: Session, tmp_path: Path) -> None:
    """The same award renders byte-stable content rows across two generations (deterministic)."""

    seed = _seed_cycle_and_run(db_session)
    award_id = freeze_award(
        db_session,
        cycle_id=seed["cycle_id"],
        analysis_run_id=seed["run_id"],
        scenario_code="B",
        award_code="AWD-DET-01",
        frozen_by="tester",
    )
    add_adjustment(
        db_session,
        award_id=award_id,
        adjustment_type="MARKET_HIKE",
        effective_date=date(2026, 4, 1),
        reason="layer one",
        created_by="tester",
        line_changes=[(seed["dc1"], seed["lot1"], seed["tf1"], seed["sup1"], _D("105.00"))],
    )

    def _body(path: Path) -> list[tuple[object, ...]]:
        write_post_award_adjustments_xlsx(db_session, award_id=award_id, output_path=path)
        ws = load_workbook(path)["Current Effective Prices"]
        return [
            tuple(ws.cell(row=r, column=c).value for c in range(1, 8))
            for r in range(6, ws.max_row + 1)
        ]

    assert _body(tmp_path / "a.xlsx") == _body(tmp_path / "b.xlsx")
