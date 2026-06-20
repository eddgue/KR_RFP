"""End-to-end "see it working" demo (D19 prototype fidelity) — the WHOLE loop on Postgres.

Runs the entire decision-support loop against a REAL local governed Postgres store with SYNTHETIC
data, then produces tangible, client-openable output files:

  seed (synthetic cycle: client/commodity, ~3 DCs, ~4 lots/items, 2 TFs, 3 rounds, ~6 suppliers,
        a strategy config, volumes, incumbents — every entity carries a clearly-SYNTHETIC but
        READABLE name/description so the outputs are legible, D23)
    -> generate the OWNED bid template for the cycle scope (intake generator; keys embedded, D21)
    -> simulate supplier returns (fill the template with varied synthetic bids across the rounds)
    -> ingest via the KEY-VALIDATED path (D21) -> bid.bid_line rows
    -> run the ENGINE RUNNER on the final round -> sealed eng.analysis_run + scores + scenarios
       + split awards (decision-support; never asserts an award, ADR-0006)
    -> generate the pre-award decision-support view FROM THE RECORDS:
         demo/output/RECOMMENDATION.md   — cycle + strategy + scenario comparison + per-cell split
                                           award recommendation (DC x lot x TF -> supplier(s) with
                                           volume_share %, awarded price, savings vs baseline)
    -> SIMULATE the human selecting Scenario B -> promote to an AWARD (a simple in-memory award
       selection for the demo; the real flow gates this through award -> freeze -> sign-off before
       any output is generated, D22) -> generate the FINAL post-award booking outputs FROM THE
       AWARD (D22 — booking guide is the LAST step, after awards, not straight off a scenario):
         demo/output/BOOKING_GUIDE_INTERNAL.xlsx  — the buyers/pricing master: awarded supplier
                                           (NAME) per DC x lot x item x TF with FOB/landed $/case,
                                           volume, routing — what pricing uses to update the system
         demo/output/SUPPLIER_AWARD_GUIDES.xlsx   — one sheet per awarded supplier: "here is what
                                           you've been awarded" (that supplier's lots/DCs/volumes/
                                           prices only)

D23 — human-facing outputs render RESOLVED NAMES, never key IDs. The keys JOIN (D21); the NAMES
DISPLAY. Every readable cell shows the seeded supplier/DC/lot/item/TF name resolved from the result
keys; a trailing "key" reference column is kept for traceability but the readable columns lead.

SYNTHETIC ONLY — every name is a clearly-fictional placeholder (e.g. "Green Valley Farms (DEMO)",
"Atlanta DC (ATL)") and every price is invented. This output is shown to a client, so it contains
NO real supplier names or prices. This script is pragmatic (it seeds via raw SQL for the FK-heavy
governed ref/cyc/norm spine) but stays clean; the runner it drives is the real service. Run with
`DATABASE_URL` pointed at a fresh DB that has had `alembic upgrade head`.
"""

from __future__ import annotations

import uuid
from collections import defaultdict
from dataclasses import dataclass
from datetime import UTC, date, datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.core.db.session import unit_of_work
from app.domain.bid.bid_ingester import Completeness, ingest_template
from app.domain.bid.models import BidLine
from app.domain.bid.template_generator import generate_template_bytes
from app.domain.bid.template_schema import (
    BODY_START_ROW,
    HEADER_ROW,
    SHEET_BIDS,
    BidColumn,
    CycleScope,
    ScopeRow,
)
from app.domain.eng.models import (
    AnalysisRun,
    AnalysisScenario,
    AnalysisScenarioAward,
)
from app.domain.eng.runner import EngineRunner, IncumbentRow
from app.engine.interface import EngineConfig, WeightPreset
from app.output.booking_guide import (
    write_booking_guide_internal_xlsx,
    write_supplier_award_guides_xlsx,
)
from app.output.scenario_workbook import write_scenario_workbook_xlsx
from app.output.synthetic import (
    DC_NAMES,
    REGION_FREIGHT,
    VEGCOOL_SURCHARGE_CASE,
    WEEKS_PER_TF,
    _dc_region,
)
from app.output.types import Entity, SeededCycle

OUTPUT_DIR = Path(__file__).resolve().parent / "output"

# --- Synthetic scope sizes (placeholders only — NO real data). ---
N_DCS = 3
N_LOTS = 4  # one item per lot (the engine's lot grain == one item)
N_TFS = 2
N_ROUNDS = 3
N_SUPPLIERS = 6

# --- Clearly-SYNTHETIC but READABLE names (D23). Every name is OBVIOUSLY fictional — flagged
#     (DEMO) on suppliers, parenthetical demo codes on DCs — so a reader can tell at a glance the
#     output is synthetic, while still reading a name instead of a key. NO real supplier names. ---
SUPPLIER_NAMES: tuple[str, ...] = (
    "Green Valley Farms (DEMO)",
    "Sunbelt Produce (DEMO)",
    "Harvest Ridge Growers (DEMO)",
    "Blue Sky Packing (DEMO)",
    "Cornerstone Fresh (DEMO)",
    "Riverbend Organics (DEMO)",
)
# Lot/item readable descriptions (a lot == one item at the engine grain) + a pack descriptor.
ITEM_DESCRIPTIONS: tuple[tuple[str, str], ...] = (
    ("Premium Grape Tomato 10oz", "10oz clamshell"),
    ("Roma Tomato Bulk 25lb", "25lb carton"),
    ("Vine-Ripe Tomato 4x5", "25lb 2-layer"),
    ("Cherry Tomato Pint", "12x1pt flat"),
)
LOT_NAMES: tuple[str, ...] = (
    "Lot 1 — Grape Tomato",
    "Lot 2 — Roma Bulk",
    "Lot 3 — Vine-Ripe",
    "Lot 4 — Cherry",
)
# Timeframe readable names (a season window with the fiscal periods it spans).
TF_NAMES: tuple[str, ...] = (
    "Spring 2026 (P4-P6)",
    "Summer 2026 (P7-P9)",
)
# Round readable labels (final = last).
ROUND_NAMES: tuple[str, ...] = (
    "Round 1 — Opening",
    "Round 2 — Negotiation",
    "Round 3 — Final",
)


def _id() -> str:
    return str(uuid.uuid4())


# ---------------------------------------------------------------------------
# 1) SEED — a synthetic cycle in the governed store (raw SQL; FK-heavy spine)
# ---------------------------------------------------------------------------
def seed_cycle(session: Session) -> SeededCycle:
    """Insert a synthetic cycle + its scope into the governed ref/cyc/norm tables."""

    now = datetime.now(UTC).replace(tzinfo=None)

    client_id = _id()
    commodity_id = _id()
    subcommodity_id = _id()
    cycle_id = _id()
    cycle_code = f"CYC-{now:%Y%m%d}-{cycle_id[:4].upper()}"
    cycle_name = "Field Tomatoes Sourcing Cycle — Spring/Summer 2026 (DEMO)"

    session.execute(
        text(
            "INSERT INTO ref.client (id, client_code, client_name, is_active) "
            "VALUES (gen_random_uuid(), :code, :name, true)"
        ),
        {"code": f"CLIENT-{cycle_id[:6].upper()}", "name": "Demo Sourcing Org (DEMO)"},
    )
    session.execute(
        text(
            "INSERT INTO ref.commodity (id, client_id, commodity_code, commodity_name) "
            "VALUES (:cid, NULL, :code, :name)"
        ),
        {"cid": commodity_id, "code": "COMM-DEMO", "name": "Field Tomatoes (DEMO)"},
    )
    # ref.commodity.id is uuid; store a varchar mirror for the cyc FK chain (text commodity_id).
    commodity_text_id = commodity_id  # the cyc spine uses text commodity_id keys
    session.execute(
        text(
            "INSERT INTO ref.subcommodity "
            "(subcommodity_id, commodity_id, subcommodity_code, subcommodity_name, active_flag) "
            "VALUES (:sid, :cid, :code, :name, true)"
        ),
        {
            "sid": subcommodity_id,
            "cid": commodity_text_id,
            "code": "SUBCOMM-DEMO",
            "name": "Field Tomatoes — Round/Vine (DEMO)",
        },
    )

    session.execute(
        text(
            "INSERT INTO cyc.cycle (cycle_id, cycle_code, cycle_name, commodity_id, "
            "subcommodity_id, status, why_now, target_effective_date, round_count, "
            "created_at, created_by) VALUES (:cyc, :code, :name, :cid, :sid, 'OPEN', "
            "'Synthetic demo cycle', :ted, :rc, :now, 'demo-seed')"
        ),
        {
            "cyc": cycle_id,
            "code": cycle_code,
            "name": cycle_name,
            "cid": commodity_text_id,
            "sid": subcommodity_id,
            "ted": date(now.year, 12, 31),
            "rc": N_ROUNDS,
            "now": now,
        },
    )

    # DCs (ref.dc) — readable demo names ("Atlanta DC (ATL)", ...). `code` is the short
    # demo region code (DC key reference); `name` is what every human-facing output renders (D23).
    dcs: list[Entity] = []
    for i in range(1, N_DCS + 1):
        dc_id = _id()
        dc_name, region_code = DC_NAMES[(i - 1) % len(DC_NAMES)]
        code = f"DC{i:02d}"
        session.execute(
            text(
                "INSERT INTO ref.dc (dc_id, dc_code, dc_name, region, division, active_flag) "
                "VALUES (:id, :code, :name, :region, 'Produce (DEMO)', true)"
            ),
            {"id": dc_id, "code": code, "name": dc_name, "region": region_code},
        )
        dcs.append(Entity(dc_id, code, dc_name))

    # Suppliers (ref.supplier) — clearly-fictional readable names ("Green Valley Farms (DEMO)").
    # `name` (canonical_name) is what every output renders; `code` keeps a short key reference.
    suppliers: list[Entity] = []
    for i in range(1, N_SUPPLIERS + 1):
        sup_id = _id()
        name = SUPPLIER_NAMES[(i - 1) % len(SUPPLIER_NAMES)]
        session.execute(
            text(
                "INSERT INTO ref.supplier (supplier_id, canonical_name, active_flag, created_at) "
                "VALUES (:id, :name, true, :now)"
            ),
            {"id": sup_id, "name": name, "now": now},
        )
        suppliers.append(Entity(sup_id, f"SUP-{i:02d}", name))

    # Items (ref.item) — one per lot. Readable descriptions ("Premium Grape Tomato 10oz").
    items: list[Entity] = []
    for i in range(1, N_LOTS + 1):
        item_id = _id()
        desc, pack = ITEM_DESCRIPTIONS[(i - 1) % len(ITEM_DESCRIPTIONS)]
        session.execute(
            text(
                "INSERT INTO ref.item (item_id, item_code, description, pack_desc, commodity_id, "
                "subcommodity_id) VALUES (:id, :code, :desc, :pack, :cid, :sid)"
            ),
            {
                "id": item_id,
                "code": f"ITEM-{i:02d}",
                "desc": desc,
                "pack": pack,
                "cid": commodity_text_id,
                "sid": subcommodity_id,
            },
        )
        items.append(Entity(item_id, f"ITEM-{i:02d}", desc))

    # Timeframes (cyc.cycle_timeframe) — readable season names ("Spring 2026 (P4-P6)").
    tfs: list[Entity] = []
    for i in range(1, N_TFS + 1):
        tf_id = _id()
        code = f"TF{i:02d}"
        tf_name = TF_NAMES[(i - 1) % len(TF_NAMES)]
        start = date(now.year, 1 + (i - 1) * 3, 1)
        end = date(now.year, 3 + (i - 1) * 3, 28)
        session.execute(
            text(
                "INSERT INTO cyc.cycle_timeframe (tf_id, cycle_id, tf_code, tf_name, "
                "start_date, end_date, week_count) VALUES (:id, :cyc, :code, :name, :s, :e, :w)"
            ),
            {
                "id": tf_id,
                "cyc": cycle_id,
                "code": code,
                "name": tf_name,
                "s": start,
                "e": end,
                "w": WEEKS_PER_TF,
            },
        )
        # `name` holds the readable season; `code` keeps the TF key reference (TF01/TF02).
        tfs.append(Entity(tf_id, code, tf_name))

    # Rounds (cyc.cycle_round) — readable labels ("Round 3 — Final"); final = last.
    rounds: list[Entity] = []
    for i in range(1, N_ROUNDS + 1):
        round_id = _id()
        session.execute(
            text(
                "INSERT INTO cyc.cycle_round (round_id, cycle_id, round_number, status, "
                "round_status, is_final) VALUES (:id, :cyc, :n, 'OPEN', 'OPEN', :final)"
            ),
            {"id": round_id, "cyc": cycle_id, "n": i, "final": i == N_ROUNDS},
        )
        # `code` is the R-token the runner/template use to JOIN; `name` is the readable label.
        rounds.append(Entity(round_id, f"R{i}", ROUND_NAMES[(i - 1) % len(ROUND_NAMES)]))

    # Lots (cyc.cycle_lot) + item scope + lot<->item link (one item per lot).
    lots: list[Entity] = []
    for i in range(1, N_LOTS + 1):
        lot_id = _id()
        code = f"LOT-{i:02d}"
        lot_name = LOT_NAMES[(i - 1) % len(LOT_NAMES)]
        item = items[i - 1]
        session.execute(
            text(
                "INSERT INTO cyc.cycle_lot (lot_id, cycle_id, lot_code, lot_name, active_flag) "
                "VALUES (:id, :cyc, :code, :name, true)"
            ),
            {"id": lot_id, "cyc": cycle_id, "code": code, "name": lot_name},
        )
        session.execute(
            text(
                "INSERT INTO cyc.cycle_item_scope (cycle_id, item_id, commodity_id, "
                "subcommodity_id, inclusion_status, added_at, added_by) "
                "VALUES (:cyc, :item, :cid, :sid, 'IN_SCOPE', :now, 'demo-seed')"
            ),
            {
                "cyc": cycle_id,
                "item": item.id,
                "cid": commodity_text_id,
                "sid": subcommodity_id,
                "now": now,
            },
        )
        session.execute(
            text(
                "INSERT INTO cyc.cycle_lot_item (lot_item_id, cycle_id, lot_id, item_id, "
                "required_flag, sort_order) VALUES (:lid, :cyc, :lot, :item, true, :so)"
            ),
            {"lid": _id(), "cyc": cycle_id, "lot": lot_id, "item": item.id, "so": i},
        )
        lots.append(Entity(lot_id, code, lot_name))

    # Invited suppliers (the submitted-vs-missing denominator).
    for sup in suppliers:
        session.execute(
            text(
                "INSERT INTO cyc.cycle_invited_supplier (cycle_id, supplier_id, invited_at, "
                "invited_by) VALUES (:cyc, :sup, :now, 'demo-seed')"
            ),
            {"cyc": cycle_id, "sup": sup.id, "now": now},
        )

    # Projected volumes (cyc.cycle_projected_volume) at DC x item x tf — synthetic cases.
    period_cases_by_cell: dict[tuple[str, str, str], Decimal] = {}
    for di, dc in enumerate(dcs):
        for li, item in enumerate(items):
            for ti, tf in enumerate(tfs):
                weekly = Decimal(400 + di * 50 + li * 30 + ti * 20)
                period = weekly * Decimal(WEEKS_PER_TF)
                session.execute(
                    text(
                        "INSERT INTO cyc.cycle_projected_volume (volume_id, cycle_id, dc_id, "
                        "item_id, tf_id, volume_input_method, projected_weekly_cases, "
                        "projected_period_cases) VALUES (:id, :cyc, :dc, :item, :tf, "
                        "'WEEKLY_X_WEEKS', :wk, :pd)"
                    ),
                    {
                        "id": _id(),
                        "cyc": cycle_id,
                        "dc": dc.id,
                        "item": item.id,
                        "tf": tf.id,
                        "wk": weekly,
                        "pd": period,
                    },
                )
                # cell key uses lot_id (one lot per item) for the engine's lot grain.
                lot = lots[li]
                period_cases_by_cell[(dc.id, lot.id, tf.id)] = period

    # Incumbents (perf.historical_award_assignment) + a routing baseline per (dc, lot).
    # The incumbent is SUP-01 on each (dc, lot); routing baseline ~ the demo's mid price.
    incumbent_by_dc_lot: dict[tuple[str, str], str] = {}
    incumbent_routing: dict[tuple[str, str], Decimal] = {}
    inc_run_id = _id()
    session.execute(
        text(
            "INSERT INTO norm.normalization_run (normalization_run_id, dataset_type, cycle_id, "
            "status) VALUES (:id, 'HISTORICAL_AWARD', :cyc, 'APPROVED')"
        ),
        {"id": inc_run_id, "cyc": cycle_id},
    )
    incumbent = suppliers[0]
    for di, dc in enumerate(dcs):
        for li, item in enumerate(items):
            lot = lots[li]
            # Incumbent routing = prior-period actual-paid (the iTrade baseline, D11). Modelled as
            # the incumbent's own final-round bid + a margin the RFP captures (~7%), so the cycle
            # shows realistic SAVINGS vs what we paid last period (the headline buyers sign).
            routing = (_synthetic_price(N_ROUNDS - 1, di, li, 0) * Decimal("1.07")).quantize(
                Decimal("0.01")
            )
            assignment_id = _id()
            session.execute(
                text(
                    "INSERT INTO perf.historical_award_assignment (assignment_id, cycle_id, "
                    "dc_id, item_id, supplier_id, effective_start_date, effective_end_date, "
                    "awarded_volume_cases, ingestion_run_id, incumbent_flag, created_at, "
                    "created_by) VALUES (:id, :cyc, :dc, :item, :sup, :s, :e, :vol, :run, "
                    "true, :now, 'demo-seed')"
                ),
                {
                    "id": assignment_id,
                    "cyc": cycle_id,
                    "dc": dc.id,
                    "item": item.id,
                    "sup": incumbent.id,
                    "s": date(now.year - 1, 1, 1),
                    "e": date(now.year - 1, 12, 31),
                    "vol": Decimal("100000"),
                    "run": inc_run_id,
                    "now": now,
                },
            )
            # Persist the routing baseline as the preferred price-basis row
            # (perf.historical_awarded_price_basis) — the governed home of the iTrade actual-paid
            # baseline (D11). `app.cycle.loader.load_cycle` reads this back as `incumbent_routing`
            # for a REAL cycle, so the generator runs identically whether the CycleView is seeded
            # in-memory or reconstructed from the records.
            session.execute(
                text(
                    "INSERT INTO perf.historical_awarded_price_basis (price_basis_id, "
                    "assignment_id, routing_basis, awarded_price_per_case, preferred_basis_flag, "
                    "preferred_basis_source, created_at) VALUES (:id, :aid, 'DELIVERED', :price, "
                    "true, 'demo-seed', :now)"
                ),
                {"id": _id(), "aid": assignment_id, "price": routing, "now": now},
            )
            incumbent_by_dc_lot[(dc.id, lot.id)] = incumbent.id
            incumbent_routing[(dc.id, lot.id)] = routing

    session.flush()
    return SeededCycle(
        cycle_id=cycle_id,
        cycle_code=cycle_code,
        cycle_name=cycle_name,
        client_id=client_id,
        commodity_id=commodity_text_id,
        dcs=dcs,
        lots=lots,
        items=items,
        tfs=tfs,
        rounds=rounds,
        suppliers=suppliers,
        incumbent_by_dc_lot=incumbent_by_dc_lot,
        incumbent_routing=incumbent_routing,
        period_cases_by_cell=period_cases_by_cell,
    )


# ---------------------------------------------------------------------------
# 2) GENERATE the owned bid template for the cycle scope (intake generator, D21)
# ---------------------------------------------------------------------------
def build_scope(seeded: SeededCycle, round_entity: Entity) -> CycleScope:
    """Build the intake CycleScope (embedded keys) for ONE round across all cells x suppliers."""

    rows: list[ScopeRow] = []
    for dc in seeded.dcs:
        for li, item in enumerate(seeded.items):
            lot = seeded.lots[li]
            for tf in seeded.tfs:
                for sup in seeded.suppliers:
                    rows.append(
                        ScopeRow(
                            round_code=round_entity.code,
                            bid_type="STANDARD",
                            round_id=round_entity.id,
                            tf_id=tf.id,
                            supplier_id=sup.id,
                            dc_id=dc.id,
                            lot_id=lot.id,
                            item_id=item.id,
                            supplier_label=sup.name,
                            dc_label=dc.name,
                            lot_label=lot.name,
                            item_label=item.name,
                            tf_code=tf.code,
                        )
                    )
    return CycleScope(
        cycle_id=seeded.cycle_id,
        cycle_code=seeded.cycle_code,
        cycle_name=seeded.cycle_name,
        window_label=f"{round_entity.code} window (SYNTHETIC)",
        rows=tuple(rows),
    )


# ---------------------------------------------------------------------------
# 3) SIMULATE supplier returns — fill the generated template with varied prices
# ---------------------------------------------------------------------------
def _synthetic_price(round_idx: int, dc_idx: int, lot_idx: int, sup_idx: int) -> Decimal:
    """A deterministic, varied synthetic All-In $/case (placeholder economics only).

    Tuned so a real DC-level split emerges: different suppliers are strongest on different lots, so
    within one DC the engine's max-2-per-DC lens awards two DIFFERENT suppliers across the lots (the
    V3 split semantic). `_lot_specialist(lot)` is each lot's keenest supplier; others fan upward.
    Prices also drift down slightly each round (competitive tension across the 3 rounds).
    """

    base = Decimal("10.00") + Decimal(lot_idx) * Decimal("0.50") + Decimal(dc_idx) * Decimal("0.20")
    # Each lot has a "specialist" supplier (rotates by lot) who bids keenest on that lot.
    specialist = _lot_specialist(lot_idx)
    distance = abs(sup_idx - specialist)
    spread = Decimal(distance) * Decimal("0.30")  # keenest at the specialist, fanning upward
    # Round-over-round CONCESSION varies by supplier (the negotiation/fairness lens, pillar 4):
    # the incumbent (idx 0) HOLDS the installed base (barely moves); challengers concede more,
    # the hungriest most. This asymmetry is what "are you being treated fairly?" reads.
    if sup_idx == 0:
        drift_rate = Decimal("0.04")  # incumbent leans on tenure — small concession
    else:
        drift_rate = Decimal("0.16") + Decimal(sup_idx) * Decimal("0.02")  # challengers move more
    round_drift = Decimal(round_idx) * drift_rate
    price = base + spread - round_drift
    return price.quantize(Decimal("0.01"))


def _lot_specialist(lot_idx: int) -> int:
    """The supplier index bidding keenest on a lot (rotates so DCs split across suppliers)."""

    # Alternate the keenest supplier between SUP-01 (idx 0) and SUP-02 (idx 1) by lot so each DC's
    # lots are won by two different suppliers -> a genuine 2-supplier DC split in scenarios B/D.
    rotation = (0, 1, 0, 1, 2, 1)
    return rotation[lot_idx % len(rotation)]


def fill_template(template_bytes: bytes, scope: SeededCycle, round_idx: int) -> bytes:
    """Open the generated template and write synthetic All-In + volume cells for each row."""

    wb = load_workbook(BytesIO(template_bytes))
    ws = wb[SHEET_BIDS]
    headers = _header_map(ws)

    dc_idx = {dc.id: i for i, dc in enumerate(scope.dcs)}
    lot_idx = {lot.id: i for i, lot in enumerate(scope.lots)}
    sup_idx = {sup.id: i for i, sup in enumerate(scope.suppliers)}

    all_in_col = headers[BidColumn.ALL_IN.value]
    fob_col = headers[BidColumn.FOB.value]
    deliv_col = headers[BidColumn.DELIVERY_SURCHARGE.value]
    vegcool_col = headers[BidColumn.VEGCOOL_SURCHARGE.value]
    weekly_col = headers[BidColumn.WEEKLY_VOL_OFFERED.value]
    total_col = headers[BidColumn.TOTAL_VOL_OFFERED.value]
    dc_id_col = headers[BidColumn.DC_ID.value]
    lot_id_col = headers[BidColumn.LOT_ID.value]
    sup_id_col = headers[BidColumn.SUPPLIER_ID.value]

    for row in range(BODY_START_ROW, ws.max_row + 1):
        dc_id = _cell_str(ws, row, dc_id_col)
        lot_id = _cell_str(ws, row, lot_id_col)
        sup_id = _cell_str(ws, row, sup_id_col)
        if not (dc_id and lot_id and sup_id):
            continue
        # A couple of suppliers decline a couple of cells (No-Bid: leave all price cells blank).
        si = sup_idx.get(sup_id, 0)
        if si >= 5 and lot_idx.get(lot_id, 0) == (N_LOTS - 1):
            continue  # SUP-06 declines the last lot -> a genuine No-Bid row
        di = dc_idx.get(dc_id, 0)
        price = _synthetic_price(round_idx, di, lot_idx.get(lot_id, 0), si)
        # Decompose the landed All-In into FOB (farm-gate) + Delivery (lane freight, by region) +
        # VegCool (cold-chain). All-In stays the value the engine scores (§7 primary path); the
        # components ride along for the FOB-vs-All-In freight view. NO Lot Discount -> the
        # `ck_bid_line_no_double_discount` guard is satisfied (All-In present, discount 0).
        delivery = REGION_FREIGHT[_dc_region(di)]
        vegcool = VEGCOOL_SURCHARGE_CASE
        fob = (price - delivery - vegcool).quantize(Decimal("0.01"))
        # Offer full coverage (weekly ~ the demand band) so coverage gates pass.
        weekly = Decimal(600)
        total = weekly * Decimal(WEEKS_PER_TF)
        ws.cell(row=row, column=all_in_col, value=float(price))
        ws.cell(row=row, column=fob_col, value=float(fob))
        ws.cell(row=row, column=deliv_col, value=float(delivery))
        ws.cell(row=row, column=vegcool_col, value=float(vegcool))
        ws.cell(row=row, column=weekly_col, value=float(weekly))
        ws.cell(row=row, column=total_col, value=float(total))

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# 4) INGEST (key-validated, D21) -> persist bid.bid_line rows
# ---------------------------------------------------------------------------
def ingest_and_persist(
    session: Session,
    filled_bytes: bytes,
    scope: CycleScope,
    seeded: SeededCycle,
    round_entity: Entity,
) -> int:
    """Ingest OUR returned template (key-validated) and write bid.bid_line rows. Returns count."""

    now = datetime.now(UTC).replace(tzinfo=None)

    result = ingest_template(filled_bytes, scope)
    if result.quarantined:
        # Surface but do not crash — a robust ingest tolerates declines/quarantine.
        print(f"   ingest quarantined {len(result.quarantined)} row(s)")

    # One submission per supplier for this round (FK chain: source_artifact -> bid_submission).
    submission_by_sup: dict[str, str] = {}
    for sup in seeded.suppliers:
        artifact_id = _id()
        session.execute(
            text(
                "INSERT INTO norm.source_artifact (artifact_id, artifact_type, file_name, "
                "file_hash_sha256, received_at, status, cycle_id, round_id, supplier_id, "
                "created_by) VALUES (:aid, 'BID_SUBMISSION', :fn, :hash, :now, 'RECEIVED', "
                ":cyc, :rnd, :sup, 'demo-seed')"
            ),
            {
                "aid": artifact_id,
                "fn": f"bid_{sup.code}_{round_entity.code}.xlsx",
                "hash": _id().replace("-", "")[:64].ljust(64, "0"),
                "now": now,
                "cyc": seeded.cycle_id,
                "rnd": round_entity.id,
                "sup": sup.id,
            },
        )
        submission_id = _id()
        session.execute(
            text(
                "INSERT INTO bid.bid_submission (submission_id, cycle_id, round_id, supplier_id, "
                "source_artifact_id, submitted_at, version_number, overall_status, "
                "standard_terms_accepted) VALUES (:sid, :cyc, :rnd, :sup, :aid, :now, 1, "
                "'SUBMITTED', true)"
            ),
            {
                "sid": submission_id,
                "cyc": seeded.cycle_id,
                "rnd": round_entity.id,
                "sup": sup.id,
                "aid": artifact_id,
                "now": now,
            },
        )
        submission_by_sup[sup.id] = submission_id

    count = 0
    for line in result.lines:
        if line.completeness is not Completeness.BID:
            continue  # only persist priced lines (no_bid / incomplete are not scoreable)
        ident = line.identity
        bid_line = BidLine(
            bid_line_id=_id(),
            submission_id=submission_by_sup[ident.supplier_id],
            cycle_id=seeded.cycle_id,
            round_id=round_entity.id,
            supplier_id=ident.supplier_id,
            dc_id=ident.dc_id,
            lot_id=ident.lot_id,
            item_id=ident.item_id,
            tf_id=ident.tf_id,
            currency_code="USD",
            price_basis=line.price_basis or "ALL_IN",
            submitted_all_in_case=line.components.all_in,
            fob_case=line.components.fob,
            delivery_surcharge_case=line.components.delivery_surcharge or None,
            vegcool_surcharge_case=line.components.vegcool_surcharge or None,
            lot_discount_case=line.components.lot_discount or None,
            price_basis_resolved=line.price_basis or None,
            volume_minimum_cases=line.total_vol_offered,
            exclusivity_required_flag=False,
            validity_status="VALID",
            source_row_number=line.source_row_number,
            created_at=now,
            is_scoreable=True,
            is_awardable=True,
        )
        session.add(bid_line)
        count += 1
    session.flush()
    return count


# ---------------------------------------------------------------------------
# small openpyxl helpers
# ---------------------------------------------------------------------------
def _header_map(ws: Worksheet) -> dict[str, int]:
    out: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=HEADER_ROW, column=col).value
        if value is not None:
            out[str(value).strip()] = col
    return out


def _cell_str(ws: Worksheet, row: int, col: int) -> str:
    value = ws.cell(row=row, column=col).value
    return "" if value is None else str(value).strip()


# ---------------------------------------------------------------------------
# 6) GENERATE human-readable output FROM THE RECORDS
# ---------------------------------------------------------------------------
def write_recommendation_md(
    session: Session,
    seeded: SeededCycle,
    analysis_run_id: str,
    config: EngineConfig,
) -> Path:
    """Render demo/output/RECOMMENDATION.md purely from the sealed eng.* records."""

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    run = session.get(AnalysisRun, analysis_run_id)
    assert run is not None  # noqa: S101

    # Display maps (id -> RESOLVED READABLE NAME — D23; the readable columns lead).
    dc_name = {dc.id: dc.name for dc in seeded.dcs}
    lot_name = {lot.id: lot.name for lot in seeded.lots}
    sup_name = {sup.id: sup.name for sup in seeded.suppliers}
    tf_name = {tf.id: tf.name for tf in seeded.tfs}
    item_name_for_lot = {seeded.lots[i].id: seeded.items[i].name for i in range(len(seeded.lots))}
    # Compact KEY-reference codes (id -> short code) — a trailing traceability column, not the lead.
    dc_code = {dc.id: dc.code for dc in seeded.dcs}
    lot_code = {lot.id: lot.code for lot in seeded.lots}
    sup_code = {sup.id: sup.code for sup in seeded.suppliers}
    lot_routing_avg = {
        lot.id: (
            sum(
                (seeded.incumbent_routing[(dc.id, lot.id)] for dc in seeded.dcs),
                Decimal("0"),
            )
            / Decimal(len(seeded.dcs))
        )
        for lot in seeded.lots
    }

    scenarios = (
        session.query(AnalysisScenario)
        .filter(AnalysisScenario.analysis_run_id == analysis_run_id)
        .order_by(AnalysisScenario.scenario_code)
        .all()
    )
    scen_by_code = {s.scenario_code: s for s in scenarios}
    scen_id_by_code = {s.scenario_code: s.analysis_scenario_id for s in scenarios}

    def awards_for(code: str) -> list[AnalysisScenarioAward]:
        sid = scen_id_by_code.get(code)
        if sid is None:
            return []
        return (
            session.query(AnalysisScenarioAward)
            .filter(AnalysisScenarioAward.analysis_scenario_id == sid)
            .order_by(
                AnalysisScenarioAward.dc_id,
                AnalysisScenarioAward.lot_id,
                AnalysisScenarioAward.tf_id,
            )
            .all()
        )

    spend_a = scen_by_code["A"].objective_total_spend if "A" in scen_by_code else None
    spend_b = scen_by_code["B"].objective_total_spend if "B" in scen_by_code else None
    delta_pct = None
    if spend_a and spend_b and spend_a > 0:
        delta_pct = (spend_b - spend_a) / spend_a * Decimal("100")

    lines: list[str] = []
    lines.append("# Sourcing Recommendation (DECISION-SUPPORT)\n")
    lines.append(
        "> This is the **pre-award** decision-support view. It **recommends**; it does not assert "
        "an award. A human reviewer selects a scenario, which is then promoted to an award, frozen "
        "and signed off before any booking output is generated (ADR-0006, D22). Every supplier / "
        "DC / lot / item / timeframe below is shown by its **resolved NAME** (D23 — keys join, "
        "names display). All names and prices are **clearly-fictional SYNTHETIC placeholders** "
        '(e.g. "Green Valley Farms (DEMO)", "Atlanta DC (ATL)") — no real suppliers/prices.\n'
    )
    lines.append("## Cycle\n")
    lines.append(f"- **Cycle:** {seeded.cycle_code} — {seeded.cycle_name}")
    lines.append(f"- **Round analysed:** {run.round_id[:8]}… (final round)")
    lines.append(
        f"- **Scope:** {len(seeded.dcs)} DCs x {len(seeded.lots)} lots x {len(seeded.tfs)} "
        f"timeframes; {len(seeded.suppliers)} invited suppliers"
    )
    lines.append(f"- **Engine:** `{run.engine_version}` (sealed run `{run.analysis_run_id[:8]}…`)")
    lines.append(
        f"- **Sealed manifest:** input `sha256:{run.input_hash_manifest[:16]}…`, "
        f"output `sha256:{run.output_hash_manifest[:16]}…`\n"
    )

    lines.append("## Strategy (config-driven — D18 strategy-agnostic)\n")
    lines.append(f"- **Weights preset:** {config.preset.value}")
    lines.append(
        f"- **Five-factor weights:** price {config.weight_price}, coverage "
        f"{config.weight_coverage}, historical {config.weight_historical}, z-risk "
        f"{config.weight_zrisk}, continuity {config.weight_continuity}"
    )
    lines.append(f"- **Max suppliers per DC (split cap):** {config.max_sup_dc}")
    lines.append(
        f"- **Thresholds:** premium ceiling {config.global_premium_threshold}, coverage floor "
        f"{config.coverage_floor}, concentration {config.conc_thresh}\n"
    )

    lines.append("## Scenario comparison (the lenses A–G)\n")
    lines.append("| Lens | Label | Objective spend (synthetic) |")
    lines.append("| --- | --- | --- |")
    for s in scenarios:
        spend = f"${s.objective_total_spend:,.2f}" if s.objective_total_spend is not None else "—"
        lines.append(f"| {s.scenario_code} | {s.label} | {spend} |")
    lines.append("")
    if delta_pct is not None:
        direction = "above" if delta_pct >= 0 else "below"
        lines.append(
            f"**Headline:** Scenario **B (risk-adjusted recommendation)** lands "
            f"**{abs(delta_pct):.2f}% {direction}** the Scenario **A (lowest-cost benchmark)** "
            f"spend — the risk-adjusted premium the recommendation trades for coverage/continuity. "
            f"Scenario A is a benchmark only and is never auto-applied.\n"
        )

    lines.append("## Recommended split award — per DC × lot × timeframe (Scenario B)\n")
    lines.append(
        "Each cell below is a **recommendation**. `volume_share` is the proposed split fraction; "
        "savings are vs. the incumbent routing baseline. Flags surface (they never auto-reject).\n"
    )
    lines.append(
        "| DC | Lot | Item | TF | Recommended supplier(s) | Volume share | Awarded $/case | "
        "Savings vs baseline | Flags | Key ref (DC·lot·sup) |"
    )
    lines.append("| --- | --- | --- | --- | --- | --- | --- | --- | --- | --- |")
    b_awards = awards_for("B")
    # Group by cell to show the split shares together.
    by_cell: dict[tuple[str, str, str], list[AnalysisScenarioAward]] = defaultdict(list)
    for a in b_awards:
        by_cell[(a.dc_id, a.lot_id, a.tf_id)].append(a)
    for (dc_id, lot_id, tf_id), cell_awards in by_cell.items():
        for a in cell_awards:
            baseline = lot_routing_avg.get(lot_id, Decimal("0"))
            savings = (
                (baseline - a.awarded_price) / baseline * Decimal("100")
                if baseline > 0
                else Decimal("0")
            )
            flags = []
            if a.cap_breach_flag:
                flags.append("CAP-BREACH")
            if a.is_fallback:
                flags.append("FALLBACK")
            flag_text = ", ".join(flags) if flags else "—"
            # NAMES lead (D23); a compact code reference trails for traceability.
            sup_disp = sup_name.get(a.supplier_id, a.supplier_id[:6])
            key_ref = (
                f"{dc_code.get(dc_id, dc_id[:6])}·{lot_code.get(lot_id, lot_id[:6])}·"
                f"{sup_code.get(a.supplier_id, a.supplier_id[:6])}"
            )
            lines.append(
                f"| {dc_name.get(dc_id, dc_id[:6])} | {lot_name.get(lot_id, lot_id[:6])} | "
                f"{item_name_for_lot.get(lot_id, '')} | {tf_name.get(tf_id, tf_id[:6])} | "
                f"{sup_disp} | {a.volume_share * 100:.0f}% | ${a.awarded_price:,.2f} | "
                f"{savings:+.1f}% | {flag_text} | {key_ref} |"
            )
    lines.append("")

    # Per-DC supplier split (the real V3 split semantic: a DC's lots spread across its top-N
    # suppliers). Volume-weighted share per supplier within each DC, from the Scenario B awards.
    lines.append("## DC-level supplier split (Scenario B) — volume-weighted share per DC\n")
    lines.append(
        "The recommendation spreads each DC's volume across suppliers (max "
        f"{config.max_sup_dc} per DC). Shares below are **volume-weighted** across the DC's "
        "lots/timeframes — the decision-support split a buyer reviews:\n"
    )
    lines.append("| DC | Supplier | Volume share of DC | Awarded lots |")
    lines.append("| --- | --- | --- | --- |")
    dc_split = _dc_supplier_split(b_awards, seeded.period_cases_by_cell)
    for dc_id in sorted(dc_split, key=lambda d: dc_name.get(d, d)):
        sup_shares = dc_split[dc_id]
        total = sum(v for _s, (v, _n) in sup_shares.items()) or Decimal("1")
        ordered = sorted(sup_shares.items(), key=lambda kv: -kv[1][0])
        for sup_id, (vol, lots_n) in ordered:
            share = vol / total * Decimal("100")
            lines.append(
                f"| {dc_name.get(dc_id, dc_id[:6])} | {sup_name.get(sup_id, sup_id[:6])} | "
                f"{share:.0f}% | {lots_n} |"
            )
    lines.append("")

    # Scenario D split highlight (the explicit max-N-per-DC split lens).
    d_awards = awards_for("D")
    split_cells = _split_cells(d_awards)
    if split_cells:
        lines.append("## Split highlight — Scenario D (max-N per DC)\n")
        lines.append(
            "Cells where the **max-N-per-DC** lens spreads volume across more than one supplier "
            "(decision-support; `FALLBACK` marks a lot filled from the wider field):\n"
        )
        lines.append("| DC | Lot | TF | Suppliers (share) | Fallback |")
        lines.append("| --- | --- | --- | --- | --- |")
        for (dc_id, lot_id, tf_id), cell_awards in split_cells.items():
            sup_text = ", ".join(
                f"{sup_name.get(a.supplier_id, a.supplier_id[:6])} ({a.volume_share * 100:.0f}%)"
                for a in cell_awards
            )
            fb = "yes" if any(a.is_fallback for a in cell_awards) else "no"
            lines.append(
                f"| {dc_name.get(dc_id, dc_id[:6])} | {lot_name.get(lot_id, lot_id[:6])} | "
                f"{tf_name.get(tf_id, tf_id[:6])} | {sup_text} | {fb} |"
            )
        lines.append("")

    lines.append("---\n")
    lines.append(
        "_Generated from the sealed `eng.analysis_run` records on the governed Postgres store. "
        "Decision-support only — recommends, does not assert (ADR-0006)._\n"
    )

    path = OUTPUT_DIR / "RECOMMENDATION.md"
    path.write_text("\n".join(lines), encoding="utf-8")
    return path


def _split_cells(
    awards: list[AnalysisScenarioAward],
) -> dict[tuple[str, str, str], list[AnalysisScenarioAward]]:
    """Cells (dc, lot, tf) under a lens with more than one supplier award (a real split)."""

    by_cell: dict[tuple[str, str, str], list[AnalysisScenarioAward]] = defaultdict(list)
    for a in awards:
        by_cell[(a.dc_id, a.lot_id, a.tf_id)].append(a)
    return {cell: rows for cell, rows in by_cell.items() if len(rows) > 1}


def _dc_supplier_split(
    awards: list[AnalysisScenarioAward],
    period_cases_by_cell: dict[tuple[str, str, str], Decimal],
) -> dict[str, dict[str, tuple[Decimal, int]]]:
    """Per DC: supplier -> (volume-weighted cases awarded, distinct lot count) across its cells.

    This is the real V3 split view — how a DC's volume spreads across its (up to max_sup_dc)
    suppliers, the decision-support figure a buyer reviews. Volume per cell = projected cases x the
    cell's volume_share (1.0 for the winner-take-cell B lens).
    """

    vol_by: dict[str, dict[str, Decimal]] = defaultdict(lambda: defaultdict(lambda: Decimal("0")))
    lots_by: dict[str, dict[str, set[str]]] = defaultdict(lambda: defaultdict(set))
    for a in awards:
        cases = period_cases_by_cell.get((a.dc_id, a.lot_id, a.tf_id), Decimal("0"))
        vol_by[a.dc_id][a.supplier_id] += cases * a.volume_share
        lots_by[a.dc_id][a.supplier_id].add(a.lot_id)
    return {
        dc_id: {sup_id: (vol, len(lots_by[dc_id][sup_id])) for sup_id, vol in sups.items()}
        for dc_id, sups in vol_by.items()
    }


# ---------------------------------------------------------------------------
# 7) AWARD SELECTION (D22) — simulate the human selecting Scenario B and promoting it to an
#    award. The real flow routes scenario -> human selects -> awd.award -> FREEZE -> SIGN-OFF
#    -> booking guide. The `awd.*` tables are a later phase, so the demo promotes the selected
#    scenario's award rows into a simple in-memory AwardedCell set here (clearly noting where the
#    freeze/sign-off gates would sit) and generates the booking outputs FROM THIS AWARD — never
#    straight off the scenario.
# ---------------------------------------------------------------------------
@dataclass(frozen=True)
class AwardedCell:
    """One frozen award line: a supplier's awarded share of a (dc, lot, item, tf) cell.

    The in-memory stand-in for an `awd.award` row (the demo defers the `awd.*` spine). Carries the
    JOIN keys AND the resolved display names (D23) so both booking guides render names off one
    award, not the raw scenario keys.
    """

    dc_id: str
    lot_id: str
    item_id: str
    tf_id: str
    supplier_id: str
    volume_share: Decimal
    awarded_price: Decimal
    period_cases: Decimal
    routing_baseline: Decimal


@dataclass(frozen=True)
class SelectedAward:
    """The promoted award: the selected scenario + its frozen awarded cells (the booking basis)."""

    scenario_code: str
    scenario_label: str
    cells: tuple[AwardedCell, ...]


def select_award_from_scenario(
    session: Session,
    seeded: SeededCycle,
    analysis_run_id: str,
    selected_scenario_code: str = "B",
) -> SelectedAward:
    """Simulate the human selecting a scenario and promoting it to a (frozen) award.

    Real sequence (D22): the engine writes decision-support scenarios; a human SELECTS one
    (here: Scenario B, the risk-adjusted recommendation); selection promotes it to `awd.award`;
    the award is then FROZEN and SIGNED OFF; only then are the booking outputs generated. The
    `awd.*` spine lands in a later phase, so this returns an in-memory frozen award assembled from
    the selected scenario's `eng.analysis_scenario_award` rows, with the keys resolved to the
    seeded item + routing baseline. >>> FREEZE + SIGN-OFF GATES WOULD SIT HERE <<< before any
    output is generated; the demo notes the gate rather than enforcing it.
    """

    item_for_lot = {seeded.lots[i].id: seeded.items[i] for i in range(len(seeded.lots))}

    scen = (
        session.query(AnalysisScenario)
        .filter(
            AnalysisScenario.analysis_run_id == analysis_run_id,
            AnalysisScenario.scenario_code == selected_scenario_code,
        )
        .one()
    )
    award_rows = (
        session.query(AnalysisScenarioAward)
        .filter(AnalysisScenarioAward.analysis_scenario_id == scen.analysis_scenario_id)
        .order_by(
            AnalysisScenarioAward.dc_id,
            AnalysisScenarioAward.lot_id,
            AnalysisScenarioAward.tf_id,
        )
        .all()
    )

    cells: list[AwardedCell] = []
    for a in award_rows:
        item = item_for_lot.get(a.lot_id)
        cells.append(
            AwardedCell(
                dc_id=a.dc_id,
                lot_id=a.lot_id,
                item_id=item.id if item else a.lot_id,
                tf_id=a.tf_id,
                supplier_id=a.supplier_id,
                volume_share=a.volume_share,
                awarded_price=a.awarded_price,
                period_cases=seeded.period_cases_by_cell.get(
                    (a.dc_id, a.lot_id, a.tf_id), Decimal("0")
                ),
                routing_baseline=seeded.incumbent_routing.get((a.dc_id, a.lot_id), Decimal("0")),
            )
        )
    return SelectedAward(
        scenario_code=scen.scenario_code,
        scenario_label=scen.label,
        cells=tuple(cells),
    )


# ---------------------------------------------------------------------------
# Orchestration — the whole loop
# ---------------------------------------------------------------------------
def main() -> None:
    print("=== KR RFP — end-to-end cycle demo (SYNTHETIC) ===")
    config = EngineConfig(
        preset=WeightPreset.BALANCED,
        weight_price=Decimal("0.35"),
        weight_coverage=Decimal("0.25"),
        weight_historical=Decimal("0.20"),
        weight_zrisk=Decimal("0.10"),
        weight_continuity=Decimal("0.10"),
        max_sup_dc=2,
        conc_thresh=Decimal("0.40"),
        global_premium_threshold=Decimal("0.12"),
        coverage_floor=Decimal("0.80"),
    )

    with unit_of_work() as session:
        print(
            "[1/8] Seeding synthetic cycle (client/commodity, DCs, lots, items, TFs, rounds, "
            "suppliers, volumes, incumbents — readable DEMO names, D23)…"
        )
        seeded = seed_cycle(session)
        print(
            f"   cycle {seeded.cycle_code}: {len(seeded.dcs)} DCs, {len(seeded.lots)} lots, "
            f"{len(seeded.tfs)} TFs, {len(seeded.rounds)} rounds, {len(seeded.suppliers)} suppliers"
        )

        total_lines = 0
        for round_idx, round_entity in enumerate(seeded.rounds):
            scope = build_scope(seeded, round_entity)
            template_bytes = generate_template_bytes(scope)
            if round_idx == 0:
                print(
                    f"[2/8] Generated owned bid template for {round_entity.code} "
                    f"({len(scope.rows)} scope rows, keys embedded — D21)"
                )
            filled = fill_template(template_bytes, seeded, round_idx)
            n = ingest_and_persist(session, filled, scope, seeded, round_entity)
            total_lines += n
            print(f"[3/8] {round_entity.code}: ingested (key-validated) -> {n} bid.bid_line rows")
        print(f"   total bid_line rows across {len(seeded.rounds)} rounds: {total_lines}")

        final_round = seeded.rounds[-1]
        incumbents = tuple(
            IncumbentRow(
                dc_id=dc_id,
                lot_id=lot_id,
                supplier_id=sup_id,
                routing_cost_per_case=seeded.incumbent_routing[(dc_id, lot_id)],
            )
            for (dc_id, lot_id), sup_id in seeded.incumbent_by_dc_lot.items()
        )

        print(
            f"[4/8] Running engine runner on final round {final_round.code} "
            f"(read-by-key -> assemble -> V3Engine.run -> seal)…"
        )
        runner = EngineRunner(session)
        run_result = runner.run_analysis(
            cycle_id=seeded.cycle_id,
            round_id=final_round.id,
            config=config,
            incumbents=incumbents,
            run_by="demo-runner",
        )
        print(
            f"   sealed run {run_result.analysis_run_id[:8]}… — {run_result.score_count} scores, "
            f"{run_result.scenario_count} scenarios, {run_result.award_count} split award rows"
        )
        print(f"   input  manifest sha256: {run_result.input_hash[:24]}…")
        print(f"   output manifest sha256: {run_result.output_hash[:24]}…")

        print("[5/8] Generating RECOMMENDATION.md (pre-award decision-support, names not keys)…")
        rec_path = write_recommendation_md(session, seeded, run_result.analysis_run_id, config)

        print(
            "[6/8] Simulating the human selecting Scenario B -> promote to award "
            "(real flow gates this through award -> FREEZE -> SIGN-OFF before any output, D22)…"
        )
        award = select_award_from_scenario(
            session, seeded, run_result.analysis_run_id, selected_scenario_code="B"
        )
        awarded_suppliers = {c.supplier_id for c in award.cells}
        print(
            f"   selected Scenario {award.scenario_code} ({award.scenario_label}) -> award of "
            f"{len(award.cells)} cells across {len(awarded_suppliers)} suppliers "
            "[freeze + sign-off gates noted, deferred to the awd.* phase]"
        )

        print(
            "[7/9] Generating BOOKING_GUIDE_INTERNAL.xlsx FROM THE AWARD "
            "(buyers/pricing; presentation-formatted — D24)…"
        )
        internal_path = write_booking_guide_internal_xlsx(
            seeded, award, output_path=OUTPUT_DIR / "BOOKING_GUIDE_INTERNAL.xlsx"
        )
        print(
            "[8/9] Generating SUPPLIER_AWARD_GUIDES.xlsx FROM THE AWARD "
            "(one sheet per awarded supplier; presentation-formatted — D24)…"
        )
        supplier_path = write_supplier_award_guides_xlsx(
            seeded, award, output_path=OUTPUT_DIR / "SUPPLIER_AWARD_GUIDES.xlsx"
        )
        print(
            "[9/9] Generating SCENARIO_WORKBOOK.xlsx — ALIGNMENT/COMPARISON tool (D26/D27): "
            "Scenario Comparison (lenses side by side + LIVE Custom column + EXPANDABLE "
            "scenario→DC→supplier drill) + interactive Custom Scenario (D25) + a flat "
            "Data (pivot me) Excel Table (D27)…"
        )
        workbook_path = write_scenario_workbook_xlsx(
            session, seeded, config, run_result.analysis_run_id, final_round.id, award
        )

    print("=== DONE ===")
    print(f"   {rec_path}  ({rec_path.stat().st_size} bytes)")
    print(f"   {internal_path}  ({internal_path.stat().st_size} bytes)")
    print(f"   {supplier_path}  ({supplier_path.stat().st_size} bytes)")
    print(f"   {workbook_path}  ({workbook_path.stat().st_size} bytes)")


if __name__ == "__main__":
    main()
