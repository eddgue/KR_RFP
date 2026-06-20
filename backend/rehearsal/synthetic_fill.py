"""Synthetic data filler for the HARNESS REHEARSAL (clearly synthetic — never real RFP data).

The orchestrator/secretary generate BLANK fill-out docs; this fills them with a small synthetic
scope so a rehearsal can run a full cycle + edge cases without hand-typing and without touching any
real commercial data. Commodity is "Test Greens" so nothing here can be mistaken for a live cycle.

Usage (run on the file the harness just generated, in the run's inputs/):
    python -m rehearsal.synthetic_fill setup  <inputs>/01_setup_kickoff.xlsx
    python -m rehearsal.synthetic_fill bids   <inputs>/0X_roundN_bid_template.xlsx  N
    python -m rehearsal.synthetic_fill messy  <a generated bid template>  <out.xlsx>

`setup` and `bids` fill IN PLACE (re-upload the same file); `messy` writes a supplier-format file
for the ingest_any edge case.
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.pilot.setup_template import EXAMPLE_START_ROW, HEADER_ROW

# --- the synthetic scope (fixed, obviously-fake) --------------------------------------------------
COMMODITY = "Test Greens"
DCS = [
    ("Atlanta Test DC", "East", "GA"),
    ("Dallas Test DC", "South", "TX"),
    ("Denver Test DC", "West", "CO"),
]
LOTS = [
    ("Lot 1 — Spring Mix", "Spring Mix 5oz", "5oz clamshell", "Conventional", "Greens"),
    ("Lot 2 — Romaine", "Romaine Hearts 3ct", "3ct sleeve", "Conventional", "Greens"),
    ("Lot 3 — Spinach", "Baby Spinach 1lb", "1lb clamshell", "Conventional", "Greens"),
]
SUPPLIERS = [("Alpha Farms", "East"), ("Beta Growers", "South"), ("Gamma Produce", "West"),
             ("Delta Fresh", "East")]
INCUMBENT = "Delta Fresh"  # the incumbent across all cells (so retention has something to chew on)
TF = ("Spring Test", "2026-04-01", "2026-06-30", 13)
ROUNDS = 3
WEEKLY = {0: 400, 1: 300, 2: 200}  # weekly cases per lot index
# base all-in price per (supplier index -> [price per lot index]); the incumbent (Delta=3) is HIGH.
BASE = {0: [12.0, 14.0, 11.0], 1: [12.5, 13.5, 11.5], 2: [13.0, 14.5, 12.0], 3: [22.0, 24.0, 20.0]}


def _hdrmap(ws: Worksheet, header_row: int) -> dict[str, int]:
    return {
        ws.cell(row=header_row, column=c).value: c
        for c in range(1, ws.max_column + 1)
        if ws.cell(row=header_row, column=c).value
    }


def _write_tab(ws: Worksheet, cols: dict[str, list[object]]) -> None:
    hdr = _hdrmap(ws, HEADER_ROW)
    n = max(len(v) for v in cols.values())
    for i in range(n):
        for label, vals in cols.items():
            value = vals[i] if i < len(vals) else None
            ws.cell(row=EXAMPLE_START_ROW + i, column=hdr[label], value=value)
    for r in range(EXAMPLE_START_ROW + n, EXAMPLE_START_ROW + n + 4):  # clear leftover example rows
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c, value=None)


def fill_setup(path: Path) -> None:
    wb = load_workbook(path)
    _write_tab(wb["Cycle"], {
        "Cycle Label": ["Test Greens — REHEARSAL"], "Commodity": [COMMODITY],
        "Sub-commodity": [COMMODITY], "Horizon (weeks)": [13], "Rounds": [ROUNDS],
        "Target Effective Date": ["2026-04-01"], "Weight Preset": ["balanced"],
        "Max Suppliers / DC": [2], "Premium Ceiling": [0.15], "Concentration Threshold": [0.40],
        "Coverage Floor": [0.80]})
    _write_tab(wb["DCs"], {"DC Name": [d[0] for d in DCS], "Region": [d[1] for d in DCS],
                           "State": [d[2] for d in DCS]})
    _write_tab(wb["Lots and Items"], {
        "Lot Name": [lt[0] for lt in LOTS], "Item Description": [lt[1] for lt in LOTS],
        "Pack Size / UOM": [lt[2] for lt in LOTS], "Product Type": [lt[3] for lt in LOTS],
        "Category": [lt[4] for lt in LOTS]})
    _write_tab(wb["Suppliers"], {
        "Supplier Name": [s[0] for s in SUPPLIERS], "Region / Origin": [s[1] for s in SUPPLIERS],
        "Notes": ["incumbent" if s[0] == INCUMBENT else "" for s in SUPPLIERS]})
    _write_tab(wb["Timeframes"], {"Timeframe Label": [TF[0]], "Start Date": [TF[1]],
                                  "End Date": [TF[2]], "Week Count": [TF[3]]})
    vols = [(d[0], lt[0], WEEKLY[i]) for d in DCS for i, lt in enumerate(LOTS)]
    _write_tab(wb["Volumes"], {
        "DC Name": [v[0] for v in vols], "Lot Name": [v[1] for v in vols],
        "Timeframe": [TF[0]] * len(vols), "Method": ["WEEKLY_X_WEEKS"] * len(vols),
        "Weekly Cases": [v[2] for v in vols], "Weeks": [TF[3]] * len(vols)})
    incs = [(d[0], lt[0], INCUMBENT, round(BASE[3][i] * 1.1, 2))
            for d in DCS for i, lt in enumerate(LOTS)]
    _write_tab(wb["Incumbents"], {
        "DC Name": [x[0] for x in incs], "Lot Name": [x[1] for x in incs],
        "Incumbent Supplier": [x[2] for x in incs],
        "Routing Baseline $/case": [x[3] for x in incs],
        "Contract Notes": ["auto-renew"] * len(incs)})
    wb.save(path)
    print(f"filled setup ({len(DCS)} DCs, {len(LOTS)} lots, {len(SUPPLIERS)} suppliers, "
          f"{ROUNDS} rounds) -> {path.name}")


def fill_bids(path: Path, round_no: int) -> None:
    wb = load_workbook(path)
    ws = wb["Bids"]
    col = _hdrmap(ws, 2)
    sup_idx = {s[0]: i for i, s in enumerate(SUPPLIERS)}
    lot_idx = {lt[0]: i for i, lt in enumerate(LOTS)}
    drift = 1.0 - 0.03 * (round_no - 1)  # bids drift down ~3%/round; incumbent holds firmer
    filled = 0
    for r in range(3, ws.max_row + 1):
        sup = str(ws.cell(r, col["Supplier"]).value or "").strip()
        lot = str(ws.cell(r, col["Lot"]).value or "").strip()
        if sup not in sup_idx or lot not in lot_idx:
            continue
        si, li = sup_idx[sup], lot_idx[lot]
        if sup == "Gamma Produce" and lot == "Lot 3 — Spinach":
            continue  # EDGE: Gamma declines this cell entirely (a NO-BID) — leave it blank
        firm = 1.0 if sup == INCUMBENT else drift
        price = round(BASE[si][li] * firm, 2)
        ws.cell(r, col["All-In $/case"], price)
        ws.cell(r, col["Weekly Vol Offered"], WEEKLY[li])
        ws.cell(r, col["Total Vol Offered"], WEEKLY[li] * TF[3])
        if sup in ("Beta Growers", "Delta Fresh"):  # exercise the component columns (optional)
            ws.cell(r, col["FOB $/case"], round(price * 0.85, 2))
            ws.cell(r, col["Delivery Surcharge"], round(price * 0.15, 2))
        if "Transit Days" in col:
            ws.cell(r, col["Transit Days"], 2 + (si + li) % 5)  # synthetic lane transit
        if sup == INCUMBENT:
            ws.cell(r, col["Pricing Comments"], "Firm — incumbent program pricing")
        filled += 1
    wb.save(path)
    print(f"filled round {round_no} bids: {filled} priced lines -> {path.name}")


def fill_messy(template_path: Path, out_path: Path) -> None:
    """A single supplier's OWN-format file (odd headers, shuffled cols) for the ingest_any case."""
    src = load_workbook(template_path)
    ws = src["Bids"]
    col = _hdrmap(ws, 2)
    rows = []
    for r in range(3, ws.max_row + 1):
        sup = str(ws.cell(r, col["Supplier"]).value or "").strip()
        if sup != "Alpha Farms":
            continue
        rows.append((str(ws.cell(r, col["Lot"]).value or "").strip(),
                     str(ws.cell(r, col["DC Name"]).value or "").strip(), sup))
    out = Workbook()
    o = out.active
    o.title = "Our Bid"
    o.cell(row=1, column=1, value="Alpha Farms — Test Greens bid (their own sheet)")
    for ci, h in enumerate(["Product", "Warehouse", "Vendor", "Cases/Week", "Delivered Price"], 1):
        o.cell(row=2, column=ci, value=h)
    lot_idx = {lt[0]: i for i, lt in enumerate(LOTS)}
    rr = 3
    for lot, dc, sup in rows:
        li = lot_idx.get(lot, 0)
        o.cell(rr, 1, lot)
        o.cell(rr, 2, dc)
        o.cell(rr, 3, sup)
        o.cell(rr, 4, WEEKLY[li])
        o.cell(rr, 5, round(BASE[0][li] * 0.95, 2))
        rr += 1
    out.save(out_path)
    print(f"wrote messy supplier file ({len(rows)} lines) -> {out_path}")


def main() -> None:
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)
    cmd = sys.argv[1]
    if cmd == "setup":
        fill_setup(Path(sys.argv[2]))
    elif cmd == "bids":
        fill_bids(Path(sys.argv[2]), int(sys.argv[3]))
    elif cmd == "messy":
        fill_messy(Path(sys.argv[2]), Path(sys.argv[3]))
    else:
        print(f"unknown command {cmd!r}")
        sys.exit(1)


if __name__ == "__main__":
    main()
